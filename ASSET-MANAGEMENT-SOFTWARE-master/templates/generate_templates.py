"""
OCP Maintenance AI MVP — Excel Template Generator
Generates 13 .xlsx data-loading templates with:
  - OCP green headers (#1B5E20 fill, white bold font)
  - Frozen panes, auto-sized columns
  - Data validation dropdowns for all enum columns
  - 3-5 realistic phosphate example rows
  - Instructions sheet per template
"""

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Import enum values from shared constants (single source of truth)
from tools.engines._template_constants import (  # noqa: E402
    ABC_CLASSES,
    BUDGETED_AS_VALUES,
    BUDGETED_LIFE_UNITS,
    BUDGET_TYPES,
    CAPTURE_TYPES,
    CAUSES,
    CRIT_CATEGORIES,
    CRIT_METHODS,
    CRITICALITIES,
    EQUIP_STATUSES,
    FAILURE_CONSEQUENCES,
    FAILURE_PATTERNS,
    FAILURE_TYPES,
    FM_COMBOS,
    FREQUENCY_UNITS,
    FSN_CLASSES,
    FUNCTION_TYPES,
    JUSTIFICATION_CATEGORIES,
    LABOUR_SPECIALTIES,
    LANGUAGES,
    MECHANISMS,
    NODE_TYPES,
    ORDER_TYPES,
    PRIORITIES,
    RCA_LEVELS,
    SHIFT_TYPES,
    SHUTDOWN_TYPES,
    STRATEGY_RULES,
    STRATEGY_STATUSES,
    STRATEGY_TYPES,
    TASK_CONSTRAINTS,
    TASK_TYPES,
    UNITS_OF_MEASURE,
    VED_CLASSES,
    WO_STATUSES,
    WP_CONSTRAINTS,
    WP_TYPES,
)

# ── Branding ──────────────────────────────────────────────────────
OCP_GREEN = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
INSTR_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1B5E20")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

OUTPUT_DIR = Path(__file__).parent

# Template versioning
TEMPLATE_VERSION = "1.0"
TEMPLATE_DATE = "2026-03-05"


# ── Helpers ───────────────────────────────────────────────────────
def _stamp_version(wb):
    """Stamp version and date metadata into workbook properties."""
    wb.properties.creator = "VSC AMS"
    wb.properties.description = f"Template v{TEMPLATE_VERSION} ({TEMPLATE_DATE})"
    wb.properties.lastModifiedBy = "VSC AMS"


def _save_workbook(wb, path):
    """Stamp version metadata and save workbook."""
    _stamp_version(wb)
    wb.save(path)


def _apply_header_style(ws, num_cols: int):
    """Apply OCP green header styling to row 1."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = OCP_GREEN
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_size_columns(ws, min_width=12, max_width=40):
    """Auto-size columns based on header length."""
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value or ""
        width = min(max(len(str(header_val)) + 4, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_data_validation(ws, col_idx: int, values: list, num_rows: int = 100):
    """Add dropdown data validation to a column.

    If the inline formula exceeds 255 chars (Excel limit), automatically
    creates a hidden 'Lookups' sheet and uses a cell-range reference instead.
    """
    inline_formula = ",".join(values)
    if len(inline_formula) + 2 <= 255:  # +2 for surrounding quotes
        formula = '"' + inline_formula + '"'
    else:
        # Use lookup sheet for long lists
        wb = ws.parent
        if "Lookups" not in wb.sheetnames:
            lk = wb.create_sheet("Lookups")
            lk.sheet_state = "hidden"
            lk._lookup_next_col = 1
        else:
            lk = wb["Lookups"]
        lk_col = getattr(lk, "_lookup_next_col", 1)
        for i, val in enumerate(values, start=1):
            lk.cell(row=i, column=lk_col, value=val)
        lk_letter = get_column_letter(lk_col)
        formula = f"=Lookups!${lk_letter}$1:${lk_letter}${len(values)}"
        lk._lookup_next_col = lk_col + 1

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a valid value from the dropdown."
    dv.errorTitle = "Invalid Value"
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}2:{col_letter}{num_rows}")
    ws.add_data_validation(dv)


def _write_data_rows(ws, rows: list[list]):
    """Write data rows starting at row 2."""
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def _add_instructions_sheet(wb, field_descriptions: list[tuple]):
    """Add an Instructions sheet with field descriptions."""
    ws = wb.create_sheet("Instructions")
    headers = ["Field Name", "Data Type", "Required", "Valid Values / Constraints", "Description"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = OCP_GREEN
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"

    for row_idx, (name, dtype, required, valid_values, desc) in enumerate(field_descriptions, start=2):
        ws.cell(row=row_idx, column=1, value=name).font = INSTR_HEADER_FONT
        ws.cell(row=row_idx, column=2, value=dtype).font = DATA_FONT
        ws.cell(row=row_idx, column=3, value=required).font = DATA_FONT
        ws.cell(row=row_idx, column=4, value=valid_values).font = DATA_FONT
        ws.cell(row=row_idx, column=5, value=desc).font = DATA_FONT
        for c in range(1, 6):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            ws.cell(row=row_idx, column=c).alignment = WRAP_ALIGN

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 50


# ══════════════════════════════════════════════════════════════════
# TEMPLATE GENERATORS
# ══════════════════════════════════════════════════════════════════

def generate_01_equipment_hierarchy():
    """01: Equipment Hierarchy — plant/area/system/equipment tree."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Hierarchy"

    headers = [
        "plant_id", "plant_name", "area_code", "area_name", "system_code",
        "system_name", "equipment_tag", "equipment_description", "equipment_type",
        "manufacturer", "model", "serial_number", "power_kw", "weight_kg",
        "criticality", "status", "sap_func_loc", "installation_date",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["OCP-JFC1", "Jorf Fertilizer Complex 1", "BRY", "Broyage (Grinding)", "SAG",
         "SAG Mill Circuit", "BRY-SAG-ML-001", "SAG Mill 36x20 Primary Grinding", "SAG_MILL",
         "FLSmidth", "SAG 36x20", "SN-SAG-2019-001", 8500, 285000,
         "AA", "ACTIVE", "JFC1-MIN-BRY-SAG-01", "2019-06-15"],
        ["OCP-JFC1", "Jorf Fertilizer Complex 1", "BRY", "Broyage (Grinding)", "BML",
         "Ball Mill Circuit", "BRY-BML-ML-001", "Ball Mill 22x36 Secondary Grinding", "BALL_MILL",
         "Metso", "BM 22x36", "SN-BML-2019-002", 5000, 180000,
         "AA", "ACTIVE", "JFC1-MIN-BRY-BML-01", "2019-08-20"],
        ["OCP-JFC1", "Jorf Fertilizer Complex 1", "FLT", "Flottation", "FCL",
         "Flotation Cell Bank", "FLT-FCL-CL-001", "Flotation Cell Bank A Unit 1", "FLOTATION_CELL",
         "Outotec", "TankCell e500", "SN-FCL-2020-001", 250, 45000,
         "A+", "ACTIVE", "JFC1-MIN-FLT-FCL-01", "2020-03-10"],
        ["OCP-JFC1", "Jorf Fertilizer Complex 1", "PMP", "Pompage (Pumping)", "SLP",
         "Slurry Pump Station", "PMP-SLP-PP-001", "Slurry Pump 16x14 Mill Discharge", "SLURRY_PUMP",
         "Weir Minerals", "Warman 16/14 AH", "SN-SLP-2019-005", 350, 8500,
         "A+", "ACTIVE", "JFC1-MIN-PMP-SLP-01", "2019-07-01"],
        ["OCP-JFC1", "Jorf Fertilizer Complex 1", "CVY", "Convoyage (Conveying)", "CVB",
         "Conveyor Belt System", "CVY-CVB-CV-001", "Overland Conveyor 2000mm x 3.5km", "BELT_CONVEYOR",
         "Continental", "PIPE2000", "SN-CVB-2020-010", 200, 95000,
         "A", "ACTIVE", "JFC1-MIN-CVY-CVB-01", "2020-01-15"],
    ]
    _write_data_rows(ws, rows)

    # Data validations
    _add_data_validation(ws, 15, CRITICALITIES)
    _add_data_validation(ws, 16, EQUIP_STATUSES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "SAP Plant code identifier"),
        ("plant_name", "Text", "Yes", "Free text", "Full plant name"),
        ("area_code", "Text", "Yes", "3-4 chars, e.g. BRY, FLT", "Area abbreviation code"),
        ("area_name", "Text", "Yes", "Free text", "Full area name (FR/EN)"),
        ("system_code", "Text", "Yes", "3-4 chars, e.g. SAG, BML", "System abbreviation code"),
        ("system_name", "Text", "Yes", "Free text", "Full system name"),
        ("equipment_tag", "Text", "Yes", "Format: AREA-SYS-TYPE-NNN", "Unique equipment tag identifier"),
        ("equipment_description", "Text", "Yes", "Max 80 chars", "Equipment description"),
        ("equipment_type", "Text", "Yes", "e.g. SAG_MILL, SLURRY_PUMP", "Equipment type category"),
        ("manufacturer", "Text", "No", "Free text", "Equipment manufacturer/OEM"),
        ("model", "Text", "No", "Free text", "Equipment model designation"),
        ("serial_number", "Text", "No", "Free text", "Equipment serial number"),
        ("power_kw", "Number", "No", ">=0", "Installed power in kilowatts"),
        ("weight_kg", "Number", "No", ">=0", "Equipment weight in kilograms"),
        ("criticality", "Text", "Yes", ", ".join(CRITICALITIES), "Equipment criticality class"),
        ("status", "Text", "Yes", ", ".join(EQUIP_STATUSES), "Equipment operational status"),
        ("sap_func_loc", "Text", "No", "SAP TPLNR format", "SAP Functional Location code"),
        ("installation_date", "Date", "No", "YYYY-MM-DD", "Installation date"),
    ])

    path = OUTPUT_DIR / "01_equipment_hierarchy.xlsx"
    _save_workbook(wb, path)
    return path


def generate_02_criticality_assessment():
    """02: Criticality Assessment — 11 consequence categories + probability."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Criticality Assessment"

    headers = ["equipment_tag", "method"] + [cat.lower() for cat in CRIT_CATEGORIES] + ["probability"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["BRY-SAG-ML-001", "FULL_MATRIX", 5, 4, 4, 5, 4, 3, 4, 5, 3, 4, 4, 4],
        ["BRY-BML-ML-001", "FULL_MATRIX", 4, 3, 4, 5, 4, 3, 4, 4, 3, 3, 3, 4],
        ["FLT-FCL-CL-001", "FULL_MATRIX", 3, 3, 3, 4, 3, 2, 3, 4, 2, 3, 3, 3],
        ["PMP-SLP-PP-001", "FULL_MATRIX", 4, 3, 3, 4, 3, 3, 3, 4, 2, 3, 3, 4],
        ["CVY-CVB-CV-001", "FULL_MATRIX", 3, 2, 3, 4, 3, 2, 3, 3, 2, 2, 3, 3],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 2, CRIT_METHODS)
    for col_idx in range(3, 14):
        _add_data_validation(ws, col_idx, ["1", "2", "3", "4", "5"])

    _auto_size_columns(ws)

    instr_fields = [
        ("equipment_tag", "Text", "Yes", "Must exist in hierarchy", "Equipment tag from Template 01"),
        ("method", "Text", "Yes", ", ".join(CRIT_METHODS), "Assessment method"),
    ]
    for cat in CRIT_CATEGORIES:
        instr_fields.append((cat.lower(), "Integer", "Yes", "1-5", f"{cat} consequence level (1=lowest, 5=highest)"))
    instr_fields.append(("probability", "Integer", "Yes", "1-5", "Failure probability (1=rare, 5=almost certain)"))

    _add_instructions_sheet(wb, instr_fields)

    path = OUTPUT_DIR / "02_criticality_assessment.xlsx"
    _save_workbook(wb, path)
    return path


def generate_03_failure_modes():
    """03: Failure Modes — FMEA data with 72-combo validation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Failure Modes"

    headers = [
        "equipment_tag", "function_description", "function_type", "failure_type",
        "what_component", "mechanism", "cause", "failure_pattern",
        "failure_consequence", "evidence", "downtime_hours",
        "detection_method", "rpn_severity", "rpn_occurrence", "rpn_detection",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["BRY-SAG-ML-001", "Grind ore to 80% passing 150 microns", "PRIMARY", "PARTIAL",
         "Mill Bearings", "WEARS", "BREAKDOWN_OF_LUBRICATION", "B_AGE",
         "EVIDENT_OPERATIONAL", "Vibration increase, temperature rise", 48,
         "Vibration analysis", 8, 4, 3],
        ["BRY-SAG-ML-001", "Grind ore to 80% passing 150 microns", "PRIMARY", "TOTAL",
         "Motor Stator", "SHORT_CIRCUITS", "BREAKDOWN_IN_INSULATION", "E_RANDOM",
         "EVIDENT_SAFETY", "Insulation resistance decline", 120,
         "Insulation testing", 9, 3, 4],
        ["BRY-SAG-ML-001", "Protect personnel from rotating hazards", "PROTECTIVE", "TOTAL",
         "Guard Interlock", "DEGRADES", "AGE", "B_AGE",
         "HIDDEN_SAFETY", "Interlock fails to trip", 4,
         "Functional test", 10, 2, 2],
        ["PMP-SLP-PP-001", "Transfer slurry at 400 m3/h", "PRIMARY", "PARTIAL",
         "Impeller", "SEVERS", "ABRASION", "C_FATIGUE",
         "EVIDENT_OPERATIONAL", "Reduced flow rate and efficiency", 8,
         "Flow monitoring", 6, 5, 3],
        ["PMP-SLP-PP-001", "Prevent slurry leakage to environment", "SECONDARY", "TOTAL",
         "Mechanical Seal", "DEGRADES", "CHEMICAL_ATTACK", "B_AGE",
         "EVIDENT_ENVIRONMENTAL", "Visible seal leakage", 6,
         "Visual inspection", 7, 4, 2],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 3, FUNCTION_TYPES)
    _add_data_validation(ws, 4, FAILURE_TYPES)
    _add_data_validation(ws, 6, MECHANISMS)
    _add_data_validation(ws, 7, CAUSES)
    _add_data_validation(ws, 8, FAILURE_PATTERNS)
    _add_data_validation(ws, 9, FAILURE_CONSEQUENCES)

    _auto_size_columns(ws)

    # Add 72-combo reference sheet
    ws_combo = wb.create_sheet("Valid FM Combinations")
    ws_combo.cell(row=1, column=1, value="Mechanism").fill = OCP_GREEN
    ws_combo.cell(row=1, column=1).font = HEADER_FONT
    ws_combo.cell(row=1, column=2, value="Cause").fill = OCP_GREEN
    ws_combo.cell(row=1, column=2).font = HEADER_FONT
    ws_combo.freeze_panes = "A2"

    for row_idx, (mech, cause) in enumerate(FM_COMBOS, start=2):
        ws_combo.cell(row=row_idx, column=1, value=mech).font = DATA_FONT
        ws_combo.cell(row=row_idx, column=2, value=cause).font = DATA_FONT
    ws_combo.column_dimensions["A"].width = 30
    ws_combo.column_dimensions["B"].width = 35

    _add_instructions_sheet(wb, [
        ("equipment_tag", "Text", "Yes", "Must exist in hierarchy", "Equipment tag from Template 01"),
        ("function_description", "Text", "Yes", "Max 200 chars", "Equipment function being analyzed"),
        ("function_type", "Text", "Yes", ", ".join(FUNCTION_TYPES), "Function classification"),
        ("failure_type", "Text", "Yes", ", ".join(FAILURE_TYPES), "Total or partial failure"),
        ("what_component", "Text", "Yes", "Free text", "Component that fails (the 'what')"),
        ("mechanism", "Text", "Yes", "18 valid values (see FM Combinations sheet)", "How it fails"),
        ("cause", "Text", "Yes", "44 valid values (see FM Combinations sheet)", "Root cause of mechanism"),
        ("failure_pattern", "Text", "Yes", ", ".join(FAILURE_PATTERNS), "Nowlan-Heap failure pattern (A-F)"),
        ("failure_consequence", "Text", "Yes", ", ".join(FAILURE_CONSEQUENCES), "RCM consequence classification"),
        ("evidence", "Text", "No", "Free text", "Observable evidence of failure progression"),
        ("downtime_hours", "Number", "No", ">=0", "Expected downtime if failure occurs (hours)"),
        ("detection_method", "Text", "No", "Free text", "How the failure mode is detected"),
        ("rpn_severity", "Integer", "No", "1-10", "FMECA: Severity rating (1=no effect, 10=catastrophic)"),
        ("rpn_occurrence", "Integer", "No", "1-10", "FMECA: Occurrence rating (1=rare, 10=almost certain)"),
        ("rpn_detection", "Integer", "No", "1-10", "FMECA: Detection rating (1=always detected, 10=undetectable)"),
    ])

    path = OUTPUT_DIR / "03_failure_modes.xlsx"
    _save_workbook(wb, path)
    return path


def generate_04_maintenance_tasks():
    """04: Maintenance Tasks — task catalog (NO frequency). Multi-sheet: Tasks + Labour + Materials + Tools.

    Tasks are standalone action definitions. Frequency, acceptable limits,
    and conditional comments belong to the maintenance STRATEGY (Template 14).
    A task can be reused across multiple strategies.
    """
    wb = Workbook()

    # ── Sheet 1: Tasks (one row per task — no frequency) ──
    ws = wb.active
    ws.title = "Tasks"

    task_headers = [
        "task_id", "task_name", "task_name_fr", "task_type",
        "constraint", "access_time_hours",
        "budgeted_as", "budgeted_life", "budgeted_life_time_units",
        "budgeted_life_operational_units",
        "consequences", "justification", "origin", "notes",
    ]
    for col_idx, h in enumerate(task_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(task_headers))

    task_rows = [
        # Primary tasks (proactive — used with frequency in strategy template)
        ["T-001", "Inspect SAG Mill Bearings for Wear",
         "Inspecter les roulements du broyeur SAG pour usure", "INSPECT",
         "ONLINE", 0, None, None, None, None,
         "Undetected bearing degradation leading to catastrophic failure",
         "RCM: CBM for AA-class rotating equipment", "R8_LIBRARY", None],
        ["T-002", "Lubricate SAG Mill Main Bearings",
         "Lubrifier les paliers principaux du broyeur SAG", "LUBRICATE",
         "ONLINE", 0, None, None, None, None,
         "Bearing seizure from lubrication breakdown",
         "OEM requirement + RCM analysis", "R8_LIBRARY", None],
        ["T-003", "Perform Functional Test of Guard Interlock",
         "Effectuer test fonctionnel de l'interverrouillage du carter", "TEST",
         "TEST_MODE", 0.5, None, None, None, None,
         "Hidden failure of safety interlock undetected",
         "RCM: FFI for hidden safety failure", "R8_LIBRARY", None],
        ["T-004", "Inspect Slurry Pump Impeller Wear",
         "Inspecter l'usure de la roue de la pompe a boue", "INSPECT",
         "ONLINE", 0, None, None, None, None,
         "Reduced pump efficiency and flow rate",
         "RCM: CBM for high-wear wet-end components", "R8_LIBRARY", None],
        ["T-005", "Clean Drive Motor Fins and Cooling Fan",
         "Nettoyer les ailettes du moteur et le ventilateur", "CLEAN",
         "OFFLINE", 0.1, None, None, None, None,
         "Motor overheating from restricted airflow",
         "OEM cooling system maintenance", "R8_LIBRARY", None],
        # Secondary tasks (corrective — triggered by condition, no frequency)
        ["T-006", "Replace SAG Mill Trunnion Bearing",
         "Remplacer le palier tourillon du broyeur SAG", "REPLACE",
         "OFFLINE", 96, "REPLACE", 5, "YEARS", "OPERATING_HOURS",
         "Catastrophic bearing failure causing extended shutdown",
         "Secondary task: triggered when T-001 finds limits exceeded",
         "RCM_ANALYSIS", None],
        ["T-007", "Replace SAG Mill Liner Set",
         "Remplacer le jeu de blindages du broyeur SAG", "REPLACE",
         "OFFLINE", 72, "REPLACE", None, None, "OPERATING_HOURS",
         "Loss of grinding efficiency, liner failure risk",
         "Scheduled replacement based on wear measurement",
         "R8_LIBRARY", None],
        ["T-008", "Repair Slurry Pump Impeller",
         "Reparer la roue de la pompe a boue", "REPAIR",
         "OFFLINE", 8, "REPAIR", None, None, None,
         "Flow reduction below design capacity",
         "Secondary task: triggered when T-004 finds wear beyond limits",
         "RCM_ANALYSIS", None],
        ["T-009", "Replace Slurry Pump Mechanical Seal",
         "Remplacer le joint mecanique de la pompe a boue", "REPLACE",
         "OFFLINE", 4, "REPLACE", 2, "YEARS", None,
         "Environmental contamination from seal failure",
         "RTF secondary: replace on failure",
         "HISTORICAL_MTBF", None],
        ["T-010", "Replace Guard Interlock Mechanism",
         "Remplacer le mecanisme d'interverrouillage du carter", "REPLACE",
         "OFFLINE", 2, "REPLACE", None, None, None,
         "Safety interlock inoperable",
         "Secondary task: triggered when T-003 test fails",
         "RCM_ANALYSIS", None],
    ]
    _write_data_rows(ws, task_rows)

    _add_data_validation(ws, 4, TASK_TYPES)
    _add_data_validation(ws, 5, TASK_CONSTRAINTS)
    _add_data_validation(ws, 7, BUDGETED_AS_VALUES)
    _add_data_validation(ws, 9, BUDGETED_LIFE_UNITS)
    _add_data_validation(ws, 10, FREQUENCY_UNITS)

    _auto_size_columns(ws)

    # ── Sheet 2: Task_Labour (multiple rows per task_id) ──
    ws_labour = wb.create_sheet("Task_Labour")
    labour_headers = [
        "labour_id", "task_id", "worker_id", "specialty", "quantity",
        "hours_per_person", "hourly_rate_usd", "company",
    ]
    for col_idx, h in enumerate(labour_headers, start=1):
        ws_labour.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_labour, len(labour_headers))

    labour_rows = [
        # T-001: 1 fitter + 1 conmon specialist (vibration)
        ["LR-001", "T-001", "TEC-001", "FITTER", 1, 0.5, None, "OCP"],
        ["LR-002", "T-001", "TEC-009", "CONMON_SPECIALIST", 1, 0.5, None, "OCP"],
        # T-002: 1 lubricator
        ["LR-003", "T-002", "TEC-005", "LUBRICATOR", 1, 0.5, None, "OCP"],
        # T-003: 1 instrumentist + 1 electrician (functional test)
        ["LR-004", "T-003", "TEC-003", "INSTRUMENTIST", 1, 0.5, None, "OCP"],
        ["LR-005", "T-003", "TEC-002", "ELECTRICIAN", 1, 0.5, None, "OCP"],
        # T-004: 1 fitter
        ["LR-006", "T-004", "TEC-001", "FITTER", 1, 0.5, None, "OCP"],
        # T-005: 1 fitter (cleaning)
        ["LR-007", "T-005", "TEC-010", "FITTER", 1, 0.1, None, "OCP"],
        # T-006: 4 fitters + 2 operators + 1 electrician (major replacement)
        ["LR-008", "T-006", None, "FITTER", 4, 96.0, None, "OCP"],
        ["LR-009", "T-006", None, "OPERATOR", 2, 96.0, None, "OCP"],
        ["LR-010", "T-006", "TEC-002", "ELECTRICIAN", 1, 8.0, None, "OCP"],
        # T-007: 4 fitters + 2 operators + 1 electrician (liner change)
        ["LR-011", "T-007", None, "FITTER", 4, 72.0, None, "OCP"],
        ["LR-012", "T-007", None, "OPERATOR", 2, 72.0, None, "OCP"],
        ["LR-013", "T-007", "TEC-008", "ELECTRICIAN", 1, 8.0, None, "OCP"],
        # T-008: 2 fitters (impeller repair)
        ["LR-014", "T-008", None, "FITTER", 2, 8.0, None, "Contractor-A"],
        # T-009: 2 fitters + 1 electrician
        ["LR-015", "T-009", None, "FITTER", 2, 4.0, None, "Contractor-A"],
        ["LR-016", "T-009", "TEC-008", "ELECTRICIAN", 1, 1.0, None, "OCP"],
    ]
    _write_data_rows(ws_labour, labour_rows)
    _add_data_validation(ws_labour, 4, LABOUR_SPECIALTIES)
    _auto_size_columns(ws_labour)

    # ── Sheet 3: Task_Materials (multiple rows per task_id) ──
    ws_materials = wb.create_sheet("Task_Materials")
    mat_headers = [
        "material_line_id", "task_id", "material_code", "description",
        "manufacturer", "part_number", "quantity", "unit_of_measure",
        "unit_price_usd", "equipment_bom_ref",
    ]
    for col_idx, h in enumerate(mat_headers, start=1):
        ws_materials.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_materials, len(mat_headers))

    mat_rows = [
        ["MR-001", "T-002", "LUB-GRS-001", "SKF LGMT 2 Bearing Grease",
         "SKF", "LGMT 2/1", 2, "KG", 45.0, "BOM-SAG-001"],
        ["MR-002", "T-006", "BRG-SKF-22330", "SKF 22330 CC/W33 Spherical Roller Bearing",
         "SKF", "22330 CC/W33", 1, "EA", 8500.0, "BOM-SAG-001"],
        ["MR-003", "T-007", "LNR-SAG-001", "SAG Mill Liner Plate Hi-Chrome",
         "FLSmidth", "FLS-LNR-36", 48, "EA", 2800.0, "BOM-SAG-001"],
        ["MR-004", "T-007", "LNR-SAG-002", "SAG Mill Lifter Bar",
         "FLSmidth", "FLS-LFT-36", 24, "EA", 3500.0, "BOM-SAG-001"],
        ["MR-005", "T-007", "BLT-SAG-001", "Liner Bolt M48x200 Grade 10.9",
         "Hilti", "BLT-M48-200", 192, "EA", 35.0, "BOM-SAG-001"],
        ["MR-006", "T-009", "SEL-MEC-001", "Mechanical Seal Assembly 100mm",
         "Weir Minerals", "WM-SEAL-100", 1, "EA", 4500.0, "BOM-SLP-001"],
        ["MR-007", "T-009", "GAS-VIT-001", "Viton O-Ring Seal Kit",
         "Parker", "VIT-ORING-100", 1, "EA", 120.0, "BOM-SLP-001"],
    ]
    _write_data_rows(ws_materials, mat_rows)
    _add_data_validation(ws_materials, 8, UNITS_OF_MEASURE)
    _auto_size_columns(ws_materials)

    # ── Sheet 4: Task_Tools (multiple rows per task_id) ──
    ws_tools = wb.create_sheet("Task_Tools")
    tools_headers = [
        "tool_line_id", "task_id", "item_type", "tool_code", "description",
    ]
    for col_idx, h in enumerate(tools_headers, start=1):
        ws_tools.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_tools, len(tools_headers))

    tool_rows = [
        ["TL-001", "T-001", "TOOL", "TL-VIB-001", "Vibration analyzer (SKF CMAS 100-SL)"],
        ["TL-002", "T-001", "TOOL", "TL-IR-001", "Infrared thermometer"],
        ["TL-003", "T-002", "TOOL", "TL-GRG-001", "Grease gun (pneumatic)"],
        ["TL-004", "T-003", "TOOL", "TL-FTK-001", "Interlock function test kit"],
        ["TL-005", "T-006", "SPECIAL_EQUIPMENT", "SE-CRN-050", "Overhead crane 50T"],
        ["TL-006", "T-007", "SPECIAL_EQUIPMENT", "SE-CRN-050", "Overhead crane 50T"],
        ["TL-007", "T-007", "SPECIAL_EQUIPMENT", "SE-LNR-001", "Liner handler attachment"],
        ["TL-008", "T-007", "TOOL", "TL-TRQ-001", "Torque wrench 200-800 Nm"],
        ["TL-009", "T-007", "TOOL", "TL-HBT-001", "Hydraulic bolt tensioner M48"],
        ["TL-010", "T-009", "TOOL", "TL-SLK-001", "Seal installation tool kit"],
        ["TL-011", "T-009", "SPECIAL_EQUIPMENT", "SE-HST-002", "Portable hoist 2T"],
    ]
    _write_data_rows(ws_tools, tool_rows)
    _add_data_validation(ws_tools, 3, ["TOOL", "SPECIAL_EQUIPMENT"])
    _auto_size_columns(ws_tools)

    _add_instructions_sheet(wb, [
        ("== TASKS SHEET ==", "", "", "", "Task catalog — one row per maintenance task. Tasks do NOT have frequency (frequency belongs to the maintenance strategy in Template 14)."),
        ("task_id", "Text", "Yes", "Unique, e.g. T-001", "Unique task identifier. Reusable across multiple strategies."),
        ("task_name", "Text", "Yes", "Max 72 chars", "Task name (EN). Format: 'Inspect [MI] for [evidence]', 'Replace [MI]', etc."),
        ("task_name_fr", "Text", "Yes", "Max 72 chars", "Task name in French"),
        ("task_type", "Text", "Yes", ", ".join(TASK_TYPES), "Type of maintenance action"),
        ("constraint", "Text", "Yes", ", ".join(TASK_CONSTRAINTS), "Default execution constraint"),
        ("access_time_hours", "Number", "Yes", ">=0 (0 for ONLINE)", "Equipment access/preparation time in hours"),
        ("budgeted_as", "Text", "No", ", ".join(BUDGETED_AS_VALUES), "For corrective/replacement tasks: REPAIR (sub-component) or REPLACE (entire MI). NOT_BUDGETED for primary inspection tasks."),
        ("budgeted_life", "Number", "No", ">0", "Estimated useful life of the MI (for budgeted tasks)"),
        ("budgeted_life_time_units", "Text", "No", ", ".join(BUDGETED_LIFE_UNITS), "Time units for budgeted life"),
        ("budgeted_life_operational_units", "Text", "No", ", ".join(FREQUENCY_UNITS), "Operational units for budgeted life"),
        ("consequences", "Text", "Yes", "Free text", "What happens if the task is not performed"),
        ("justification", "Text", "No", "Free text", "RCM/engineering justification for the task"),
        ("origin", "Text", "No", "Free text", "Source: OEM, R8_LIBRARY, RCM_ANALYSIS, HISTORICAL_MTBF, WORKSHOP"),
        ("notes", "Text", "No", "Free text", "Additional notes"),
        ("", "", "", "", ""),
        ("== TASK_LABOUR SHEET ==", "", "", "", "One row per specialty per task (linked by task_id)"),
        ("labour_id", "Text", "Yes", "Unique, e.g. LR-001", "Unique labour resource line ID"),
        ("task_id", "Text", "Yes", "Must match Tasks sheet", "Foreign key to Tasks sheet"),
        ("worker_id", "Text", "No", "From Template 09", "Specific worker ID (optional, for named assignments)"),
        ("specialty", "Text", "Yes", ", ".join(LABOUR_SPECIALTIES), "Labour specialty"),
        ("quantity", "Integer", "Yes", ">=1", "Number of workers of this specialty"),
        ("hours_per_person", "Number", "Yes", ">0", "Hours per person for this task"),
        ("hourly_rate_usd", "Number", "No", ">=0", "Hourly labour rate (optional)"),
        ("company", "Text", "No", "Free text", "Company/contractor name (OCP or contractor)"),
        ("", "", "", "", ""),
        ("== TASK_MATERIALS SHEET ==", "", "", "", "One row per material per task (linked by task_id)"),
        ("material_line_id", "Text", "Yes", "Unique, e.g. MR-001", "Unique material resource line ID"),
        ("task_id", "Text", "Yes", "Must match Tasks sheet", "Foreign key to Tasks sheet"),
        ("material_code", "Text", "Yes", "SAP material code", "Links to Template 07 and SAP material master"),
        ("description", "Text", "Yes", "Free text", "Material description"),
        ("manufacturer", "Text", "No", "Free text", "Material manufacturer"),
        ("part_number", "Text", "No", "Manufacturer part #", "Manufacturer part number"),
        ("quantity", "Number", "Yes", ">0", "Quantity needed"),
        ("unit_of_measure", "Text", "Yes", ", ".join(UNITS_OF_MEASURE), "Unit of measure"),
        ("unit_price_usd", "Number", "No", ">=0", "Unit price in USD"),
        ("equipment_bom_ref", "Text", "No", "BOM reference", "SAP Equipment BOM reference"),
        ("", "", "", "", ""),
        ("== TASK_TOOLS SHEET ==", "", "", "", "One row per tool/equipment per task (linked by task_id)"),
        ("tool_line_id", "Text", "Yes", "Unique, e.g. TL-001", "Unique tool/equipment line ID"),
        ("task_id", "Text", "Yes", "Must match Tasks sheet", "Foreign key to Tasks sheet"),
        ("item_type", "Text", "Yes", "TOOL, SPECIAL_EQUIPMENT", "Tool or special equipment"),
        ("tool_code", "Text", "No", "Unique code", "Tool/equipment catalog code"),
        ("description", "Text", "Yes", "Free text", "Tool/equipment description"),
    ])

    path = OUTPUT_DIR / "04_maintenance_tasks.xlsx"
    _save_workbook(wb, path)
    return path


def generate_05_work_packages():
    """05: Work Packages — SAP-ready maintenance packages."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Packages"

    headers = [
        "wp_name", "wp_code", "equipment_tag", "frequency_value", "frequency_unit",
        "constraint", "wp_type", "access_time_hours", "task_ids_csv",
        "estimated_total_hours", "crew_size",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["4W BRY-SAG FITTER INSPECT ONLINE", "WP-SAG-4W-FIT-ONL", "BRY-SAG-ML-001",
         4, "WEEKS", "ONLINE", "STANDALONE", 0, "T-001,T-002",
         1.5, 2],
        ["8KH BRY-SAG LINER REPLACE OFFLINE", "WP-SAG-8KH-LNR-OFF", "BRY-SAG-ML-001",
         8000, "HOURS_RUN", "OFFLINE", "STANDALONE", 72, "T-003",
         576, 7],
        ["2W PMP-SLP FITTER INSPECT ONLINE", "WP-SLP-2W-FIT-ONL", "PMP-SLP-PP-001",
         2, "WEEKS", "ONLINE", "STANDALONE", 0, "T-004",
         0.5, 1],
        ["6M PMP-SLP SEAL REPLACE OFFLINE", "WP-SLP-6M-SEL-OFF", "PMP-SLP-PP-001",
         6, "MONTHS", "OFFLINE", "STANDALONE", 4, "T-005",
         13, 3],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 5, FREQUENCY_UNITS)
    _add_data_validation(ws, 6, WP_CONSTRAINTS)
    _add_data_validation(ws, 7, WP_TYPES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("wp_name", "Text", "Yes", "Max 40 chars, ALL CAPS", "Work package name (SAP format)"),
        ("wp_code", "Text", "Yes", "Unique code", "Work package identifier"),
        ("equipment_tag", "Text", "Yes", "Must exist in hierarchy", "Primary equipment tag"),
        ("frequency_value", "Number", "Yes", ">0", "Package execution frequency"),
        ("frequency_unit", "Text", "Yes", ", ".join(FREQUENCY_UNITS), "Frequency unit"),
        ("constraint", "Text", "Yes", ", ".join(WP_CONSTRAINTS), "Online or offline execution"),
        ("wp_type", "Text", "Yes", ", ".join(WP_TYPES), "Package type"),
        ("access_time_hours", "Number", "Yes", ">=0", "Equipment access time (0 for ONLINE)"),
        ("task_ids_csv", "Text", "Yes", "Comma-separated task IDs", "Task IDs from Template 04"),
        ("estimated_total_hours", "Number", "No", ">0", "Total estimated labour hours"),
        ("crew_size", "Integer", "No", ">=1", "Total crew size for the package"),
    ])

    path = OUTPUT_DIR / "05_work_packages.xlsx"
    _save_workbook(wb, path)
    return path


def generate_06_work_order_history():
    """06: Work Order History — 24 months of realistic WO data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Order History"

    headers = [
        "wo_id", "order_type", "equipment_tag", "sap_func_loc", "priority",
        "description", "created_date", "planned_start", "planned_end",
        "actual_start", "actual_end", "duration_hours", "status",
        "cost_labour_usd", "cost_materials_usd", "failure_found",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["WO-2025-00142", "PM02", "BRY-SAG-ML-001", "JFC1-MIN-BRY-SAG-01", "4_PLANNED",
         "4W SAG Mill bearing inspection", str(today - timedelta(days=120)),
         str(today - timedelta(days=118)), str(today - timedelta(days=118)),
         str(today - timedelta(days=118)), str(today - timedelta(days=118)),
         1.5, "COMPLETED", 150, 0, "No"],
        ["WO-2025-00287", "PM03", "PMP-SLP-PP-001", "JFC1-MIN-PMP-SLP-01", "2_URGENT",
         "Slurry pump seal replacement — active leak", str(today - timedelta(days=90)),
         str(today - timedelta(days=89)), str(today - timedelta(days=89)),
         str(today - timedelta(days=89)), str(today - timedelta(days=89)),
         6, "COMPLETED", 800, 4620, "Yes — mechanical seal failure"],
        ["WO-2025-00334", "PM02", "BRY-SAG-ML-001", "JFC1-MIN-BRY-SAG-01", "4_PLANNED",
         "SAG Mill lubrication — main bearings", str(today - timedelta(days=60)),
         str(today - timedelta(days=58)), str(today - timedelta(days=58)),
         str(today - timedelta(days=58)), str(today - timedelta(days=58)),
         0.5, "COMPLETED", 50, 90, "No"],
        ["WO-2025-00501", "PM01", "CVY-CVB-CV-001", "JFC1-MIN-CVY-CVB-01", "3_NORMAL",
         "Belt conveyor alignment inspection", str(today - timedelta(days=30)),
         str(today - timedelta(days=28)), str(today - timedelta(days=28)),
         str(today - timedelta(days=28)), str(today - timedelta(days=28)),
         2, "COMPLETED", 200, 0, "Belt tracking 15mm off-center"],
        ["WO-2025-00612", "PM02", "FLT-FCL-CL-001", "JFC1-MIN-FLT-FCL-01", "4_PLANNED",
         "Flotation cell agitator inspection", str(today - timedelta(days=7)),
         str(today - timedelta(days=5)), str(today - timedelta(days=5)),
         None, None, None, "RELEASED", None, None, None],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 2, ORDER_TYPES)
    _add_data_validation(ws, 5, PRIORITIES)
    _add_data_validation(ws, 13, WO_STATUSES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("wo_id", "Text", "Yes", "Unique WO number", "Work order identifier"),
        ("order_type", "Text", "Yes", ", ".join(ORDER_TYPES), "SAP order type"),
        ("equipment_tag", "Text", "Yes", "Must exist in hierarchy", "Equipment tag"),
        ("sap_func_loc", "Text", "No", "SAP TPLNR", "SAP functional location"),
        ("priority", "Text", "Yes", ", ".join(PRIORITIES), "Work order priority"),
        ("description", "Text", "Yes", "Free text", "Work order description"),
        ("created_date", "Date", "Yes", "YYYY-MM-DD", "Date WO was created"),
        ("planned_start", "Date", "No", "YYYY-MM-DD", "Planned start date"),
        ("planned_end", "Date", "No", "YYYY-MM-DD", "Planned end date"),
        ("actual_start", "Date", "No", "YYYY-MM-DD", "Actual start date"),
        ("actual_end", "Date", "No", "YYYY-MM-DD", "Actual end date"),
        ("duration_hours", "Number", "No", ">0", "Actual duration in hours"),
        ("status", "Text", "Yes", ", ".join(WO_STATUSES), "Work order status"),
        ("cost_labour_usd", "Number", "No", ">=0", "Labour cost in USD"),
        ("cost_materials_usd", "Number", "No", ">=0", "Materials cost in USD"),
        ("failure_found", "Text", "No", "Free text", "Failure description if found during PM"),
    ])

    path = OUTPUT_DIR / "06_work_order_history.xlsx"
    _save_workbook(wb, path)
    return path


def generate_07_spare_parts_inventory():
    """07: Spare Parts Inventory — VED/FSN/ABC classifications."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Spare Parts Inventory"

    headers = [
        "material_code", "description", "manufacturer", "part_number",
        "ved_class", "fsn_class", "abc_class",
        "quantity_on_hand", "min_stock", "max_stock", "reorder_point",
        "lead_time_days", "unit_cost_usd", "unit_of_measure",
        "applicable_equipment_csv", "warehouse_location",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["BRG-SKF-22330", "SKF 22330 CC/W33 Spherical Roller Bearing", "SKF", "22330 CC/W33",
         "VITAL", "SLOW_MOVING", "A_HIGH",
         2, 1, 4, 2, 90, 8500.0, "EA",
         "BRY-SAG-ML-001,BRY-BML-ML-001", "WH-01-A3-R2"],
        ["SEL-MEC-100", "Mechanical Seal Assembly 100mm Warman", "Weir Minerals", "WM-SEAL-100",
         "VITAL", "FAST_MOVING", "A_HIGH",
         4, 2, 8, 3, 60, 4500.0, "EA",
         "PMP-SLP-PP-001", "WH-01-B1-R5"],
        ["LNR-SAG-HC36", "SAG Mill Liner Plate Hi-Chrome 36ft", "FLSmidth", "FLS-LNR-36",
         "ESSENTIAL", "SLOW_MOVING", "A_HIGH",
         48, 24, 96, 48, 120, 2800.0, "EA",
         "BRY-SAG-ML-001", "WH-02-YARD-01"],
        ["LUB-GRS-LGMT2", "SKF LGMT 2 General Purpose Grease 1kg", "SKF", "LGMT 2/1",
         "DESIRABLE", "FAST_MOVING", "C_LOW",
         20, 10, 50, 15, 14, 45.0, "KG",
         "BRY-SAG-ML-001,BRY-BML-ML-001,PMP-SLP-PP-001", "WH-01-C2-R1"],
        ["IMP-WAR-16", "Warman 16/14 AH Impeller A05", "Weir Minerals", "WM-IMP-16-A05",
         "VITAL", "SLOW_MOVING", "A_HIGH",
         1, 1, 3, 1, 75, 12000.0, "EA",
         "PMP-SLP-PP-001", "WH-02-YARD-02"],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 5, VED_CLASSES)
    _add_data_validation(ws, 6, FSN_CLASSES)
    _add_data_validation(ws, 7, ABC_CLASSES)
    _add_data_validation(ws, 14, UNITS_OF_MEASURE)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("material_code", "Text", "Yes", "Unique code", "SAP material code"),
        ("description", "Text", "Yes", "Max 80 chars", "Material description"),
        ("manufacturer", "Text", "No", "Free text", "Part manufacturer"),
        ("part_number", "Text", "No", "Free text", "Manufacturer part number"),
        ("ved_class", "Text", "Yes", ", ".join(VED_CLASSES), "VED analysis: Vital/Essential/Desirable"),
        ("fsn_class", "Text", "Yes", ", ".join(FSN_CLASSES), "FSN analysis: Fast/Slow/Non-moving"),
        ("abc_class", "Text", "Yes", ", ".join(ABC_CLASSES), "ABC analysis: cost-based A/B/C"),
        ("quantity_on_hand", "Integer", "Yes", ">=0", "Current stock quantity"),
        ("min_stock", "Integer", "Yes", ">=0", "Minimum stock level"),
        ("max_stock", "Integer", "Yes", ">=min_stock", "Maximum stock level"),
        ("reorder_point", "Integer", "Yes", ">=min_stock", "Reorder trigger point"),
        ("lead_time_days", "Integer", "Yes", ">=0", "Procurement lead time in days"),
        ("unit_cost_usd", "Number", "Yes", ">=0", "Unit cost in USD"),
        ("unit_of_measure", "Text", "Yes", ", ".join(UNITS_OF_MEASURE), "Unit of measure"),
        ("applicable_equipment_csv", "Text", "No", "Comma-separated tags", "Equipment tags this part applies to"),
        ("warehouse_location", "Text", "No", "WH-xx-xx-xx format", "Warehouse bin location"),
    ])

    path = OUTPUT_DIR / "07_spare_parts_inventory.xlsx"
    _save_workbook(wb, path)
    return path


def generate_08_shutdown_calendar():
    """08: Shutdown Calendar — planned shutdown events."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shutdown Calendar"

    headers = [
        "plant_id", "shutdown_name", "shutdown_type",
        "planned_start", "planned_end", "planned_hours",
        "areas_csv", "description", "work_orders_csv",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["OCP-JFC1", "BRY Minor Shutdown Q1", "MINOR_8H",
         str(today + timedelta(days=30)), str(today + timedelta(days=30)), 8,
         "BRY", "Minor shutdown for bearing inspections and greasing", ""],
        ["OCP-JFC1", "BRY Major Shutdown Annual", "MAJOR_20H_PLUS",
         str(today + timedelta(days=90)), str(today + timedelta(days=93)), 72,
         "BRY,PMP", "Annual major shutdown: liner change, pump overhaul", ""],
        ["OCP-JFC1", "FLT Minor Shutdown Q2", "MINOR_8H",
         str(today + timedelta(days=60)), str(today + timedelta(days=60)), 8,
         "FLT", "Flotation cell inspection and agitator service", ""],
        ["OCP-JFC1", "CVY Minor Shutdown Q1", "MINOR_8H",
         str(today + timedelta(days=45)), str(today + timedelta(days=45)), 8,
         "CVY", "Belt splice inspection and idler replacement", ""],
        ["OCP-JFC1", "Plant-Wide Major Shutdown", "MAJOR_20H_PLUS",
         str(today + timedelta(days=180)), str(today + timedelta(days=187)), 168,
         "BRY,FLT,PMP,CVY,SED,FIL", "Annual plant-wide turnaround", ""],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 3, SHUTDOWN_TYPES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "Plant identifier"),
        ("shutdown_name", "Text", "Yes", "Free text", "Shutdown event name"),
        ("shutdown_type", "Text", "Yes", ", ".join(SHUTDOWN_TYPES), "Shutdown type"),
        ("planned_start", "Date", "Yes", "YYYY-MM-DD", "Planned start date"),
        ("planned_end", "Date", "Yes", "YYYY-MM-DD", "Planned end date"),
        ("planned_hours", "Number", "Yes", ">0", "Total planned shutdown hours"),
        ("areas_csv", "Text", "Yes", "Comma-separated area codes", "Affected areas"),
        ("description", "Text", "No", "Free text", "Shutdown scope description"),
        ("work_orders_csv", "Text", "No", "Comma-separated WO IDs", "Associated work orders"),
    ])

    path = OUTPUT_DIR / "08_shutdown_calendar.xlsx"
    _save_workbook(wb, path)
    return path


def generate_09_workforce():
    """09: Workforce — technician roster with specialties and shifts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Workforce"

    headers = [
        "worker_id", "name", "specialty", "shift", "plant_id",
        "available", "certifications_csv", "phone", "email",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        ["TEC-001", "Ahmed Benali", "FITTER", "MORNING", "OCP-JFC1",
         True, "Mechanical Fitting Level 3,Crane Operation", "+212-600-001001", "a.benali@ocp.ma"],
        ["TEC-002", "Youssef El Amrani", "ELECTRICIAN", "MORNING", "OCP-JFC1",
         True, "HV Switching,Electrical Installation", "+212-600-001002", "y.elamrani@ocp.ma"],
        ["TEC-003", "Fatima Zahra Idrissi", "INSTRUMENTIST", "MORNING", "OCP-JFC1",
         True, "ISA Certified,PLC Programming", "+212-600-001003", "fz.idrissi@ocp.ma"],
        ["TEC-004", "Mohammed Tazi", "FITTER", "AFTERNOON", "OCP-JFC1",
         True, "Welding ASME IX,Rigging", "+212-600-001004", "m.tazi@ocp.ma"],
        ["TEC-005", "Rachid Ouazzani", "LUBRICATOR", "MORNING", "OCP-JFC1",
         True, "ICML MLT-I,Oil Analysis", "+212-600-001005", "r.ouazzani@ocp.ma"],
        ["TEC-006", "Karim Bouazza", "OPERATOR", "MORNING", "OCP-JFC1",
         True, "SAG Mill Operation,Process Control", "+212-600-001006", "k.bouazza@ocp.ma"],
        ["TEC-007", "Hassan Chaoui", "FITTER", "NIGHT", "OCP-JFC1",
         True, "Pump Overhaul,Alignment", "+212-600-001007", "h.chaoui@ocp.ma"],
        ["TEC-008", "Omar Fassi", "ELECTRICIAN", "AFTERNOON", "OCP-JFC1",
         True, "MV Motors,VFD Commissioning", "+212-600-001008", "o.fassi@ocp.ma"],
        ["TEC-009", "Abdellah Rhazi", "CONMON_SPECIALIST", "MORNING", "OCP-JFC1",
         True, "Vibration Analysis Cat III,Thermography Level 2", "+212-600-001009", "a.rhazi@ocp.ma"],
        ["TEC-010", "Samir Benjelloun", "FITTER", "MORNING", "OCP-JFC1",
         True, "Conveyor Systems,Belt Splicing", "+212-600-001010", "s.benjelloun@ocp.ma"],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 3, LABOUR_SPECIALTIES)
    _add_data_validation(ws, 4, SHIFT_TYPES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("worker_id", "Text", "Yes", "Unique, e.g. TEC-001", "Worker identifier"),
        ("name", "Text", "Yes", "Full name", "Worker full name"),
        ("specialty", "Text", "Yes", ", ".join(LABOUR_SPECIALTIES), "Primary labour specialty"),
        ("shift", "Text", "Yes", ", ".join(SHIFT_TYPES), "Assigned shift"),
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "Assigned plant"),
        ("available", "Boolean", "Yes", "TRUE/FALSE", "Currently available for scheduling"),
        ("certifications_csv", "Text", "No", "Comma-separated", "Professional certifications"),
        ("phone", "Text", "No", "Phone number", "Contact phone"),
        ("email", "Text", "No", "Email address", "Contact email"),
    ])

    path = OUTPUT_DIR / "09_workforce.xlsx"
    _save_workbook(wb, path)
    return path


def generate_10_field_capture():
    """10: Field Capture — technician observations (voice/text in FR/EN/AR)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Captures"

    headers = [
        "technician_id", "capture_type", "language", "equipment_tag",
        "location_hint", "raw_text", "timestamp",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["TEC-001", "VOICE", "fr", "BRY-SAG-ML-001",
         "Coté palier de commande", "J'ai remarqué une vibration anormale sur le palier principal du broyeur SAG. La température est montée à 82 degrés.",
         f"{today - timedelta(days=3)} 08:30:00"],
        ["TEC-002", "TEXT", "fr", "PMP-SLP-PP-001",
         "Station de pompage zone B", "Fuite visible au niveau du joint mécanique de la pompe à boue. Débit réduit d'environ 15%.",
         f"{today - timedelta(days=2)} 14:15:00"],
        ["TEC-005", "VOICE", "fr", "BRY-BML-ML-001",
         "Système de lubrification", "Le niveau d'huile dans le réservoir de lubrification du broyeur à boulets est bas. Il faut faire l'appoint.",
         f"{today - timedelta(days=1)} 07:45:00"],
        ["TEC-009", "TEXT", "en", "CVY-CVB-CV-001",
         "Head pulley area", "Belt tracking 15mm off-center at head pulley. Scraper blade worn, material buildup on return side.",
         f"{today} 10:00:00"],
        ["TEC-006", "VOICE", "ar", "FLT-FCL-CL-001",
         "خلية التعويم رقم 1", "لاحظت تسرب هواء من وصلة المحرك في خلية التعويم. الرغوة غير مستقرة.",
         f"{today} 06:30:00"],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 2, CAPTURE_TYPES)
    _add_data_validation(ws, 3, LANGUAGES)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("technician_id", "Text", "Yes", "Must exist in Template 09", "Technician identifier"),
        ("capture_type", "Text", "Yes", ", ".join(CAPTURE_TYPES), "Capture input type"),
        ("language", "Text", "Yes", ", ".join(LANGUAGES), "Language of the observation"),
        ("equipment_tag", "Text", "No", "From hierarchy", "Equipment tag (if known)"),
        ("location_hint", "Text", "No", "Free text", "Physical location description"),
        ("raw_text", "Text", "Yes", "Free text (any language)", "Raw observation text or voice transcription"),
        ("timestamp", "DateTime", "Yes", "YYYY-MM-DD HH:MM:SS", "Capture timestamp"),
    ])

    path = OUTPUT_DIR / "10_field_capture.xlsx"
    _save_workbook(wb, path)
    return path


def generate_11_rca_events():
    """11: RCA Events — failure events for root cause analysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RCA Events"

    headers = [
        "event_description", "plant_id", "equipment_tag", "level",
        "max_consequence", "frequency", "event_date", "downtime_hours",
        "production_loss_tonnes", "direct_cost_usd",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["SAG Mill trunnion bearing failure causing unplanned shutdown", "OCP-JFC1", "BRY-SAG-ML-001",
         "3", 5, 2, str(today - timedelta(days=180)), 96, 4800, 185000],
        ["Slurry pump mechanical seal catastrophic failure", "OCP-JFC1", "PMP-SLP-PP-001",
         "2", 4, 4, str(today - timedelta(days=90)), 8, 400, 12000],
        ["Conveyor belt splice failure causing 3-hour stoppage", "OCP-JFC1", "CVY-CVB-CV-001",
         "1", 3, 3, str(today - timedelta(days=60)), 3, 150, 8000],
        ["Flotation cell agitator motor overheating trip", "OCP-JFC1", "FLT-FCL-CL-001",
         "2", 3, 2, str(today - timedelta(days=45)), 4, 200, 5000],
        ["Ball mill gearbox oil contamination", "OCP-JFC1", "BRY-BML-ML-001",
         "2", 4, 2, str(today - timedelta(days=30)), 24, 1200, 45000],
    ]
    _write_data_rows(ws, rows)

    _add_data_validation(ws, 4, RCA_LEVELS)
    _add_data_validation(ws, 5, ["1", "2", "3", "4", "5"])
    _add_data_validation(ws, 6, ["1", "2", "3", "4", "5"])

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("event_description", "Text", "Yes", "Free text", "Description of the failure event"),
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "Plant where event occurred"),
        ("equipment_tag", "Text", "No", "From hierarchy", "Equipment tag involved"),
        ("level", "Text", "Yes", ", ".join(RCA_LEVELS), "RCA investigation level (1=simple, 3=full)"),
        ("max_consequence", "Integer", "No", "1-5", "Maximum consequence severity"),
        ("frequency", "Integer", "No", "1-5", "Event recurrence frequency"),
        ("event_date", "Date", "No", "YYYY-MM-DD", "Date of the event"),
        ("downtime_hours", "Number", "No", ">=0", "Total downtime caused (hours)"),
        ("production_loss_tonnes", "Number", "No", ">=0", "Production loss in tonnes"),
        ("direct_cost_usd", "Number", "No", ">=0", "Direct cost of the event in USD"),
    ])

    path = OUTPUT_DIR / "11_rca_events.xlsx"
    _save_workbook(wb, path)
    return path


def generate_12_planning_kpi_input():
    """12: Planning KPI Input — GFSN planning metrics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning KPI Input"

    headers = [
        "plant_id", "period_start", "period_end",
        "wo_planned", "wo_completed",
        "manhours_planned", "manhours_actual",
        "pm_planned", "pm_executed",
        "backlog_hours", "weekly_capacity_hours",
        "corrective_count", "total_wo",
        "schedule_compliance_planned", "schedule_compliance_executed",
        "release_horizon_days",
        "pending_notices", "total_notices",
        "scheduled_capacity_hours", "total_capacity_hours",
        "proactive_wo", "planned_wo",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["OCP-JFC1", str(today - timedelta(days=7)), str(today),
         100, 92, 800.0, 780.0, 50, 48, 300.0, 200.0,
         15, 100, 90, 82, 5, 12, 100, 170.0, 200.0, 75, 88],
        ["OCP-JFC1", str(today - timedelta(days=14)), str(today - timedelta(days=7)),
         95, 88, 760.0, 740.0, 48, 45, 320.0, 200.0,
         18, 95, 88, 80, 4, 15, 95, 165.0, 200.0, 70, 85],
        ["OCP-JFC1", str(today - timedelta(days=21)), str(today - timedelta(days=14)),
         105, 95, 840.0, 820.0, 52, 50, 280.0, 200.0,
         12, 105, 92, 85, 5, 10, 105, 175.0, 200.0, 80, 92],
    ]
    _write_data_rows(ws, rows)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "Plant identifier"),
        ("period_start", "Date", "Yes", "YYYY-MM-DD", "KPI period start date"),
        ("period_end", "Date", "Yes", "YYYY-MM-DD", "KPI period end date"),
        ("wo_planned", "Integer", "Yes", ">=0", "Work orders planned for the period"),
        ("wo_completed", "Integer", "Yes", ">=0", "Work orders completed in the period"),
        ("manhours_planned", "Number", "Yes", ">=0", "Total manhours planned"),
        ("manhours_actual", "Number", "Yes", ">=0", "Total manhours actually used"),
        ("pm_planned", "Integer", "Yes", ">=0", "Preventive maintenance tasks planned"),
        ("pm_executed", "Integer", "Yes", ">=0", "Preventive maintenance tasks executed"),
        ("backlog_hours", "Number", "Yes", ">=0", "Total backlog in hours"),
        ("weekly_capacity_hours", "Number", "Yes", ">0", "Weekly maintenance capacity (hours)"),
        ("corrective_count", "Integer", "Yes", ">=0", "Number of corrective work orders"),
        ("total_wo", "Integer", "Yes", ">0", "Total work orders (denominator)"),
        ("schedule_compliance_planned", "Integer", "Yes", ">=0", "Scheduled WOs planned"),
        ("schedule_compliance_executed", "Integer", "Yes", ">=0", "Scheduled WOs executed on time"),
        ("release_horizon_days", "Integer", "Yes", ">=0", "Avg days WO released before execution"),
        ("pending_notices", "Integer", "Yes", ">=0", "Pending notification count"),
        ("total_notices", "Integer", "Yes", ">0", "Total notifications received"),
        ("scheduled_capacity_hours", "Number", "Yes", ">=0", "Hours consumed by scheduled work"),
        ("total_capacity_hours", "Number", "Yes", ">0", "Total available capacity hours"),
        ("proactive_wo", "Integer", "Yes", ">=0", "Proactive (PM/PdM) work orders"),
        ("planned_wo", "Integer", "Yes", ">=0", "Planned work orders (vs reactive)"),
    ])

    path = OUTPUT_DIR / "12_planning_kpi_input.xlsx"
    _save_workbook(wb, path)
    return path


def generate_13_de_kpi_input():
    """13: DE KPI Input — Defect Elimination program metrics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DE KPI Input"

    headers = [
        "plant_id", "period_start", "period_end",
        "events_reported", "events_required",
        "meetings_held", "meetings_required",
        "actions_implemented", "actions_planned",
        "savings_achieved", "savings_target",
        "failures_current", "failures_previous",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    today = date.today()
    rows = [
        ["OCP-JFC1", str(today - timedelta(days=30)), str(today),
         18, 20, 9, 10, 14, 16, 85000.0, 100000.0, 8, 12],
        ["OCP-JFC1", str(today - timedelta(days=60)), str(today - timedelta(days=30)),
         16, 20, 8, 10, 12, 15, 72000.0, 100000.0, 10, 14],
        ["OCP-JFC1", str(today - timedelta(days=90)), str(today - timedelta(days=60)),
         20, 20, 10, 10, 15, 15, 95000.0, 100000.0, 12, 18],
    ]
    _write_data_rows(ws, rows)

    _auto_size_columns(ws)

    _add_instructions_sheet(wb, [
        ("plant_id", "Text", "Yes", "e.g. OCP-JFC1", "Plant identifier"),
        ("period_start", "Date", "Yes", "YYYY-MM-DD", "KPI period start date"),
        ("period_end", "Date", "Yes", "YYYY-MM-DD", "KPI period end date"),
        ("events_reported", "Integer", "Yes", ">=0", "DE events reported in the period"),
        ("events_required", "Integer", "Yes", ">0", "DE events required (target)"),
        ("meetings_held", "Integer", "Yes", ">=0", "DE review meetings held"),
        ("meetings_required", "Integer", "Yes", ">0", "DE review meetings required"),
        ("actions_implemented", "Integer", "Yes", ">=0", "CAPA actions implemented"),
        ("actions_planned", "Integer", "Yes", ">0", "CAPA actions planned"),
        ("savings_achieved", "Number", "Yes", ">=0", "Cost savings achieved (USD)"),
        ("savings_target", "Number", "Yes", ">0", "Cost savings target (USD)"),
        ("failures_current", "Integer", "Yes", ">=0", "Repeat failure count (current period)"),
        ("failures_previous", "Integer", "Yes", ">=0", "Repeat failure count (previous period)"),
    ])

    path = OUTPUT_DIR / "13_de_kpi_input.xlsx"
    _save_workbook(wb, path)
    return path


def generate_14_maintenance_strategy():
    """14: Maintenance Strategy — links equipment + failure mode + strategy type + tasks.

    Based on real-world R8/AssetTactics export structure. Each row is one
    failure mode with its maintenance strategy decision: tactics type,
    primary task (with interval and acceptable limits), and secondary task
    (corrective, triggered by condition or failure). Follows REF-01 §3.5.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Strategies"

    headers = [
        # Identity
        "strategy_id", "equipment_tag", "maintainable_item",
        "function_and_failure",
        # Failure mode
        "what", "mechanism", "cause",
        # Strategy decision
        "status", "tactics_type",
        # Primary task (proactive — has interval)
        "primary_task_id", "primary_task_interval", "operational_units",
        "time_units", "primary_task_acceptable_limits",
        "primary_task_conditional_comments", "primary_task_constraint",
        "primary_task_task_type", "primary_task_access_time",
        # Secondary task (corrective — triggered by condition)
        "secondary_task_id", "secondary_task_constraint",
        "secondary_task_task_type", "secondary_task_access_time",
        "secondary_task_comments",
        # Budget (for secondary/replacement tasks)
        "budgeted_as", "budgeted_life", "budgeted_life_time_units",
        "budgeted_life_operational_units",
        # Metadata
        "existing_task", "justification_category", "justification", "notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))

    rows = [
        # S-001: CB — SAG Mill bearings (primary=inspect, secondary=replace)
        ["S-001", "BRY-SAG-ML-001", "Trunnion Bearing Assembly",
         "PRODUCTION-Grind ore to 80% passing 150um-Bearing fails-PRIMARY",
         "Trunnion Bearing", "WEARS", "BREAKDOWN_OF_LUBRICATION",
         "RECOMMENDED", "CONDITION_BASED",
         "T-001", 4, "OPERATING_HOURS", "WEEKS",
         "Vibration < 4.5 mm/s, Temperature < 85C",
         "If vibration > 3.5 mm/s, increase frequency to weekly",
         "ONLINE", "INSPECT", 0,
         "T-006", "OFFLINE", "REPLACE", 96,
         "Full bearing replacement with alignment verification",
         "REPLACE", 5, "YEARS", "OPERATING_HOURS",
         "R8_LIBRARY", None, None, None],
        # S-002: FT — SAG Mill liner (primary=replace, no secondary)
        ["S-002", "BRY-SAG-ML-001", "Grinding Chamber",
         "PRODUCTION-Grind ore to 80% passing 150um-Liner worn-PRIMARY",
         "Liner Set", "WEARS", "MECHANICAL_OVERLOAD",
         "RECOMMENDED", "FIXED_TIME",
         "T-007", 8000, "OPERATING_HOURS", None,
         None, None,
         "OFFLINE", "REPLACE", 72,
         None, None, None, None,
         None,
         "REPLACE", None, None, "OPERATING_HOURS",
         "R8_LIBRARY", None, "Wear-out pattern B — age-related", None],
        # S-003: CB — Slurry pump impeller (primary=inspect, secondary=repair)
        ["S-003", "PMP-SLP-PP-001", "Wet End Assembly",
         "PRODUCTION-Transfer slurry at 400 m3/h-Reduced flow-PRIMARY",
         "Impeller", "SEVERS", "ABRASION",
         "RECOMMENDED", "CONDITION_BASED",
         "T-004", 2, None, "WEEKS",
         "Impeller OD > 340mm, no visible cracking",
         "If wear rate > 1mm/week, plan impeller repair",
         "ONLINE", "INSPECT", 0,
         "T-008", "OFFLINE", "REPAIR", 8,
         "Impeller rebuilding or replacement per wear assessment",
         "REPAIR", None, None, None,
         "R8_LIBRARY", None, None, None],
        # S-004: RTF — Slurry pump seal (no primary, only secondary=replace on failure)
        ["S-004", "PMP-SLP-PP-001", "Sealing System",
         "ENVIRONMENT-Prevent slurry leakage-Seal fails-PRIMARY",
         "Mechanical Seal", "DEGRADES", "CHEMICAL_ATTACK",
         "RECOMMENDED", "RUN_TO_FAILURE",
         None, None, None, None,
         None, None, None, None, None,
         "T-009", "OFFLINE", "REPLACE", 4,
         "Replace on failure — cost of prevention exceeds cost of failure",
         "REPLACE", 2, "YEARS", None,
         None, None, "Cost analysis: prevention > failure cost", None],
        # S-005: FFI — Guard interlock (primary=functional test, secondary=replace)
        ["S-005", "BRY-SAG-ML-001", "Safety Guard System",
         "SAFETY-Protect personnel from rotating hazards-Interlock fails-PROTECTIVE",
         "Guard Interlock", "DEGRADES", "AGE",
         "RECOMMENDED", "FAULT_FINDING",
         "T-003", 12, None, "WEEKS",
         "Interlock trips within 2 seconds of guard opening",
         "If test fails, immediately isolate and replace interlock",
         "TEST_MODE", "TEST", 0.5,
         "T-010", "OFFLINE", "REPLACE", 2,
         None,
         "REPLACE", None, None, None,
         "R8_LIBRARY", None, "Hidden safety failure — mandatory FFI", None],
        # S-006: Redundant example
        ["S-006", "BRY-SAG-ML-001", "Trunnion Bearing Assembly",
         "PRODUCTION-Grind ore to 80% passing 150um-Bearing fails-PRIMARY",
         "Trunnion Bearing", "WEARS", "METAL_TO_METAL_CONTACT",
         "REDUNDANT", "CONDITION_BASED",
         "T-001", 4, None, "WEEKS",
         "Vibration < 4.5 mm/s", None,
         "ONLINE", "INSPECT", 0,
         "T-006", "OFFLINE", "REPLACE", 96,
         None,
         None, None, None, None,
         "R8_LIBRARY", "ELIMINATED",
         "Failure mode already covered by S-001 (same inspection task covers both causes)",
         "STRATEGY: Redundant — same primary task already assigned"],
    ]
    _write_data_rows(ws, rows)

    # Data validations
    _add_data_validation(ws, 6, MECHANISMS)         # mechanism
    _add_data_validation(ws, 7, CAUSES)             # cause (uses Lookups sheet)
    _add_data_validation(ws, 8, STRATEGY_STATUSES)  # status
    _add_data_validation(ws, 9, STRATEGY_TYPES)     # tactics_type
    _add_data_validation(ws, 12, FREQUENCY_UNITS)   # operational_units
    _add_data_validation(ws, 13, FREQUENCY_UNITS)   # time_units
    _add_data_validation(ws, 16, TASK_CONSTRAINTS)   # primary_task_constraint
    _add_data_validation(ws, 17, TASK_TYPES)         # primary_task_task_type
    _add_data_validation(ws, 20, TASK_CONSTRAINTS)   # secondary_task_constraint
    _add_data_validation(ws, 21, TASK_TYPES)         # secondary_task_task_type
    _add_data_validation(ws, 24, BUDGETED_AS_VALUES)  # budgeted_as
    _add_data_validation(ws, 26, BUDGETED_LIFE_UNITS)  # budgeted_life_time_units
    _add_data_validation(ws, 27, FREQUENCY_UNITS)   # budgeted_life_operational_units
    _add_data_validation(ws, 29, JUSTIFICATION_CATEGORIES)  # justification_category

    _auto_size_columns(ws)

    # ── Valid FM Combinations reference sheet ──
    ws_combo = wb.create_sheet("Valid FM Combinations")
    ws_combo.cell(row=1, column=1, value="Mechanism").fill = OCP_GREEN
    ws_combo.cell(row=1, column=1).font = HEADER_FONT
    ws_combo.cell(row=1, column=2, value="Cause").fill = OCP_GREEN
    ws_combo.cell(row=1, column=2).font = HEADER_FONT
    ws_combo.freeze_panes = "A2"

    for row_idx, (mech, cause) in enumerate(FM_COMBOS, start=2):
        ws_combo.cell(row=row_idx, column=1, value=mech).font = DATA_FONT
        ws_combo.cell(row=row_idx, column=2, value=cause).font = DATA_FONT
    ws_combo.column_dimensions["A"].width = 30
    ws_combo.column_dimensions["B"].width = 35

    # ── Strategy Type Rules reference sheet ──
    ws_rules = wb.create_sheet("Strategy Type Rules")
    rule_headers = ["Tactics Type", "Has Primary Task?", "Has Secondary Task?",
                    "Has Acceptable Limits?", "Has Interval?", "Description"]
    for col_idx, h in enumerate(rule_headers, start=1):
        cell = ws_rules.cell(row=1, column=col_idx, value=h)
        cell.fill = OCP_GREEN
        cell.font = HEADER_FONT
    ws_rules.freeze_panes = "A2"
    for row_idx, row_data in enumerate(STRATEGY_RULES, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_rules.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = WRAP_ALIGN
    for col_idx in range(1, 7):
        ws_rules.column_dimensions[get_column_letter(col_idx)].width = 25

    _add_instructions_sheet(wb, [
        ("== IDENTITY ==", "", "", "", "Strategy identification fields"),
        ("strategy_id", "Text", "Yes", "Unique, e.g. S-001", "Unique strategy identifier"),
        ("equipment_tag", "Text", "Yes", "From Template 01", "Equipment tag"),
        ("maintainable_item", "Text", "Yes", "Free text", "Name of the maintainable item (MI)"),
        ("function_and_failure", "Text", "No", "{Category}-{Function}-{Failure}-{FunctionType}", "Structured function and failure description (optional)"),
        ("", "", "", "", ""),
        ("== FAILURE MODE ==", "", "", "", "What + Mechanism + Cause (must be valid 72-combo)"),
        ("what", "Text", "Yes", "Capital letter, singular", "Component sub-part that fails"),
        ("mechanism", "Text", "Yes", "18 valid values", "How it fails (see FM Combinations sheet)"),
        ("cause", "Text", "Yes", "44 valid values", "Why it fails (see FM Combinations sheet)"),
        ("", "", "", "", ""),
        ("== STRATEGY DECISION ==", "", "", "", "Maintenance strategy type selection"),
        ("status", "Text", "Yes", ", ".join(STRATEGY_STATUSES), "RECOMMENDED or REDUNDANT"),
        ("tactics_type", "Text", "Yes", ", ".join(STRATEGY_TYPES), "Strategy type (see Strategy Type Rules sheet)"),
        ("", "", "", "", ""),
        ("== PRIMARY TASK ==", "", "", "", "Proactive task with fixed interval. NULL for RTF."),
        ("primary_task_id", "Text", "Conditional", "FK to Template 04", "Primary task ID (NULL for RTF/REDESIGN)"),
        ("primary_task_interval", "Number", "Conditional", ">0", "Frequency value (NULL for RTF/REDESIGN)"),
        ("operational_units", "Text", "Conditional", ", ".join(FREQUENCY_UNITS), "Operational freq units (for operational causes)"),
        ("time_units", "Text", "Conditional", "DAYS,WEEKS,MONTHS,YEARS", "Calendar freq units (for calendar causes)"),
        ("primary_task_acceptable_limits", "Text", "Conditional", "Free text", "Acceptable condition thresholds (CB/FFI only)"),
        ("primary_task_conditional_comments", "Text", "No", "Free text", "Action when limits exceeded (CB/FFI only)"),
        ("primary_task_constraint", "Text", "Conditional", ", ".join(TASK_CONSTRAINTS), "Constraint override for this strategy"),
        ("primary_task_task_type", "Text", "Conditional", ", ".join(TASK_TYPES), "Task type (can override task default)"),
        ("primary_task_access_time", "Number", "No", ">=0", "Access time override (hours)"),
        ("", "", "", "", ""),
        ("== SECONDARY TASK ==", "", "", "", "Corrective task triggered by condition/failure. NULL for FT."),
        ("secondary_task_id", "Text", "Conditional", "FK to Template 04", "Secondary task ID (NULL for FT/REDESIGN)"),
        ("secondary_task_constraint", "Text", "Conditional", ", ".join(TASK_CONSTRAINTS), "Constraint for corrective task"),
        ("secondary_task_task_type", "Text", "Conditional", ", ".join(TASK_TYPES), "Task type for corrective task"),
        ("secondary_task_access_time", "Number", "No", ">=0", "Access time for corrective task (hours)"),
        ("secondary_task_comments", "Text", "No", "Free text", "Additional corrective action notes"),
        ("", "", "", "", ""),
        ("== BUDGET ==", "", "", "", "Budget classification for replacement/repair tasks"),
        ("budgeted_as", "Text", "No", ", ".join(BUDGETED_AS_VALUES), "NOT_BUDGETED, REPAIR, or REPLACE"),
        ("budgeted_life", "Number", "No", ">0", "Expected useful life of MI"),
        ("budgeted_life_time_units", "Text", "No", ", ".join(BUDGETED_LIFE_UNITS), "Time units for budgeted life"),
        ("budgeted_life_operational_units", "Text", "No", ", ".join(FREQUENCY_UNITS), "Operational units for budgeted life"),
        ("", "", "", "", ""),
        ("== METADATA ==", "", "", "", "Traceability and change management"),
        ("existing_task", "Text", "No", "Free text", "Source library/workshop: R8_LIBRARY, Anglo Tactics Library, WS [date]"),
        ("justification_category", "Text", "No", ", ".join(JUSTIFICATION_CATEGORIES), "MSO change category (for strategy optimization)"),
        ("justification", "Text", "No", "Free text", "Rationale for strategy selection or elimination"),
        ("notes", "Text", "No", "Free text", "Additional notes (prefix with ASSET: or STRATEGY: for context)"),
    ])

    path = OUTPUT_DIR / "14_maintenance_strategy.xlsx"
    _save_workbook(wb, path)
    return path


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

GENERATORS = [
    generate_01_equipment_hierarchy,
    generate_02_criticality_assessment,
    generate_03_failure_modes,
    generate_04_maintenance_tasks,
    generate_05_work_packages,
    generate_06_work_order_history,
    generate_07_spare_parts_inventory,
    generate_08_shutdown_calendar,
    generate_09_workforce,
    generate_10_field_capture,
    generate_11_rca_events,
    generate_12_planning_kpi_input,
    generate_13_de_kpi_input,
    generate_14_maintenance_strategy,
]


def generate_all(output_dir: Path | None = None):
    """Generate all 14 Excel templates.

    Args:
        output_dir: Target directory. Defaults to OUTPUT_DIR (templates/).
                    Pass a client project's 5-templates/ dir for client-specific generation.
    """
    global OUTPUT_DIR
    target = output_dir or OUTPUT_DIR
    target.mkdir(parents=True, exist_ok=True)
    paths = []
    # Temporarily swap OUTPUT_DIR for generators that reference it
    original = OUTPUT_DIR
    try:
        OUTPUT_DIR = target
        for gen_fn in GENERATORS:
            path = gen_fn()
            paths.append(path)
            print(f"  [OK] {path.name}")
    finally:
        OUTPUT_DIR = original
    return paths


if __name__ == "__main__":
    import sys

    target = Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT_DIR
    print("Generating OCP Maintenance AI data-loading templates...")
    print(f"Output directory: {target}")
    print()
    generated = generate_all(target)
    print(f"\nDone! {len(generated)} templates generated.")
