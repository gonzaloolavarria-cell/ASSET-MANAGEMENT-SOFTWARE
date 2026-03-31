import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='1B5E20')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
dfont = Font(name='Calibri', size=11)
dalign = Alignment(vertical='center')

# --- Sheet 1: Equipment Hierarchy ---
df = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment Hierarchy')

wb = Workbook()
ws1 = wb.active
ws1.title = 'Equipment Hierarchy'

new_headers = [
    'iso14224_level_1_business', 'iso14224_level_2_installation',
    'iso14224_level_3_plant_unit', 'iso14224_level_4_section',
] + list(df.columns)

for c, h in enumerate(new_headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

for i, (_, row) in enumerate(df.iterrows(), 2):
    vals = [
        'Mining & Minerals Processing',
        row['plant_name'],
        row['area_name'],
        row['system_name'],
    ] + [row[col] if pd.notna(row[col]) else None for col in df.columns]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=i, column=c, value=v)
        cell.font = dfont
        cell.alignment = dalign

widths_s1 = [30, 30, 25, 25, 12, 28, 12, 24, 12, 24, 17, 25, 15, 14, 10, 15, 12, 12, 22, 16, 12, 16]
for c, w in enumerate(widths_s1, 1):
    if c <= len(widths_s1):
        ws1.column_dimensions[get_column_letter(c)].width = w
ws1.auto_filter.ref = f'A1:{get_column_letter(len(new_headers))}{len(df)+1}'
ws1.freeze_panes = 'A2'

# --- Sheet 2: Equipment BOM ---
df_bom = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment BOM')

ws2 = wb.create_sheet('Equipment BOM')
LEVEL_MAP = {4: 5, 5: 6, 6: 7}
NTYPE_MAP = {'SUB_ASSEMBLY': 'SUBUNIT'}
LEVEL_NAME = {5: 'Equipment Class/Unit', 6: 'Subunit', 7: 'Component/Maintainable Item'}

bom_headers = list(df_bom.columns) + ['iso14224_level', 'iso14224_level_name']
for c, h in enumerate(bom_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

for i, (_, row) in enumerate(df_bom.iterrows(), 2):
    old_level = int(row['level'])
    new_level = LEVEL_MAP.get(old_level, old_level)
    ntype = NTYPE_MAP.get(row['node_type'], row['node_type'])
    for c, col in enumerate(df_bom.columns, 1):
        if col == 'level':
            v = new_level
        elif col == 'node_type':
            v = ntype
        else:
            v = row[col] if pd.notna(row[col]) else None
        cell = ws2.cell(row=i, column=c, value=v)
        cell.font = dfont
        cell.alignment = dalign
    ws2.cell(row=i, column=len(df_bom.columns)+1, value=new_level).font = dfont
    ws2.cell(row=i, column=len(df_bom.columns)+2, value=LEVEL_NAME.get(new_level, '')).font = dfont

widths_s2 = [17, 7, 18, 28, 28, 28, 24, 6, 16, 14, 14, 28]
for c, w in enumerate(widths_s2, 1):
    if c <= len(widths_s2):
        ws2.column_dimensions[get_column_letter(c)].width = w
ws2.auto_filter.ref = f'A1:{get_column_letter(len(bom_headers))}{len(df_bom)+1}'
ws2.freeze_panes = 'A2'

wb.save('templates/01_equipment_hierarchy.xlsx')
print(f'T01 done. Sheet1: {len(df)} rows x {len(new_headers)} cols. Sheet2: {len(df_bom)} rows x {len(bom_headers)} cols.')

# Verify
from collections import Counter
levels = [LEVEL_MAP.get(int(r), int(r)) for r in df_bom['level']]
cnt = Counter(levels)
print(f'BOM levels: {dict(cnt)}')
