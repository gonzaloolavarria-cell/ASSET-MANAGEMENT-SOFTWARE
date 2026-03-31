"""
OCP Maintenance AI — Comprehensive Excel Template Generator V2
Generates 14 .xlsx templates with FULL realistic synthetic data
for Jorf Fertilizer Complex 1 (OCP-JFC1).
~60 equipment, ~300 failure modes, ~150 tasks, ~300 strategies.
"""
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

random.seed(42)
OUTPUT_DIR = Path(__file__).parent

# ── Branding ─────────────────────────────────────────────────
OCP_GREEN = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# ── Enum Values ──────────────────────────────────────────────
CRITICALITIES = ["AA", "A+", "A", "B", "C", "D"]
EQUIP_STATUSES = ["ACTIVE", "INACTIVE", "DECOMMISSIONED"]
NODE_TYPES = ["PLANT", "AREA", "SYSTEM", "EQUIPMENT", "SUB_ASSEMBLY", "MAINTAINABLE_ITEM"]
CRIT_CATEGORIES = [
    "SAFETY", "HEALTH", "ENVIRONMENT", "PRODUCTION", "OPERATING_COST",
    "CAPITAL_COST", "SCHEDULE", "REVENUE", "COMMUNICATIONS", "COMPLIANCE", "REPUTATION",
]
CRIT_METHODS = ["FULL_MATRIX", "SIMPLIFIED"]
MECHANISMS = [
    "ARCS", "BLOCKS", "BREAKS_FRACTURE_SEPARATES", "CORRODES", "CRACKS",
    "DEGRADES", "DISTORTS", "DRIFTS", "EXPIRES", "IMMOBILISED", "LOOSES_PRELOAD",
    "OPEN_CIRCUIT", "OVERHEATS_MELTS", "SEVERS", "SHORT_CIRCUITS",
    "THERMALLY_OVERLOADS", "WASHES_OFF", "WEARS",
]
CAUSES = [
    "ABRASION", "AGE", "BREAKDOWN_IN_INSULATION", "BREAKDOWN_OF_LUBRICATION",
    "BIO_ORGANISMS", "CHEMICAL_ATTACK", "CHEMICAL_REACTION", "CONTAMINATION",
    "CORROSIVE_ENVIRONMENT", "CREEP", "CREVICE", "CYCLIC_LOADING",
    "DISSIMILAR_METALS_CONTACT", "ELECTRICAL_ARCING", "ELECTRICAL_OVERLOAD",
    "ENTRAINED_AIR", "EXCESSIVE_FLUID_VELOCITY", "EXCESSIVE_PARTICLE_SIZE",
    "EXCESSIVE_TEMPERATURE", "EXPOSURE_TO_ATMOSPHERE", "EXPOSURE_TO_HIGH_TEMP_CORROSIVE",
    "EXPOSURE_TO_HIGH_TEMP", "EXPOSURE_TO_LIQUID_METAL", "HIGH_TEMP_CORROSIVE",
    "IMPACT_SHOCK_LOADING", "INSUFFICIENT_FLUID_VELOCITY", "LACK_OF_LUBRICATION",
    "LOW_PRESSURE", "LUBRICANT_CONTAMINATION", "MECHANICAL_OVERLOAD",
    "METAL_TO_METAL_CONTACT", "OFF_CENTER_LOADING", "OVERCURRENT",
    "POOR_ELECTRICAL_CONNECTIONS", "POOR_ELECTRICAL_INSULATION", "RADIATION",
    "RELATIVE_MOVEMENT", "RUBBING", "STRAY_CURRENT", "THERMAL_OVERLOAD",
    "THERMAL_STRESSES", "UNEVEN_LOADING", "USE", "VIBRATION",
]
FAILURE_PATTERNS = ["A_BATHTUB", "B_AGE", "C_FATIGUE", "D_STRESS", "E_RANDOM", "F_EARLY_LIFE"]
FAILURE_CONSEQUENCES = [
    "HIDDEN_SAFETY", "HIDDEN_NONSAFETY", "EVIDENT_SAFETY",
    "EVIDENT_ENVIRONMENTAL", "EVIDENT_OPERATIONAL", "EVIDENT_NONOPERATIONAL",
]
STRATEGY_TYPES = ["CONDITION_BASED", "FIXED_TIME", "RUN_TO_FAILURE", "FAULT_FINDING", "REDESIGN", "OEM"]
FAILURE_TYPES = ["TOTAL", "PARTIAL"]
FUNCTION_TYPES = ["PRIMARY", "SECONDARY", "PROTECTIVE"]
TASK_TYPES = ["INSPECT", "CHECK", "TEST", "LUBRICATE", "CLEAN", "REPLACE", "REPAIR", "CALIBRATE"]
TASK_CONSTRAINTS = ["ONLINE", "OFFLINE", "TEST_MODE"]
LABOUR_SPECIALTIES = ["FITTER", "ELECTRICIAN", "INSTRUMENTIST", "OPERATOR", "CONMON_SPECIALIST", "LUBRICATOR"]
UNITS_OF_MEASURE = ["EA", "L", "KG", "M"]
FREQUENCY_UNITS = ["HOURS", "DAYS", "WEEKS", "MONTHS", "YEARS", "HOURS_RUN", "OPERATING_HOURS", "TONNES", "CYCLES"]
BUDGET_TYPES = ["REPAIR", "REPLACE"]
WP_TYPES = ["STANDALONE", "SUPPRESSIVE", "SEQUENTIAL"]
WP_CONSTRAINTS = ["ONLINE", "OFFLINE"]
ORDER_TYPES = ["PM01", "PM02", "PM03"]
PRIORITIES = ["1_EMERGENCY", "2_URGENT", "3_NORMAL", "4_PLANNED"]
WO_STATUSES = ["CREATED", "RELEASED", "IN_PROGRESS", "COMPLETED", "CLOSED", "CANCELLED"]
VED_CLASSES = ["VITAL", "ESSENTIAL", "DESIRABLE"]
FSN_CLASSES = ["FAST_MOVING", "SLOW_MOVING", "NON_MOVING"]
ABC_CLASSES = ["A_HIGH", "B_MEDIUM", "C_LOW"]
SHUTDOWN_TYPES = ["MINOR_8H", "MAJOR_20H_PLUS"]
SHIFT_TYPES = ["MORNING", "AFTERNOON", "NIGHT"]
CAPTURE_TYPES = ["VOICE", "TEXT", "IMAGE", "VOICE+IMAGE"]
LANGUAGES = ["fr", "en", "ar"]
STRATEGY_STATUSES = ["RECOMMENDED", "REDUNDANT"]
BUDGETED_AS_VALUES = ["NOT_BUDGETED", "REPAIR", "REPLACE"]
BUDGETED_LIFE_UNITS = ["YEARS", "MONTHS", "WEEKS"]
JUSTIFICATION_CATEGORIES = [
    "MODIFIED", "ELIMINATED", "FREQUENCY_CHANGE",
    "TACTIC_CHANGE", "MAINTAINED", "NEW_TASK",
]
RCA_LEVELS = ["1", "2", "3"]


# ── Excel Helpers ────────────────────────────────────────────
def _apply_header_style(ws, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = OCP_GREEN
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_size_columns(ws, min_width=12, max_width=40):
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value or ""
        width = min(max(len(str(header_val)) + 4, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_data_validation(ws, col_idx, values, num_rows=500):
    inline_formula = ",".join(values)
    if len(inline_formula) + 2 <= 255:
        formula = '"' + inline_formula + '"'
    else:
        wb = ws.parent
        if "Lookups" not in wb.sheetnames:
            lk = wb.create_sheet("Lookups")
            lk.sheet_state = "hidden"
            lk._lookup_next_col = 1
        else:
            lk = wb["Lookups"]
        lk_col = getattr(lk, "_lookup_next_col", 1)
        for i, val in enumerate(values, start=1):
            lk.cell(row=i, column=lk_col, value=val)
        lk_letter = get_column_letter(lk_col)
        formula = f"=Lookups!${lk_letter}$1:${lk_letter}${len(values)}"
        lk._lookup_next_col = lk_col + 1
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a valid value from the dropdown."
    dv.errorTitle = "Invalid Value"
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}2:{col_letter}{num_rows}")
    ws.add_data_validation(dv)


def _write_data_rows(ws, rows):
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def _hdr(ws, headers):
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws, len(headers))


# ══════════════════════════════════════════════════════════════
# PLANT DATA — Complete Jorf Fertilizer Complex 1
# ══════════════════════════════════════════════════════════════

# (area_code, area_name_en, area_name_fr, systems)
# system = (sys_code, sys_name_en, sys_name_fr, equipment_list)
# equipment = (tag, desc_en, desc_fr, eq_type, manufacturer, model, serial, power_kw, weight_kg, crit, install_date)

PLANT = {
    "plant_id": "OCP-JFC1",
    "plant_name": "Jorf Fertilizer Complex 1",
    "plant_name_fr": "Complexe d'engrais de Jorf 1",
    "location": "El Jadida, Morocco",
}

# Sub-assembly and MI definitions per equipment type
# Format: {eq_type: [(sub_assy_name, [(mi_name, mi_category), ...]), ...]}
EQUIPMENT_BOM = {
    "SAG_MILL": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX"), ("Motor Coupling", "COUPLING"), ("Pinion", "GEAR")]),
        ("Grinding System", [("Mill Liner", "LINER"), ("Lifter Bar", "LINER"), ("Shell", "STRUCTURE")]),
        ("Feed System", [("Feed Chute", "STRUCTURE"), ("Feed Trunnion Bearing", "BEARING")]),
        ("Discharge System", [("Discharge Trunnion Bearing", "BEARING"), ("Trommel Screen", "SCREEN")]),
        ("Lubrication System", [("Lube Oil Pump", "PUMP"), ("Lube Oil Filter", "FILTER"), ("Oil Cooler", "HEAT_EXCHANGER"), ("Oil Reservoir", "VESSEL")]),
        ("Instrumentation", [("Vibration Sensor", "SENSOR"), ("Temperature Sensor", "SENSOR"), ("Speed Sensor", "SENSOR")]),
    ],
    "BALL_MILL": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX"), ("Motor Coupling", "COUPLING")]),
        ("Grinding System", [("Mill Liner", "LINER"), ("Lifter Bar", "LINER")]),
        ("Feed System", [("Feed Chute", "STRUCTURE")]),
        ("Discharge System", [("Discharge Grate", "SCREEN"), ("Trommel Screen", "SCREEN")]),
        ("Lubrication System", [("Lube Oil Pump", "PUMP"), ("Lube Oil Filter", "FILTER"), ("Oil Cooler", "HEAT_EXCHANGER")]),
    ],
    "SLURRY_PUMP": [
        ("Wet End", [("Impeller", "IMPELLER"), ("Volute Liner", "LINER"), ("Throat Bush", "LINER"), ("Frame Plate Liner", "LINER")]),
        ("Bearing Assembly", [("DE Bearing", "BEARING"), ("NDE Bearing", "BEARING"), ("Bearing Housing", "STRUCTURE")]),
        ("Drive System", [("Pump Motor", "MOTOR"), ("Motor Coupling", "COUPLING")]),
        ("Sealing System", [("Mechanical Seal", "SEAL"), ("Shaft Sleeve", "SLEEVE")]),
    ],
    "WATER_PUMP": [
        ("Hydraulic End", [("Impeller", "IMPELLER"), ("Wear Ring", "LINER"), ("Casing", "STRUCTURE")]),
        ("Bearing Assembly", [("DE Bearing", "BEARING"), ("NDE Bearing", "BEARING")]),
        ("Drive System", [("Pump Motor", "MOTOR"), ("Motor Coupling", "COUPLING")]),
        ("Sealing System", [("Mechanical Seal", "SEAL")]),
    ],
    "FLOTATION_CELL": [
        ("Agitator System", [("Agitator Motor", "MOTOR"), ("Impeller", "IMPELLER"), ("Agitator Shaft", "SHAFT")]),
        ("Air System", [("Blower", "BLOWER"), ("Air Pipe", "PIPE"), ("Sparger", "NOZZLE")]),
        ("Overflow System", [("Launder", "STRUCTURE"), ("Weir Plate", "STRUCTURE")]),
    ],
    "REAGENT_PUMP": [
        ("Pump Assembly", [("Diaphragm", "SEAL"), ("Check Valve", "VALVE"), ("Pump Head", "STRUCTURE")]),
        ("Drive System", [("Pump Motor", "MOTOR")]),
    ],
    "HYDROCYCLONE": [
        ("Cyclone Body", [("Vortex Finder", "LINER"), ("Spigot", "LINER"), ("Cylinder Section", "STRUCTURE")]),
        ("Feed System", [("Feed Inlet", "PIPE"), ("Feed Box", "STRUCTURE")]),
    ],
    "THICKENER": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Thickener Gearbox", "GEARBOX")]),
        ("Rake System", [("Rake Arm", "STRUCTURE"), ("Rake Blade", "LINER"), ("Torque Limiter", "COUPLING")]),
        ("Feed System", [("Feedwell", "STRUCTURE"), ("Distributor", "NOZZLE")]),
        ("Underflow System", [("Cone Liner", "LINER"), ("Underflow Valve", "VALVE")]),
    ],
    "BELT_FILTER": [
        ("Vacuum System", [("Vacuum Pump", "PUMP"), ("Vacuum Receiver", "VESSEL")]),
        ("Filter Belt", [("Filter Cloth", "BELT"), ("Belt Tracking Roller", "BEARING")]),
        ("Drive System", [("Drive Motor", "MOTOR"), ("Filter Gearbox", "GEARBOX")]),
        ("Wash System", [("Spray Bar", "PIPE"), ("Wash Nozzle", "NOZZLE")]),
    ],
    "DISC_FILTER": [
        ("Filter Assembly", [("Filter Disc", "SCREEN"), ("Disc Shaft", "SHAFT"), ("Filter Cloth", "BELT")]),
        ("Vacuum System", [("Vacuum Pump", "PUMP")]),
        ("Drive System", [("Drive Motor", "MOTOR"), ("Filter Gearbox", "GEARBOX")]),
    ],
    "ROTARY_DRYER": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX"), ("Pinion", "GEAR"), ("Girth Gear", "GEAR")]),
        ("Shell System", [("Dryer Shell", "STRUCTURE"), ("Lifting Flight", "STRUCTURE")]),
        ("Burner System", [("Main Burner", "BURNER"), ("Ignition System", "ELECTRICAL"), ("Fuel Valve", "VALVE")]),
        ("Seal System", [("Inlet Seal", "SEAL"), ("Outlet Seal", "SEAL")]),
        ("Support System", [("Riding Ring", "BEARING"), ("Support Roller", "BEARING"), ("Thrust Roller", "BEARING")]),
    ],
    "ROTARY_COOLER": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX"), ("Pinion", "GEAR")]),
        ("Shell System", [("Cooler Shell", "STRUCTURE"), ("Lifting Flight", "STRUCTURE")]),
        ("Support System", [("Riding Ring", "BEARING"), ("Support Roller", "BEARING")]),
    ],
    "BELT_CONVEYOR": [
        ("Drive System", [("Drive Motor", "MOTOR"), ("Conveyor Gearbox", "GEARBOX"), ("Motor Coupling", "COUPLING")]),
        ("Head Pulley", [("Head Pulley Drum", "BEARING"), ("Head Pulley Bearing", "BEARING"), ("Pulley Lagging", "LINER")]),
        ("Tail Pulley", [("Tail Pulley Drum", "BEARING"), ("Tail Pulley Bearing", "BEARING"), ("Belt Scraper", "LINER")]),
        ("Belt", [("Conveyor Belt", "BELT"), ("Belt Splice", "BELT")]),
        ("Idlers", [("Carry Idler Set", "BEARING"), ("Return Idler Set", "BEARING"), ("Impact Idler Set", "BEARING")]),
    ],
    "STACKER": [
        ("Travel System", [("Travel Motor", "MOTOR"), ("Travel Gearbox", "GEARBOX"), ("Travel Wheel", "BEARING")]),
        ("Boom System", [("Boom Conveyor Motor", "MOTOR"), ("Boom Structure", "STRUCTURE")]),
        ("Slew System", [("Slew Motor", "MOTOR"), ("Slew Bearing", "BEARING")]),
    ],
    "RECLAIMER": [
        ("Travel System", [("Travel Motor", "MOTOR"), ("Travel Gearbox", "GEARBOX"), ("Travel Wheel", "BEARING")]),
        ("Bucket Wheel", [("Bucket Wheel Motor", "MOTOR"), ("Bucket", "LINER"), ("Wheel Bearing", "BEARING")]),
        ("Boom System", [("Boom Structure", "STRUCTURE"), ("Boom Hoist Motor", "MOTOR")]),
    ],
    "SILO": [
        ("Discharge System", [("Discharge Gate", "VALVE"), ("Vibrating Feeder", "MOTOR")]),
        ("Instrumentation", [("Level Sensor", "SENSOR"), ("Weight Sensor", "SENSOR")]),
    ],
    "GRANULATOR": [
        ("Drum System", [("Drum Shell", "STRUCTURE"), ("Drum Liner", "LINER"), ("Drum Seal", "SEAL")]),
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX"), ("Girth Gear", "GEAR")]),
        ("Spray System", [("Spray Nozzle", "NOZZLE"), ("Spray Pipe", "PIPE")]),
    ],
    "AMMONIATOR": [
        ("Reactor Drum", [("Drum Shell", "STRUCTURE"), ("Rubber Lining", "LINER"), ("Drum Seal", "SEAL")]),
        ("Drive System", [("Drive Motor", "MOTOR"), ("Main Gearbox", "GEARBOX")]),
        ("Ammonia System", [("Ammonia Valve", "VALVE"), ("Distributor Pipe", "PIPE")]),
    ],
    "SCREEN": [
        ("Vibrating System", [("Vibrator Motor", "MOTOR"), ("Vibrator Bearing", "BEARING"), ("Screen Deck", "SCREEN")]),
        ("Structure", [("Screen Frame", "STRUCTURE"), ("Spring Mounts", "STRUCTURE")]),
    ],
    "CRUSHER": [
        ("Crushing System", [("Crushing Roll", "LINER"), ("Roll Bearing", "BEARING"), ("Roll Shaft", "SHAFT")]),
        ("Drive System", [("Drive Motor", "MOTOR"), ("Crusher Gearbox", "GEARBOX"), ("Motor Coupling", "COUPLING")]),
    ],
    "REACTOR_VESSEL": [
        ("Agitator System", [("Agitator Motor", "MOTOR"), ("Agitator Gearbox", "GEARBOX"), ("Agitator Impeller", "IMPELLER"), ("Agitator Shaft", "SHAFT")]),
        ("Cooling System", [("Cooling Coil", "HEAT_EXCHANGER"), ("Expansion Joint", "SEAL")]),
        ("Structure", [("Vessel Shell", "VESSEL"), ("Rubber Lining", "LINER"), ("Vessel Nozzle", "NOZZLE")]),
    ],
    "AGITATOR": [
        ("Drive System", [("Agitator Motor", "MOTOR"), ("Agitator Gearbox", "GEARBOX")]),
        ("Wet End", [("Agitator Impeller", "IMPELLER"), ("Agitator Shaft", "SHAFT"), ("Shaft Seal", "SEAL")]),
    ],
    "FLASH_COOLER": [
        ("Vessel", [("Flash Vessel", "VESSEL"), ("Rubber Lining", "LINER")]),
        ("Piping", [("Inlet Pipe", "PIPE"), ("Outlet Valve", "VALVE")]),
    ],
    "HEAT_EXCHANGER": [
        ("Tube Bundle", [("Tube Bundle", "HEAT_EXCHANGER"), ("Tube Sheet", "STRUCTURE")]),
        ("Shell Side", [("HX Shell", "VESSEL"), ("Baffle Plates", "STRUCTURE")]),
        ("Gaskets", [("Flange Gasket", "SEAL"), ("Header Gasket", "SEAL")]),
    ],
    "ACID_PUMP": [
        ("Hydraulic End", [("Impeller", "IMPELLER"), ("Casing Liner", "LINER"), ("Wear Ring", "LINER")]),
        ("Drive System", [("Pump Motor", "MOTOR"), ("Motor Coupling", "COUPLING")]),
        ("Sealing System", [("Mechanical Seal", "SEAL"), ("Shaft Sleeve", "SLEEVE")]),
    ],
    "AIR_COMPRESSOR": [
        ("Compression", [("Rotor Element", "IMPELLER"), ("Inlet Valve", "VALVE"), ("Discharge Valve", "VALVE")]),
        ("Drive System", [("Compressor Motor", "MOTOR"), ("Motor Coupling", "COUPLING")]),
        ("Cooling System", [("Aftercooler", "HEAT_EXCHANGER"), ("Cooling Fan", "BLOWER")]),
        ("Lubrication", [("Oil Pump", "PUMP"), ("Oil Filter", "FILTER"), ("Oil Separator", "FILTER")]),
    ],
    "AIR_DRYER": [
        ("Desiccant System", [("Desiccant Bed", "FILTER"), ("Switching Valve", "VALVE")]),
        ("Control System", [("PLC Controller", "ELECTRICAL"), ("Pressure Sensor", "SENSOR")]),
    ],
    "COOLING_TOWER": [
        ("Fan System", [("Fan Motor", "MOTOR"), ("Fan Blades", "BLOWER"), ("Fan Gearbox", "GEARBOX")]),
        ("Water Distribution", [("Distribution Nozzle", "NOZZLE"), ("Distribution Basin", "STRUCTURE")]),
        ("Fill Media", [("Fill Pack", "STRUCTURE"), ("Drift Eliminator", "STRUCTURE")]),
    ],
    "TRANSFORMER": [
        ("Core Assembly", [("Transformer Core", "ELECTRICAL"), ("Primary Winding", "ELECTRICAL"), ("Secondary Winding", "ELECTRICAL")]),
        ("Cooling System", [("Radiator", "HEAT_EXCHANGER"), ("Cooling Fan", "BLOWER"), ("Oil Pump", "PUMP")]),
        ("Protection", [("Buchholz Relay", "SENSOR"), ("PRD Valve", "VALVE"), ("Temperature Gauge", "SENSOR")]),
    ],
    "GENERATOR": [
        ("Engine System", [("Engine Block", "STRUCTURE"), ("Fuel Injection Pump", "PUMP"), ("Exhaust System", "PIPE")]),
        ("Alternator", [("Stator Winding", "ELECTRICAL"), ("Rotor Assembly", "ELECTRICAL"), ("AVR", "ELECTRICAL")]),
        ("Cooling System", [("Radiator", "HEAT_EXCHANGER"), ("Coolant Pump", "PUMP")]),
    ],
}

# Failure mode templates per MI category
# (mechanism, cause, pattern, consequence, detection, downtime_h, strategy, primary_task_type, secondary_task_type)
FM_TEMPLATES = {
    "MOTOR": [
        ("WEARS", "BREAKDOWN_OF_LUBRICATION", "B_AGE", "EVIDENT_OPERATIONAL", "Vibration analysis", 24, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("SHORT_CIRCUITS", "BREAKDOWN_IN_INSULATION", "E_RANDOM", "EVIDENT_SAFETY", "Insulation testing", 48, "CONDITION_BASED", "TEST", "REPLACE"),
        ("OVERHEATS_MELTS", "OVERCURRENT", "E_RANDOM", "EVIDENT_OPERATIONAL", "Thermal imaging", 8, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "GEARBOX": [
        ("WEARS", "BREAKDOWN_OF_LUBRICATION", "B_AGE", "EVIDENT_OPERATIONAL", "Oil analysis", 72, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("CRACKS", "CYCLIC_LOADING", "C_FATIGUE", "EVIDENT_OPERATIONAL", "Vibration analysis", 96, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "COUPLING": [
        ("LOOSES_PRELOAD", "VIBRATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Vibration check", 4, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("WEARS", "RELATIVE_MOVEMENT", "B_AGE", "EVIDENT_NONOPERATIONAL", "Visual inspection", 2, "FIXED_TIME", "REPLACE", None),
    ],
    "GEAR": [
        ("WEARS", "MECHANICAL_OVERLOAD", "B_AGE", "EVIDENT_OPERATIONAL", "Oil analysis + vibration", 96, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "BEARING": [
        ("WEARS", "BREAKDOWN_OF_LUBRICATION", "B_AGE", "EVIDENT_OPERATIONAL", "Vibration analysis", 24, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("OVERHEATS_MELTS", "LACK_OF_LUBRICATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Temperature monitoring", 12, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "IMPELLER": [
        ("SEVERS", "ABRASION", "D_STRESS", "EVIDENT_OPERATIONAL", "Flow/pressure monitoring", 8, "CONDITION_BASED", "INSPECT", "REPAIR"),
        ("CORRODES", "CHEMICAL_ATTACK", "B_AGE", "EVIDENT_OPERATIONAL", "Visual inspection", 8, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "LINER": [
        ("WEARS", "MECHANICAL_OVERLOAD", "B_AGE", "EVIDENT_OPERATIONAL", "Thickness measurement", 72, "FIXED_TIME", "REPLACE", None),
    ],
    "SEAL": [
        ("DEGRADES", "AGE", "B_AGE", "EVIDENT_ENVIRONMENTAL", "Visual inspection for leaks", 4, "FIXED_TIME", "REPLACE", None),
        ("DEGRADES", "CHEMICAL_ATTACK", "B_AGE", "EVIDENT_ENVIRONMENTAL", "Leak detection", 6, "RUN_TO_FAILURE", None, "REPLACE"),
    ],
    "SLEEVE": [
        ("WEARS", "RELATIVE_MOVEMENT", "B_AGE", "EVIDENT_NONOPERATIONAL", "Dimensional measurement", 4, "FIXED_TIME", "REPLACE", None),
    ],
    "BELT": [
        ("WEARS", "RELATIVE_MOVEMENT", "B_AGE", "EVIDENT_OPERATIONAL", "Belt thickness measurement", 8, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("BREAKS_FRACTURE_SEPARATES", "MECHANICAL_OVERLOAD", "E_RANDOM", "EVIDENT_OPERATIONAL", "Visual inspection", 12, "RUN_TO_FAILURE", None, "REPLACE"),
    ],
    "SCREEN": [
        ("WEARS", "IMPACT_SHOCK_LOADING", "D_STRESS", "EVIDENT_OPERATIONAL", "Screen inspection", 4, "FIXED_TIME", "REPLACE", None),
    ],
    "FILTER": [
        ("BLOCKS", "CONTAMINATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Differential pressure", 2, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "VALVE": [
        ("CORRODES", "CORROSIVE_ENVIRONMENT", "B_AGE", "EVIDENT_OPERATIONAL", "Valve stroke test", 4, "CONDITION_BASED", "TEST", "REPLACE"),
        ("BLOCKS", "CONTAMINATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Flow check", 2, "CONDITION_BASED", "INSPECT", "CLEAN"),
    ],
    "SENSOR": [
        ("DRIFTS", "USE", "B_AGE", "HIDDEN_NONSAFETY", "Calibration check", 1, "FIXED_TIME", "CALIBRATE", None),
        ("EXPIRES", "AGE", "B_AGE", "HIDDEN_NONSAFETY", "Functional test", 1, "FAULT_FINDING", "TEST", "REPLACE"),
    ],
    "STRUCTURE": [
        ("CORRODES", "CORROSIVE_ENVIRONMENT", "B_AGE", "EVIDENT_NONOPERATIONAL", "Visual/UT inspection", 0, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "VESSEL": [
        ("CORRODES", "CHEMICAL_ATTACK", "B_AGE", "EVIDENT_ENVIRONMENTAL", "UT thickness measurement", 0, "CONDITION_BASED", "INSPECT", "REPAIR"),
        ("CRACKS", "THERMAL_STRESSES", "C_FATIGUE", "EVIDENT_SAFETY", "NDT inspection", 0, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "PIPE": [
        ("CORRODES", "CORROSIVE_ENVIRONMENT", "B_AGE", "EVIDENT_ENVIRONMENTAL", "UT thickness", 4, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "NOZZLE": [
        ("WEARS", "EXCESSIVE_FLUID_VELOCITY", "D_STRESS", "EVIDENT_OPERATIONAL", "Flow measurement", 2, "FIXED_TIME", "REPLACE", None),
        ("BLOCKS", "CONTAMINATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Flow check", 1, "CONDITION_BASED", "INSPECT", "CLEAN"),
    ],
    "HEAT_EXCHANGER": [
        ("CORRODES", "CHEMICAL_ATTACK", "B_AGE", "EVIDENT_OPERATIONAL", "UT thickness", 8, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("BLOCKS", "CONTAMINATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Pressure drop monitoring", 4, "CONDITION_BASED", "INSPECT", "CLEAN"),
    ],
    "BLOWER": [
        ("WEARS", "RELATIVE_MOVEMENT", "B_AGE", "EVIDENT_OPERATIONAL", "Vibration analysis", 4, "CONDITION_BASED", "INSPECT", "REPLACE"),
        ("LOOSES_PRELOAD", "VIBRATION", "E_RANDOM", "EVIDENT_OPERATIONAL", "Vibration check", 2, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "SHAFT": [
        ("WEARS", "RELATIVE_MOVEMENT", "B_AGE", "EVIDENT_OPERATIONAL", "Runout measurement", 24, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
    "PUMP": [
        ("WEARS", "ENTRAINED_AIR", "E_RANDOM", "EVIDENT_OPERATIONAL", "Pressure/flow monitoring", 4, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "ELECTRICAL": [
        ("SHORT_CIRCUITS", "BREAKDOWN_IN_INSULATION", "E_RANDOM", "EVIDENT_SAFETY", "Insulation resistance test", 8, "CONDITION_BASED", "TEST", "REPLACE"),
        ("THERMALLY_OVERLOADS", "OVERCURRENT", "E_RANDOM", "EVIDENT_OPERATIONAL", "Thermal scan", 4, "CONDITION_BASED", "INSPECT", "REPAIR"),
    ],
    "BURNER": [
        ("DEGRADES", "EXCESSIVE_TEMPERATURE", "B_AGE", "EVIDENT_SAFETY", "Flame monitoring", 8, "CONDITION_BASED", "INSPECT", "REPLACE"),
    ],
}
# Default fallback
FM_TEMPLATES["DEFAULT"] = [
    ("WEARS", "USE", "B_AGE", "EVIDENT_OPERATIONAL", "Visual inspection", 4, "CONDITION_BASED", "INSPECT", "REPAIR"),
]


AREAS = [
    ("BRY", "Grinding", "Broyage", [
        ("SAG", "SAG Mill Circuit", "Circuit Broyeur SAG", [
            ("BRY-SAG-ML-001", "SAG Mill 36x20 Primary #1", "Broyeur SAG 36x20 Primaire #1", "SAG_MILL", "FLSmidth", "SAG 36x20", "SN-SAG-2019-001", 8500, 285000, "AA", "2019-06-15"),
            ("BRY-SAG-ML-002", "SAG Mill 36x20 Primary #2", "Broyeur SAG 36x20 Primaire #2", "SAG_MILL", "FLSmidth", "SAG 36x20", "SN-SAG-2019-002", 8500, 285000, "AA", "2019-08-20"),
        ]),
        ("BML", "Ball Mill Circuit", "Circuit Broyeur a Boulets", [
            ("BRY-BML-ML-001", "Ball Mill 22x36 Secondary #1", "Broyeur a Boulets 22x36 #1", "BALL_MILL", "Metso Outotec", "BM 22x36", "SN-BML-2019-003", 5000, 180000, "AA", "2019-08-25"),
            ("BRY-BML-ML-002", "Ball Mill 22x36 Secondary #2", "Broyeur a Boulets 22x36 #2", "BALL_MILL", "Metso Outotec", "BM 22x36", "SN-BML-2019-004", 5000, 180000, "AA", "2019-10-10"),
        ]),
        ("CYC", "Classification Circuit", "Circuit de Classification", [
            ("BRY-CYC-HC-001", "Hydrocyclone Cluster #1", "Cluster Hydrocyclone #1", "HYDROCYCLONE", "Weir Minerals", "Cavex 650", "SN-CYC-2019-005", 0, 8000, "A", "2019-07-01"),
            ("BRY-CYC-HC-002", "Hydrocyclone Cluster #2", "Cluster Hydrocyclone #2", "HYDROCYCLONE", "Weir Minerals", "Cavex 650", "SN-CYC-2019-006", 0, 8000, "A", "2019-07-01"),
            ("BRY-CYC-PP-001", "Cyclone Feed Pump #1", "Pompe Alimentation Cyclone #1", "SLURRY_PUMP", "Weir Minerals", "Warman 14/12 AH", "SN-CYP-2019-007", 300, 7500, "A+", "2019-07-10"),
            ("BRY-CYC-PP-002", "Cyclone Feed Pump #2", "Pompe Alimentation Cyclone #2", "SLURRY_PUMP", "Weir Minerals", "Warman 14/12 AH", "SN-CYP-2019-008", 300, 7500, "A+", "2019-07-10"),
        ]),
    ]),
    ("FLT", "Flotation", "Flottation", [
        ("RGH", "Rougher Flotation", "Flottation Ebauchage", [
            ("FLT-RGH-FC-001", "Rougher Flotation Cell #1", "Cellule Flottation Ebauchage #1", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-001", 250, 45000, "A+", "2020-03-10"),
            ("FLT-RGH-FC-002", "Rougher Flotation Cell #2", "Cellule Flottation Ebauchage #2", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-002", 250, 45000, "A+", "2020-03-10"),
            ("FLT-RGH-FC-003", "Rougher Flotation Cell #3", "Cellule Flottation Ebauchage #3", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-003", 250, 45000, "A+", "2020-03-15"),
            ("FLT-RGH-FC-004", "Rougher Flotation Cell #4", "Cellule Flottation Ebauchage #4", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-004", 250, 45000, "A+", "2020-03-15"),
            ("FLT-RGH-FC-005", "Rougher Flotation Cell #5", "Cellule Flottation Ebauchage #5", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-005", 250, 45000, "A+", "2020-03-20"),
            ("FLT-RGH-FC-006", "Rougher Flotation Cell #6", "Cellule Flottation Ebauchage #6", "FLOTATION_CELL", "Outotec", "TankCell e500", "SN-RFC-2020-006", 250, 45000, "A+", "2020-03-20"),
        ]),
        ("CLN", "Cleaner Flotation", "Flottation Finissage", [
            ("FLT-CLN-FC-001", "Cleaner Flotation Cell #1", "Cellule Flottation Finissage #1", "FLOTATION_CELL", "Outotec", "TankCell e300", "SN-CFC-2020-007", 150, 30000, "A", "2020-04-01"),
            ("FLT-CLN-FC-002", "Cleaner Flotation Cell #2", "Cellule Flottation Finissage #2", "FLOTATION_CELL", "Outotec", "TankCell e300", "SN-CFC-2020-008", 150, 30000, "A", "2020-04-01"),
            ("FLT-CLN-FC-003", "Cleaner Flotation Cell #3", "Cellule Flottation Finissage #3", "FLOTATION_CELL", "Outotec", "TankCell e300", "SN-CFC-2020-009", 150, 30000, "A", "2020-04-05"),
            ("FLT-CLN-FC-004", "Cleaner Flotation Cell #4", "Cellule Flottation Finissage #4", "FLOTATION_CELL", "Outotec", "TankCell e300", "SN-CFC-2020-010", 150, 30000, "A", "2020-04-05"),
        ]),
        ("REA", "Reagent System", "Systeme Reactifs", [
            ("FLT-REA-DP-001", "Reagent Dosing Pump #1", "Pompe Dosage Reactif #1", "REAGENT_PUMP", "ProMinent", "Sigma X", "SN-RDP-2020-011", 2, 80, "B", "2020-03-01"),
            ("FLT-REA-DP-002", "Reagent Dosing Pump #2", "Pompe Dosage Reactif #2", "REAGENT_PUMP", "ProMinent", "Sigma X", "SN-RDP-2020-012", 2, 80, "B", "2020-03-01"),
            ("FLT-REA-DP-003", "Reagent Dosing Pump #3", "Pompe Dosage Reactif #3", "REAGENT_PUMP", "ProMinent", "Sigma X", "SN-RDP-2020-013", 2, 80, "B", "2020-03-01"),
        ]),
    ]),
    ("SED", "Sedimentation", "Sedimentation", [
        ("THK", "Thickening", "Epaississement", [
            ("SED-THK-TH-001", "High-Rate Thickener 45m #1", "Epaississeur 45m #1", "THICKENER", "FLSmidth", "HRT-45", "SN-THK-2019-001", 45, 250000, "A", "2019-09-01"),
            ("SED-THK-TH-002", "High-Rate Thickener 45m #2", "Epaississeur 45m #2", "THICKENER", "FLSmidth", "HRT-45", "SN-THK-2019-002", 45, 250000, "A", "2019-09-15"),
            ("SED-THK-PP-001", "Thickener Underflow Pump #1", "Pompe Soutirage Epaississeur #1", "SLURRY_PUMP", "Weir Minerals", "Warman 10/8 AH", "SN-TUP-2019-003", 150, 4500, "A+", "2019-09-01"),
            ("SED-THK-PP-002", "Thickener Underflow Pump #2", "Pompe Soutirage Epaississeur #2", "SLURRY_PUMP", "Weir Minerals", "Warman 10/8 AH", "SN-TUP-2019-004", 150, 4500, "A+", "2019-09-15"),
        ]),
    ]),
    ("FIL", "Filtration", "Filtration", [
        ("BLT", "Belt Filtration", "Filtration a Bande", [
            ("FIL-BLT-BF-001", "Belt Filter Press #1", "Filtre Presse a Bande #1", "BELT_FILTER", "ANDRITZ", "BF-3000", "SN-BFP-2020-001", 45, 25000, "A+", "2020-01-15"),
            ("FIL-BLT-BF-002", "Belt Filter Press #2", "Filtre Presse a Bande #2", "BELT_FILTER", "ANDRITZ", "BF-3000", "SN-BFP-2020-002", 45, 25000, "A+", "2020-01-20"),
            ("FIL-BLT-BF-003", "Belt Filter Press #3", "Filtre Presse a Bande #3", "BELT_FILTER", "ANDRITZ", "BF-3000", "SN-BFP-2020-003", 45, 25000, "A+", "2020-02-01"),
        ]),
        ("DSC", "Disc Filtration", "Filtration a Disque", [
            ("FIL-DSC-DF-001", "Disc Filter #1", "Filtre a Disque #1", "DISC_FILTER", "ANDRITZ", "DF-120", "SN-DSC-2020-004", 30, 18000, "A", "2020-02-10"),
            ("FIL-DSC-DF-002", "Disc Filter #2", "Filtre a Disque #2", "DISC_FILTER", "ANDRITZ", "DF-120", "SN-DSC-2020-005", 30, 18000, "A", "2020-02-15"),
        ]),
    ]),
    ("SEQ", "Drying", "Sechage", [
        ("DRY", "Rotary Drying", "Sechage Rotatif", [
            ("SEQ-DRY-RD-001", "Rotary Dryer 4.5m x 30m #1", "Secheur Rotatif 4.5x30 #1", "ROTARY_DRYER", "FLSmidth", "RD 4.5x30", "SN-DRY-2019-001", 500, 180000, "A+", "2019-11-01"),
            ("SEQ-DRY-RD-002", "Rotary Dryer 4.5m x 30m #2", "Secheur Rotatif 4.5x30 #2", "ROTARY_DRYER", "FLSmidth", "RD 4.5x30", "SN-DRY-2019-002", 500, 180000, "A+", "2019-12-01"),
        ]),
        ("CLR", "Product Cooling", "Refroidissement Produit", [
            ("SEQ-CLR-RC-001", "Rotary Cooler 3.5m x 25m", "Refroidisseur Rotatif 3.5x25", "ROTARY_COOLER", "FLSmidth", "RC 3.5x25", "SN-CLR-2019-003", 200, 120000, "A", "2019-12-15"),
        ]),
    ]),
    ("CVY", "Conveying", "Convoyage", [
        ("OLC", "Ore Feed Conveyors", "Convoyeurs Alimentation Minerai", [
            ("CVY-OLC-CV-001", "Overland Conveyor 2000mm x 3.5km", "Convoyeur Terrestre 2000mm x 3.5km", "BELT_CONVEYOR", "Continental", "PIPE2000", "SN-OLC-2020-001", 800, 95000, "A", "2020-01-15"),
            ("CVY-OLC-CV-002", "Transfer Conveyor #1 1400mm", "Convoyeur Transfert #1 1400mm", "BELT_CONVEYOR", "Continental", "CV-1400", "SN-TRF-2020-002", 200, 35000, "A", "2020-02-01"),
            ("CVY-OLC-CV-003", "Transfer Conveyor #2 1400mm", "Convoyeur Transfert #2 1400mm", "BELT_CONVEYOR", "Continental", "CV-1400", "SN-TRF-2020-003", 200, 35000, "A", "2020-02-15"),
        ]),
        ("PRD", "Product Conveyors", "Convoyeurs Produit", [
            ("CVY-PRD-CV-001", "Product Conveyor #1 1000mm", "Convoyeur Produit #1 1000mm", "BELT_CONVEYOR", "Rulmeca", "CV-1000", "SN-PRD-2020-004", 110, 20000, "B", "2020-03-01"),
            ("CVY-PRD-CV-002", "Product Conveyor #2 1000mm", "Convoyeur Produit #2 1000mm", "BELT_CONVEYOR", "Rulmeca", "CV-1000", "SN-PRD-2020-005", 110, 20000, "B", "2020-03-10"),
        ]),
    ]),
    ("STK", "Storage", "Stockage", [
        ("STK", "Ore Storage", "Stockage Minerai", [
            ("STK-STK-SK-001", "Stacker 2000 t/h", "Empileur 2000 t/h", "STACKER", "ThyssenKrupp", "SK-2000", "SN-STK-2020-001", 150, 120000, "B", "2020-05-01"),
            ("STK-RCL-RC-001", "Reclaimer 1500 t/h", "Machine de Reprise 1500 t/h", "RECLAIMER", "ThyssenKrupp", "RC-1500", "SN-RCL-2020-002", 200, 150000, "B", "2020-05-15"),
        ]),
        ("SIL", "Product Silos", "Silos Produit", [
            ("STK-SIL-SI-001", "Product Storage Silo #1 5000t", "Silo Stockage Produit #1 5000t", "SILO", "Local", "SIL-5000", "SN-SIL-2019-001", 5, 80000, "C", "2019-06-01"),
            ("STK-SIL-SI-002", "Product Storage Silo #2 5000t", "Silo Stockage Produit #2 5000t", "SILO", "Local", "SIL-5000", "SN-SIL-2019-002", 5, 80000, "C", "2019-06-01"),
        ]),
    ]),
    ("PMP", "Pumping", "Pompage", [
        ("MDP", "Mill Discharge Pumps", "Pompes Decharge Broyeur", [
            ("PMP-MDP-PP-001", "Slurry Pump 16x14 Mill Discharge #1", "Pompe Boue 16x14 Decharge Broyeur #1", "SLURRY_PUMP", "Weir Minerals", "Warman 16/14 AH", "SN-MDP-2019-001", 350, 8500, "A+", "2019-07-01"),
            ("PMP-MDP-PP-002", "Slurry Pump 16x14 Mill Discharge #2", "Pompe Boue 16x14 Decharge Broyeur #2", "SLURRY_PUMP", "Weir Minerals", "Warman 16/14 AH", "SN-MDP-2019-002", 350, 8500, "A+", "2019-07-15"),
        ]),
        ("PWS", "Process Water System", "Systeme Eau Process", [
            ("PMP-PWS-PP-001", "Process Water Pump #1", "Pompe Eau Process #1", "WATER_PUMP", "KSB", "Omega 150-500", "SN-PWS-2019-003", 75, 2500, "B", "2019-08-01"),
            ("PMP-PWS-PP-002", "Process Water Pump #2", "Pompe Eau Process #2", "WATER_PUMP", "KSB", "Omega 150-500", "SN-PWS-2019-004", 75, 2500, "B", "2019-08-01"),
            ("PMP-PWS-PP-003", "Process Water Pump #3", "Pompe Eau Process #3", "WATER_PUMP", "KSB", "Omega 150-500", "SN-PWS-2019-005", 75, 2500, "B", "2019-08-15"),
        ]),
    ]),
    ("GRN", "Granulation", "Granulation", [
        ("GRD", "Granulation Circuit", "Circuit Granulation", [
            ("GRN-GRD-GD-001", "Granulator Drum 4m x 12m", "Tambour Granulateur 4x12m", "GRANULATOR", "FEECO", "GD 4x12", "SN-GRD-2020-001", 300, 95000, "A+", "2020-06-01"),
            ("GRN-GRD-AG-001", "Ammoniator-Granulator", "Ammoniteur-Granulateur", "AMMONIATOR", "FEECO", "AG 3.5x10", "SN-AMG-2020-002", 250, 80000, "A+", "2020-06-15"),
        ]),
        ("SCR", "Product Sizing", "Calibrage Produit", [
            ("GRN-SCR-SC-001", "Vibrating Screen Double Deck", "Crible Vibrant Double Etage", "SCREEN", "Metso Outotec", "CVB 2060", "SN-SCR-2020-003", 45, 12000, "A", "2020-07-01"),
            ("GRN-SCR-CR-001", "Chain Crusher", "Concasseur a Chaines", "CRUSHER", "Stedman", "Grand Slam", "SN-CRS-2020-004", 110, 15000, "A", "2020-07-01"),
        ]),
    ]),
    ("ACR", "Acid Production", "Production Acide", [
        ("REA", "Phosphoric Acid Reactor", "Reacteur Acide Phosphorique", [
            ("ACR-REA-RV-001", "Reactor Vessel 500m3", "Cuve Reacteur 500m3", "REACTOR_VESSEL", "Prayon", "RV-500", "SN-REA-2019-001", 200, 120000, "AA", "2019-05-01"),
            ("ACR-REA-AG-001", "Reactor Agitator", "Agitateur Reacteur", "AGITATOR", "Lightnin", "A-510", "SN-RAG-2019-002", 150, 8000, "A+", "2019-05-01"),
        ]),
        ("HEX", "Heat Exchange", "Echange Thermique", [
            ("ACR-HEX-FC-001", "Flash Cooler", "Refroidisseur Flash", "FLASH_COOLER", "Prayon", "FC-200", "SN-FLC-2019-003", 0, 15000, "A", "2019-05-15"),
            ("ACR-HEX-HE-001", "Shell & Tube Heat Exchanger", "Echangeur Tubes Calandre", "HEAT_EXCHANGER", "Alfa Laval", "M15-BFG", "SN-HEX-2019-004", 0, 8000, "A", "2019-06-01"),
        ]),
        ("SAC", "Sulfuric Acid System", "Systeme Acide Sulfurique", [
            ("ACR-SAC-PP-001", "Sulfuric Acid Pump", "Pompe Acide Sulfurique", "ACID_PUMP", "KSB", "MegaCPK 125-80", "SN-SAC-2019-005", 90, 3000, "A+", "2019-05-01"),
        ]),
    ]),
    ("UTL", "Utilities", "Utilites", [
        ("AIR", "Compressed Air", "Air Comprime", [
            ("UTL-AIR-AC-001", "Rotary Screw Compressor #1", "Compresseur a Vis #1", "AIR_COMPRESSOR", "Atlas Copco", "GA 250", "SN-AIR-2019-001", 250, 5000, "B", "2019-06-01"),
            ("UTL-AIR-AC-002", "Rotary Screw Compressor #2", "Compresseur a Vis #2", "AIR_COMPRESSOR", "Atlas Copco", "GA 250", "SN-AIR-2019-002", 250, 5000, "B", "2019-06-01"),
            ("UTL-AIR-AD-001", "Heatless Desiccant Dryer", "Secheur Dessiccant", "AIR_DRYER", "Atlas Copco", "CD 800", "SN-ADR-2019-003", 15, 2000, "C", "2019-06-15"),
        ]),
        ("CWT", "Cooling Water", "Eau de Refroidissement", [
            ("UTL-CWT-CT-001", "Cooling Tower 4-Cell", "Tour de Refroidissement 4 Cellules", "COOLING_TOWER", "SPX Cooling", "CT-4000", "SN-CWT-2019-001", 120, 45000, "B", "2019-07-01"),
            ("UTL-CWT-PP-001", "Cooling Water Pump", "Pompe Eau Refroidissement", "WATER_PUMP", "KSB", "Omega 200-400", "SN-CWP-2019-002", 110, 3500, "B", "2019-07-01"),
        ]),
        ("ELC", "Electrical Power", "Alimentation Electrique", [
            ("UTL-ELC-TR-001", "Main Power Transformer 40MVA", "Transformateur Principal 40MVA", "TRANSFORMER", "ABB", "RESIBLOC 40MVA", "SN-TRF-2019-001", 0, 85000, "A", "2019-04-01"),
            ("UTL-ELC-GN-001", "Emergency Diesel Generator 2MW", "Groupe Electrogene Diesel 2MW", "GENERATOR", "Caterpillar", "3516B", "SN-GEN-2019-002", 2000, 25000, "A", "2019-04-15"),
        ]),
    ]),
]


# ══════════════════════════════════════════════════════════════
# DATA BUILDERS — Generate flat rows from hierarchy
# ══════════════════════════════════════════════════════════════

def build_all_data():
    """Build all interconnected data sets from the plant hierarchy."""
    equipment_rows = []   # For T01
    bom_rows = []         # For T01 BOM tab (Levels 4-5-6)
    crit_rows = []        # For T02
    fm_rows = []          # For T03
    task_rows = []        # For T04
    task_labour = []
    task_materials = []
    task_tools = []
    strategy_rows = []    # For T14
    wp_rows = []          # For T05
    spare_rows = []       # For T07

    task_counter = [0]
    strategy_counter = [0]
    wp_counter = [0]
    spare_set = {}  # dedup
    task_info = {}  # task_id -> {name, type, constraint, hours} for WP detail lookup
    wp_task_detail_rows = []  # For T05 second sheet
    sap_counter = [10000000]  # SAP material number counter

    # Criticality score templates by criticality class
    CRIT_SCORES = {
        "AA": {"base": [5, 4, 4, 5, 4, 4, 4, 5, 3, 4, 4], "prob": 4},
        "A+": {"base": [4, 3, 3, 4, 3, 3, 3, 4, 2, 3, 3], "prob": 4},
        "A":  {"base": [3, 3, 3, 4, 3, 2, 3, 3, 2, 3, 3], "prob": 3},
        "B":  {"base": [2, 2, 2, 3, 2, 2, 2, 3, 2, 2, 2], "prob": 3},
        "C":  {"base": [2, 1, 2, 2, 2, 1, 2, 2, 1, 2, 1], "prob": 2},
        "D":  {"base": [1, 1, 1, 2, 1, 1, 1, 1, 1, 1, 1], "prob": 2},
    }

    def next_task_id():
        task_counter[0] += 1
        return f"T-{task_counter[0]:04d}"

    def next_strat_id():
        strategy_counter[0] += 1
        return f"S-{strategy_counter[0]:04d}"

    def next_wp_id():
        wp_counter[0] += 1
        return f"WP-{wp_counter[0]:04d}"

    # Interval templates by strategy+task_type
    def get_interval(strategy_type, task_type, eq_crit):
        if strategy_type == "RUN_TO_FAILURE":
            return None, None, None
        base_map = {
            "INSPECT": (4, None, "WEEKS"),
            "CHECK": (1, None, "WEEKS"),
            "TEST": (12, None, "WEEKS"),
            "LUBRICATE": (2, None, "WEEKS"),
            "CLEAN": (4, None, "WEEKS"),
            "CALIBRATE": (26, None, "WEEKS"),
            "REPLACE": (8000, "OPERATING_HOURS", None),
            "REPAIR": (4000, "OPERATING_HOURS", None),
        }
        val, op_unit, time_unit = base_map.get(task_type, (4, None, "WEEKS"))
        # Reduce interval for higher criticality
        if eq_crit in ("AA", "A+") and time_unit == "WEEKS":
            val = max(1, val // 2)
        return val, op_unit, time_unit

    # Acceptable limits templates
    LIMITS = {
        "MOTOR": {"INSPECT": "Vibration < 4.5 mm/s RMS, Temperature < 80C", "TEST": "Insulation resistance > 5 MOhm"},
        "GEARBOX": {"INSPECT": "Oil Fe < 100 ppm, Vibration < 3 mm/s"},
        "BEARING": {"INSPECT": "Vibration < 4.5 mm/s, Temperature < 85C"},
        "IMPELLER": {"INSPECT": "OD within 5mm of nominal, no visible cracking"},
        "VALVE": {"TEST": "Valve opens/closes within 5 seconds, no leakage"},
        "SENSOR": {"CALIBRATE": "Reading within +/- 2% of reference"},
        "VESSEL": {"INSPECT": "Wall thickness > 6mm minimum"},
        "HEAT_EXCHANGER": {"INSPECT": "Tube wall > 1.5mm, pressure drop < 0.5 bar"},
        "DEFAULT": {"INSPECT": "Within OEM specifications", "TEST": "Function verified OK"},
    }

    for area_code, area_en, area_fr, systems in AREAS:
        for sys_code, sys_en, sys_fr, equipment_list in systems:
            for eq_tag, eq_desc, eq_desc_fr, eq_type, mfr, model, serial, power, weight, crit, install in equipment_list:
                sap_fl = f"JFC1-MIN-{area_code}-{sys_code}-{eq_tag.split('-')[-1]}"

                # T01 row
                equipment_rows.append([
                    PLANT["plant_id"], PLANT["plant_name"], area_code, f"{area_fr} ({area_en})",
                    sys_code, sys_en, eq_tag, eq_desc, eq_type,
                    mfr, model, serial, power, weight, crit, "ACTIVE", sap_fl, install,
                ])

                # T02 row
                scores = CRIT_SCORES.get(crit, CRIT_SCORES["B"])
                varied = [min(5, max(1, s + random.choice([-1, 0, 0, 0, 1]))) for s in scores["base"]]
                crit_rows.append([eq_tag, "FULL_MATRIX"] + varied + [scores["prob"]])

                # Get BOM for this equipment type
                bom = EQUIPMENT_BOM.get(eq_type, [("Main Assembly", [("Main Component", "DEFAULT")])])
                eq_tasks_by_freq = {}  # (freq_val, freq_unit, constraint) -> [task_ids]

                # BOM rows: Level 4 (equipment) → Level 5 (sub-assembly) → Level 6 (MI)
                bom_rows.append([
                    eq_tag, 4, "EQUIPMENT", eq_tag, eq_desc, eq_desc_fr,
                    None, 0, None, mfr,
                ])
                for sa_idx, (sa_name, mi_list) in enumerate(bom, 1):
                    sa_tag = f"{eq_tag}-{sa_name[:3].upper()}"
                    bom_rows.append([
                        eq_tag, 5, "SUB_ASSEMBLY", sa_tag, sa_name, sa_name,
                        eq_tag, sa_idx, None, None,
                    ])
                    for mi_idx, (mi_name, mi_cat) in enumerate(mi_list, 1):
                        mi_tag = f"{sa_tag}-{mi_name[:3].upper()}{mi_idx:02d}"
                        bom_rows.append([
                            eq_tag, 6, "MAINTAINABLE_ITEM", mi_tag, mi_name, mi_name,
                            sa_tag, mi_idx, mi_cat, None,
                        ])

                for sa_name, mi_list in bom:
                    for mi_name, mi_cat in mi_list:
                        # Get failure modes for this MI category
                        fms = FM_TEMPLATES.get(mi_cat, FM_TEMPLATES["DEFAULT"])
                        for mech, cause, pattern, consequence, detection, dt_h, strat_type, pri_tt, sec_tt in fms:
                            # T03 row — failure mode
                            func_desc = f"Maintain {eq_desc} operational performance"
                            func_type = "PROTECTIVE" if consequence.startswith("HIDDEN") else "PRIMARY"
                            fail_type = random.choice(["TOTAL", "PARTIAL"])
                            rpn_s = random.randint(4, 9)
                            rpn_o = random.randint(2, 7)
                            rpn_d = random.randint(2, 6)
                            fm_rows.append([
                                eq_tag, func_desc, func_type, fail_type,
                                mi_name, mech, cause, pattern,
                                consequence, detection, dt_h,
                                detection, rpn_s, rpn_o, rpn_d,
                            ])

                            # Build tasks
                            pri_task_id = None
                            pri_task_name = None
                            sec_task_id = None
                            sec_task_name = None

                            if pri_tt and strat_type != "RUN_TO_FAILURE":
                                pri_task_id = next_task_id()
                                interval, op_u, time_u = get_interval(strat_type, pri_tt, crit)
                                constraint = "TEST_MODE" if pri_tt == "TEST" else "ONLINE"
                                access_t = 0.5 if constraint == "TEST_MODE" else 0
                                task_name = f"{pri_tt.title()} {mi_name} on {eq_tag}"
                                task_name_fr = f"{pri_tt.title()} {mi_name} sur {eq_tag}"
                                pri_task_name = task_name

                                # Acceptable limits and conditional comments (CB/FFI tasks)
                                limits_dict = LIMITS.get(mi_cat, LIMITS["DEFAULT"])
                                acc_limits = limits_dict.get(pri_tt, limits_dict.get("INSPECT"))
                                cond_comments = f"If outside limits: schedule {mi_name} replacement/repair" if acc_limits and strat_type in ("CONDITION_BASED", "FAULT_FINDING") else None
                                if strat_type not in ("CONDITION_BASED", "FAULT_FINDING"):
                                    acc_limits = None

                                task_rows.append([
                                    pri_task_id, task_name, task_name_fr, pri_tt,
                                    constraint, access_t, acc_limits, cond_comments,
                                    None, None, None, None,
                                    f"Undetected degradation of {mi_name}",
                                    f"RCM: {strat_type} for {mi_cat}", "R8_LIBRARY", None,
                                ])
                                task_info[pri_task_id] = {"name": task_name, "type": pri_tt, "constraint": constraint, "hours": 0.5}

                                # Labour for primary task
                                spec = "CONMON_SPECIALIST" if pri_tt in ("INSPECT", "CHECK") else (
                                    "INSTRUMENTIST" if pri_tt in ("TEST", "CALIBRATE") else (
                                    "LUBRICATOR" if pri_tt == "LUBRICATE" else "FITTER"))
                                hrs = 0.5 if pri_tt in ("INSPECT", "CHECK", "LUBRICATE", "CALIBRATE") else 1.0
                                lr_id = f"LR-{len(task_labour)+1:04d}"
                                task_labour.append([lr_id, pri_task_id, None, spec, 1, hrs, None, "OCP"])

                                # Group for WP
                                key = (interval, time_u or op_u, constraint, eq_tag)
                                eq_tasks_by_freq.setdefault(key, []).append(pri_task_id)

                            if sec_tt:
                                sec_task_id = next_task_id()
                                sec_constraint = "OFFLINE"
                                sec_access = max(dt_h * 0.5, 2) if dt_h > 0 else 2
                                budget = "REPLACE" if sec_tt == "REPLACE" else "REPAIR"
                                task_name = f"{sec_tt.title()} {mi_name} on {eq_tag}"
                                task_name_fr = f"{sec_tt.title()} {mi_name} sur {eq_tag}"
                                sec_task_name = task_name

                                task_rows.append([
                                    sec_task_id, task_name, task_name_fr, sec_tt,
                                    sec_constraint, sec_access, None, None, budget,
                                    random.choice([2, 3, 5]) if sec_tt == "REPLACE" else None,
                                    "YEARS" if sec_tt == "REPLACE" else None,
                                    "OPERATING_HOURS" if sec_tt == "REPLACE" else None,
                                    f"Failure of {mi_name} causing downtime",
                                    f"Secondary: triggered by condition on {mi_name}",
                                    "RCM_ANALYSIS", None,
                                ])
                                task_info[sec_task_id] = {"name": task_name, "type": sec_tt, "constraint": sec_constraint, "hours": round(sec_access, 1)}

                                # Labour for secondary
                                sec_spec = "ELECTRICIAN" if mi_cat in ("MOTOR", "ELECTRICAL") else "FITTER"
                                sec_qty = 2 if dt_h > 12 else 1
                                sec_hrs = max(sec_access, 2)
                                lr_id2 = f"LR-{len(task_labour)+1:04d}"
                                task_labour.append([lr_id2, sec_task_id, None, sec_spec, sec_qty, sec_hrs, None, "OCP"])

                                # Material for secondary REPLACE tasks
                                if sec_tt == "REPLACE":
                                    mat_code = f"MAT-{mi_cat[:4]}-{mi_name.upper().replace(' ', '-')[:15]}"
                                    if mat_code not in spare_set:
                                        price = random.choice([50, 200, 500, 1500, 3000, 8000, 12000])
                                        sap_mat = f"{sap_counter[0]:010d}"
                                        sap_counter[0] += 1
                                        spare_set[mat_code] = {
                                            "desc": f"Replacement {mi_name}",
                                            "mfr": mfr,
                                            "pn": f"{mfr[:3].upper()}-{mi_name[:8].upper().replace(' ', '')}",
                                            "price": price,
                                            "equip": [],
                                            "sap_code": sap_mat,
                                        }
                                    spare_set[mat_code]["equip"].append(eq_tag)
                                    mr_id = f"MR-{len(task_materials)+1:04d}"
                                    task_materials.append([
                                        mr_id, sec_task_id, mat_code, spare_set[mat_code]["sap_code"],
                                        f"Replacement {mi_name}",
                                        mfr, spare_set[mat_code]["pn"], 1, "EA",
                                        spare_set[mat_code]["price"], None,
                                    ])

                            # T14 — Strategy row
                            strat_id = next_strat_id()
                            interval_val, op_u, time_u = (None, None, None)
                            limits_text = None
                            if pri_task_id:
                                interval_val, op_u, time_u = get_interval(strat_type, pri_tt, crit)
                                limits_dict = LIMITS.get(mi_cat, LIMITS["DEFAULT"])
                                limits_text = limits_dict.get(pri_tt, limits_dict.get("INSPECT", "Within OEM specifications"))

                            strategy_rows.append([
                                strat_id, eq_tag, f"{sa_name} — {mi_name}",
                                f"{consequence.split('_')[-1]}-{func_desc[:40]}-{mi_name} fails-{func_type}",
                                mi_name, mech, cause,
                                "RECOMMENDED", strat_type,
                                pri_task_id, pri_task_name, interval_val, op_u, time_u,
                                limits_text if strat_type in ("CONDITION_BASED", "FAULT_FINDING") else None,
                                None,
                                "ONLINE" if pri_task_id and pri_tt not in ("TEST",) else ("TEST_MODE" if pri_tt == "TEST" else None),
                                pri_tt, 0 if pri_task_id else None,
                                sec_task_id, sec_task_name, "OFFLINE" if sec_task_id else None,
                                sec_tt, max(dt_h * 0.5, 2) if sec_task_id and dt_h > 0 else (2 if sec_task_id else None),
                                None,
                                "REPLACE" if sec_tt == "REPLACE" else ("REPAIR" if sec_tt == "REPAIR" else None),
                                random.choice([2, 3, 5]) if sec_tt == "REPLACE" else None,
                                "YEARS" if sec_tt == "REPLACE" else None,
                                "OPERATING_HOURS" if sec_tt == "REPLACE" else None,
                                "R8_LIBRARY", None, None, None,
                            ])

                # Build WPs from grouped tasks
                EXEC_ORDER = {"CHECK": 1, "INSPECT": 2, "LUBRICATE": 3, "CLEAN": 4, "CALIBRATE": 5, "TEST": 6, "REPLACE": 7, "REPAIR": 8}
                for (freq_val, freq_unit, constraint, etag), task_ids in eq_tasks_by_freq.items():
                    if not freq_val or not task_ids:
                        continue
                    wp_id = next_wp_id()
                    wp_code = f"WP-{etag}-{freq_val}{(freq_unit or 'W')[:1]}"
                    wp_name = f"{freq_val}{(freq_unit or 'W')[:1]} {etag} {constraint[:3]} INSP"
                    if len(wp_name) > 40:
                        wp_name = wp_name[:40]
                    est_hours = len(task_ids) * 0.5
                    crew = max(1, len(task_ids) // 3)
                    wp_rows.append([
                        wp_name.upper(), wp_code, etag, freq_val, freq_unit or "WEEKS",
                        constraint, "STANDALONE", 0 if constraint == "ONLINE" else 2,
                        ",".join(task_ids[:10]), round(est_hours, 1), crew,
                    ])

                    # WP Task Details — sorted by execution sequence
                    sorted_tids = sorted(
                        task_ids[:10],
                        key=lambda tid: EXEC_ORDER.get(task_info.get(tid, {}).get("type", "INSPECT"), 5),
                    )
                    for seq, tid in enumerate(sorted_tids, 1):
                        ti = task_info.get(tid, {})
                        wp_task_detail_rows.append([
                            wp_code, wp_name.upper(), etag, seq, tid,
                            ti.get("name", ""), ti.get("type", ""),
                            ti.get("constraint", ""), ti.get("hours", 0.5),
                        ])

    # Build spare parts (T07)
    for mat_code, info in spare_set.items():
        ved = "VITAL" if info["price"] > 5000 else ("ESSENTIAL" if info["price"] > 500 else "DESIRABLE")
        fsn = random.choice(["FAST_MOVING", "SLOW_MOVING", "SLOW_MOVING"])
        abc = "A_HIGH" if info["price"] > 3000 else ("B_MEDIUM" if info["price"] > 200 else "C_LOW")
        qty = random.randint(1, 10)
        spare_rows.append([
            mat_code, info["sap_code"], info["desc"], info["mfr"], info["pn"],
            ved, fsn, abc,
            qty, max(1, qty // 2), qty * 3, qty,
            random.choice([30, 60, 90, 120]), info["price"], "EA",
            ",".join(list(set(info["equip"]))[:5]),
            f"WH-01-{random.choice('ABCDE')}{random.randint(1,9)}-R{random.randint(1,9)}",
        ])

    return {
        "equipment": equipment_rows,
        "bom": bom_rows,
        "criticality": crit_rows,
        "failure_modes": fm_rows,
        "tasks": task_rows,
        "task_labour": task_labour,
        "task_materials": task_materials,
        "task_tools": task_tools,
        "strategies": strategy_rows,
        "work_packages": wp_rows,
        "wp_task_details": wp_task_detail_rows,
        "spare_parts": spare_rows,
    }


# ══════════════════════════════════════════════════════════════
# TEMPLATE GENERATORS
# ══════════════════════════════════════════════════════════════

def gen_01(data):
    wb = Workbook(); ws = wb.active; ws.title = "Equipment Hierarchy"
    headers = [
        "plant_id", "plant_name", "area_code", "area_name", "system_code",
        "system_name", "equipment_tag", "equipment_description", "equipment_type",
        "manufacturer", "model", "serial_number", "power_kw", "weight_kg",
        "criticality", "status", "sap_func_loc", "installation_date",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["equipment"])
    _add_data_validation(ws, 15, CRITICALITIES)
    _add_data_validation(ws, 16, EQUIP_STATUSES)
    _auto_size_columns(ws)

    # BOM tab: Level 4 (Equipment) → Level 5 (Sub-Assembly) → Level 6 (Maintainable Item)
    ws2 = wb.create_sheet("Equipment BOM")
    bom_headers = [
        "equipment_tag", "level", "node_type", "code", "name", "name_fr",
        "parent_code", "order", "mi_category", "manufacturer",
    ]
    _hdr(ws2, bom_headers)
    _write_data_rows(ws2, data["bom"])
    _add_data_validation(ws2, 3, ["EQUIPMENT", "SUB_ASSEMBLY", "MAINTAINABLE_ITEM"])
    _auto_size_columns(ws2)

    wb.save(OUTPUT_DIR / "01_equipment_hierarchy.xlsx")
    return len(data["equipment"])


def gen_02(data):
    wb = Workbook(); ws = wb.active; ws.title = "Criticality Assessment"
    headers = ["equipment_tag", "method"] + [c.lower() for c in CRIT_CATEGORIES] + ["probability"]
    _hdr(ws, headers)
    _write_data_rows(ws, data["criticality"])
    _add_data_validation(ws, 2, CRIT_METHODS)
    for ci in range(3, 14):
        _add_data_validation(ws, ci, ["1", "2", "3", "4", "5"])
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "02_criticality_assessment.xlsx")
    return len(data["criticality"])


def gen_03(data):
    wb = Workbook(); ws = wb.active; ws.title = "Failure Modes"
    headers = [
        "equipment_tag", "function_description", "function_type", "failure_type",
        "what_component", "mechanism", "cause", "failure_pattern",
        "failure_consequence", "evidence", "downtime_hours",
        "detection_method", "rpn_severity", "rpn_occurrence", "rpn_detection",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["failure_modes"])
    _add_data_validation(ws, 3, FUNCTION_TYPES)
    _add_data_validation(ws, 4, FAILURE_TYPES)
    _add_data_validation(ws, 6, MECHANISMS)
    _add_data_validation(ws, 7, CAUSES)
    _add_data_validation(ws, 8, FAILURE_PATTERNS)
    _add_data_validation(ws, 9, FAILURE_CONSEQUENCES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "03_failure_modes.xlsx")
    return len(data["failure_modes"])


def gen_04(data):
    wb = Workbook(); ws = wb.active; ws.title = "Tasks"
    headers = [
        "task_id", "task_name", "task_name_fr", "task_type",
        "constraint", "access_time_hours",
        "acceptable_limits", "conditional_comments",
        "budgeted_as", "budgeted_life", "budgeted_life_time_units",
        "budgeted_life_operational_units",
        "consequences", "justification", "origin", "notes",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["tasks"])
    _add_data_validation(ws, 4, TASK_TYPES)
    _add_data_validation(ws, 5, TASK_CONSTRAINTS)
    _add_data_validation(ws, 9, BUDGETED_AS_VALUES)
    _auto_size_columns(ws)

    # Labour sheet
    ws2 = wb.create_sheet("Task_Labour")
    lh = ["labour_id", "task_id", "worker_id", "specialty", "quantity", "hours_per_person", "hourly_rate_usd", "company"]
    _hdr(ws2, lh)
    _write_data_rows(ws2, data["task_labour"])
    _add_data_validation(ws2, 4, LABOUR_SPECIALTIES)
    _auto_size_columns(ws2)

    # Materials sheet
    ws3 = wb.create_sheet("Task_Materials")
    mh = ["material_line_id", "task_id", "material_code", "sap_material_number", "description", "manufacturer", "part_number", "quantity", "unit_of_measure", "unit_price_usd", "equipment_bom_ref"]
    _hdr(ws3, mh)
    _write_data_rows(ws3, data["task_materials"])
    _add_data_validation(ws3, 9, UNITS_OF_MEASURE)
    _auto_size_columns(ws3)

    wb.save(OUTPUT_DIR / "04_maintenance_tasks.xlsx")
    return len(data["tasks"])


def gen_05(data):
    wb = Workbook(); ws = wb.active; ws.title = "Work Packages"
    headers = [
        "wp_name", "wp_code", "equipment_tag", "frequency_value", "frequency_unit",
        "constraint", "wp_type", "access_time_hours", "task_ids_csv",
        "estimated_total_hours", "crew_size",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["work_packages"])
    _add_data_validation(ws, 5, FREQUENCY_UNITS)
    _add_data_validation(ws, 6, WP_CONSTRAINTS)
    _add_data_validation(ws, 7, WP_TYPES)
    _auto_size_columns(ws)

    # WP Task Details sheet — individual tasks with execution sequence
    ws2 = wb.create_sheet("WP Task Details")
    detail_headers = [
        "wp_code", "wp_name", "equipment_tag", "execution_sequence",
        "task_id", "task_name", "task_type", "constraint", "estimated_hours",
    ]
    _hdr(ws2, detail_headers)
    _write_data_rows(ws2, data["wp_task_details"])
    _add_data_validation(ws2, 7, TASK_TYPES)
    _add_data_validation(ws2, 8, TASK_CONSTRAINTS)
    _auto_size_columns(ws2)

    wb.save(OUTPUT_DIR / "05_work_packages.xlsx")
    return len(data["work_packages"])


def gen_06(data):
    """Work Order History — 24 months for all equipment."""
    wb = Workbook(); ws = wb.active; ws.title = "Work Order History"
    headers = [
        "wo_id", "order_type", "equipment_tag", "sap_func_loc", "priority",
        "description", "created_date", "planned_start", "planned_end",
        "actual_start", "actual_end", "duration_hours", "status",
        "cost_labour_usd", "cost_materials_usd", "failure_found",
    ]
    _hdr(ws, headers)

    rows = []
    today = date.today()
    wo_num = 100000
    for eq_row in data["equipment"]:
        eq_tag = eq_row[6]
        sap_fl = eq_row[16]
        crit = eq_row[14]
        # More WOs for higher criticality
        num_wo = {"AA": 24, "A+": 18, "A": 12, "B": 8, "C": 4, "D": 2}.get(crit, 6)
        for i in range(num_wo):
            wo_num += 1
            days_ago = random.randint(1, 730)
            d = today - timedelta(days=days_ago)
            ot = random.choice(["PM02", "PM02", "PM02", "PM03", "PM01"])
            pri = random.choice(["4_PLANNED", "4_PLANNED", "3_NORMAL", "3_NORMAL", "2_URGENT"])
            dur = random.choice([0.5, 1, 2, 4, 8, 12, 24, 48])
            st = "COMPLETED" if days_ago > 14 else random.choice(["COMPLETED", "RELEASED", "IN_PROGRESS"])
            cost_l = round(dur * random.randint(80, 150), 0)
            cost_m = round(random.choice([0, 0, 0, 200, 500, 1500, 5000, 12000]), 0) if ot == "PM03" else 0
            fail = random.choice([None, None, None, "Bearing wear detected", "Seal leak", "Vibration high", "Corrosion found"])

            rows.append([
                f"WO-{d.year}-{wo_num}", ot, eq_tag, sap_fl, pri,
                f"{ot} maintenance on {eq_tag}", str(d), str(d), str(d + timedelta(days=1)),
                str(d) if st in ("COMPLETED", "IN_PROGRESS") else None,
                str(d + timedelta(hours=int(dur))) if st == "COMPLETED" else None,
                dur if st == "COMPLETED" else None, st,
                cost_l if st == "COMPLETED" else None,
                cost_m if st == "COMPLETED" else None,
                fail,
            ])

    _write_data_rows(ws, rows)
    _add_data_validation(ws, 2, ORDER_TYPES)
    _add_data_validation(ws, 5, PRIORITIES)
    _add_data_validation(ws, 13, WO_STATUSES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "06_work_order_history.xlsx")
    return len(rows)


def gen_07(data):
    wb = Workbook(); ws = wb.active; ws.title = "Spare Parts Inventory"
    headers = [
        "material_code", "sap_material_number", "description", "manufacturer", "part_number",
        "ved_class", "fsn_class", "abc_class",
        "quantity_on_hand", "min_stock", "max_stock", "reorder_point",
        "lead_time_days", "unit_cost_usd", "unit_of_measure",
        "applicable_equipment_csv", "warehouse_location",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["spare_parts"])
    _add_data_validation(ws, 6, VED_CLASSES)
    _add_data_validation(ws, 7, FSN_CLASSES)
    _add_data_validation(ws, 8, ABC_CLASSES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "07_spare_parts_inventory.xlsx")
    return len(data["spare_parts"])


def gen_08(data):
    wb = Workbook(); ws = wb.active; ws.title = "Shutdown Calendar"
    headers = [
        "plant_id", "shutdown_name", "shutdown_type",
        "planned_start", "planned_end", "planned_hours",
        "areas_csv", "description", "work_packages_csv",
    ]
    _hdr(ws, headers)
    today = date.today()

    # Equipment tags per area for WP generation
    area_equip = {}
    for eq_row in data["equipment"]:
        area = eq_row[2]  # area_code
        tag = eq_row[6]   # equipment_tag
        area_equip.setdefault(area, []).append(tag)

    shutdowns = [
        ("BRY Minor Shutdown Q1", "MINOR_8H", 30, 0, 8, "BRY", "SAG mill bearing inspections and greasing"),
        ("FLT Minor Shutdown Q1", "MINOR_8H", 45, 0, 8, "FLT", "Flotation cell agitator and launder inspection"),
        ("CVY Minor Shutdown Q1", "MINOR_8H", 50, 0, 8, "CVY", "Belt splice and idler roller inspection"),
        ("PMP Minor Shutdown Q1", "MINOR_8H", 55, 0, 8, "PMP", "Pump mechanical seal and impeller check"),
        ("SED Minor Shutdown Q1", "MINOR_8H", 60, 0, 8, "SED", "Thickener rake arm and drive inspection"),
        ("ACR Minor Shutdown Q1", "MINOR_8H", 70, 0, 8, "ACR", "Reactor lining and agitator inspection"),
        ("BRY Major Shutdown Annual", "MAJOR_20H_PLUS", 90, 3, 72, "BRY,PMP", "SAG liner change + mill bearing replacement + pump overhaul"),
        ("FIL Minor Shutdown Q2", "MINOR_8H", 110, 0, 8, "FIL", "Filter belt tracking and vacuum pump service"),
        ("GRN Minor Shutdown Q2", "MINOR_8H", 120, 0, 8, "GRN", "Granulator drum liner and screen inspection"),
        ("SEQ Major Shutdown", "MAJOR_20H_PLUS", 130, 3, 72, "SEQ", "Dryer refractory reline and girth gear inspection"),
        ("UTL Minor Shutdown Q2", "MINOR_8H", 140, 0, 8, "UTL", "Compressor valve overhaul and cooling tower service"),
        ("STK Minor Shutdown Q2", "MINOR_8H", 150, 0, 8, "STK", "Stacker-reclaimer boom and slewing gear inspection"),
        ("FLT Minor Shutdown Q2", "MINOR_8H", 155, 0, 8, "FLT", "Cleaner flotation cell overhaul"),
        ("CVY Minor Shutdown Q2", "MINOR_8H", 160, 0, 8, "CVY", "Conveyor drive gearbox oil change and alignment"),
        ("Plant-Wide Major Shutdown", "MAJOR_20H_PLUS", 180, 7, 168, "BRY,FLT,PMP,CVY,SED,FIL,SEQ,GRN,ACR,UTL,STK", "Annual plant turnaround — all areas"),
        ("BRY Minor Shutdown Q3", "MINOR_8H", 210, 0, 8, "BRY", "Post-turnaround vibration baseline and alignment verify"),
        ("PMP Minor Shutdown Q3", "MINOR_8H", 220, 0, 8, "PMP", "Pump performance test and seal inspection"),
        ("ACR Minor Shutdown Q3", "MINOR_8H", 240, 0, 8, "ACR", "Acid pump seal check and reactor vessel inspection"),
        ("FLT Minor Shutdown Q3", "MINOR_8H", 260, 0, 8, "FLT", "Flotation air blower and cell level control check"),
        ("SEQ Minor Shutdown Q3", "MINOR_8H", 270, 0, 8, "SEQ", "Dryer seal inspection and kiln alignment check"),
    ]

    # Build shutdown rows and WP detail rows
    sd_rows = []
    sd_wp_rows = []  # For second sheet
    sd_wp_counter = 0

    # WP task templates per area
    AREA_TASKS = {
        "BRY": [("Inspect SAG mill bearings", 4, 2), ("Grease SAG mill pinion", 2, 1), ("Check ball mill liner wear", 3, 2), ("Inspect cyclone cluster", 2, 1)],
        "FLT": [("Inspect flotation agitator", 3, 2), ("Check launder condition", 2, 1), ("Service air blower", 4, 2), ("Inspect cell level control", 2, 1)],
        "CVY": [("Inspect belt splice", 3, 2), ("Replace idler rollers", 4, 2), ("Check belt tracking", 2, 1), ("Lubricate drive components", 2, 1)],
        "PMP": [("Replace mechanical seal", 6, 2), ("Inspect impeller wear", 4, 2), ("Check bearing vibration", 2, 1), ("Test pump performance", 3, 1)],
        "SED": [("Inspect rake arm", 4, 2), ("Check drive torque", 2, 1), ("Inspect underflow pump", 3, 2)],
        "FIL": [("Inspect filter belt", 3, 2), ("Service vacuum pump", 4, 2), ("Check belt tracking system", 2, 1)],
        "SEQ": [("Inspect dryer refractory", 8, 3), ("Check girth gear mesh", 4, 2), ("Inspect riding ring", 3, 2), ("Service seal assembly", 4, 2)],
        "GRN": [("Inspect drum liner", 4, 2), ("Check screen condition", 3, 2), ("Inspect scraper blades", 2, 1)],
        "ACR": [("Inspect reactor lining", 6, 3), ("Check agitator shaft", 4, 2), ("Inspect acid pump seals", 3, 2)],
        "UTL": [("Service compressor valves", 4, 2), ("Inspect cooling tower fill", 3, 2), ("Check transformer oil", 2, 1), ("Test emergency generator", 3, 2)],
        "STK": [("Inspect boom structure", 4, 2), ("Check slewing gear", 3, 2), ("Lubricate travel wheels", 2, 1)],
    }

    for sd_name, sd_type, start_offset, dur_days, hours, areas_csv, desc in shutdowns:
        start = today + timedelta(days=start_offset)
        end = start + timedelta(days=dur_days) if dur_days > 0 else start

        # Generate WPs for this shutdown
        wp_codes = []
        for area in areas_csv.split(","):
            area = area.strip()
            tasks = AREA_TASKS.get(area, [("General inspection", 2, 1)])
            equip_list = area_equip.get(area, [])
            # Pick 2-5 tasks depending on shutdown type
            n_tasks = min(len(tasks), 5 if sd_type == "MAJOR_20H_PLUS" else 3)
            for task_desc, est_h, crew in tasks[:n_tasks]:
                sd_wp_counter += 1
                wp_code = f"SD-WP-{sd_wp_counter:03d}"
                wp_codes.append(wp_code)
                eq_tag = equip_list[sd_wp_counter % len(equip_list)] if equip_list else f"{area}-GEN-001"
                priority = "1_CRITICAL" if sd_type == "MAJOR_20H_PLUS" else "2_HIGH"
                sd_wp_rows.append([
                    sd_name, wp_code, eq_tag, task_desc, est_h, crew, priority,
                ])

        sd_rows.append([
            "OCP-JFC1", sd_name, sd_type, str(start), str(end), hours,
            areas_csv, desc, ",".join(wp_codes),
        ])

    _write_data_rows(ws, sd_rows)
    _add_data_validation(ws, 3, SHUTDOWN_TYPES)
    _auto_size_columns(ws)

    # Shutdown Work Packages detail sheet
    ws2 = wb.create_sheet("Shutdown Work Packages")
    wp_headers = [
        "shutdown_name", "wp_code", "equipment_tag", "task_description",
        "estimated_hours", "crew_size", "priority",
    ]
    _hdr(ws2, wp_headers)
    _write_data_rows(ws2, sd_wp_rows)
    _auto_size_columns(ws2)

    wb.save(OUTPUT_DIR / "08_shutdown_calendar.xlsx")
    return len(sd_rows)


def gen_09(data):
    wb = Workbook(); ws = wb.active; ws.title = "Workforce"
    headers = ["worker_id", "name", "specialty", "shift", "plant_id", "available", "certifications_csv", "phone", "email"]
    _hdr(ws, headers)
    workers = [
        ("TEC-001", "Ahmed Benali", "FITTER", "MORNING", True, "Mechanical Fitting L3,Crane Op"),
        ("TEC-002", "Youssef El Amrani", "ELECTRICIAN", "MORNING", True, "HV Switching,Electrical Install"),
        ("TEC-003", "Fatima Zahra Idrissi", "INSTRUMENTIST", "MORNING", True, "ISA Certified,PLC Programming"),
        ("TEC-004", "Mohammed Tazi", "FITTER", "AFTERNOON", True, "Welding ASME IX,Rigging"),
        ("TEC-005", "Rachid Ouazzani", "LUBRICATOR", "MORNING", True, "ICML MLT-I,Oil Analysis"),
        ("TEC-006", "Karim Bouazza", "OPERATOR", "MORNING", True, "SAG Mill Operation,Process Ctrl"),
        ("TEC-007", "Hassan Chaoui", "FITTER", "NIGHT", True, "Pump Overhaul,Alignment"),
        ("TEC-008", "Omar Fassi", "ELECTRICIAN", "AFTERNOON", True, "MV Motors,VFD Commissioning"),
        ("TEC-009", "Abdellah Rhazi", "CONMON_SPECIALIST", "MORNING", True, "Vibration Cat III,Thermo L2"),
        ("TEC-010", "Samir Benjelloun", "FITTER", "MORNING", True, "Conveyor Systems,Belt Splicing"),
        ("TEC-011", "Nadia Alaoui", "INSTRUMENTIST", "AFTERNOON", True, "DCS Configuration,HART"),
        ("TEC-012", "Mustapha Kettani", "FITTER", "MORNING", True, "Hydraulics,Pneumatics"),
        ("TEC-013", "Khalid Berrada", "ELECTRICIAN", "NIGHT", True, "HV Cables,Protection Relays"),
        ("TEC-014", "Ibrahim Zouiten", "OPERATOR", "AFTERNOON", True, "Flotation Operation,Reagents"),
        ("TEC-015", "Hamid Lahlou", "FITTER", "AFTERNOON", True, "Valve Overhaul,Pipe Fitting"),
        ("TEC-016", "Samira El Fassi", "CONMON_SPECIALIST", "MORNING", True, "Oil Analysis ICML MLA-I,Ultrasonics"),
        ("TEC-017", "Brahim Tahiri", "FITTER", "NIGHT", True, "Gearbox Overhaul,Bearing Install"),
        ("TEC-018", "Adil Mansouri", "ELECTRICIAN", "MORNING", True, "Motor Rewind,Insulation Test"),
        ("TEC-019", "Zineb Chakir", "INSTRUMENTIST", "MORNING", True, "Calibration,Flow Meters"),
        ("TEC-020", "Driss Filali", "LUBRICATOR", "AFTERNOON", True, "ICML MLT-II,Grease Analysis"),
        ("TEC-021", "Mourad Bennani", "FITTER", "MORNING", True, "Crusher Maint,Screen Repair"),
        ("TEC-022", "Salim Hajji", "OPERATOR", "NIGHT", True, "Acid Plant Operation,Safety"),
        ("TEC-023", "Hafida Zerouali", "FITTER", "MORNING", True, "Filter Press,Vacuum Systems"),
        ("TEC-024", "Tarik Ammor", "ELECTRICIAN", "AFTERNOON", True, "Transformer Maint,Switchgear"),
        ("TEC-025", "Jamal Oukacha", "FITTER", "NIGHT", True, "Dryer Systems,Burner Maint"),
        ("TEC-026", "Leila Benmoussa", "CONMON_SPECIALIST", "AFTERNOON", True, "Vibration Cat II,Alignment"),
        ("TEC-027", "Amine Sabri", "FITTER", "MORNING", False, "General Mechanical (On Leave)"),
        ("TEC-028", "Othmane Lagrini", "OPERATOR", "MORNING", True, "Utilities Operation,HVAC"),
        ("TEC-029", "Soukaina Mernissi", "INSTRUMENTIST", "AFTERNOON", True, "Analyzer Maint,pH/Conductivity"),
        ("TEC-030", "Mehdi Raissouni", "FITTER", "AFTERNOON", True, "Compressor Maint,Air Systems"),
    ]
    rows = []
    for i, (wid, name, spec, shift, avail, certs) in enumerate(workers):
        rows.append([wid, name, spec, shift, "OCP-JFC1", avail, certs, f"+212-600-00{1001+i}", f"{name.split()[0].lower()[0]}.{name.split()[-1].lower()}@ocp.ma"])
    _write_data_rows(ws, rows)
    _add_data_validation(ws, 3, LABOUR_SPECIALTIES)
    _add_data_validation(ws, 4, SHIFT_TYPES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "09_workforce.xlsx")
    return len(rows)


def gen_10(data):
    wb = Workbook(); ws = wb.active; ws.title = "Field Captures"
    headers = ["technician_id", "capture_type", "language", "equipment_tag", "location_hint", "raw_text", "timestamp"]
    _hdr(ws, headers)
    today = date.today()
    rows = [
        ["TEC-001", "VOICE", "fr", "BRY-SAG-ML-001", "Cote palier de commande", "J'ai remarque une vibration anormale sur le palier principal du broyeur SAG. La temperature est montee a 82 degres.", f"{today-timedelta(21)} 08:30:00"],
        ["TEC-002", "TEXT", "fr", "PMP-MDP-PP-001", "Station de pompage zone B", "Fuite visible au niveau du joint mecanique de la pompe a boue. Debit reduit d'environ 15%.", f"{today-timedelta(20)} 14:15:00"],
        ["TEC-005", "VOICE", "fr", "BRY-BML-ML-001", "Systeme de lubrification", "Le niveau d'huile dans le reservoir de lubrification du broyeur a boulets est bas. Il faut faire l'appoint.", f"{today-timedelta(19)} 07:45:00"],
        ["TEC-009", "TEXT", "en", "CVY-OLC-CV-001", "Head pulley area", "Belt tracking 15mm off-center at head pulley. Scraper blade worn, material buildup on return side.", f"{today-timedelta(18)} 10:00:00"],
        ["TEC-006", "VOICE", "ar", "FLT-RGH-FC-001", "خلية التعويم رقم 1", "لاحظت تسرب هواء من وصلة المحرك في خلية التعويم. الرغوة غير مستقرة.", f"{today-timedelta(17)} 06:30:00"],
        ["TEC-004", "TEXT", "fr", "SED-THK-TH-001", "Zone epaississeur", "Le couple de l'entrainement de l'epaississeur est eleve. Rateau potentiellement bloque.", f"{today-timedelta(16)} 15:00:00"],
        ["TEC-007", "VOICE", "fr", "PMP-MDP-PP-002", "Pompe decharge broyeur #2", "Vibration excessive detectee cote aspiration. Le joint est probablement use.", f"{today-timedelta(15)} 22:30:00"],
        ["TEC-003", "TEXT", "en", "UTL-ELC-TR-001", "Main transformer room", "Oil temperature reading 85C, above normal operating range. Cooling fans running at full speed.", f"{today-timedelta(14)} 09:00:00"],
        ["TEC-008", "VOICE", "fr", "SEQ-DRY-RD-001", "Secheur rotatif #1", "Bruit metallique au niveau du pignon d'entrainement. Possible jeu excessif dans l'engrenage.", f"{today-timedelta(13)} 16:30:00"],
        ["TEC-010", "TEXT", "en", "CVY-PRD-CV-001", "Product conveyor discharge", "Belt splice showing signs of separation. Recommend replacement at next shutdown.", f"{today-timedelta(12)} 11:00:00"],
        ["TEC-011", "TEXT", "fr", "ACR-REA-RV-001", "Reacteur acide phosphorique", "Le capteur de temperature de la cuve affiche des valeurs erratiques. Calibrage necessaire.", f"{today-timedelta(11)} 08:00:00"],
        ["TEC-012", "VOICE", "fr", "FIL-BLT-BF-001", "Filtre presse a bande #1", "La bande du filtre est decentree. Le systeme de guidage ne corrige plus automatiquement.", f"{today-timedelta(10)} 13:45:00"],
        ["TEC-009", "IMAGE", "en", "GRN-GRD-GD-001", "Granulator drum inlet", "Significant wear visible on drum liner plates near inlet. Estimated 40% remaining life.", f"{today-timedelta(9)} 07:30:00"],
        ["TEC-016", "TEXT", "en", "UTL-AIR-AC-001", "Compressor room", "Oil analysis shows elevated iron particles at 85 ppm. Recommend increased monitoring frequency.", f"{today-timedelta(8)} 10:30:00"],
        ["TEC-013", "VOICE", "fr", "UTL-ELC-GN-001", "Groupe electrogene", "Le test de demarrage automatique a echoue. Le moteur ne demarre pas au premier essai.", f"{today-timedelta(7)} 23:00:00"],
        ["TEC-014", "TEXT", "fr", "FLT-RGH-FC-003", "Cellule flottation #3", "Moteur agitateur surchauffe. Temperature bobinage 145C. Arret necessaire pour refroidissement.", f"{today-timedelta(6)} 14:00:00"],
        ["TEC-015", "VOICE", "fr", "ACR-SAC-PP-001", "Pompe acide sulfurique", "Le joint mecanique presente des signes de fuite. Legere trace d'acide sur la plaque de base.", f"{today-timedelta(5)} 09:30:00"],
        ["TEC-017", "TEXT", "en", "BRY-SAG-ML-002", "SAG Mill #2 gearbox", "Gearbox oil sample shows water contamination at 0.3%. Source suspected to be seal deterioration.", f"{today-timedelta(4)} 07:00:00"],
        ["TEC-018", "VOICE", "fr", "FLT-CLN-FC-002", "Cellule finissage #2", "Court-circuit detecte sur le cablage du moteur de l'agitateur. Isolation degradee.", f"{today-timedelta(3)} 16:00:00"],
        ["TEC-019", "TEXT", "en", "SED-THK-PP-001", "Thickener underflow pump", "Flow meter reading 15% below setpoint. Impeller wear suspected based on vibration signature.", f"{today-timedelta(2)} 11:30:00"],
        ["TEC-020", "VOICE", "fr", "SEQ-DRY-RD-002", "Secheur #2 systeme lubrification", "Niveau d'huile tres bas sur le palier de support. Fuite detectee au niveau du joint.", f"{today-timedelta(1)} 08:15:00"],
        ["TEC-021", "TEXT", "en", "GRN-SCR-SC-001", "Vibrating screen", "Screen deck panel #3 has torn mesh. Production fines escaping to oversize. Needs immediate replacement.", f"{today} 06:45:00"],
        ["TEC-022", "VOICE", "ar", "ACR-REA-AG-001", "محرك خلاط المفاعل", "سمعت صوت غير طبيعي من محرك الخلاط. الاهتزاز مرتفع.", f"{today} 07:00:00"],
        ["TEC-023", "TEXT", "fr", "FIL-DSC-DF-001", "Filtre a disque #1", "Toile filtrante endommagee sur 3 secteurs. Filtrat trouble. Remplacement necessaire.", f"{today} 09:00:00"],
        ["TEC-024", "VOICE+IMAGE", "en", "UTL-CWT-CT-001", "Cooling tower cell #2", "Fan blade showing cracks at root. Captured image for reference. Recommend grounding fan immediately.", f"{today} 10:30:00"],
    ]
    _write_data_rows(ws, rows)
    _add_data_validation(ws, 2, CAPTURE_TYPES)
    _add_data_validation(ws, 3, LANGUAGES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "10_field_capture.xlsx")
    return len(rows)


def gen_11(data):
    wb = Workbook(); ws = wb.active; ws.title = "RCA Events"
    headers = [
        "event_description", "plant_id", "equipment_tag", "level",
        "max_consequence", "frequency", "event_date", "downtime_hours",
        "production_loss_tonnes", "direct_cost_usd",
    ]
    _hdr(ws, headers)
    today = date.today()
    rows = [
        ["SAG Mill #1 trunnion bearing catastrophic failure", "OCP-JFC1", "BRY-SAG-ML-001", "3", 5, 2, str(today-timedelta(360)), 96, 4800, 185000],
        ["Slurry pump mechanical seal failure causing slurry spill", "OCP-JFC1", "PMP-MDP-PP-001", "2", 4, 4, str(today-timedelta(270)), 8, 400, 12000],
        ["Overland conveyor belt splice failure", "OCP-JFC1", "CVY-OLC-CV-001", "1", 3, 3, str(today-timedelta(240)), 6, 300, 15000],
        ["Rougher flotation cell agitator motor burnout", "OCP-JFC1", "FLT-RGH-FC-003", "2", 3, 2, str(today-timedelta(210)), 48, 2400, 35000],
        ["Ball Mill #1 gearbox tooth fracture", "OCP-JFC1", "BRY-BML-ML-001", "3", 4, 1, str(today-timedelta(180)), 168, 8400, 250000],
        ["Reactor vessel rubber lining failure", "OCP-JFC1", "ACR-REA-RV-001", "3", 5, 1, str(today-timedelta(150)), 72, 3600, 120000],
        ["Rotary dryer girth gear crack propagation", "OCP-JFC1", "SEQ-DRY-RD-001", "2", 4, 1, str(today-timedelta(120)), 120, 6000, 180000],
        ["Thickener rake arm structural failure", "OCP-JFC1", "SED-THK-TH-001", "2", 3, 1, str(today-timedelta(90)), 48, 2400, 65000],
        ["Belt filter vacuum pump seizure", "OCP-JFC1", "FIL-BLT-BF-002", "1", 3, 3, str(today-timedelta(60)), 12, 600, 18000],
        ["Main transformer cooling system failure", "OCP-JFC1", "UTL-ELC-TR-001", "2", 5, 1, str(today-timedelta(45)), 4, 200, 25000],
        ["Sulfuric acid pump seal failure — acid leak", "OCP-JFC1", "ACR-SAC-PP-001", "2", 5, 2, str(today-timedelta(30)), 6, 300, 22000],
        ["Emergency generator failed to start on demand", "OCP-JFC1", "UTL-ELC-GN-001", "2", 5, 1, str(today-timedelta(15)), 0, 0, 5000],
    ]
    _write_data_rows(ws, rows)
    _add_data_validation(ws, 4, RCA_LEVELS)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "11_rca_events.xlsx")
    return len(rows)


def gen_12(data):
    wb = Workbook(); ws = wb.active; ws.title = "Planning KPI Input"
    headers = [
        "plant_id", "period_start", "period_end",
        "wo_planned", "wo_completed", "manhours_planned", "manhours_actual",
        "pm_planned", "pm_executed", "backlog_hours", "weekly_capacity_hours",
        "corrective_count", "total_wo", "schedule_compliance_planned", "schedule_compliance_executed",
        "release_horizon_days", "pending_notices", "total_notices",
        "scheduled_capacity_hours", "total_capacity_hours", "proactive_wo", "planned_wo",
    ]
    _hdr(ws, headers)
    today = date.today()
    rows = []
    for w in range(12):
        end = today - timedelta(days=7 * w)
        start = end - timedelta(days=7)
        planned = random.randint(85, 120)
        completed = int(planned * random.uniform(0.82, 0.96))
        mh_p = round(planned * random.uniform(6, 9), 1)
        mh_a = round(mh_p * random.uniform(0.9, 1.1), 1)
        pm_p = random.randint(40, 65)
        pm_e = int(pm_p * random.uniform(0.85, 0.98))
        rows.append([
            "OCP-JFC1", str(start), str(end),
            planned, completed, mh_p, mh_a,
            pm_p, pm_e, round(random.uniform(200, 450), 0), 200.0,
            random.randint(8, 25), planned, random.randint(80, 100), random.randint(70, 92),
            random.randint(3, 7), random.randint(5, 20), planned,
            round(random.uniform(150, 190), 0), 200.0,
            random.randint(60, 90), random.randint(70, 100),
        ])
    _write_data_rows(ws, rows)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "12_planning_kpi_input.xlsx")
    return len(rows)


def gen_13(data):
    wb = Workbook(); ws = wb.active; ws.title = "DE KPI Input"
    headers = [
        "plant_id", "period_start", "period_end",
        "events_reported", "events_required", "meetings_held", "meetings_required",
        "actions_implemented", "actions_planned", "savings_achieved", "savings_target",
        "failures_current", "failures_previous",
    ]
    _hdr(ws, headers)
    today = date.today()
    rows = []
    for m in range(12):
        end = today - timedelta(days=30 * m)
        start = end - timedelta(days=30)
        rows.append([
            "OCP-JFC1", str(start), str(end),
            random.randint(14, 22), 20, random.randint(7, 10), 10,
            random.randint(10, 18), random.randint(14, 20),
            round(random.uniform(60000, 110000), 0), 100000.0,
            random.randint(5, 15), random.randint(8, 20),
        ])
    _write_data_rows(ws, rows)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "13_de_kpi_input.xlsx")
    return len(rows)


def gen_14(data):
    wb = Workbook(); ws = wb.active; ws.title = "Strategies"
    headers = [
        "strategy_id", "equipment_tag", "maintainable_item",
        "function_and_failure",
        "what", "mechanism", "cause",
        "status", "tactics_type",
        "primary_task_id", "primary_task_name", "primary_task_interval",
        "operational_units", "time_units",
        "primary_task_acceptable_limits",
        "primary_task_conditional_comments", "primary_task_constraint",
        "primary_task_task_type", "primary_task_access_time",
        "secondary_task_id", "secondary_task_name",
        "secondary_task_constraint",
        "secondary_task_task_type", "secondary_task_access_time",
        "secondary_task_comments",
        "budgeted_as", "budgeted_life", "budgeted_life_time_units",
        "budgeted_life_operational_units",
        "existing_task", "justification_category", "justification", "notes",
    ]
    _hdr(ws, headers)
    _write_data_rows(ws, data["strategies"])
    _add_data_validation(ws, 6, MECHANISMS)
    _add_data_validation(ws, 7, CAUSES)
    _add_data_validation(ws, 8, STRATEGY_STATUSES)
    _add_data_validation(ws, 9, STRATEGY_TYPES)
    _add_data_validation(ws, 13, FREQUENCY_UNITS)
    _add_data_validation(ws, 14, FREQUENCY_UNITS)
    _add_data_validation(ws, 17, TASK_CONSTRAINTS)
    _add_data_validation(ws, 18, TASK_TYPES)
    _add_data_validation(ws, 22, TASK_CONSTRAINTS)
    _add_data_validation(ws, 23, TASK_TYPES)
    _add_data_validation(ws, 26, BUDGETED_AS_VALUES)
    _add_data_validation(ws, 31, JUSTIFICATION_CATEGORIES)
    _auto_size_columns(ws)
    wb.save(OUTPUT_DIR / "14_maintenance_strategy.xlsx")
    return len(data["strategies"])


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def generate_all():
    print("Building comprehensive plant data...")
    data = build_all_data()

    generators = [
        ("01_equipment_hierarchy", gen_01),
        ("02_criticality_assessment", gen_02),
        ("03_failure_modes", gen_03),
        ("04_maintenance_tasks", gen_04),
        ("05_work_packages", gen_05),
        ("06_work_order_history", gen_06),
        ("07_spare_parts_inventory", gen_07),
        ("08_shutdown_calendar", gen_08),
        ("09_workforce", gen_09),
        ("10_field_capture", gen_10),
        ("11_rca_events", gen_11),
        ("12_planning_kpi_input", gen_12),
        ("13_de_kpi_input", gen_13),
        ("14_maintenance_strategy", gen_14),
    ]

    for name, fn in generators:
        count = fn(data)
        print(f"  [OK] {name}.xlsx — {count} rows")

    print(f"\nDone! 14 templates generated in {OUTPUT_DIR}")


if __name__ == "__main__":
    generate_all()
