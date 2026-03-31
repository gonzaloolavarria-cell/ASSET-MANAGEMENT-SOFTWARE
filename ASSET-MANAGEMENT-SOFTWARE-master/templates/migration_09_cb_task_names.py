import pandas as pd
from openpyxl import load_workbook

# === CONFIGURATION: CBM Technique Selection ===
# Maps (mi_category, mechanism) -> technique type
TECHNIQUE_MAP = {
    ('MOTOR', 'WEARS'): 'VIBRATION',
    ('MOTOR', 'SHORT_CIRCUITS'): 'MEASUREMENT',
    ('MOTOR', 'OVERHEATS_MELTS'): 'THERMOGRAPHY',
    ('GEARBOX', 'WEARS'): 'OIL_ANALYSIS',
    ('GEARBOX', 'CRACKS'): 'VIBRATION',
    ('BEARING', 'WEARS'): 'VIBRATION',
    ('BEARING', 'OVERHEATS_MELTS'): 'THERMOGRAPHY',
    ('COUPLING', 'LOOSES_PRELOAD'): 'VISUAL',
    ('GEAR', 'WEARS'): 'VISUAL',
    ('STRUCTURE', 'CORRODES'): 'VISUAL',
    ('PUMP', 'WEARS'): 'VISUAL',
    ('FILTER', 'BLOCKS'): 'VISUAL',
    ('HEAT_EXCHANGER', 'CORRODES'): 'ULTRASOUND',
    ('HEAT_EXCHANGER', 'BLOCKS'): 'MEASUREMENT',
    ('VESSEL', 'CORRODES'): 'ULTRASOUND',
    ('VESSEL', 'CRACKS'): 'ULTRASOUND',
    ('PIPE', 'CORRODES'): 'VISUAL',
    ('IMPELLER', 'SEVERS'): 'MEASUREMENT',
    ('IMPELLER', 'CORRODES'): 'VISUAL',
    ('SHAFT', 'WEARS'): 'VISUAL',
    ('BLOWER', 'WEARS'): 'VIBRATION',
    ('BLOWER', 'LOOSES_PRELOAD'): 'VISUAL',
    ('NOZZLE', 'BLOCKS'): 'VISUAL',
    ('VALVE', 'CORRODES'): 'VISUAL',
    ('VALVE', 'BLOCKS'): 'VISUAL',
    ('BELT', 'WEARS'): 'VISUAL',
    ('BURNER', 'DEGRADES'): 'VISUAL',
    ('ELECTRICAL', 'SHORT_CIRCUITS'): 'MEASUREMENT',
    ('ELECTRICAL', 'THERMALLY_OVERLOADS'): 'THERMOGRAPHY',
}

# Evidence description per failure mechanism (for visual inspections)
EVIDENCE = {
    'WEARS': 'wear, scoring, and material loss',
    'CORRODES': 'corrosion, pitting, and material degradation',
    'CRACKS': 'cracking and crack propagation',
    'OVERHEATS_MELTS': 'overheating, discoloration, and thermal damage',
    'SHORT_CIRCUITS': 'insulation degradation and electrical damage',
    'BLOCKS': 'blockage and contamination buildup',
    'LEAKS': 'leakage and fluid loss',
    'DISTORTS': 'deformation and misalignment',
    'FRACTURES': 'fracture and structural failure',
    'SEVERS': 'erosion and material thinning',
    'LOOSES_PRELOAD': 'looseness and loss of preload',
    'DEGRADES': 'degradation and material deterioration',
    'ARCS': 'arcing and burn marks',
    'THERMALLY_OVERLOADS': 'thermal overload and hot spots',
}

# What to measure per (mi_category, mechanism)
MEASURE_WHAT = {
    ('MOTOR', 'SHORT_CIRCUITS'): 'insulation resistance',
    ('ELECTRICAL', 'SHORT_CIRCUITS'): 'insulation resistance',
    ('IMPELLER', 'SEVERS'): 'impeller outer diameter',
    ('HEAT_EXCHANGER', 'BLOCKS'): 'pressure drop across tubes',
}

def build_task_name(technique, mi_name, tag, mechanism, mi_cat):
    evidence = EVIDENCE.get(mechanism, 'abnormal condition')
    if technique == 'VISUAL':
        return f'Inspect {mi_name} for {evidence} [{tag}]'
    elif technique == 'VIBRATION':
        return f'Perform vibration analysis on {mi_name} [{tag}]'
    elif technique == 'OIL_ANALYSIS':
        return f'Take oil sample on {mi_name} [{tag}]'
    elif technique == 'ULTRASOUND':
        return f'Perform ultrasound inspection on {mi_name} [{tag}]'
    elif technique == 'THERMOGRAPHY':
        return f'Perform thermography on {mi_name} [{tag}]'
    elif technique == 'MEASUREMENT':
        what = MEASURE_WHAT.get((mi_cat, mechanism), 'parameters')
        return f'Measure {what} on {mi_name} [{tag}]'
    return f'Inspect {mi_name} for {evidence} [{tag}]'

# === LOAD DATA ===
bom = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment BOM')
mi_rows = bom[bom['node_type'] == 'MAINTAINABLE_ITEM']
mi_cat_lookup = dict(zip(zip(mi_rows['equipment_tag'], mi_rows['name']), mi_rows['mi_category']))

t14 = pd.read_excel('templates/14_maintenance_strategy.xlsx')
t04 = pd.read_excel('templates/04_maintenance_tasks.xlsx')

# === BUILD TASK NAME UPDATES ===
task_id_to_new_name = {}
changes = 0

for idx, row in t14.iterrows():
    if row['tactics_type'] != 'CONDITION_BASED':
        continue

    tag = row['equipment_tag']
    mi_name = row['what']
    mechanism = row['mechanism']
    task_id = row['primary_task_id']
    mi_cat = mi_cat_lookup.get((tag, mi_name), 'UNKNOWN')

    technique = TECHNIQUE_MAP.get((mi_cat, mechanism), 'VISUAL')
    new_name = build_task_name(technique, mi_name, tag, mechanism, mi_cat)

    old_name = row['primary_task_name']
    if old_name != new_name:
        t14.at[idx, 'primary_task_name'] = new_name
        task_id_to_new_name[task_id] = new_name
        changes += 1

print(f'T14: {changes} primary_task_name values updated')

# === UPDATE T14 FILE ===
wb14 = load_workbook('templates/14_maintenance_strategy.xlsx')
ws14 = wb14.active
headers14 = [ws14.cell(1, c).value for c in range(1, ws14.max_column + 1)]
name_col = headers14.index('primary_task_name') + 1

for idx, row in t14.iterrows():
    if row['tactics_type'] == 'CONDITION_BASED':
        ws14.cell(row=idx + 2, column=name_col, value=row['primary_task_name'])

wb14.save('templates/14_maintenance_strategy.xlsx')
print('T14 saved.')

# === UPDATE T04 FILE ===
wb04 = load_workbook('templates/04_maintenance_tasks.xlsx')
ws04 = wb04.active
headers04 = [ws04.cell(1, c).value for c in range(1, ws04.max_column + 1)]
tid_col = headers04.index('task_id') + 1
tname_col = headers04.index('task_name') + 1

t04_changes = 0
for r in range(2, ws04.max_row + 1):
    tid = ws04.cell(r, tid_col).value
    if tid in task_id_to_new_name:
        ws04.cell(r, tname_col, value=task_id_to_new_name[tid])
        t04_changes += 1

wb04.save('templates/04_maintenance_tasks.xlsx')
print(f'T04: {t04_changes} task_name values updated. Saved.')

# === SUMMARY ===
print(f'\n=== TECHNIQUE DISTRIBUTION ===')
technique_counts = {}
for idx, row in t14.iterrows():
    if row['tactics_type'] != 'CONDITION_BASED':
        continue
    mi_cat = mi_cat_lookup.get((row['equipment_tag'], row['what']), 'UNKNOWN')
    tech = TECHNIQUE_MAP.get((mi_cat, row['mechanism']), 'VISUAL')
    technique_counts[tech] = technique_counts.get(tech, 0) + 1

for tech, count in sorted(technique_counts.items(), key=lambda x: -x[1]):
    print(f'  {tech:20s}: {count}')

print(f'\n=== SAMPLE NEW TASK NAMES ===')
cb = t14[t14['tactics_type'] == 'CONDITION_BASED']
for _, r in cb.head(20).iterrows():
    print(f'  {r["primary_task_name"]}')
