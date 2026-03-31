import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='1B5E20')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
dfont = Font(name='Calibri', size=11)
dalign = Alignment(vertical='center', wrap_text=True)

# Build subunit lookup from T01 BOM
df_bom = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment BOM')
subunit_codes = {}
mi_to_subunit = {}
for _, r in df_bom.iterrows():
    ntype = r['node_type']
    if ntype == 'SUBUNIT':
        subunit_codes[r['code']] = r['name']
    elif ntype == 'MAINTAINABLE_ITEM':
        parent = r['parent_code']
        mi_to_subunit[(r['equipment_tag'], r['name'])] = subunit_codes.get(parent, 'Unknown')

# Read T03
df = pd.read_excel('templates/03_failure_modes.xlsx')

# Map subunit for each row
def get_subunit(row):
    key = (row['equipment_tag'], row['Maintainable_item'])
    return mi_to_subunit.get(key, 'Unknown')

df['subunit'] = df.apply(get_subunit, axis=1)

unknown = df[df['subunit'] == 'Unknown']
if len(unknown) > 0:
    print(f'WARNING: {len(unknown)} rows have unknown subunit:')
    for _, r in unknown.drop_duplicates(subset=['equipment_tag', 'Maintainable_item']).iterrows():
        print(f'  {r["equipment_tag"]} / {r["Maintainable_item"]}')

# Build new spreadsheet with normalized headers
wb = Workbook()
ws = wb.active
ws.title = 'failure_modes'

new_headers = [
    'equipment_tag', 'equipment_function_description', 'equipment_functional_failure',
    'function_type', 'failure_type', 'subunit', 'maintainable_item',
    'maintainable_item_function_description', 'maintainable_item_functional_failure',
    'mechanism', 'cause', 'failure_pattern', 'failure_consequence',
    'evidence', 'downtime_hours', 'detection_method',
    'rpn_severity', 'rpn_occurrence', 'rpn_detection'
]

old_to_new = {
    'equipment_tag': 'equipment_tag',
    'Equipment_function_description': 'equipment_function_description',
    'Equipment_functional_failure': 'equipment_functional_failure',
    'function_type': 'function_type',
    'failure_type': 'failure_type',
    'Maintainable_item': 'maintainable_item',
    'Maintainable_item_function_description': 'maintainable_item_function_description',
    'Maintainable_item_functional_failure': 'maintainable_item_functional_failure',
}

for c, h in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

col_src = {
    0: 'equipment_tag', 1: 'Equipment_function_description', 2: 'Equipment_functional_failure',
    3: 'function_type', 4: 'failure_type', 5: 'subunit', 6: 'Maintainable_item',
    7: 'Maintainable_item_function_description', 8: 'Maintainable_item_functional_failure',
    9: 'mechanism', 10: 'cause', 11: 'failure_pattern', 12: 'failure_consequence',
    13: 'evidence', 14: 'downtime_hours', 15: 'detection_method',
    16: 'rpn_severity', 17: 'rpn_occurrence', 18: 'rpn_detection'
}

for i, (_, row) in enumerate(df.iterrows(), 2):
    for c_idx, src_col in col_src.items():
        v = row[src_col] if pd.notna(row[src_col]) else None
        cell = ws.cell(row=i, column=c_idx+1, value=v)
        cell.font = dfont
        cell.alignment = dalign

widths = [17, 50, 50, 15, 14, 22, 24, 50, 50, 18, 26, 19, 23, 12, 16, 20, 14, 16, 14]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.auto_filter.ref = f'A1:S{len(df)+1}'
ws.freeze_panes = 'A2'

wb.save('templates/03_failure_modes.xlsx')
unknowns = (df['subunit'] == 'Unknown').sum()
print(f'T03 done. {len(df)} rows x {len(new_headers)} cols. {unknowns} unknown subunits.')
