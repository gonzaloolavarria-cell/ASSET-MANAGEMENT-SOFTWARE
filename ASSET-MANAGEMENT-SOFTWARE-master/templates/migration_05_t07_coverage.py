import pandas as pd
from openpyxl import load_workbook

# Read T01 for all equipment and their BOM
t01 = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment Hierarchy')
all_tags = set(t01['equipment_tag'].unique())
tag_to_type = dict(zip(t01['equipment_tag'], t01['equipment_type']))

df_bom = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment BOM')
mi_bom = df_bom[df_bom['node_type'] == 'MAINTAINABLE_ITEM']

# Build: equipment_type -> set of MI categories
type_to_mi_cats = {}
for _, r in mi_bom.iterrows():
    tag = r['equipment_tag']
    etype = tag_to_type.get(tag, '')
    cat = r['mi_category']
    if pd.notna(cat):
        type_to_mi_cats.setdefault(etype, set()).add(cat)

# Read T07
df07 = pd.read_excel('templates/07_spare_parts_inventory.xlsx')

# Find which equipment are covered
covered_tags = set()
for _, r in df07.iterrows():
    csv = str(r['applicable_equipment_csv'])
    for tag in csv.split(','):
        tag = tag.strip()
        if tag:
            covered_tags.add(tag)

missing_tags = all_tags - covered_tags
print(f'Equipment not in any spare part: {sorted(missing_tags)}')

# For each missing tag, find matching spare parts by MI category
# Build: mi_category -> list of material_code rows
wb = load_workbook('templates/07_spare_parts_inventory.xlsx')
ws = wb.active

# Find column index for applicable_equipment_csv
header_row = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
eq_col = header_row.index('applicable_equipment_csv') + 1

# For each spare part row, get its MI category from material_code
# Material codes follow pattern: MAT-{MI_CAT}-{COMPONENT}
# We need to check if any existing equipment in that row shares the same type as a missing one
for r in range(2, ws.max_row + 1):
    mat_code = ws.cell(r, 1).value
    desc = ws.cell(r, 3).value
    csv_val = str(ws.cell(r, eq_col).value)
    current_tags = [t.strip() for t in csv_val.split(',') if t.strip()]

    if not current_tags:
        continue

    # Get equipment types of current tags
    current_types = set(tag_to_type.get(t, '') for t in current_tags)

    # For each missing tag, check if its type matches any current type
    tags_to_add = []
    for mt in missing_tags:
        mt_type = tag_to_type.get(mt, '')
        if mt_type in current_types and mt not in current_tags:
            tags_to_add.append(mt)

    if tags_to_add:
        new_csv = ','.join(current_tags + sorted(tags_to_add))
        ws.cell(r, eq_col, value=new_csv)

wb.save('templates/07_spare_parts_inventory.xlsx')

# Verify coverage
df07_new = pd.read_excel('templates/07_spare_parts_inventory.xlsx')
covered_after = set()
for _, r in df07_new.iterrows():
    csv = str(r['applicable_equipment_csv'])
    for tag in csv.split(','):
        tag = tag.strip()
        if tag:
            covered_after.add(tag)

still_missing = all_tags - covered_after
print(f'After fix - covered: {len(covered_after)}/{len(all_tags)}')
if still_missing:
    print(f'Still missing: {sorted(still_missing)}')
else:
    print('All 63 equipment now have spare parts coverage.')
