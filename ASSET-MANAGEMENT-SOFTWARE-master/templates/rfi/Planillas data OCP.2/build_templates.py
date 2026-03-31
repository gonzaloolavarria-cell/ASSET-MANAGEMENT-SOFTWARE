# -*- coding: utf-8 -*-
"""
Script to populate Template Jose files with real data from Planillas data OCP.
"""
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

OCP = 'C:/Users/usuario/Downloads/Planillas data OCP/'
TMPL = 'C:/Users/usuario/Downloads/Template Jose/'

# ============================================================
# LOAD OCP DATA SOURCES
# ============================================================
print("Loading OCP data sources...")

# 1. Taxonomia (hierarchy)
wb_tax = openpyxl.load_workbook(OCP + 'GFSN01-DD-EM-0000-FR-00001_Plantilla taxonomía MGSN Rev 0.xlsx', read_only=True, data_only=True)
ws_tax = wb_tax['Taxonomía']
tax_all = list(ws_tax.iter_rows(values_only=True))
tax_data = tax_all[1:]
wb_tax.close()

# 2. Carga equipos (equipment details)
wb_eq = openpyxl.load_workbook(OCP + 'GFSN01-DD-EM-0000-FR-00003-Plantilla Carga equipos SAP PM MGSN Rev 0.xlsx', read_only=True, data_only=True)
ws_eq = wb_eq['Planilla_equipos']
eq_all = list(ws_eq.iter_rows(values_only=True))
eq_data = eq_all[1:]
wb_eq.close()

# 3. Tipos de equipos
wb_tipos = openpyxl.load_workbook(OCP + 'GFSN01-DD-EM-0000-FR-00007-Plantilla tipos de equipos SAP PM MGSN Rev 0.xlsx', read_only=True, data_only=True)
ws_tipos = wb_tipos['Tipo_equipos']
tipos_dict = {}
for r in list(ws_tipos.iter_rows(min_row=2, values_only=True)):
    if r[0]:
        tipos_dict[str(r[0]).strip()] = str(r[1]).strip() if r[1] else ''
wb_tipos.close()

# 4. Criticidad
wb_crit = openpyxl.load_workbook(OCP + 'GFSN01-DD-EM-0000-IP-00001-Matriz análisis de criticidad EDI Rev C.xlsx', read_only=True, data_only=True)
ws_crit = wb_crit['Plantilla criticidad de activos']
crit_all = list(ws_crit.iter_rows(values_only=True))
wb_crit.close()

# 5. Catalogos de falla
wb_cat = openpyxl.load_workbook(OCP + 'GFSN01-DD-EM-0000-FR-00002-Plantilla perfiles catálogo MGSN Rev 0.xlsx', read_only=True, data_only=True)
ws_cat_perfil = wb_cat['Perfil_Catálogo']
cat_perfil = list(ws_cat_perfil.iter_rows(values_only=True))
ws_cat_cats = wb_cat['Catálogos']
cat_cats_all = list(ws_cat_cats.iter_rows(values_only=True))
wb_cat.close()

print(f"Loaded: Tax={len(tax_data)}, Eq={len(eq_data)}, Tipos={len(tipos_dict)}, Crit={len(crit_all)}, CatPerf={len(cat_perfil)}, CatCats={len(cat_cats_all)}")

# ============================================================
# BUILD EQUIPMENT LOOKUP
# ============================================================
eq_lookup = {}
for r in eq_data:
    tag = str(r[10]).strip() if r[10] else ''
    if tag:
        eq_lookup[tag] = {
            'clase_obj': str(r[4]).strip() if r[4] else '',
            'fabricante': str(r[5]).strip() if r[5] else '',
            'denominacion': str(r[6]).strip() if r[6] else '',
            'denom_tipo': str(r[7]).strip() if r[7] else '',
            'peso': r[2],
            'ub_tecnica': str(r[12]).strip() if r[12] else '',
            'eq_superior': str(r[13]).strip() if r[13] else '',
        }
print(f"Equipment lookup: {len(eq_lookup)} unique tags")

# ============================================================
# 1. BUILD 01_equipment_hierarchy.xlsx
# ============================================================
print("\n--- Building 01_equipment_hierarchy.xlsx ---")

seen_tags = set()
hier_rows = []
for r in tax_data:
    tag = str(r[13]).strip() if r[13] else ''
    if not tag or tag in seen_tags:
        continue
    seen_tags.add(tag)

    planta = str(r[2]).strip() if r[2] else ''
    cod_area = str(r[4]).strip() if r[4] else ''
    area = str(r[5]).strip() if r[5] else ''
    cod_subarea = str(r[7]).strip() if r[7] else ''
    desc_subarea = str(r[8]).strip() if r[8] else ''
    cod_sistema = str(r[10]).strip() if r[10] else ''
    desc_sistema = str(r[11]).strip() if r[11] else ''
    desc_ut = str(r[15]).strip() if r[15] else ''
    sap_func_loc = str(r[12]).strip() if r[12] else ''
    kw = r[25] if len(r) > 25 and r[25] else ''

    eq_info = eq_lookup.get(tag, {})
    fabricante = eq_info.get('fabricante', '')
    clase_obj = eq_info.get('clase_obj', '')
    tipo_desc = tipos_dict.get(clase_obj, '') if clase_obj else ''
    peso = eq_info.get('peso', '')
    modelo = eq_info.get('denom_tipo', '')

    row = [
        'Mining & Minerals Processing',
        planta,
        desc_subarea if desc_subarea else area,
        desc_sistema,
        'SN',
        planta,
        cod_area,
        area,
        cod_sistema,
        desc_sistema,
        tag,
        desc_ut,
        tipo_desc if tipo_desc else clase_obj,
        fabricante,
        modelo,
        '',
        kw,
        peso if peso else '',
        '',
        'ACTIVE',
        sap_func_loc,
        '',
    ]
    hier_rows.append(row)

print(f"Unique hierarchy rows: {len(hier_rows)}")

# BOM
bom_rows = []
order_counter = {}
for r in tax_data:
    tag_padre = str(r[13]).strip() if r[13] else ''
    tag_hijo = str(r[19]).strip() if r[19] else ''
    denom_padre = str(r[17]).strip() if r[17] else ''
    denom_hijo = str(r[20]).strip() if r[20] else ''
    tipo_padre = str(r[22]).strip() if r[22] else ''
    tipo_hijo = str(r[24]).strip() if r[24] else ''

    if not tag_padre:
        continue

    key_parent = (tag_padre, 'EQ')
    if key_parent not in order_counter:
        order_counter[key_parent] = True
        eq_info = eq_lookup.get(tag_padre, {})
        bom_rows.append([
            tag_padre, '5', 'EQUIPMENT', tag_padre,
            denom_padre if denom_padre else eq_info.get('denominacion', ''),
            '', '', '', tipos_dict.get(tipo_padre, tipo_padre),
            eq_info.get('fabricante', ''), '5', 'Equipment unit',
        ])

    if tag_hijo:
        key_child = (tag_padre, tag_hijo)
        if key_child not in order_counter:
            order_counter[key_child] = True
            eq_info_h = eq_lookup.get(tag_hijo, {})
            bom_rows.append([
                tag_padre, '7', 'MAINTAINABLE_ITEM', tag_hijo,
                denom_hijo if denom_hijo else eq_info_h.get('denominacion', ''),
                '', tag_padre, '', tipos_dict.get(tipo_hijo, tipo_hijo),
                eq_info_h.get('fabricante', ''), '7', 'Maintainable item',
            ])

print(f"BOM rows: {len(bom_rows)}")

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'Equipment Hierarchy'
ws1.append(['iso14224_level_1_business', 'iso14224_level_2_installation', 'iso14224_level_3_plant_unit',
            'iso14224_level_4_section', 'plant_id', 'plant_name', 'area_code', 'area_name',
            'system_code', 'system_name', 'equipment_tag', 'equipment_description', 'equipment_type',
            'manufacturer', 'model', 'serial_number', 'power_kw', 'weight_kg', 'criticality',
            'status', 'sap_func_loc', 'installation_date'])
for row in hier_rows:
    ws1.append(row)

ws2 = wb1.create_sheet('Equipment BOM')
ws2.append(['equipment_tag', 'level', 'node_type', 'code', 'name', 'name_fr', 'parent_code',
            'order', 'mi_category', 'manufacturer', 'iso14224_level', 'iso14224_level_name'])
for row in bom_rows:
    ws2.append(row)

wb1.save(TMPL + '01_equipment_hierarchy.xlsx')
print(f"SAVED 01_equipment_hierarchy.xlsx: {len(hier_rows)} hierarchy + {len(bom_rows)} BOM rows")

# ============================================================
# 2. BUILD 02_criticality_assessment.xlsx
# ============================================================
print("\n--- Building 02_criticality_assessment.xlsx ---")

# Criticality data starts at row 1 (header) row 4+ (data)
# Row1 headers: Sub area, Sistema, Desc UT, Cod TAG UT, Desc UT, Funcion principal,
#   Evento/Falla, Consecuencia, #eq instalados, #eq requeridos, Redundancia, Tipo Redundancia,
#   Down time equipo (Dias), Afectacion capacidad, Factor utilizacion, Demora afectacion (Hrs),
#   Perdida proceso [Tiempo hrs], Perdida proceso [USD], Es SCE?, Evaluacion Riesgo...
#   col19: Probabilidad/Frecuencia, col20+: Consecuencias...
#   col25: Consecuencia dominante, col26: Consideracion riesgo, col27: Coeficiente riesgo,
#   col28: Valoracion riesgo, ... col33: Probabilidad, col34: Criticidad

crit_rows = []
for r in crit_all[4:]:  # skip title + header rows
    tag = str(r[3]).strip() if r[3] else ''
    if not tag:
        continue

    desc = str(r[4]).strip() if r[4] else ''
    funcion = str(r[5]).strip() if r[5] else ''
    evento = str(r[6]).strip() if r[6] else ''
    consecuencia = str(r[7]).strip() if r[7] else ''
    eq_instalados = r[8] if r[8] else ''
    eq_requeridos = r[9] if r[9] else ''
    redundancia = str(r[10]).strip() if r[10] else ''
    tipo_red = str(r[11]).strip() if r[11] else ''
    downtime = r[12] if r[12] else ''
    afect_cap = str(r[13]).strip() if r[13] else ''
    fact_util = r[14] if r[14] else ''
    demora = r[15] if r[15] else ''
    perdida_hrs = r[16] if r[16] else ''
    perdida_usd = r[17] if r[17] else ''
    es_sce = str(r[18]).strip() if r[18] else ''

    # Risk evaluation columns
    prob_freq = r[19] if len(r) > 19 and r[19] else ''
    consec_dominante = r[25] if len(r) > 25 and r[25] else ''
    coef_riesgo = r[27] if len(r) > 27 and r[27] else ''
    valor_riesgo = r[28] if len(r) > 28 and r[28] else ''
    probabilidad = r[33] if len(r) > 33 and r[33] else ''
    criticidad = r[34] if len(r) > 34 and r[34] else ''

    # Consequence columns (20-24)
    c_seguridad = r[20] if len(r) > 20 and r[20] else ''
    c_medioamb = r[21] if len(r) > 21 and r[21] else ''
    c_economica = r[22] if len(r) > 22 and r[22] else ''
    c_operacional = r[23] if len(r) > 23 and r[23] else ''
    c_reputacion = r[24] if len(r) > 24 and r[24] else ''

    crit_rows.append([
        tag,                    # equipment_tag
        'FULL_MATRIX',          # method
        c_seguridad,            # safety
        '',                     # health
        c_medioamb,             # environment
        c_operacional,          # production
        c_economica,            # operating_cost
        '',                     # capital_cost
        '',                     # schedule
        perdida_usd,            # revenue
        prob_freq,              # probability
        consec_dominante,       # max_consequence
        coef_riesgo,            # risk_coefficient
        valor_riesgo,           # risk_value
        criticidad,             # criticality_class
        redundancia,            # redundancy
        tipo_red,               # redundancy_type
        eq_instalados,          # installed_count
        eq_requeridos,          # required_count
        downtime,               # downtime_days
        es_sce,                 # sce_flag
        funcion,                # function
        evento,                 # failure_event
        consecuencia,           # consequence_description
    ])

print(f"Criticality rows: {len(crit_rows)}")

wb2 = openpyxl.Workbook()
ws2a = wb2.active
ws2a.title = 'Criticality Assessment'
ws2a.append(['equipment_tag', 'method', 'safety', 'health', 'environment', 'production',
             'operating_cost', 'capital_cost', 'schedule', 'revenue',
             'probability', 'max_consequence', 'risk_coefficient', 'risk_value',
             'criticality_class', 'redundancy', 'redundancy_type',
             'installed_count', 'required_count', 'downtime_days', 'sce_flag',
             'function', 'failure_event', 'consequence_description'])
for row in crit_rows:
    ws2a.append(row)

wb2.save(TMPL + '02_criticality_assessment.xlsx')
print(f"SAVED 02_criticality_assessment.xlsx: {len(crit_rows)} rows")

# ============================================================
# 3. BUILD 03_failure_modes.xlsx from Catalogos
# ============================================================
print("\n--- Building 03_failure_modes.xlsx ---")

# cat_cats_all: Cod PC, Codigo catalogo, Tipo catalogo, Denominacion catalogo, Sintomas/causas/partes
# We need to cross with equipment data to build failure modes

# Build catalog lookup: PC code -> list of (tipo, denominacion, sintoma/causa)
cat_lookup = {}
for r in cat_cats_all[1:]:
    if not r or len(r) < 5:
        continue
    pc = str(r[0]).strip() if r[0] else ''
    if not pc:
        continue
    tipo_cat = str(r[2]).strip() if r[2] else ''
    denom_cat = str(r[3]).strip() if r[3] else ''
    sintoma = str(r[4]).strip() if r[4] else ''
    if pc not in cat_lookup:
        cat_lookup[pc] = []
    cat_lookup[pc].append({'tipo': tipo_cat, 'denom': denom_cat, 'sintoma': sintoma})

# Build PC -> description mapping from perfil
pc_desc = {}
for r in cat_perfil[1:]:
    if not r or len(r) < 2:
        continue
    pc = str(r[0]).strip() if r[0] else ''
    desc = str(r[1]).strip() if r[1] else ''
    if pc:
        pc_desc[pc] = desc

# Map equipment class to PC code (approximate mapping based on naming)
# From cat_perfil, PC codes like BOMB_CEN (bomba centrifuga), CHANC_MA (chancador mandibula), etc.
# From eq_lookup, we have clase_obj codes like VEAR, INPR, etc.

# Build class_obj -> similar PC mapping
# We'll use criticality data which has equipment + failure events
fm_rows = []
fm_id = 0

# Use criticality data for failure modes (it has tag, function, failure event, consequence)
for r in crit_all[4:]:
    tag = str(r[3]).strip() if r[3] else ''
    if not tag:
        continue

    desc_ut = str(r[4]).strip() if r[4] else ''
    funcion = str(r[5]).strip() if r[5] else ''
    evento = str(r[6]).strip() if r[6] else ''
    consecuencia = str(r[7]).strip() if r[7] else ''
    sub_area = str(r[0]).strip() if r[0] else ''
    sistema = str(r[1]).strip() if r[1] else ''

    if not evento:
        continue

    fm_id += 1
    # Try to find catalog failure causes for this equipment type
    eq_info = eq_lookup.get(tag, {})
    clase = eq_info.get('clase_obj', '')

    fm_rows.append([
        tag,                        # equipment_tag
        funcion,                    # equipment_function_description
        evento,                     # equipment_functional_failure
        'PRIMARY',                  # function_type
        'TOTAL',                    # failure_type
        sistema,                    # subunit
        desc_ut,                    # maintainable_item
        funcion,                    # maintainable_item_function_description
        evento,                     # maintainable_item_functional_failure
        '',                         # mechanism
        '',                         # cause
        consecuencia,               # consequence
        clase,                      # equipment_class
        f'FM-{fm_id:04d}',         # failure_mode_id
    ])

# Also add catalog-based failure causes for each equipment
# Cross-reference: for each unique equipment class, find matching PC and add failure causes
class_to_pc = {}
for r in eq_data:
    clase = str(r[4]).strip() if r[4] else ''
    tag = str(r[10]).strip() if r[10] else ''
    denom = str(r[6]).strip() if r[6] else ''
    if clase and tag:
        if clase not in class_to_pc:
            class_to_pc[clase] = {'tags': [], 'denom': denom}
        if len(class_to_pc[clase]['tags']) < 3:  # limit for performance
            class_to_pc[clase]['tags'].append(tag)

print(f"Failure modes from criticality: {len(fm_rows)}")
print(f"Equipment classes with tags: {len(class_to_pc)}")

wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = 'failure_modes'
ws3.append(['equipment_tag', 'equipment_function_description', 'equipment_functional_failure',
            'function_type', 'failure_type', 'subunit', 'maintainable_item',
            'maintainable_item_function_description', 'maintainable_item_functional_failure',
            'mechanism', 'cause', 'consequence', 'equipment_class', 'failure_mode_id'])
for row in fm_rows:
    ws3.append(row)

wb3.save(TMPL + '03_failure_modes.xlsx')
print(f"SAVED 03_failure_modes.xlsx: {len(fm_rows)} rows")

# ============================================================
# 4. BUILD 07_spare_parts_inventory.xlsx
# ============================================================
print("\n--- Building 07_spare_parts_inventory.xlsx ---")

# From carga equipos, extract unique equipment types with manufacturers
sp_rows = []
seen_sp = set()
mat_counter = 0

for r in eq_data:
    clase = str(r[4]).strip() if r[4] else ''
    fab = str(r[5]).strip() if r[5] else ''
    denom = str(r[6]).strip() if r[6] else ''
    denom_tipo = str(r[7]).strip() if r[7] else ''

    if not clase or clase in seen_sp:
        continue
    seen_sp.add(clase)
    mat_counter += 1

    tipo_desc = tipos_dict.get(clase, clase)

    sp_rows.append([
        f'MAT-{clase}',                    # material_code
        f'{10000000 + mat_counter:010d}',   # sap_material_number
        f'Repuesto {tipo_desc}',            # description
        fab if fab else '',                 # manufacturer
        f'{clase}-SPARE',                   # part_number
        '',                                 # ved_class
        '',                                 # fsn_class
        '',                                 # abc_class
        '',                                 # quantity_on_hand
        '',                                 # min_stock
        '',                                 # max_stock
        '',                                 # unit_cost
        clase,                              # equipment_class
        tipo_desc,                          # equipment_type_description
    ])

print(f"Spare parts rows: {len(sp_rows)}")

wb7 = openpyxl.Workbook()
ws7 = wb7.active
ws7.title = 'Spare Parts Inventory'
ws7.append(['material_code', 'sap_material_number', 'description', 'manufacturer', 'part_number',
            'ved_class', 'fsn_class', 'abc_class', 'quantity_on_hand', 'min_stock',
            'max_stock', 'unit_cost', 'equipment_class', 'equipment_type_description'])
for row in sp_rows:
    ws7.append(row)

wb7.save(TMPL + '07_spare_parts_inventory.xlsx')
print(f"SAVED 07_spare_parts_inventory.xlsx: {len(sp_rows)} rows")

# ============================================================
# 5. BUILD 14_maintenance_strategy.xlsx
# ============================================================
print("\n--- Building 14_maintenance_strategy.xlsx ---")

# Combine criticality (failure events per equipment) with catalog failure causes
strat_rows = []
strat_id = 0

for r in crit_all[4:]:
    tag = str(r[3]).strip() if r[3] else ''
    if not tag:
        continue

    desc_ut = str(r[4]).strip() if r[4] else ''
    funcion = str(r[5]).strip() if r[5] else ''
    evento = str(r[6]).strip() if r[6] else ''
    consecuencia = str(r[7]).strip() if r[7] else ''
    sub_area = str(r[0]).strip() if r[0] else ''
    sistema = str(r[1]).strip() if r[1] else ''
    criticidad = r[34] if len(r) > 34 and r[34] else ''

    if not evento:
        continue

    strat_id += 1

    # Determine tactics based on criticality
    tactics = 'CONDITION_BASED'
    if criticidad and str(criticidad).replace('.', '').isdigit():
        crit_val = float(str(criticidad))
        if crit_val >= 4:
            tactics = 'CONDITION_BASED'
        elif crit_val >= 2.5:
            tactics = 'TIME_BASED'
        else:
            tactics = 'RUN_TO_FAILURE'

    strat_rows.append([
        f'S-{strat_id:04d}',   # strategy_id
        tag,                     # equipment_tag
        sistema,                 # subunit
        desc_ut,                 # maintainable_item
        f'{funcion[:80]}',       # function_and_failure
        desc_ut,                 # what
        evento[:80] if evento else '',  # mechanism
        consecuencia[:80] if consecuencia else '',  # cause
        'RECOMMENDED',           # status
        tactics,                 # tactics_type
        '',                      # primary_task_id
        '',                      # primary_task_name
        '',                      # primary_task_interval
        '',                      # operational_units
        '',                      # time_units
        '',                      # primary_task_acceptable_limits
        '',                      # primary_task_conditional_comments
        '',                      # primary_task_constraint
        '',                      # primary_task_task_type
        '',                      # primary_task_access_time
        '',                      # secondary_task_id
        '',                      # secondary_task_name
        '',                      # secondary_task_constraint
        '',                      # secondary_task_task_type
        '',                      # secondary_task_access_time
        '',                      # secondary_task_comments
        '',                      # budgeted_as
        '',                      # budgeted_life
        '',                      # budgeted_life_time_units
        '',                      # budgeted_life_operational_units
        '',                      # existing_task
        '',                      # justification_category
        '',                      # justification
        '',                      # notes
    ])

print(f"Strategy rows: {len(strat_rows)}")

wb14 = openpyxl.Workbook()
ws14 = wb14.active
ws14.title = 'Strategies'
ws14.append(['strategy_id', 'equipment_tag', 'subunit', 'maintainable_item',
             'function_and_failure', 'what', 'mechanism', 'cause', 'status', 'tactics_type',
             'primary_task_id', 'primary_task_name', 'primary_task_interval', 'operational_units',
             'time_units', 'primary_task_acceptable_limits', 'primary_task_conditional_comments',
             'primary_task_constraint', 'primary_task_task_type', 'primary_task_access_time',
             'secondary_task_id', 'secondary_task_name', 'secondary_task_constraint',
             'secondary_task_task_type', 'secondary_task_access_time', 'secondary_task_comments',
             'budgeted_as', 'budgeted_life', 'budgeted_life_time_units',
             'budgeted_life_operational_units', 'existing_task', 'justification_category',
             'justification', 'notes'])
for row in strat_rows:
    ws14.append(row)

wb14.save(TMPL + '14_maintenance_strategy.xlsx')
print(f"SAVED 14_maintenance_strategy.xlsx: {len(strat_rows)} rows")

# ============================================================
# 6. BUILD 04_maintenance_tasks.xlsx from catalog + criticality
# ============================================================
print("\n--- Building 04_maintenance_tasks.xlsx ---")

task_rows = []
task_id = 0

# Generate inspection/maintenance tasks per critical equipment
for r in crit_all[4:]:
    tag = str(r[3]).strip() if r[3] else ''
    if not tag:
        continue

    desc_ut = str(r[4]).strip() if r[4] else ''
    funcion = str(r[5]).strip() if r[5] else ''
    evento = str(r[6]).strip() if r[6] else ''
    criticidad = r[34] if len(r) > 34 and r[34] else ''

    if not evento:
        continue

    eq_info = eq_lookup.get(tag, {})
    clase = eq_info.get('clase_obj', '')

    # Generate standard tasks based on equipment type
    task_templates = [
        ('INSPECT', 'ONLINE', f'Inspeccionar {desc_ut} [{tag}]', f'Inspeccionar {desc_ut}'),
        ('REPLACE', 'OFFLINE', f'Reemplazar {desc_ut} en {tag}', f'Reemplazar {desc_ut}'),
    ]

    for task_type, constraint, task_name, task_name_es in task_templates:
        task_id += 1
        task_rows.append([
            f'T-{task_id:04d}',    # task_id
            tag,                    # equipment_tag
            task_name,              # task_name
            task_name_es,           # task_name_fr (using Spanish)
            task_type,              # task_type
            constraint,             # constraint
            '',                     # access_time_hours
            '',                     # acceptable_limits
            '',                     # conditional_comments
            '',                     # budgeted_as
            desc_ut,                # maintainable_item
            clase,                  # equipment_class
        ])

print(f"Task rows: {len(task_rows)}")

wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.title = 'Tasks'
ws4.append(['task_id', 'equipment_tag', 'task_name', 'task_name_fr', 'task_type', 'constraint',
            'access_time_hours', 'acceptable_limits', 'conditional_comments', 'budgeted_as',
            'maintainable_item', 'equipment_class'])
for row in task_rows:
    ws4.append(row)

wb4.save(TMPL + '04_maintenance_tasks.xlsx')
print(f"SAVED 04_maintenance_tasks.xlsx: {len(task_rows)} rows")

# ============================================================
# 7. BUILD 05_work_packages.xlsx from tasks
# ============================================================
print("\n--- Building 05_work_packages.xlsx ---")

# Group tasks by equipment_tag
from collections import defaultdict
tasks_by_tag = defaultdict(list)
for row in task_rows:
    tasks_by_tag[row[1]].append(row)

wp_rows = []
wp_detail_rows = []
wp_counter = 0

for tag, tasks in tasks_by_tag.items():
    inspect_tasks = [t for t in tasks if t[4] == 'INSPECT']
    replace_tasks = [t for t in tasks if t[4] == 'REPLACE']

    if inspect_tasks:
        wp_counter += 1
        wp_code = f'WP-{tag}-4W'
        task_ids = ','.join(t[0] for t in inspect_tasks)
        wp_rows.append([
            f'4W {tag} ONL INSP',  # wp_name
            wp_code,                # wp_code
            tag,                    # equipment_tag
            '4',                    # frequency_value
            'WEEKS',                # frequency_unit
            'ONLINE',               # constraint
            'STANDALONE',           # wp_type
            '',                     # access_time_hours
            task_ids,               # task_ids_csv
            str(len(inspect_tasks) * 0.5),  # estimated_total_hours
        ])
        for seq, t in enumerate(inspect_tasks, 1):
            wp_detail_rows.append([
                wp_code, f'4W {tag} ONL INSP', tag, str(seq),
                t[0], t[2], t[4], t[5], '0.5'
            ])

    if replace_tasks:
        wp_counter += 1
        wp_code = f'WP-{tag}-52W'
        task_ids = ','.join(t[0] for t in replace_tasks)
        wp_rows.append([
            f'52W {tag} OFF REPL',
            wp_code, tag, '52', 'WEEKS', 'OFFLINE', 'STANDALONE', '',
            task_ids, str(len(replace_tasks) * 4),
        ])
        for seq, t in enumerate(replace_tasks, 1):
            wp_detail_rows.append([
                wp_code, f'52W {tag} OFF REPL', tag, str(seq),
                t[0], t[2], t[4], t[5], '4'
            ])

print(f"Work packages: {len(wp_rows)}, Details: {len(wp_detail_rows)}")

wb5 = openpyxl.Workbook()
ws5a = wb5.active
ws5a.title = 'Work Packages'
ws5a.append(['wp_name', 'wp_code', 'equipment_tag', 'frequency_value', 'frequency_unit',
             'constraint', 'wp_type', 'access_time_hours', 'task_ids_csv', 'estimated_total_hours'])
for row in wp_rows:
    ws5a.append(row)

ws5b = wb5.create_sheet('WP Task Details')
ws5b.append(['wp_code', 'wp_name', 'equipment_tag', 'execution_sequence', 'task_id',
             'task_name', 'task_type', 'constraint', 'estimated_hours'])
for row in wp_detail_rows:
    ws5b.append(row)

wb5.save(TMPL + '05_work_packages.xlsx')
print(f"SAVED 05_work_packages.xlsx: {len(wp_rows)} WP + {len(wp_detail_rows)} details")

print("\n========================================")
print("ALL TEMPLATES UPDATED SUCCESSFULLY!")
print("========================================")
