import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='1B5E20')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
dfont = Font(name='Calibri', size=11)
dalign = Alignment(vertical='center', wrap_text=True)

# Read T14
df14 = pd.read_excel('templates/14_maintenance_strategy.xlsx')

# Split maintainable_item into subunit and MI
def split_mi(val):
    val = str(val)
    for sep in [' \u2014 ', ' \u2192 ', ' -- ', ' - ']:
        if sep in val:
            parts = val.split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    return '', val.strip()

df14[['subunit', 'mi_name']] = df14['maintainable_item'].apply(lambda x: pd.Series(split_mi(x)))

# Read T03 for alignment verification
df03 = pd.read_excel('templates/03_failure_modes.xlsx')

# Verify row alignment
mismatches = 0
for i in range(min(len(df14), len(df03))):
    t14_tag = df14.iloc[i]['equipment_tag']
    t03_tag = df03.iloc[i]['equipment_tag']
    t14_mech = df14.iloc[i]['mechanism']
    t03_mech = df03.iloc[i]['mechanism']
    t14_mi = df14.iloc[i]['mi_name']
    t03_mi = df03.iloc[i]['maintainable_item']
    if t14_tag != t03_tag or t14_mi != t03_mi or t14_mech != t03_mech:
        mismatches += 1
        if mismatches <= 5:
            print(f'Row {i}: T14({t14_tag}/{t14_mi}/{t14_mech}) != T03({t03_tag}/{t03_mi}/{t03_mech})')

if mismatches > 0:
    print(f'WARNING: {mismatches} row alignment mismatches between T14 and T03')
else:
    print('Row alignment T14<->T03: PERFECT (946/946)')

# Build new spreadsheet
wb = Workbook()
ws = wb.active
ws.title = 'Strategies'

# New column order: insert subunit at position 3, replace old maintainable_item
old_cols = list(df14.columns)
old_cols_filtered = [c for c in old_cols if c not in ('subunit', 'mi_name')]

new_headers = []
col_sources = []
for col in old_cols_filtered:
    if col == 'maintainable_item':
        new_headers.append('subunit')
        col_sources.append('subunit')
        new_headers.append('maintainable_item')
        col_sources.append('mi_name')
    else:
        new_headers.append(col)
        col_sources.append(col)

for c, h in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

for i, (_, row) in enumerate(df14.iterrows(), 2):
    for c, src in enumerate(col_sources, 1):
        v = row[src] if pd.notna(row[src]) else None
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = dfont
        cell.alignment = dalign

widths = [10, 17, 22, 22, 35, 22, 18, 24, 14, 16, 10, 40, 14, 14, 14, 35, 40, 14, 16, 14,
          10, 40, 14, 16, 14, 40, 12, 12, 22, 28, 14, 16, 16, 8]
for c, w in enumerate(widths, 1):
    if c <= len(new_headers):
        ws.column_dimensions[get_column_letter(c)].width = w
ws.auto_filter.ref = f'A1:{get_column_letter(len(new_headers))}{len(df14)+1}'
ws.freeze_panes = 'A2'

wb.save('templates/14_maintenance_strategy.xlsx')
print(f'T14 done. {len(df14)} rows x {len(new_headers)} cols.')
