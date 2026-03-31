import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

hier = pd.read_excel('templates/01_equipment_hierarchy.xlsx')
fm = pd.read_excel('templates/03_failure_modes.xlsx')

EI = {}
for _, r in hier.iterrows():
    EI[r['equipment_tag']] = (r['equipment_type'], int(r['power_kw']))

EF = {
    'SAG_MILL': 'To reduce ROM ore to target P80 product size at design throughput using {p} kW semi-autogenous grinding',
    'BALL_MILL': 'To grind classified slurry to final target P80 at design throughput using {p} kW ball milling',
    'HYDROCYCLONE': 'To classify ground slurry by particle size at design cut point, directing fines to overflow and coarse to underflow',
    'FLOTATION_CELL': 'To recover target mineral by froth flotation at design recovery and concentrate grade using {p} kW agitation and air injection',
    'REAGENT_PUMP': 'To meter and deliver reagent at design dosing rate and accuracy using {p} kW diaphragm pumping',
    'THICKENER': 'To separate solids from liquid by gravity sedimentation to design underflow density and overflow clarity using {p} kW rake drive',
    'SLURRY_PUMP': 'To deliver slurry at design flow rate and head using {p} kW centrifugal pumping',
    'BELT_FILTER': 'To dewater concentrate to target moisture content using {p} kW belt filter pressing',
    'DISC_FILTER': 'To dewater slurry to target cake moisture using {p} kW vacuum disc filtration',
    'ROTARY_DRYER': 'To reduce product moisture to target specification using {p} kW direct-heat rotary drying',
    'ROTARY_COOLER': 'To cool dried product to target storage temperature using {p} kW rotary cooling',
    'BELT_CONVEYOR': 'To transport material at design capacity over design distance using {p} kW belt conveying',
    'STACKER': 'To stack material onto stockpile at design rate using {p} kW stacking system',
    'RECLAIMER': 'To reclaim material from stockpile at design rate using {p} kW bucket wheel reclaiming',
    'SILO': 'To store product at design capacity and discharge on demand without degradation',
    'WATER_PUMP': 'To deliver process water at design flow rate and pressure using {p} kW centrifugal pumping',
    'GRANULATOR': 'To form granulated product of target size at design throughput using {p} kW drum granulation',
    'AMMONIATOR': 'To react ammonia with acid slurry forming granulated product at design throughput using {p} kW agitated drum',
    'SCREEN': 'To separate granulated product by size into design fractions using {p} kW vibrating screening',
    'CRUSHER': 'To reduce oversize granules to target recycle size using {p} kW chain crushing',
    'REACTOR_VESSEL': 'To contain and facilitate chemical reaction at design temperature, pressure, and residence time using {p} kW agitation',
    'AGITATOR': 'To provide uniform mixing and solids suspension in reactor at design speed using {p} kW agitation',
    'FLASH_COOLER': 'To cool reaction slurry by flash evaporation to target temperature at design flow rate',
    'HEAT_EXCHANGER': 'To transfer heat between process fluids at design duty and approach temperature without cross-contamination',
    'ACID_PUMP': 'To deliver sulfuric acid at design flow rate and pressure using {p} kW acid-resistant pumping',
    'AIR_COMPRESSOR': 'To deliver compressed air at design pressure and flow rate using {p} kW rotary screw compression',
    'AIR_DRYER': 'To remove moisture from compressed air to target dew point using {p} kW desiccant drying',
    'COOLING_TOWER': 'To cool water from design inlet to outlet temperature using {p} kW forced-draft evaporative cooling',
    'TRANSFORMER': 'To step down high voltage to distribution voltage at 40 MVA rated capacity without interruption',
    'GENERATOR': 'To generate emergency power at {p} kW rated capacity upon loss of main supply within design start time',
}

EF_TAG = {
    'CYC-PP': 'To deliver slurry to hydrocyclone cluster at design flow rate and pressure using {p} kW centrifugal pumping',
    'THK-PP': 'To transfer thickener underflow at design flow rate and density using {p} kW centrifugal pumping',
    'MDP-PP': 'To transfer mill discharge slurry to classification at design flow rate using {p} kW centrifugal pumping',
    'PWS-PP': 'To supply process water to plant at design flow rate and pressure using {p} kW centrifugal pumping',
    'CWT-PP': 'To circulate cooling water at design flow rate and pressure using {p} kW centrifugal pumping',
    'RGH-FC': 'To recover target mineral in rougher stage at design recovery using {p} kW agitation and air injection',
    'CLN-FC': 'To upgrade rougher concentrate to final grade in cleaner stage using {p} kW agitation and air injection',
    'OLC-CV-001': 'To transport ore 3.5 km overland at design capacity using {p} kW belt conveying',
    'PRD-CV': 'To convey finished product to storage at design capacity using {p} kW belt conveying',
}

def ef(tag):
    et, p = EI.get(tag, ('', 0))
    for pat, t in EF_TAG.items():
        if pat in tag:
            return t.format(p=p)
    return EF.get(et, 'To operate at design parameters').format(p=p)

def eff(tag, ft):
    f = ef(tag)
    c = f[3:] if f.startswith('To ') else f
    return f'Unable to {c}' if ft == 'TOTAL' else f'{c[0].upper()}{c[1:]} outside acceptable performance limits'

MF = {
    'Drive Motor': 'To convert electrical energy into rotational power at rated speed and torque to drive the equipment',
    'Pump Motor': 'To provide rotational power at rated speed and torque to drive the pump impeller',
    'Agitator Motor': 'To provide rotational power at rated speed to drive the agitator shaft and impeller',
    'Compressor Motor': 'To provide rotational power at rated speed and torque to drive the compressor rotor',
    'Fan Motor': 'To provide rotational power at rated speed to drive the cooling fan blades',
    'Vibrator Motor': 'To generate controlled vibration at design frequency and amplitude to drive the screen deck',
    'Boom Conveyor Motor': 'To provide rotational power to drive the boom conveyor belt at design speed',
    'Slew Motor': 'To provide rotational power to slew the boom at design angular velocity',
    'Travel Motor': 'To provide rotational power to propel the equipment along rails at design speed',
    'Boom Hoist Motor': 'To raise and lower the boom at design speed and load capacity',
    'Bucket Wheel Motor': 'To rotate the bucket wheel at design speed for material reclaiming',
    'Blower': 'To deliver air at design flow rate and pressure to the flotation cell sparger',
    'Main Gearbox': 'To reduce motor speed and multiply torque to design output for the driven equipment',
    'Conveyor Gearbox': 'To reduce motor speed to design belt speed with required torque multiplication',
    'Filter Gearbox': 'To reduce motor speed to design filter operating speed with required torque',
    'Thickener Gearbox': 'To reduce motor speed and multiply torque for rake rotation at design speed',
    'Travel Gearbox': 'To reduce motor speed to design travel speed with required torque multiplication',
    'Agitator Gearbox': 'To reduce motor speed to design agitator speed with required torque',
    'Crusher Gearbox': 'To reduce motor speed to design crushing roll speed with required torque',
    'Fan Gearbox': 'To reduce motor speed to design fan speed with required torque',
    'Motor Coupling': 'To transmit torque from motor to driven equipment while accommodating design misalignment',
    'Mill Liner': 'To protect the mill shell from wear and provide ore lifting action for grinding efficiency',
    'Lifter Bar': 'To lift and cascade grinding media and ore for effective size reduction',
    'Shell': 'To contain ore, grinding media, and slurry within the mill during operation',
    'Pinion': 'To transmit rotational power from gearbox to girth gear at design torque',
    'Feed Chute': 'To direct feed material into the mill trunnion without spillage or blockage',
    'Trommel Screen': 'To screen mill discharge by size, retaining oversize for continued grinding',
    'Discharge Grate': 'To retain grinding media inside the mill while allowing ground slurry to discharge',
    'Feed Trunnion Bearing': 'To support the feed-end trunnion and allow rotation under full load with minimal friction',
    'Discharge Trunnion Bearing': 'To support the discharge-end trunnion and allow rotation under full load with minimal friction',
    'Lube Oil Pump': 'To circulate lubricating oil at design flow rate and pressure to all bearing surfaces',
    'Lube Oil Filter': 'To remove contaminants from lubricating oil to maintain target cleanliness level',
    'Oil Cooler': 'To cool lubricating oil to target return temperature for bearing protection',
    'Oil Reservoir': 'To store lubricating oil at design volume with settling and deaeration capacity',
    'Oil Filter': 'To remove contaminant particles from oil to maintain target cleanliness',
    'Oil Pump': 'To circulate oil at design flow rate and pressure through the cooling or lubrication circuit',
    'Oil Separator': 'To separate oil from compressed air to maintain target oil carryover level',
    'Speed Sensor': 'To measure and transmit rotational speed within design accuracy for monitoring and control',
    'Temperature Sensor': 'To measure and transmit temperature within design accuracy for monitoring and protection',
    'Vibration Sensor': 'To measure and transmit vibration within design accuracy for condition monitoring',
    'Level Sensor': 'To measure and transmit material level within design accuracy for inventory control',
    'Weight Sensor': 'To measure and transmit weight within design accuracy for process control',
    'Pressure Sensor': 'To measure and transmit pressure within design accuracy for monitoring and control',
    'Temperature Gauge': 'To indicate temperature within design accuracy for local monitoring',
    'Impeller': 'To impart kinetic energy to pumped fluid generating design flow and head',
    'Mechanical Seal': 'To prevent fluid leakage along the shaft while allowing rotation',
    'Shaft Sleeve': 'To protect the shaft from wear and corrosion in the seal area',
    'Bearing Housing': 'To support shaft bearings, maintain alignment, and contain lubricant',
    'DE Bearing': 'To support drive-end shaft loads and maintain alignment with minimal friction',
    'NDE Bearing': 'To support non-drive-end shaft loads and maintain alignment with minimal friction',
    'Frame Plate Liner': 'To protect the pump frame plate from erosion by abrasive slurry',
    'Volute Liner': 'To protect the pump volute casing from erosion by abrasive slurry',
    'Throat Bush': 'To minimize recirculation between impeller discharge and suction',
    'Casing': 'To contain pumped fluid at design pressure and direct flow to discharge',
    'Wear Ring': 'To maintain design clearance between impeller and casing for efficiency',
    'Casing Liner': 'To protect the pump casing from chemical attack and erosion',
    'Vortex Finder': 'To direct the inner vortex of fine particles upward to the overflow outlet',
    'Spigot': 'To discharge coarse underflow at controlled density and flow rate',
    'Cylinder Section': 'To contain slurry and develop the primary vortex at design feed pressure',
    'Feed Box': 'To distribute feed slurry evenly to hydrocyclone cluster inlets',
    'Feed Inlet': 'To introduce feed slurry tangentially into cyclone body to generate the vortex',
    'Agitator Shaft': 'To transmit rotational power from drive to impeller at design speed',
    'Air Pipe': 'To deliver air from blower to flotation cell at design flow rate',
    'Sparger': 'To disperse air into fine bubbles within the cell for mineral-bubble attachment',
    'Launder': 'To collect and convey mineralized froth from cell surface to concentrate stream',
    'Weir Plate': 'To control pulp level within the flotation cell at design operating level',
    'Agitator Impeller': 'To generate turbulent mixing and disperse air bubbles throughout reactor volume',
    'Check Valve': 'To prevent reverse flow of reagent through the pump',
    'Diaphragm': 'To displace reagent at precise metered volume per stroke without leakage',
    'Pump Head': 'To contain reagent at design pressure and direct flow through valves',
    'Rake Arm': 'To convey settled solids from periphery to central underflow discharge',
    'Rake Blade': 'To scrape settled solids along the floor toward the discharge cone',
    'Feedwell': 'To dissipate feed energy and distribute slurry into the settling zone',
    'Distributor': 'To distribute feed slurry evenly across the feedwell',
    'Torque Limiter': 'To protect rake drive from overload by disengaging at design torque threshold',
    'Cone Liner': 'To protect thickener cone from erosion by settled solids',
    'Underflow Valve': 'To control underflow discharge rate and density from the thickener',
    'Filter Cloth': 'To retain solids while passing filtrate under design pressure differential',
    'Vacuum Pump': 'To generate and maintain design vacuum for filtrate extraction',
    'Vacuum Receiver': 'To stabilize vacuum and separate filtrate from air in the vacuum circuit',
    'Spray Bar': 'To deliver wash water across the cake surface for moisture removal',
    'Wash Nozzle': 'To spray wash water onto filter cake at design pressure and coverage',
    'Belt Tracking Roller': 'To maintain filter belt alignment within design tracking tolerance',
    'Filter Disc': 'To provide filtration surface and support filter cloth during rotation',
    'Disc Shaft': 'To support and rotate filter discs at design speed through slurry trough',
    'Dryer Shell': 'To contain material during drying and transmit rotational motion',
    'Cooler Shell': 'To contain material during cooling and transmit rotational motion',
    'Lifting Flight': 'To lift and cascade material through gas stream for heat transfer',
    'Girth Gear': 'To transmit rotational power from pinion to shell at design torque',
    'Riding Ring': 'To transfer shell weight to support rollers while allowing rotation',
    'Support Roller': 'To support shell riding rings and allow rotation under full load',
    'Thrust Roller': 'To maintain axial position of the rotating shell within design limits',
    'Inlet Seal': 'To prevent air ingress and gas egress at the shell inlet',
    'Outlet Seal': 'To prevent air ingress and gas egress at the shell outlet',
    'Main Burner': 'To combust fuel and generate hot gas at design temperature for drying',
    'Fuel Valve': 'To regulate fuel flow to burner at design rate for temperature control',
    'Ignition System': 'To ignite the main burner fuel reliably on demand',
    'Conveyor Belt': 'To carry and transport material at design capacity and belt speed',
    'Head Pulley Drum': 'To drive the belt by friction at head end at design speed',
    'Tail Pulley Drum': 'To redirect and maintain belt tension at the tail end',
    'Head Pulley Bearing': 'To support head pulley shaft under full belt tension with minimal friction',
    'Tail Pulley Bearing': 'To support tail pulley shaft under belt tension with minimal friction',
    'Pulley Lagging': 'To increase friction between pulley and belt preventing slippage',
    'Carry Idler Set': 'To support loaded belt and maintain belt profile on carry side',
    'Return Idler Set': 'To support return belt and prevent sagging between support points',
    'Impact Idler Set': 'To absorb impact at loading points and protect belt from damage',
    'Belt Splice': 'To join belt ends maintaining design strength and flexibility',
    'Belt Scraper': 'To remove carryback material from belt at the head pulley',
    'Boom Structure': 'To support the boom conveyor or bucket wheel at design load',
    'Slew Bearing': 'To support slewing loads and allow boom rotation about vertical axis',
    'Travel Wheel': 'To support equipment weight and enable rail travel at design speed',
    'Bucket': 'To scoop material from stockpile face at design fill factor',
    'Wheel Bearing': 'To support bucket wheel shaft under load with minimal friction',
    'Discharge Gate': 'To control material discharge from silo at design flow rate',
    'Vibrating Feeder': 'To promote flow from silo and meter discharge at design rate',
    'Drum Shell': 'To contain material during granulation and transmit rotation',
    'Drum Liner': 'To protect drum shell from wear and chemical attack',
    'Drum Seal': 'To prevent material leakage at drum inlet and outlet during rotation',
    'Spray Nozzle': 'To atomize and deliver binder liquid at design spray rate and pattern',
    'Spray Pipe': 'To distribute liquid to spray nozzles along drum length',
    'Ammonia Valve': 'To regulate ammonia flow into drum at design rate for reaction control',
    'Distributor Pipe': 'To distribute ammonia uniformly along the drum length',
    'Rubber Lining': 'To protect internal surfaces from chemical attack and corrosion',
    'Screen Deck': 'To separate material by size at design cut point and efficiency',
    'Screen Frame': 'To support screen decks and contain material during screening',
    'Spring Mounts': 'To isolate screen vibration from structure and maintain design motion',
    'Vibrator Bearing': 'To support vibrator shaft at design speed under dynamic loads',
    'Crushing Roll': 'To fracture oversize material by compression to target size',
    'Roll Bearing': 'To support crushing roll shaft under full load with minimal friction',
    'Roll Shaft': 'To transmit power to crushing roll and withstand crushing forces',
    'Vessel Shell': 'To contain process fluids at design pressure, temperature, and corrosion resistance',
    'Vessel Nozzle': 'To provide leak-tight piping connection at design pressure and temperature',
    'Cooling Coil': 'To remove reaction heat maintaining design temperature within the vessel',
    'Expansion Joint': 'To absorb thermal expansion without imposing stress on vessel nozzles',
    'Shaft Seal': 'To prevent process fluid leakage along the agitator shaft',
    'Flash Vessel': 'To contain slurry during flash evaporation at design vacuum level',
    'Inlet Pipe': 'To deliver hot slurry to flash vessel at design flow rate',
    'Outlet Valve': 'To control cooled slurry discharge from flash vessel',
    'Tube Bundle': 'To provide heat transfer surface at design duty and pressure',
    'Tube Sheet': 'To support tube ends and separate shell-side from tube-side fluids',
    'HX Shell': 'To contain shell-side fluid at design pressure directing flow across tubes',
    'Baffle Plates': 'To direct shell-side flow across tube bundle for heat transfer',
    'Header Gasket': 'To seal between header and tube sheet at design pressure',
    'Flange Gasket': 'To seal flanged connections at design pressure without leakage',
    'Rotor Element': 'To compress air from inlet to design discharge pressure at rated flow',
    'Inlet Valve': 'To regulate air intake for compressor capacity control',
    'Discharge Valve': 'To control compressed air output and prevent backflow',
    'Aftercooler': 'To cool compressed air to target downstream temperature',
    'Cooling Fan': 'To generate airflow for cooling at design rate',
    'Desiccant Bed': 'To adsorb moisture from compressed air to target dew point',
    'Switching Valve': 'To alternate airflow between towers for regeneration cycling',
    'PLC Controller': 'To control sequencing, valve timing, and alarms per design logic',
    'Fill Pack': 'To maximize water-air contact area for evaporative heat transfer',
    'Fan Blades': 'To generate forced-draft airflow through fill pack at design volume',
    'Distribution Basin': 'To distribute hot water uniformly across fill pack surface',
    'Distribution Nozzle': 'To spray water evenly across fill pack for uniform cooling',
    'Drift Eliminator': 'To capture water droplets from discharge minimizing water loss',
    'Primary Winding': 'To receive high voltage and create magnetic flux in transformer core',
    'Secondary Winding': 'To induce and deliver output voltage at design ratio and capacity',
    'Transformer Core': 'To provide low-reluctance magnetic path between windings',
    'Buchholz Relay': 'To detect internal faults by gas or oil surge and trip the transformer',
    'PRD Valve': 'To relieve excess pressure in transformer tank preventing rupture',
    'Radiator': 'To dissipate heat to ambient and maintain design operating temperature',
    'Engine Block': 'To convert fuel combustion into rotational power at rated capacity',
    'Rotor Assembly': 'To generate rotating magnetic field for power output at design frequency',
    'Stator Winding': 'To convert magnetic field into electrical power at design voltage',
    'AVR': 'To regulate output voltage within design tolerance under varying load',
    'Fuel Injection Pump': 'To deliver precise fuel at design timing and pressure to cylinders',
    'Coolant Pump': 'To circulate engine coolant at design flow rate for temperature control',
    'Exhaust System': 'To discharge combustion gases safely and attenuate noise to design limits',
}

MF_CTX = {
    ('Impeller', 'FLOTATION_CELL'): 'To create turbulent mixing and disperse air bubbles for mineral-bubble attachment throughout the cell',
    ('Rubber Lining', 'REACTOR_VESSEL'): 'To protect reactor vessel internals from acid attack by process fluids',
    ('Rubber Lining', 'AMMONIATOR'): 'To protect drum internals from acid attack and promote material bed buildup',
}

def mf(comp, tag):
    et = EI.get(tag, ('', 0))[0]
    return MF_CTX.get((comp, et), MF.get(comp, f'To perform the design function of the {comp}'))

def mff(comp, tag, ft):
    f = mf(comp, tag)
    c = f[3:] if f.startswith('To ') else f
    return f'Unable to {c}' if ft == 'TOTAL' else f'{c[0].upper()}{c[1:]} outside acceptable performance limits'

wb = Workbook()
ws = wb.active
ws.title = 'failure_modes'

headers = ['equipment_tag', 'Equipment_function_description', 'Equipment_functional_failure',
           'function_type', 'failure_type', 'Maintainable_item',
           'Maintainable_item_function_description', 'Maintainable_item_functional_failure',
           'mechanism', 'cause', 'failure_pattern', 'failure_consequence',
           'evidence', 'downtime_hours', 'detection_method',
           'rpn_severity', 'rpn_occurrence', 'rpn_detection']

hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='1B5E20')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

dfont = Font(name='Calibri', size=11)
dalign = Alignment(vertical='center', wrap_text=True)

for i, (_, row) in enumerate(fm.iterrows(), 2):
    tag, ft, comp = row['equipment_tag'], row['failure_type'], row['what_component']
    vals = [tag, ef(tag), eff(tag, ft), row['function_type'], ft, comp, mf(comp, tag), mff(comp, tag, ft),
            row['mechanism'], row['cause'], row['failure_pattern'], row['failure_consequence'],
            row['evidence'] if pd.notna(row['evidence']) else None,
            row['downtime_hours'], row['detection_method'],
            row['rpn_severity'], row['rpn_occurrence'], row['rpn_detection']]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = dfont
        cell.alignment = dalign

widths = [17, 50, 50, 15, 14, 24, 50, 50, 18, 26, 19, 23, 12, 16, 20, 14, 16, 14]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.auto_filter.ref = f'A1:R{len(fm)+1}'
ws.freeze_panes = 'A2'

wb.save('templates/03_failure_modes.xlsx')
print(f'Done. {len(fm)} rows written with {len(headers)} columns.')
