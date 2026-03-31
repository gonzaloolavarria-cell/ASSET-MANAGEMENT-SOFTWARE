import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re

hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='1B5E20')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
dfont = Font(name='Calibri', size=11)
dalign = Alignment(vertical='center', wrap_text=True)

# Read T01 for valid tags
t01 = pd.read_excel('templates/01_equipment_hierarchy.xlsx', sheet_name='Equipment Hierarchy')
iso_offset = 0
if 'iso14224_level_1_business' in t01.columns:
    iso_offset = 4
valid_tags = set(t01['equipment_tag'].unique())

# Read T04
df = pd.read_excel('templates/04_maintenance_tasks.xlsx')

# Extract equipment_tag from task_name
TAG_PATTERN = re.compile(r'on\s+([A-Z]{3}-[A-Z]{3}-[A-Z]{2}-\d{3})')

def extract_tag(name):
    m = TAG_PATTERN.search(str(name))
    return m.group(1) if m else None

df['equipment_tag'] = df['task_name'].apply(extract_tag)
missing = df[df['equipment_tag'].isna()]
if len(missing) > 0:
    print(f'WARNING: {len(missing)} tasks have no extractable equipment_tag')
    for _, r in missing.iterrows():
        print(f'  {r["task_id"]}: {r["task_name"]}')

invalid = df[~df['equipment_tag'].isin(valid_tags) & df['equipment_tag'].notna()]
if len(invalid) > 0:
    print(f'WARNING: {len(invalid)} tasks have equipment_tag not in T01')

# Build new spreadsheet
wb = Workbook()
ws = wb.active
ws.title = 'Tasks'

old_cols = [c for c in df.columns if c != 'equipment_tag']
new_headers = ['task_id', 'equipment_tag'] + [c for c in old_cols if c != 'task_id']

for c, h in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign

for i, (_, row) in enumerate(df.iterrows(), 2):
    vals = [row['task_id'], row['equipment_tag']] + [
        row[c] if pd.notna(row[c]) else None for c in old_cols if c != 'task_id'
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = dfont
        cell.alignment = dalign

widths = [10, 17, 45, 45, 12, 12, 16, 40, 45, 12, 12, 22, 28, 35, 35, 14, 8]
for c, w in enumerate(widths, 1):
    if c <= len(widths):
        ws.column_dimensions[get_column_letter(c)].width = w
ws.auto_filter.ref = f'A1:{get_column_letter(len(new_headers))}{len(df)+1}'
ws.freeze_panes = 'A2'

wb.save('templates/04_maintenance_tasks.xlsx')
tags_found = df['equipment_tag'].nunique()
print(f'T04 done. {len(df)} rows x {len(new_headers)} cols. {tags_found} unique equipment_tags extracted.')
