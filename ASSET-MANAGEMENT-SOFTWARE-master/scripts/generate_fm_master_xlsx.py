"""
Generate FM-MASTER-REFERENCE.xlsx — Multi-sheet Excel reference for 72 Failure Modes.

Reads synthesized data from MASTER.md and produces a professional Excel workbook
with 6 sheets: 72 Failure Modes, Validation Matrix, By Pattern, By Frequency Basis,
By ISO 14224, and Cause Cross-Reference.

Usage: python scripts/generate_fm_master_xlsx.py
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. DATA — All 72 FM entries (extracted from MASTER.md)
# ---------------------------------------------------------------------------

FM_DATA = [
    # (FM#, Mechanism, Cause, FreqBasis, Pattern, ISO, WeibullBeta, WeibullEta,
    #  DegradationSummary, TopPConditions, EquipmentClasses, OCPEquipment,
    #  PrimaryCBMTechnique, PFInterval, ReferenceStd, StrategyCB, StrategyFT, StrategyRTF, KeyThreshold)

    ("FM-01", "Arcs", "Breakdown in insulation", "Calendar", "B", "4.1/4.5",
     "2.5-3.5", "15,000-25,000 h",
     "Insulation aging follows Arrhenius relationship; partial discharges create self-accelerating erosion until arc flash >5,000C. At OCP, high ambient near dryers, phosphate dust, and coastal humidity accelerate HV motor insulation degradation.",
     "Declining IR (>10%/yr), PD activity >100 pC, tan delta trending >0.5% above baseline",
     "EM, PT, SG, EG, FC, PC", "SAG/ball mill drives, substation transformers",
     "Insulation resistance testing (megger)", "6-12 mo", "IEEE 43, IEC 60085",
     "Measure insulation resistance on motor winding", "Replace cable terminations per insulation class thermal life (Class F: 20 yr)", "Only redundant non-critical supplies",
     "IR <50 MOhm: rewind in 30d; IR <1 MOhm/kV: do not energize"),

    ("FM-02", "Blocks", "Contamination", "Calendar", "C", "5.1",
     "2.0-3.0", "1,000-5,000 h",
     "Foreign particles, scale, biological growth progressively accumulate on internal surfaces, reducing effective flow cross-section until functional failure. At OCP, gypsum scale in phosphoric acid exchangers, phosphate clay deposits, and marine biofouling at Jorf Lasfar/Safi.",
     "Increasing dP (>50% above clean baseline), decreasing flow rate (>10% below design), heat exchanger approach temperature rising",
     "FS, VA, PU, HE, PI, ID, CS", "Slurry filters, seawater coolers",
     "Differential pressure measurement", "1-4 wk", "OEM max dP spec",
     "Measure dP on filter", "Replace filter element per OEM interval (slurry 1-3 mo; seawater 3-6 mo)", "Only duplex filter with auto-switchover",
     "dP 2-3x baseline: clean within 2 wk; dP >3x: clean immediately"),

    ("FM-03", "Blocks", "Excessive particle size", "Operational", "E", "5.1",
     "0.8-1.2", "Highly variable",
     "Particles exceeding design passage size lodge at restrictions, causing sudden flow obstruction. At OCP, critical at trommel screens downstream of SAG mills, hydrocyclone feed distributors, and pinch valve orifices in slurry circuits.",
     "Sudden step-change dP increase (>100% in minutes), upstream pressure spike >15%, flow rate drop >20%",
     "FS, VA, PU, HC, PI, ID, ND", "Trommel screens, slurry pinch valves",
     "dP monitoring with alarm", "Minutes-hours", "OEM max dP spec",
     "Monitor dP on suction strainer - alarm-driven", "Inspect screen panels at every shutdown (weekly vibrating, monthly trommel)", "Acceptable for duplex strainers with auto-switchover",
     "Sudden dP >100% baseline: clear blockage; >2 events/month: investigate upstream screening"),

    ("FM-04", "Blocks", "Insufficient fluid velocity", "Operational", "D", "5.1",
     "0.8-1.5", "Highly variable",
     "Flow velocity below critical deposition velocity (Vc) allows suspended solids to settle, forming a stationary bed that progressively reduces cross-section toward complete blockage. At OCP, critical for the 187 km Khouribga-Jorf Lasfar slurry pipeline.",
     "Velocity falling below design minimum, increasing pipeline dP at constant flow, density meter showing lower solids at discharge vs inlet",
     "PI, PU, VA, GF, TA, HE, HC", "Khouribga-Jorf slurry pipeline",
     "Pipeline flow velocity monitoring (electromagnetic/ultrasonic)", "Continuous", "Durand equation",
     "Monitor slurry velocity >= 1.2xVc", "Water flush at every planned shutdown (before pipeline sits idle >4 hr)", "NEVER for long-distance slurry pipelines or setting materials",
     "Velocity <1.0xVc for >30 min: initiate water flush; complete blockage: do NOT increase pump pressure"),

    ("FM-05", "Breaks/Fracture/Separates", "Cyclic loading (thermal/mechanical)", "Operational", "B", "2.6",
     "2.5-5.0", "10^6-10^8 cycles",
     "Three-stage fatigue: crack initiation at stress concentrators, stable propagation per Paris' law, then sudden final fracture. At OCP, prevalent in SAG/ball mill trunnion shafts (~5x10^6 cycles/yr), vibrating screen side plates (10^8+ cycles/yr).",
     "Surface cracks detectable by MPI/DPI at stress concentrations, vibration amplitude change (stiffness reduction), oxide staining at crack mouths",
     "ML, VS, PU, CV, PI, VE, RO, CR", "SAG mill trunnion, vibrating screens",
     "MPI at known hot spots", "6-12 mo", "ASME V Art 7, ISO 9934",
     "Perform MPI on mill trunnion shaft", "Replace vibrating screen side plates per design fatigue life (2-5 yr) with SF >= 4", "NEVER for rotating shafts, pressure boundaries, or lifting equipment",
     "Crack <25% section: monitor quarterly; 25-50%: repair within 30d; >50%: remove from service immediately"),

    ("FM-06", "Breaks/Fracture/Separates", "Mechanical overload", "Operational", "E", "2.5",
     "0.8-1.2", "Highly variable",
     "Single-event load exceeds ultimate material strength causing immediate separation. At OCP, common in crusher toggle plates (designed fuse), SAG mill liner bolts during charge cascading, pump shafts from tramp material impact.",
     "Visible deformation/bending at connections, section loss from corrosion reducing safety factor, bolt stretch >1% (ultrasonic gauging)",
     "CU, ML, PU, CV, ST, VA, GB, CR", "Crusher toggle plates, mill liner bolts",
     "UT thickness at corroded/worn sections", "3-12 mo", "API 574, ASME B31.3",
     "Inspect structural condition of crusher frame", "Replace toggle plates at every major shutdown; belt splices every 12-24 mo", "Acceptable for designed sacrificial elements (toggle plates, shear pins)",
     "Section loss >30%: de-rate and plan replacement; visible plastic deformation: repair within 30d"),

    ("FM-07", "Breaks/Fracture/Separates", "Thermal overload", "Operational", "E", "2.5",
     "0.8-1.2", "Highly variable",
     "Thermal shock fracture or high-temperature embrittlement (>425C causes graphitization in carbon steel, sigma phase in stainless). At OCP, kiln/dryer refractory failure exposing shell to >800C, cast iron pump casings during water quench.",
     "Surface temperature exceeding design rating, heat discoloration (straw 200C, blue 300C, grey 400C+), refractory spalling exposing structural shell",
     "RO, HE, VE, FU, PU, PI, VA, CS", "Kilns, acid coolers",
     "Thermal scan on kiln shell", "1-4 wk", "ISO 18434-1",
     "Thermal scan kiln shell", "Internal refractory inspection at every annual shutdown; replace sections <50% design thickness", "NEVER for pressure-containing or structural elements",
     "Shell temp >350C (CS): reduce firing, plan refractory repair; >400C: emergency repair"),

    ("FM-08", "Corrodes", "Bio-organisms", "Calendar", "B", "2.2",
     "1.5-3.0", "15,000-60,000 h",
     "MIC occurs when sulfate-reducing, iron-oxidizing, and acid-producing bacteria colonize metal surfaces, forming biofilms that create localized acidic, oxygen-depleted micro-environments driving pitting at 1-5 mm/yr. At OCP, cooling water systems at Jorf Lasfar, stagnant fire water dead legs.",
     "Black FeS deposits under tubercles (H2S odor), localized hemispherical pitting at 6 o'clock, planktonic bacterial counts >10^4 CFU/mL",
     "HE, PI, TA, PU, VA, CT, FP", "Jorf Lasfar cooling water exchangers, fire water systems",
     "Bacterial count monitoring (planktonic + sessile)", "1-3 mo", "NACE TM0194, ASTM D6974",
     "Monitor bacterial counts in cooling water", "Fire water flush/inspection annually per NFPA 25; biocide shock dosing quarterly", "Only sacrificial monitoring components (corrosion coupons)",
     "Planktonic >10^4 CFU/mL: review biocide; sessile >10^5 CFU/cm2: mechanical cleaning required"),

    ("FM-09", "Corrodes", "Chemical attack", "Calendar", "B", "2.2",
     "2.0-3.5", "10,000-50,000 h",
     "Direct chemical dissolution where process chemicals react with equipment material, causing progressive mass loss following Arrhenius kinetics. At OCP, phosphoric acid (28-54% H3PO4) attacks carbon steel at 2-10 mm/yr, sulfuric acid is aggressive to most metals.",
     "Progressive wall thinning by UT (uniform loss pattern), lining degradation (blistering, delamination), corrosion coupon weight loss >2x design rate",
     "RE, TA, PU, PI, HE, VA, FS", "Phosphoric acid attack tanks, H2SO4/H3PO4 storage",
     "UT thickness measurement at TMLs", "3-6 mo", "API 510, API 574, API 570",
     "Measure wall thickness on acid piping", "Reline acid tanks per lining life (rubber 5-8 yr, PTFE 10-15 yr, brick 10-20 yr)", "Only sacrificial wear components (corrosion coupons, gaskets)",
     "Corrosion rate >design allowance: investigate root cause; lining holiday: repair within 30d"),

    ("FM-10", "Corrodes", "Corrosive environment", "Calendar", "B", "2.2",
     "2.0-3.0", "15,000-80,000 h",
     "Environmental electrochemical attack from humidity, airborne contaminants, and process spillage; rate increases sharply above 60% RH. OCP coastal sites (Jorf Lasfar, Safi) face ISO 9223 C4-C5 marine atmosphere with chloride deposition 60-300 mg/m2/day.",
     "Visible rust and paint degradation (ASTM D714 blistering, D610 rusting), galvanized coating breakthrough, structural section loss at connections",
     "ST, PI, EI, TA, CV, CR, ID", "Pipe racks, cable trays, outdoor tanks",
     "Visual coating condition assessment", "6-12 mo", "ASTM D610, ASTM D714",
     "Inspect coating condition", "Repaint per ISO 9223 (C5 coastal 5-8 yr, C4 inland 7-10 yr); min 250 um DFT", "Only low-cost cosmetic items",
     "Coating <rating 6 per ASTM D610: plan spot repair; section loss >10%: immediate de-rate"),

    ("FM-11", "Corrodes", "Crevice", "Calendar", "B", "2.2",
     "1.5-3.0", "10,000-60,000 h",
     "Localized electrochemical attack at gasket-flange interfaces where oxygen depletion and chloride migration create acidic micro-environments. At OCP, phosphoric acid service and seawater systems at Jorf Lasfar (~19,000 ppm Cl-) drive severe crevice attack at flanges.",
     "Gasket face staining/seepage, weeping at flanged joints, under-deposit hemispherical pitting visible after cleaning",
     "PI, HE, PU, VA, TA, VE, FT", "Flanged acid piping, seawater exchangers",
     "UT thickness at flange faces and gasket lines", "6-12 mo", "API 574, API 570",
     "Inspect flange faces for crevice corrosion", "Inspect flange faces at every gasket change (2-4 yr); replace splash-zone bolting every 3-5 yr", "Only non-pressure cosmetic lap joints",
     "Crevice pitting >1 mm: engineer assessment; flange finish must be <=6.3 um Ra"),

    ("FM-12", "Corrodes", "Dissimilar metals contact", "Calendar", "B", "2.2",
     "2.0-3.5", "10,000-50,000 h",
     "Galvanic corrosion when two metals with different electrochemical potentials are electrically connected in an electrolyte. At OCP, carbon steel-to-stainless transitions, copper alloy tubes in steel tube sheets, and seawater systems at Jorf Lasfar.",
     "Accelerated corrosion of less noble metal at junction, distinctive wall thinning within 1-5 pipe diameters, galvanic potential >100 mV",
     "PI, HE, PU, VA, ST, EI, FT", "CS-to-SS transitions, copper tubes in steel shells",
     "UT thickness at dissimilar metal junctions", "6-12 mo", "API 574, API 570",
     "Measure wall thickness at dissimilar metal joint", "Insulating flange test annually per NACE SP0286; sacrificial anode replacement when >85% consumed (3-7 yr)", "Only intentionally sacrificial components (zinc anodes)",
     "Wall thinning >20% at junction: install insulating gasket kit; galvanic rate >2x general: redesign"),

    ("FM-13", "Corrodes", "Exposure to atmosphere", "Calendar", "B", "2.2",
     "1.5-2.5", "30,000-100,000 h",
     "Atmospheric corrosion under thin moisture films governed by time-of-wetness, amplified by airborne chlorides, SO2, and phosphate dust. OCP coastal sites (Jorf Lasfar, Safi) are ISO 9223 C4/C5 - among the most aggressive in North Africa.",
     "Progressive surface rust, coating chalking/cracking/blistering/delamination, galvanized white rust with red rust breakthrough",
     "ST, TA, PI, CR, CV, EI, CS", "All outdoor steelwork, pipe racks, cable trays",
     "Visual coating and corrosion condition assessment", "6-12 mo", "ASTM D610, ISO 4628",
     "Assess coating condition on outdoor structure", "Repaint per ISO 9223 (C5 coastal 6-8 yr, C3 inland 12-15 yr); min C5-M system >=320 um DFT", "Only low-cost non-structural items",
     "Coating <5 per ASTM D610 or adhesion <2.0 MPa: plan full recoating at next shutdown"),

    ("FM-14", "Corrodes", "Exposure to high temperature corrosive environment", "Calendar", "B", "2.2",
     "2.0-3.5", "8,000-40,000 h",
     "High-temperature corrosion above 200C through sulfidation, hot corrosion (molten Na2SO4+V2O5), and chloride-accelerated oxidation. At OCP, rotary kilns for phosphate calcination (800-1000C), dryers at 300-600C, and waste heat recovery.",
     "Multi-layered scale buildup (oxide+sulfide), tube metal temperature trending upward, vanadium-rich deposit accumulation",
     "RO, BO, HE, FU, PI, SK, FA", "Rotary kilns, superheater tubes, waste heat boilers",
     "Tube metal temperature monitoring (thermocouples, IR)", "Continuous/weekly", "API 530, ASME PCC-3",
     "Measure tube thickness on kiln hot zone", "Kiln refractory inspection every reline (12-24 mo); boiler tube inspection annually", "Only sacrificial replaceable internals (furnace baffles)",
     "Tube metal temp >max allowable per API 530: clean deposits; vanadium >50 ppm: add Mg-based additive"),

    ("FM-15", "Corrodes", "Exposure to high temperature environment", "Calendar", "B", "2.2",
     "2.0-3.5", "15,000-80,000 h",
     "High-temperature oxidation forming oxide scales following parabolic kinetics; carbon steel limited to ~540C, 304 SS to ~870C; thermal cycling accelerates by spalling protective scale. At OCP, kiln/dryer shell exterior, boiler steam-side magnetite growth.",
     "Progressive oxide scale with spallation, steam-side oxide exfoliation, bolt torque relaxation from thread oxidation",
     "RO, BO, FU, PI, HE, FT, FA", "Kiln/dryer shells, superheater tubes, high-temp bolting",
     "Surface temperature monitoring (thermocouples, IR pyrometry)", "Continuous/weekly", "API 530, ASME PCC-3",
     "Measure wall thickness on kiln shell hot zone", "Boiler chemical clean every 5-10 yr; kiln shell UT every reline; high-temp bolt replacement every 5-8 yr", "Only sacrificial internals (furnace baffles, heater elements)",
     "Steam-side oxide >500 um: plan chemical cleaning; high-temp bolt torque loss >20%: re-torque"),

    ("FM-16", "Corrodes", "Exposure to liquid metal", "Calendar", "E", "2.2",
     "0.8-1.5", "Hours (LME) to 10,000+ h",
     "Liquid metal embrittlement (LME) causes sudden catastrophic brittle cracking; liquid metal corrosion (LMC) is progressive dissolution. At OCP, primary risks are LME during welding on galvanized structures, galvanized high-strength bolt cracking, and babbitt bearing degradation.",
     "Intergranular cracking at galvanized weld HAZ, sudden brittle fracture of galvanized high-strength bolt, weld HAZ cracking",
     "ST, FT, BE, ID, PI, WS", "Galvanized structures requiring field welding, babbitt bearings",
     "MPI/DPI at welds on galvanized structures (post-weld)", "After each weld", "AWS D1.1, ISO 9934",
     "Inspect welds on galvanized structure for LME", "Mandatory zinc removal >=100 mm from weld zone; restrict bolt grade <=8.8 per ISO 10684", "Only sacrificial zinc anodes in CP systems",
     "Any LME crack at galvanized weld: stop operation, grind out completely, remove all zinc 100 mm, re-weld"),

    ("FM-17", "Corrodes", "Poor electrical connections", "Calendar", "C", "2.2",
     "1.5-3.0", "15,000-60,000 h",
     "Poor contact concentrates current through reduced cross-section, generating I2R resistive heating in a self-reinforcing positive feedback loop toward thermal runaway. At OCP, phosphate dust contaminates connections, coastal salt air accelerates oxidation.",
     "Elevated temperature at connections (IR dT >10C), discolored/oxidized surfaces, voltage drop >50 uOhm per bolt",
     "SG, TR, EM, CA, VD, CP", "MCC panels, motor terminal boxes, Cu-Al terminations",
     "IR thermographic survey", "3-12 mo", "NETA MTS, IEC 62446, NFPA 70B",
     "Perform IR survey on switchgear", "Re-torque MCC connections every 3-5 yr; Cu-Al inspection annually", "Only low-power control connections (<100A) with redundancy",
     "dT >70C (NETA Deficiency): de-energize and repair immediately - risk of thermal runaway/fire"),

    ("FM-18", "Corrodes", "Poor electrical insulation", "Calendar", "C", "2.2",
     "1.5-2.5", "20,000-80,000 h",
     "Stray current corrosion via Faraday's Law (1 A DC dissolves ~9 kg steel/yr at anodic discharge point). At OCP, buried piping near Jorf Lasfar CP systems, DC-powered draglines at Khouribga, and phosphate ore DC rail transport.",
     "Localized deep pitting on buried piping, pipe-to-soil potential positive shifts, insulating flange resistance <1 MOhm",
     "PI, TA, ST, CP, EI, RW", "Slurry pipelines, tank farms, Khouribga draglines and rail",
     "Pipe-to-soil potential survey", "3-12 mo", "NACE SP0169, ISO 15589-1",
     "Perform pipe-to-soil potential survey", "Insulating flange test annually; CP rectifier output monthly; CIPS every 3-5 yr", "Only non-critical above-ground structures",
     "Pipe-to-soil positive shift: investigate, install drainage bond; AC voltage >15V: install AC mitigation"),

    ("FM-19", "Cracks", "Age", "Calendar", "B", "2.5/2.6",
     "2.5-4.0", "50,000-200,000 h",
     "Age-induced cracking through SCC, hydrogen embrittlement, creep cracking, polymer chain scission. At OCP, stainless steel in hot phosphoric acid with chlorides (chloride SCC >50C), age-cracking rubber linings, and hydrogen-embrittled high-strength bolts.",
     "Surface cracks by DPI/MPI at welds and stress concentrations, elastomer crazing/hardening (Shore A change >15%), seal leakage",
     "VE, PI, VA, PU, TA, RL, SG", "SS acid vessels, rubber linings, elastomeric seals",
     "Dye penetrant inspection (DPI)", "6-24 mo", "ASME V Art 6, ISO 3452",
     "Perform DPI on vessel weld seams", "Replace elastomeric seals per life (EPDM 5-10 yr, Viton 10-15 yr, PTFE 15-20 yr); RBI per API 580/581", "Only non-critical elastomerics with minor leakage",
     "SCC crack in SS: assess by UT, plan repair; elastomer hardness >15% above spec: replace immediately"),

    ("FM-20", "Cracks", "Cyclic loading (thermal/mechanical)", "Operational", "B", "2.6",
     "3.0-5.0", "10^6-10^9 cycles",
     "Fatigue cracking at stress concentrations propagating per Paris' law; crack growth is measurable between inspections. At OCP, vibrating screen weld joints at Khouribga (10^8+ cycles/yr), small-bore piping connections, kiln shell welds.",
     "Surface cracks by MPI/DPI/eddy current, measurable crack growth between NDE inspections, oxide staining at crack mouths",
     "VS, PI, VE, ML, PU, RO, CR", "Vibrating screens, SAG/ball mill shells",
     "MPI at welds", "6-12 mo", "ASME V Art 7, ISO 9934",
     "Perform MPI on screen side plates", "Inspect per BS 7608; vibrating screens every 6-12 mo; initiate NDE at 50% design fatigue life", "Never for pressure boundaries, structural members, rotating shafts",
     "Crack >25% of section or growth exceeding prediction: remove from service"),

    ("FM-21", "Cracks", "Excessive temperature", "Operational", "E", "2.5",
     "0.8-1.2 / 2.0-3.0", "Highly variable",
     "Thermal shock cracking or overtemperature embrittlement from sustained exposure. At OCP, glass-lined reactor thermal shock at Jorf Lasfar, kiln shell hot spots from refractory failure (>400C), and cast iron pump thermal shock.",
     "Surface thermal checking (elephant skin crack network), glass lining crack by spark test, metallurgical change by replica metallography",
     "RO, VE, HE, PI, PU, FU, VA", "Kilns, glass-lined reactors, cast iron pumps",
     "Process temperature monitoring (thermocouples, RTDs)", "Continuous", "Process design spec",
     "Event-driven (inspect after thermal event)", "Refractory inspection every shutdown; glass lining spark test every 6-12 mo; kiln shell thermal scan weekly", "Only expendable refractory and non-pressure ceramics",
     "Glass lining holiday: repair if <100 cm2, reline if larger; sensitization in SS HAZ: solution anneal or replace"),

    ("FM-22", "Cracks", "High temperature in corrosive environment", "Calendar", "C", "2.5",
     "1.5-3.0", "10,000-50,000 h",
     "SCC accelerated by temperature (chloride SCC >50C, caustic SCC >65C), HTHA from hydrogen at >200C, polythionic acid cracking during cooldown. At OCP, Jorf Lasfar phosphoric acid plant SS at 80-110C with chlorides/fluorides.",
     "SCC branching crack patterns by DPI, HTHA fissuring by advanced UT, metallographic replica showing grain boundary attack",
     "VE, HE, PI, PU, VA, TA, BO", "Phosphoric acid reactors/evaporators, hot acid piping",
     "DPI for surface SCC detection", "6-12 mo", "ASME V Art 6, ISO 3452",
     "Perform NDE on acid vessel welds", "Per API 580/581 RBI (SS >50C with chlorides every 2-3 yr; CS in caustic >65C every 3-5 yr)", "NEVER - SCC/HTHA failures are sudden and catastrophic",
     "Any HTHA fissuring: de-rate per API 579; corrosion rate >0.5 mm/yr SS: review process chemistry"),

    ("FM-23", "Cracks", "Impact/shock loading", "Operational", "E", "2.5",
     "0.8-1.2", "Highly variable",
     "Sudden high-energy impacts generate stresses exceeding dynamic fracture toughness; susceptibility increases at low temperatures. At OCP, cast iron pump/valve casings from water hammer, crusher castings from tramp metal, winter conditions at Khouribga (0 to -5C).",
     "Visible surface crack from impact dent, water hammer pressure >2x normal, cast iron crack with sharp edges at impact site",
     "PU, VA, PI, CU, CS, TA, VE, CV", "Cast iron pumps, water hammer piping, crusher castings",
     "DPI/MPI at impact locations", "After event + 6-12 mo", "ASME V Arts 6/7",
     "Event-driven (inspect after impact)", "Surge vessel pre-charge check 6-12 mo; slow-closing valve timing annually; winter visual at Khouribga (Dec-Feb)", "Only sacrificial impact absorbers (toggle plates, expendable chute liners)",
     "Crack in cast iron: replace immediately; water hammer >1.5x design: install surge protection"),

    ("FM-24", "Cracks", "Thermal stresses (heat/cold)", "Operational", "B", "2.6",
     "2.0-4.0", "1,000-50,000 cycles",
     "Temperature gradients generate differential expansion stresses (100C constrained change in CS generates ~250 MPa approaching yield); Coffin-Manson governs life. At OCP, kiln/dryer cycling (dT >200C), dissimilar metal welds on acid piping, outdoor piping at Khouribga.",
     "Surface checking/crazing by DPI, cracking at dissimilar metal welds, pipe support damage from excessive thermal movement",
     "PI, RO, HE, VE, FU, VA, PU", "Expansion loops, DMW joints, kiln shell welds",
     "DPI for thermal fatigue crack detection", "6-12 mo", "ASME V Art 6, ISO 3452",
     "Perform DPI on expansion loop elbows", "DMW inspection every 3-5 yr; expansion bellows every 2-3 yr per EJMA; thermal cycle logging with alarm", "Never for pressure boundaries, expansion joints, or DMWs",
     "DMW cracking: plan weld replacement within 30d; bellows cracked: replace immediately"),

    ("FM-25", "Degrades", "Age", "Calendar", "B", "2.0",
     "2.5-4.0", "30,000-100,000 h",
     "Material properties deteriorate as a time-dependent function of environmental exposure - polymer chain scission, plasticizer migration, capacitor dielectric aging. At OCP, EPDM/nitrile seals age across all plants, outdoor conveyor belt rubber suffers oxidation.",
     "Elastomer hardness increasing (Shore A >15% above spec), surface cracking/crazing/chalking, UPS battery capacity <80% of rated",
     "SG, CV, SD, UP, CL, PI, HO", "Conveyor belts, fire extinguishers, UPS batteries, PLC modules",
     "Hardness testing of elastomeric components", "6-12 mo", "ASTM D2240, ISO 7619",
     "Test battery capacity on UPS", "Replace per calendar life (EPDM/nitrile 5-8 yr, Viton 10-15 yr, VRLA batteries 3-5 yr per IEEE 450)", "Only non-critical cosmetic polymers",
     "Battery capacity <80%: replace within 12 mo; <70%: replace within 30 days"),

    ("FM-26", "Degrades", "Chemical attack", "Calendar", "B", "2.0",
     "2.0-3.5", "5,000-30,000 h",
     "Process chemicals cause progressive deterioration of polymeric/elastomeric materials through solvent swelling, depolymerization, oxidative degradation. At OCP, phosphoric acid (28-54%), HF impurities, and sulfuric acid attack EPDM, natural rubber, and FRP linings at 80-110C.",
     "Polymer swelling/softening (>5% volume change), surface blistering/delamination of linings, FRP roughening with exposed fibers",
     "PI, TA, PU, VA, SG, HE, FS", "Rubber-lined acid piping, FRP acid tanks, slurry pump liners",
     "Hardness and dimensional measurement", "6-12 mo", "ASTM D2240, ASTM D471",
     "Inspect lining condition on acid piping", "Replace gaskets per compatibility life (EPDM in H3PO4 2-3 yr; Viton 5-8 yr; PTFE 8-15 yr)", "Never for containment linings on corrosive circuits",
     "Blistering >10% area or substrate exposed: relining within 14 days"),

    ("FM-27", "Degrades", "Chemical reaction", "Calendar", "B", "2.0",
     "2.0-3.0", "10,000-50,000 h",
     "Component's own material undergoes internal chemical transformation - oil oxidation forms varnish/sludge, rubber reversion, concrete carbonation. At OCP, mill gearbox oil at 70-90C oxidizes faster, transformer oil degrades cellulose insulation.",
     "Oil TAN >2.0 mg KOH/g, varnish/sludge visible on surfaces, transformer DGA showing CO >300 ppm",
     "GB, HY, PT, CO, PU, RL, CS", "Mill gearboxes, oil-filled transformers",
     "Oil analysis (TAN, viscosity, oxidation, wear metals)", "1-3 mo", "ASTM D974, D445, D6224, ISO 4406",
     "Analyze lubricant condition on gearbox", "Oil change per OEM or analysis (mineral 3,000-8,000 h, synthetic 8,000-15,000 h)", "Never for transformer oil or safety-critical lubricants",
     "TAN >4.0 mg KOH/g: change oil immediately (acid attack on bearing surfaces)"),

    ("FM-28", "Degrades", "Contamination", "Calendar", "C", "2.0",
     "1.5-2.5", "3,000-15,000 h",
     "Foreign particles, water, or chemicals infiltrate and progressively degrade the working medium. At OCP, phosphate dust enters lubricant systems, seawater leaks into cooling circuits, slurry ingress contaminates pump bearing oil.",
     "Particle count exceeding ISO 4406 target, water in oil >200 ppm, metallic wear debris trending upward",
     "GB, HY, PU, PT, CO, EN, CS, EM", "Mill gearboxes, mobile hydraulics, slurry pump bearings",
     "Oil analysis (particle count, water, viscosity, wear metals)", "1-3 mo", "ISO 4406, ASTM D6304",
     "Analyze lubricant contamination", "Replace oil filters per OEM or dP bypass; desiccant breathers every 3-6 mo", "Never for hydraulic servo, transformer oil, or engine oil",
     "Water >500 ppm: drain/dehydrate; particle count exceeding target by 2 ISO codes: upgrade filtration"),

    ("FM-29", "Degrades", "Electrical arcing", "Operational", "C", "2.0",
     "1.5-2.5", "5,000-50,000 operations",
     "Repetitive arcing during switching progressively erodes contact surfaces and deposits conductive carbon on insulation. At OCP, motor contactors cycling >100 times/day on slurry pump auto-start/stop suffer accelerated erosion.",
     "Contact resistance >50% above baseline, thermography dT >10C at contacts, visible contact pitting/erosion",
     "SG, CL, EM, FC, EG, UP", "LV MCC contactors, MV circuit breakers, slip ring brushes",
     "Contact resistance measurement (micro-ohm)", "6-12 mo", "NETA MTS, IEC 62271-100",
     "Measure contact resistance on circuit breaker", "Replace contacts at 80% OEM-rated endurance (LV: 100k-1M ops AC-3)", "Only low-duty contactors in non-critical circuits",
     "Contact resistance >200% baseline or pitting >50% area: replace before next operational cycle"),

    ("FM-30", "Degrades", "Entrained air", "Operational", "E", "2.0",
     "0.8-1.2", "Highly variable",
     "Gas bubbles cause cavitation erosion, accelerated oil oxidation (10x at 10% air), and micro-dieseling (>1,000C localized). At OCP, slurry pump suction vortexing, air entry through worn hydraulic seals, foam from shaft seal leaks.",
     "Cavitation noise (cracking/rattling at pump suction), foam in oil reservoirs, pump discharge pressure fluctuating >+/-5%",
     "PU, HY, GB, VA, CO, PI", "Slurry pumps, mobile equipment hydraulics, mill gearbox oil systems",
     "Vibration monitoring (cavitation detection - broadband HF)", "1-4 wk", "ISO 10816, ISO 13709",
     "Monitor pump NPSH", "Verify pump suction submergence at every outage", "Never for critical pumps or hydraulic systems",
     "NPSHa must be >= NPSHr + 0.5 m; cavitation erosion on impeller: correct NPSH deficiency first"),

    ("FM-31", "Degrades", "Exposure to excessive temperature", "Operational", "E", "2.0",
     "0.8-1.5", "Highly variable",
     "Components above rated range suffer permanent property changes - polymers decompose, lubricants crack, electronics age per Arrhenius. At OCP, instrumentation near kilns/dryers exceeds 60C (many rated 55C), PVC cable becomes brittle >70C.",
     "Material becoming brittle/discolored/chalky, cable insulation cracking/hardening, equipment temperature exceeding material rating",
     "EI, CL, ID, SG, PU, PI, SD", "Cable trays near kilns, field instruments near dryers",
     "Ambient temperature monitoring near sensitive equipment", "Continuous", "OEM rated range",
     "Monitor ambient temperature near electronics", "Inspect heat protection every 12 mo; cabinet cooling filters 3-6 mo; cable testing in hot zones every 2-3 yr", "Never for safety-critical electronics or power cables",
     "Cable IR <1 MOhm/km per IEEE 400: plan cable replacement at next outage"),

    ("FM-32", "Degrades", "Radiation", "Calendar", "B", "2.0",
     "2.0-3.5", "20,000-80,000 h",
     "UV radiation breaks C-C bonds in polymers causing chain scission, cross-linking, and photo-oxidation in the outer 100-300 um. At OCP, Morocco UV index 9-11 in summer; outdoor conveyor belt covers, FRP piping/tank surfaces, and cable insulation suffer.",
     "Surface chalking/fading/discoloration, surface cracking/crazing, FRP roughening with fiber exposure",
     "EI, PI, TA, CV, SG, ST", "Outdoor cable, FRP piping/tanks, conveyor covers",
     "Visual inspection for chalking, cracking, discoloration", "3-12 mo", "ASTM D4214, ASTM D660",
     "Inspect coating on outdoor structure", "Repaint 5-8 yr coastal, 7-10 yr inland; FRP gel coat every 5-8 yr per BS 4994", "Only cosmetic coatings on non-structural elements",
     "FRP Barcol hardness drops >20% or fibers exposed: apply UV-resistant topcoat within 6 months"),

    ("FM-33", "Distorts", "Impact/shock loading", "Operational", "E", "1.4",
     "0.8-1.2", "Highly variable",
     "Sudden high-energy impact generates stresses exceeding yield causing permanent plastic deformation without fracture. At OCP, falling rock impacts conveyor steelwork at transfer points, tramp metal distorts crusher guards, mobile equipment contacts pipe supports.",
     "Visible bending/denting/buckling, alignment deviation exceeding tolerance, increased vibration from distortion-induced unbalance",
     "CV, CU, PI, PU, TA, ST", "Conveyor stringers, transfer chutes, crusher guards",
     "Visual inspection for dents, bends, deformation", "1-4 wk", "API 574, AS 4100",
     "Inspect structural alignment", "Inspect impact protection (wear plates, deflectors) every 3-6 mo", "Only non-structural, non-safety cosmetic elements",
     "Pipe dent >6% of diameter: fitness-for-service per API 579; shaft runout >50 um: realign"),

    ("FM-34", "Distorts", "Mechanical overload", "Operational", "E", "1.4",
     "0.8-1.2", "Highly variable",
     "Sustained or repeated loads exceeding yield cause permanent plastic deformation - bending, buckling, twist; once yielded, geometry change is self-reinforcing. At OCP, conveyor stringers deflect during surges, piping distorts from thermal expansion, aging structural steel overloaded.",
     "Visible permanent deflection/bowing/twisting, buckling of thin-walled structures, bolt gap opening at flanges",
     "ST, CV, TA, PI, CR, PU, VE", "Equipment support frames, phosphoric acid tanks, crane beams",
     "Structural survey (deflection, plumb, level)", "6-12 mo", "AS 4100, AISC 360",
     "Survey structural deflection", "Comprehensive structural survey every 5 yr; tank per API 653 (5-10 yr); crane runway annually", "Never for primary structural members or pressure boundaries",
     "Beam deflection >L/200 or increasing under constant load: de-rate, plan reinforcement"),

    ("FM-35", "Distorts", "Off-center loading", "Operational", "C", "1.4",
     "1.5-2.5", "5,000-25,000 h",
     "Asymmetric loads create eccentric stresses progressively deforming components toward the loaded side; eccentricity is self-reinforcing. At OCP, SAG/ball mill charge asymmetry, conveyor belt off-centering, differential silo draw-off.",
     "Belt mistracking (consistently one side), differential bearing temperature (loaded side >10C hotter), uneven wear on bearings/liners",
     "ML, CV, CR, TA, PU, ST", "SAG/ball mills, belt conveyors at Khouribga, silos",
     "Alignment and level survey", "3-12 mo", "API 686, AS 4100",
     "Check belt tracking", "Precision level survey every 12 mo; crane rail annually per AS 1418", "Never for rotating equipment foundations or structural frames",
     "Foundation differential >5 mm between supports: re-shim; belt off-center >50 mm: adjust within 30d"),

    ("FM-36", "Distorts", "Use", "Operational", "B", "1.4",
     "2.0-3.5", "10,000-50,000 h",
     "Cumulative regular operational loads cause progressive geometry change through ratcheting, ambient-temperature creep, settling, and ovalization. At OCP, SAG/ball mill shells progressively ovalize, kiln riding rings wear and ovalize, conveyor pulleys lose crown.",
     "Mill shell ovality increasing (>0.5% of diameter), kiln riding ring gap differential, progressive dimensional change vs baseline",
     "ML, RO, CV, PU, CR, VA, HE", "SAG/ball mill shells, rotary kilns/dryers, conveyor pulleys",
     "Dimensional survey against baseline", "6-24 mo", "OEM specification, API 686",
     "Measure shell ovality on mill", "Kiln alignment every 6-12 mo; mill ovality at every reline (18-36 mo); pump impeller every 5,000-15,000 h", "Only non-precision components where geometry change is cosmetic",
     "Mill ovality >1.0%: engineer assessment, reduce load; impeller distortion >5% throat area: replace"),

    ("FM-37", "Drifts", "Excessive temperature (hot/cold)", "Operational", "E", "3.4",
     "0.8-1.2", "Highly variable",
     "Temperature excursions outside rated range cause sensing elements to shift via thermal expansion, semiconductor junction voltage shift, and spring rate change. At OCP, pressure transmitters near kilns/dryers exceed 60C, pH electrodes in acid reach 80-110C.",
     "Deviation from redundant reading >1% of span, calibration drift >0.5% since last verification, smart transmitter diagnostic flags",
     "ID, AN, CV, SD, FM, LI", "Pressure transmitters, pH analyzers, PSVs on steam",
     "Calibration verification", "3-12 mo", "ISA 67.04, IEC 61511",
     "Verify calibration", "SIS instruments 6-12 mo; harsh thermal environment 3-6 mo; pH in hot acid 1-3 mo", "Never for SIS/SIF or safety-critical instruments",
     "Drift >1.0% of span: recalibrate immediately and install thermal barrier"),

    ("FM-38", "Drifts", "Impact/shock loading", "Operational", "E", "3.4",
     "0.8-1.2", "Highly variable",
     "Sudden mechanical shock permanently displaces sensing elements; even 10g shocks can shift precision transmitters 0.1-0.5% of span. At OCP, field instruments on vibrating screens/crushers/mills suffer cumulative shock, instruments on slurry lines experience water hammer.",
     "Sudden step-change after known impact, calibration shift >0.5% between verifications, gauge needle not returning to zero",
     "ID, CV, SD, AN, FM, WE", "Gauges on vibrating equipment, belt weighers, limit switches",
     "Calibration verification", "3-12 mo", "ISA 67.04, IEC 61511",
     "Verify calibration after impact event", "Instruments on vibrating equipment 3-6 mo; belt weighers 1-3 mo per OIML R50", "Never for safety instruments or custody transfer",
     "Drift >0.5% on safety instrument: recalibrate; recurrence: relocate or upgrade"),

    ("FM-39", "Drifts", "Stray current", "Calendar", "C", "3.4",
     "1.5-2.5", "5,000-20,000 h",
     "Unwanted electrical currents through unintended paths introduce parasitic voltages. At OCP, extensive VFD installations at Jorf Lasfar generate high common-mode noise, CP on underground piping couples into nearby instruments.",
     "Reading varying with nearby equipment operation, difference between installed reading and portable calibrator, AC ripple on 4-20 mA loop >1% of span",
     "ID, AN, CL, FM, WE, LI", "Thermocouples, pH analyzers, EM flow meters, belt weighers",
     "Instrument signal quality analysis (noise, ripple)", "3-6 mo", "IEC 61326, NAMUR NE 21",
     "Measure signal quality", "Earth resistance test every 12 mo (<=10 Ohm per IEEE 142); cable shield integrity every 12-24 mo", "Never for SIS, custody transfer, or environmental compliance",
     "AC noise on 4-20 mA >1% span (>0.16 mA p-p): check shield continuity, verify single-point grounding"),

    ("FM-40", "Drifts", "Uneven loading", "Operational", "C", "3.4",
     "1.5-2.5", "5,000-20,000 h",
     "Asymmetric or non-uniform mechanical loads progressively shift calibration of weighing instruments. At OCP, belt weighers drift from off-center material, tank weighing drifts from differential thermal expansion, valve positioners drift from uneven packing friction.",
     "Weighing instrument deviating from independent check, calibration shifting with loading position, load cell output imbalance >2%",
     "WE, ID, CV, LI, FM, WE", "Belt weighers, tank weighing, truck scales, positioners",
     "Calibration verification with known reference", "3-6 mo", "OIML R50/R60, ISA 67.04",
     "Verify belt weigher calibration", "Custody belt weighers monthly; process weighers 3-6 mo; valve stroke test 6-12 mo", "Never for custody transfer or safety instruments",
     "Belt weigher accuracy >+/-0.5%: recalibrate; load cells unbalanced >5%: re-shim"),

    ("FM-41", "Drifts", "Use", "Operational", "B", "3.4",
     "1.5-3.0", "5,000-30,000 h",
     "Normal use progressively shifts instrument output through elastic after-effect, electrode aging, spring relaxation, electronic drift. At OCP, pH electrodes in phosphoric acid last 3-6 mo vs 12-24 mo in clean water, kiln thermocouples degrade.",
     "Progressive calibration drift trending (consistent direction), as-found data showing systematic shift, pH electrode response time >30 s to 95%",
     "ID, AN, SD, CV, FM, LI, WE", "All process transmitters, pH analyzers, PSVs, belt weighers",
     "Scheduled calibration verification", "3-24 mo", "ISA 67.04, IEC 61511, API 576",
     "Verify calibration of transmitter", "SIS 6-12 mo; acid transmitters 3-6 mo; pH in acid 1-3 mo; PSVs per API 576 3-5 yr", "Never for SIS/SIF, custody transfer, or environmental compliance",
     "SIS drift >0.5% of span: recalibrate and shorten interval; pH slope <85% Nernst: replace electrode"),

    ("FM-42", "Expires", "Age", "Calendar", "B", "2.0",
     "3.0-6.0", "Per manufacturer rated life",
     "Intrinsic material aging causes loss of capability regardless of use with sharp failure probability increase after rated life (beta 3.0-6.0). At OCP, fire suppression agents, emergency breathing apparatus at Jorf Lasfar, UPS batteries, electrochemical gas sensors.",
     "Approaching/exceeding manufacturer shelf life, battery capacity <80% per IEEE 450, fire suppression agent weight loss >5%",
     "VA, FG, UP, NO, CL, ID, FS", "PSVs on acid reactors, gas detectors, UPS batteries, sprinkler heads",
     "Battery capacity testing (discharge test)", "6-12 mo", "IEEE 450, IEC 60896",
     "FT - Scheduled Discard is primary", "Gas sensors 2 yr, catalytic bead 3 yr, VRLA batteries 4 yr, sprinkler heads 50 yr per NFPA 25, PSV elastomers per API 576", "NEVER for safety-critical expired components",
     "Battery capacity <80%: replace within 60d; <60%: replace immediately"),

    ("FM-43", "Immobilised (binds/jams)", "Contamination", "Calendar", "C", "1.6",
     "1.5-2.5", "2,000-10,000 h",
     "Foreign particles accumulate in clearance spaces between moving parts, progressively increasing friction until binding or seizure. At OCP, phosphate slurry penetrates valve stem seals, acid corrosion products restrict actuator pistons.",
     "Increased actuating force/torque trending, valve stroke time >25% above baseline, partial stroke test showing increased friction",
     "VA, PU, CO, CV, CR, CL", "Slurry pump discharge valves, acid circuit control valves",
     "Valve partial stroke testing (PST)", "1-4 wk", "IEC 61508, ISA 84",
     "Partial stroke test on control valve", "Clean/inspect valve stem every 6-12 mo in slurry service", "Only manual valves in non-critical service",
     "Stroke time >125% baseline: schedule cleaning within 30d; fail to achieve full stroke: immediate action"),

    ("FM-44", "Immobilised (binds/jams)", "Lack of lubrication", "Calendar", "C", "1.6",
     "2.0-3.5", "3,000-15,000 h",
     "Lubricant film depletes through evaporation, oxidation, or consumption until metal-to-metal contact creates friction, heat, and thermal expansion leading to seizure. At OCP, conveyor idler bearings in >45C ambient, SAG/ball mill trunnion bearing seizure.",
     "Elevated bearing temperature (dT >15C above baseline), vibration at bearing defect frequencies, audible grinding/squealing",
     "PU, CO, EM, CV, CR, FA", "SAG mill drives, belt conveyor idlers, slurry pump bearings",
     "Vibration analysis (envelope/demodulation)", "2-8 wk", "ISO 10816-3, ISO 15243",
     "Vibration analysis on bearing", "Relubricate per SKF formula (conveyor idlers 3-6 mo, mill motor bearings 1-3 mo)", "Only sealed-for-life bearings",
     "Vibration >7.1 mm/s RMS (ISO Zone D): schedule immediate replacement; bearing temp >90C: stop immediately"),

    ("FM-45", "Looses Preload", "Creep", "Calendar", "B", "1.5",
     "2.0-3.5", "20,000-60,000 h",
     "Materials in bolted assemblies undergo slow time-dependent plastic deformation under sustained stress, following logarithmic creep relaxation. At OCP, PTFE-lined gaskets on phosphoric acid piping at 80-110C, kiln shell flanges at 300-900C.",
     "Bolt torque <80% of spec, visible gasket extrusion at flange edges, minor weeping/seepage at gasketed joints",
     "VE, HE, PI, VA, PU, RO", "Acid reactors, kiln flanges, steam piping",
     "Ultrasonic bolt tension measurement", "6-12 mo", "ASTM E1685, EN 14399",
     "Measure bolt tension on flange", "Re-torque every 12-24 mo for joints >200C; PTFE gaskets every 6-12 mo", "Only low-consequence joints with secondary containment",
     "Preload <70%: replace gasket; active leakage: repair within 7 days"),

    ("FM-46", "Looses Preload", "Excessive temperature", "Operational", "E", "1.5",
     "0.8-1.2", "Highly variable",
     "Temperature excursions cause differential thermal expansion, bolt yield strength reduction, and gasket degradation. At OCP, kiln/dryer upsets cause flame impingement >500C on casing flanges; PTFE decomposes above 260C.",
     "DCS high-high temperature alarm, bolt head temper colors (straw 200C, blue 300C, black >400C), gasket charring/embrittlement",
     "RO, VE, HE, PI, VA, BO", "Rotary kilns, waste heat boilers, hot gas ducts",
     "Post-event bolt torque and condition inspection", "Event-based", "ASME PCC-1",
     "Inspect bolts after temperature event", "Verify thermal protection annually; inspect hot flange insulation every 12 mo", "Never for joints with hazardous fluids",
     "Bolt temper colors >350C: replace all bolts; temperature >120% gasket max: replace gasket"),

    ("FM-47", "Looses Preload", "Vibration", "Operational", "B", "1.5",
     "1.5-3.0", "5,000-30,000 h",
     "Cyclic transverse forces cause incremental nut rotation via Junker mechanism, with loosening rate accelerating as preload decreases. At OCP, vibrating screens (10-25 mm/s RMS), SAG/ball mill foundations, crushers generate severe vibration.",
     "Witness mark rotation on bolt/nut, bolt torque <70% of spec, audible rattling during operation",
     "VS, CU, ML, CV, PU, FA, ST", "Vibrating screens, crusher liners, mill foundations",
     "Torque audit (calibrated wrench)", "3-6 mo", "ASME PCC-1, VDI 2230",
     "Check bolt torque on vibrating equipment", "Re-torque screens every 500-1,000 h; crusher liners every 500 h; mill foundations every 6 mo", "Never for structural, rotating equipment, or safety guard connections",
     "Preload <60% or nut rotated >30deg: replace bolt/nut set and upgrade locking method"),

    ("FM-48", "Open-Circuit", "Electrical overload", "Operational", "E", "4.2",
     "0.8-1.2", "Highly variable",
     "Current exceeding conductor thermal capacity causes I2R heating, with localized hot spots at high-resistance connections fusing first. At OCP, SAG/ball mill DOL starts draw 600-800% FLA stressing terminations; VFD environments cause additional heating.",
     "Thermography dT >10C at connections, increasing contact resistance (>50% above initial), discoloration/heat marks at connections",
     "EM, SG, PC, FC, UP, EG", "Mill drive terminals, MV/LV switchgear, MCC panels",
     "Thermography of electrical connections", "1-3 mo", "NETA MTS, ISO 18434-1",
     "Thermography on electrical connections", "Re-torque motor terminals every 12 mo first 2 yr, then 3-5 yr", "Only designed protective elements (fuses, thermal overloads)",
     "dT >40C or absolute >105C (NETA Priority 1): de-energize and repair immediately"),

    ("FM-49", "Overheats/Melts", "Contamination", "Calendar", "C", "2.7",
     "1.5-2.5", "3,000-15,000 h",
     "Foreign material on heat dissipation surfaces creates thermal insulation (1 mm CaCO3 scale equals 30 mm steel in thermal resistance). At OCP, pervasive phosphate dust coats motor fins at Khouribga; gypsum scale and marine biofouling degrade cooling at Jorf Lasfar/Safi.",
     "Operating temperature >10C above clean baseline, heat exchanger approach temperature increasing >5C, visible contamination on cooling surfaces",
     "EM, FC, PT, HE, BE, GB, CL", "SAG mill drives, VFD panels, substation transformers",
     "Operating temperature trending (RTD/thermocouple/thermography)", "Continuous-monthly", "ISO 10816, IEC 60034-11",
     "Monitor winding temperature", "Clean motor fins 1-3 mo in dust; VFD filters 1-3 mo; HX chemical CIP 3-6 mo", "Never for critical motors, VFDs, or transformers",
     "Temperature approaching OEM max at rated load: clean immediately and reduce load"),

    ("FM-50", "Overheats/Melts", "Electrical overload", "Operational", "E", "2.7",
     "0.8-1.2", "Highly variable",
     "Current exceeding continuous thermal rating generates I2R heating beyond material limits; connection overheating is self-accelerating as temperature increases resistance. At OCP, high-inertia SAG mill starts (600% FLA for 15-30 s), loose connections from vibration.",
     "Connection thermography dT >10C per NETA, winding temperature above thermal class, cable surface >70C for PVC",
     "EM, PT, SG, PC, FC, EG", "Mill drives, MCC panels, VFDs",
     "Thermography of panels and connections", "1-3 mo", "NETA MTS, ISO 18434-1",
     "Thermography on electrical panel", "Re-torque MCC every 3-5 yr, annually near mills/crushers", "Only for fuses (designed protection function)",
     "dT >40C or absolute >105C (NETA Priority 1): de-energize and repair immediately"),

    ("FM-51", "Overheats/Melts", "Lack of lubrication", "Operational", "B", "2.7",
     "2.0-4.0 / 0.8-1.2", "3,000-20,000 h",
     "Insufficient lubricant film causes metal-to-metal contact increasing friction 60-fold; temperature escalates from 60-80C to seizure >250C within minutes. At OCP, SAG/ball mill trunnion bearings are highest-consequence; conveyor idlers are most frequent.",
     "Bearing temperature rising >10C above trend, high-frequency vibration (>5 kHz envelope), oil level below minimum",
     "PU, ML, CU, CV, CO, GB, FA, EM", "SAG mill trunnion bearings, slurry pump bearings, crusher main shaft",
     "Bearing temperature monitoring", "Continuous-weekly", "ISO 10816, ISO 281",
     "Monitor bearing temp on mill trunnion", "Relubricate per SKF (pump bearings monthly-quarterly, oil changes 3,000-8,000 h mineral)", "Only sealed-for-life on non-critical equipment",
     "Temp >20C above baseline or rate >2C/hr: shut down immediately; trip at 95C rolling, 85C babbitt"),

    ("FM-52", "Overheats/Melts", "Mechanical overload", "Operational", "E", "2.7",
     "0.8-1.2", "Highly variable",
     "Excessive loads increase contact pressures beyond lubricant film capacity causing metal-to-metal heat despite adequate lubrication. At OCP, crushers processing hard rock, slurry pumps with high-density feed, gearboxes during surge loading.",
     "Bearing temperature rising while lubricant is normal, simultaneous elevated motor current/torque, gearbox oil temperature exceeding OEM limit",
     "CU, PU, ML, GB, CV, CO, FA", "Jaw/cone crushers, slurry pumps, mill gearboxes",
     "Bearing temperature + load monitoring", "Continuous-weekly", "ISO 10816, OEM spec",
     "Monitor bearing temp and load on crusher", "Verify overload protection annually; test torque limiter at each shutdown", "Never for large rotating equipment",
     "Bearing temp rising with load increase: reduce load to rated within 1 hr"),

    ("FM-53", "Overheats/Melts", "Relative movement between contacting surfaces", "Operational", "B", "2.7",
     "1.5-3.0", "5,000-25,000 h",
     "Unintended sliding, fretting, or oscillating contact generates localized frictional heat with flash temperatures hundreds of degrees above bulk. At OCP, slurry pump seal faces with shaft deflection, mill drive couplings with foundation settlement.",
     "Hot spots at seal/coupling locations (thermography), fretting corrosion (reddish-brown oxide) at interference fits, increasing mechanical seal leakage",
     "PU, ML, CV, CO, FA, GB, EM", "Slurry pump seals, mill trunnion labyrinth seals, conveyor brakes",
     "Seal face temperature monitoring", "Continuous", "API 682",
     "Monitor seal temperature", "Laser alignment every 12 mo per API 686; mechanical seal replacement per OEM life (8,000-20,000 h slurry)", "Never for seals on hazardous fluids or brakes",
     "Alignment >0.05 mm offset: realign per API 686; seal temp trend >5C/month: investigate"),

    ("FM-54", "Overheats/Melts", "Rubbing", "Operational", "E", "2.7",
     "0.8-1.5", "Highly variable",
     "Abnormal contact between rotating and stationary parts generates >600C locally; thermal expansion further reduces clearance in a self-reinforcing cycle. At OCP, slurry pump wear rings with solids buildup, kiln seals with thermal distortion, fan impellers contacting casings.",
     "Sub-synchronous vibration (1/2x) or truncated shaft orbit indicating rub, temperature spike at rubbing location, audible metallic scraping",
     "PU, FA, CO, EM, ML, RO, CV", "Slurry pump wear rings, fan impeller tips, kiln shell seals",
     "Vibration monitoring (rub detection - 1/2x, orbit analysis)", "Continuous-weekly", "ISO 7919, API 670",
     "Monitor vibration for rub", "Measure wear ring clearance at every overhaul; replace at 2x design clearance per API 610", "Never for any rotating machinery",
     "Full rub (continuous contact, temp spike): IMMEDIATE shutdown; partial rub (1/2x >50% of 1x): shutdown within 7d"),

    ("FM-55", "Severs (cut, tear, hole)", "Abrasion", "Operational", "B", "2.3",
     "2.0-4.0", "2,000-15,000 h",
     "Hard particles progressively remove wall material per Archard's equation until through-wall penetration. At OCP, quartz content (5-15% SiO2, Mohs 7) in phosphate ore drives aggressive erosion at slurry pipeline elbows, pump cut-waters, hydrocyclone apexes.",
     "Wall thickness trending below API 574 minimum, rate-of-loss >0.5 mm/yr, visible wear grooves on internal surfaces",
     "PI, PU, VA, CV, FS, HC", "Slurry pipeline elbows, slurry pump casings, pinch valves",
     "UT thickness measurement", "1-6 mo", "API 574, ASME B31.3",
     "Measure wall thickness on slurry elbow", "Replace liners (pump casings 3,000-8,000 h, chute liners 6-12 mo, cyclone apex 1,000-4,000 h)", "Never for pressurized slurry piping",
     "Wall at or below minimum design: remove from service immediately (sudden rupture risk)"),

    ("FM-56", "Severs (cut, tear, hole)", "Impact/shock loading", "Operational", "E", "2.5",
     "0.8-1.2", "Highly variable",
     "Sudden high-energy impacts cause immediate penetration or puncture; repeated sub-critical impacts create cumulative damage. At OCP, large rock fragments (>300 mm) on conveyor belts at Khouribga is #1 cause of belt replacement.",
     "Visible dents/gouges, belt rip detection system activation, surface cracking around impact sites on brittle materials",
     "CV, CU, PU, PI, TA, FS, SC", "Conveyor loading zones, crusher liners, slurry pump casings",
     "Belt rip detection (electromagnetic loop/sensor cord)", "Continuous", "DIN 22109, ISO 15236",
     "Inspect belt at loading zone", "Replace impact idlers 12-18 mo; crusher blow bars at 20-40% weight loss; screen panels 2,000-6,000 h", "Only designed sacrificial elements",
     "Belt carcass damage: splice within 2 wk or replace within 30d; pump casing punctured: immediate shutdown"),

    ("FM-57", "Severs (cut, tear, hole)", "Mechanical overload", "Operational", "E", "2.5",
     "0.8-1.2", "Highly variable",
     "Applied forces exceed UTS causing tear, rupture, or separation; most dangerous when prior damage erodes safety factors. At OCP, wire rope severance on cranes at Khouribga, conveyor belt tears from jams, hydraulic hose bursts on mobile equipment.",
     "Wire rope broken wires >6 per lay length per ISO 4309, visible plastic deformation at stress concentrations, hose surface bulging/reinforcement exposure",
     "CR, CV, PI, PU, VE, ST, LE", "Overhead cranes, belt carcass, hydraulic hoses",
     "Wire rope inspection (visual + MRT)", "1-3 mo", "ISO 4309, AS 2759",
     "Inspect wire rope on crane", "Replace hoses every 6-8 yr per SAE J1273; wire ropes max 5-8 yr per ISO 4309", "Never for lifting equipment or pressure containment; only shear pins, V-belts",
     "Rope diameter reduction >10%: discard per ISO 4309; hose reinforcement exposed: replace immediately"),

    ("FM-58", "Short-Circuits", "Breakdown in insulation", "Calendar", "B", "4.1",
     "2.5-4.0", "12,000-30,000 h",
     "Insulation dielectric strength degrades following Arrhenius thermal aging, with micro-voids enabling partial discharge that self-accelerates until arc fault. At OCP, phosphate dust creates conductive deposits, coastal humidity promotes moisture ingress, VFD dV/dt imposes 2-3x peak voltage.",
     "Declining IR (>10%/yr per IEEE 43), PD >100 pC per IEC 60270, tan delta increasing >0.05 or >2x baseline",
     "EM, PT, SG, PC, EG, FC", "SAG/ball mill HV drives, substation transformers, MV XLPE cables",
     "Insulation resistance testing (megger)", "6-12 mo", "IEEE 43, IEC 60085",
     "Measure IR on motor winding", "Replace cable terminations per insulation class life (Class F: 20 yr, derate 30-40% coastal)", "Only redundant non-critical circuits with adequate protection",
     "IR <50 MOhm or PI <1.5: rewind within 30d; IR <1 MOhm/kV: do NOT energize"),

    ("FM-59", "Short-Circuits", "Contamination", "Calendar", "C", "4.1",
     "1.5-2.5", "3,000-15,000 h",
     "Conductive or hygroscopic contaminants on insulation surfaces absorb moisture, dropping surface resistance until tracking arcs carbonize insulation to flashover. At OCP, phosphate dust at Khouribga, acid mist and gypsum at Jorf Lasfar/Safi, with >85% coastal humidity.",
     "Decreasing surface IR (<100 MOhm at 500V DC), visible conductive deposits on bus bars/insulators, partial discharge activity (TEV/ultrasonic)",
     "SG, EM, CL, PT, FC, JB", "MV switchgear, MCC panels, PLC/DCS cabinets",
     "Surface insulation resistance measurement", "3-6 mo", "IEC 62631-3, IEEE 43",
     "Inspect switchgear cleanliness", "Clean insulation surfaces every 3-6 mo at Khouribga, 6-12 mo at Jorf Lasfar", "Only individually fused low-energy control circuits",
     "Surface IR <100 MOhm at 500V DC: schedule cleaning within 30d; active tracking/PD: de-energize and clean immediately"),

    ("FM-60", "Thermally Overloads (burns, overheats, melts)", "Mechanical overload", "Operational", "E", "2.7",
     "0.8-1.2", "Highly variable",
     "Mechanical loads >design force motors to draw excess current, generating I2R heating that degrades insulation. At OCP, SAG/ball mill motors overload during hard rock feed, crusher motors from tramp material, slurry pumps from density >1.6 SG.",
     "Motor current trending >100% FLA, winding temperature exceeding thermal class, repeated overload trips (>2/month)",
     "EM, PU, CO, CV, CU, FA", "SAG/ball mill drives, crusher drives, slurry pump drives",
     "Motor current monitoring", "Continuous-weekly", "NEMA MG-1, IEC 60034-1",
     "Monitor motor current", "Test thermal protection annually per NETA MTS", "Only small (<5 kW) non-critical individually protected motors",
     "Current 110-125% FLA: reduce load immediately; do NOT increase relay setting on repeated trips"),

    ("FM-61", "Thermally Overloads (burns, overheats, melts)", "Overcurrent", "Operational", "E", "2.7",
     "0.8-1.2", "Highly variable",
     "Electrical overcurrent from phase imbalance, single-phasing, voltage depression, or harmonic distortion; negative-sequence current heats rotor at ~6x rate. At OCP, MV overhead lines at Khouribga cause phase loss, VFD-driven pumps produce 25-40% THD-I.",
     "Phase current imbalance >5%, voltage imbalance >2% at PCC, cable/neutral temperature elevated",
     "EM, PT, PC, SG, FC, CB, UP", "Mill drives, VFD-fed pumps, distribution transformers",
     "Phase current and voltage balance monitoring", "Continuous-weekly", "NEMA MG-1, IEC 60034-26",
     "Measure phase balance", "Verify relay settings every 3 yr per NETA MTS; power quality survey annually per IEEE 519", "Only small individually protected motors",
     "Voltage imbalance >3% or single phase: trip immediately; THD-I >15%: install output reactor/sine-wave filter"),

    ("FM-62", "Washes Off", "Excessive fluid velocity", "Operational", "C", "2.3",
     "1.5-3.0", "5,000-20,000 h",
     "Fluid kinetic energy exceeding coating adhesion progressively removes protective linings via hydrodynamic shear; wall shear scales with velocity squared. At OCP, rubber-lined slurry pipes lose linings at elbows at 2-5 m/s, pump volute linings erode at the tongue.",
     "UT wall thickness loss >10% from baseline, visible lining bare spots during internal inspection, downstream coating particles in fluid",
     "PI, PU, VA, HE, FS, VE", "Rubber-lined slurry pipelines, PTFE-lined acid piping, slurry pump volutes",
     "UT thickness at high-velocity points", "1-6 mo", "ASME B31.3, API 570/574",
     "Measure wall thickness on elbow", "Replace rubber lining (elbows 12-18 mo, straight 24-36 mo, pump volutes 6-12 mo)", "Only non-pressure atmospheric components",
     "Wall <120% design minimum: schedule replacement within 30d; at or below minimum: remove immediately"),

    ("FM-63", "Washes Off", "Use", "Operational", "B", "2.3",
     "2.0-4.0", "5,000-30,000 h",
     "Normal operational exposure progressively depletes protective coatings at predictable rates through mild fluid abrasion, chemical dissolution, and weathering. At OCP, marine atmosphere at Jorf Lasfar/Safi consumes paint 30-40% faster than inland.",
     "Coating DFT <70% of original, visual chalking/blistering/rust bleeding (ASTM D610/D714), corrosion rate increase on protected surfaces",
     "PI, TA, HE, VE, PU, CV, CR", "Structural steel, cooling water piping, pump shaft sleeves",
     "Coating thickness measurement (DFT gauge)", "3-12 mo", "SSPC-PA 2, ISO 19840",
     "CB preferred for high-value coatings", "Scheduled restoration (repaint marine 5-7 yr, inland 7-10 yr; rubber linings 3-5 yr in slurry)", "Only aesthetic coatings on non-structural indoor equipment",
     "DFT <50% minimum or rust Ri 4 (>8%): full restoration within 30d; base metal exposed >10%: immediate spot-prime"),

    ("FM-64", "Wears", "Breakdown of lubrication", "Operational", "B", "2.4",
     "2.0-3.5", "5,000-25,000 h",
     "Lubricant properties degrade through oxidation, thermal cracking, additive depletion, and water contamination until film thickness drops below 1.0, transitioning to severe adhesive wear. At OCP, high gearbox temps (70-90C) and phosphate dust ingress accelerate breakdown.",
     "Oil TAN >2.0 mg KOH/g, viscosity change >20%, wear metals (Fe, Cu, Sn) increasing",
     "GB, PU, CO, EM, CV, CU, EN", "Mill gearboxes, crusher gearboxes, large motor bearings",
     "Oil analysis (oxidation, viscosity, TAN, wear metals)", "1-3 mo", "ASTM D974/D445/E2412, ISO 4406",
     "Analyze lubricant on gearbox", "Oil change (mineral 4,000-8,000 h, synthetic 8,000-16,000 h; reduce 30% if >80C)", "Only sealed-for-life non-critical small bearings",
     "TAN >4.0 mg KOH/g or wear metals step-change: change oil immediately AND investigate"),

    ("FM-65", "Wears", "Entrained air", "Operational", "E", "2.4",
     "0.8-1.2", "Highly variable",
     "Air/vapor bubbles implode near surfaces generating micro-jets at >1 GPa and >5,000C, progressively removing material with characteristic spongy pitting. At OCP, slurry pump impellers cavitate with marginal NPSH; control valves flash in hot acid; cyclone feed pumps entrain air.",
     "Distinctive crackling cavitation noise, pump head/efficiency decreasing, broadband HF vibration (>5 kHz)",
     "PU, VA, PI, HY, HC", "Slurry pumps, cooling water pumps, control valves, classification cyclones",
     "Vibration monitoring (broadband HF cavitation)", "1-4 wk", "ISO 10816, ISO 13709",
     "Monitor NPSH and vibration on pump", "Inspect impeller at overhaul (cast iron 2,000-5,000 h, high-chrome 5,000-12,000 h, duplex 10,000-25,000 h)", "Only non-critical small pumps",
     "NPSHa < NPSHr x 1.3: correct suction deficiency immediately; cavitation depth >2 mm: replace and upgrade material"),

    ("FM-66", "Wears", "Excessive fluid velocity", "Operational", "B", "2.4",
     "2.0-4.0", "3,000-20,000 h",
     "Suspended particles at high velocity impact surfaces causing erosion following power-law (rate proportional to velocity^2-4); doubling velocity increases erosion 4-16x. At OCP, slurry pipeline elbows erode at up to 4 m/s, pump cut-water sees 2-3x average velocity.",
     "UT wall thinning at elbows/tees, erosion rate above design allowance, characteristic horseshoe/cat-eye patterns",
     "PI, PU, VA, HC, HE, NO", "Khouribga-Jorf slurry pipeline, slurry pump impellers, classification cyclones",
     "UT wall thickness at erosion-prone locations", "1-6 mo", "API 574, ASME B31.3",
     "Measure wall thickness on pipeline elbow", "Replace pump impeller 3,000-10,000 h; elbows 12-36 mo; cyclone apex 1,000-5,000 h", "Only non-pressurized gravity chute liners",
     "Wall 100-120% of minimum: plan replacement within 30d; at/below minimum: remove immediately"),

    ("FM-67", "Wears", "Impact/shock loading", "Operational", "B", "2.4",
     "1.5-3.0", "3,000-15,000 h",
     "Repeated high-energy impacts cause progressive removal through surface fatigue, spalling, and plastic deformation; austenitic Mn steel work-hardens but material is consumed. At OCP, crusher liners at 5-15 m/s, vibrating screen panels, mill liners, conveyor loading zones.",
     "Progressive thickness reduction at impact zones, crusher liner profile worn beyond OEM minimum, screen apertures worn >110% nominal",
     "CU, ML, SC, CV, FE, PI", "Jaw/cone/impact crushers, SAG/ball mills, vibrating screens",
     "Liner thickness/profile measurement (template, UT)", "Monthly-quarterly", "OEM specification",
     "Measure liner thickness on crusher", "Jaw plates 500-2,000 h, cone mantle 1,500-4,000 h, blow bars 200-800 h, mill liners 6,000-18,000 h", "Acceptable for sacrificial components (chute liners, screen panels)",
     "Crusher liner at/below OEM minimum: replace within 7d (risk of frame damage)"),

    ("FM-68", "Wears", "Low pressure", "Operational", "E", "2.4",
     "0.8-1.2", "Highly variable",
     "Insufficient pressure at critical locations allows cavitation, flashing, or loss of hydrodynamic bearing film, all causing accelerated material removal. At OCP, slurry pump suction in deep Khouribga sumps, acid control valves near vapor pressure, mill lube oil film loss.",
     "Suction pressure below design minimum, cavitation noise at pump, oil supply pressure below OEM minimum",
     "PU, VA, BE, CO, PI, HY", "Deep sump slurry pumps, high-dP control valves, mill trunnion bearings",
     "Suction pressure monitoring", "Continuous", "ISO 13709, HI 9.6.1",
     "Monitor suction pressure", "Clean suction strainer every 3 mo in slurry; NPSH verification at every process change", "Only non-critical small pumps",
     "NPSHa < NPSHr x 1.3: correct immediately; bearing oil below OEM minimum: trip before damage"),

    ("FM-69", "Wears", "Lubricant contamination (particles)", "Calendar", "C", "2.4",
     "1.5-2.5", "5,000-30,000 h",
     "Hard particles (1-25 um) circulating in lubricant bridge the oil film causing three-body abrasive wear; process is autocatalytic as wear debris generates secondary particles. At OCP, phosphate dust ingresses through seals, slurry leaks through failed seals.",
     "ISO 4406 count exceeding target by >= 2 codes, wear metals trending upward, ferrography showing cutting wear (long ribbon particles)",
     "GB, PU, HY, CO, EM, EN, CV", "Mill gearboxes, hydraulic servo valves, mobile equipment",
     "Oil particle count (ISO 4406)", "1-3 mo", "ISO 4406/4407",
     "Analyze particle count on gearbox oil", "Replace filters per dP or OEM (1,000-3,000 h); desiccant breathers every 3-6 mo in dusty OCP", "Never for servo valves or precision gears",
     "Count exceeding target by >4 ISO codes: flush, replace oil, inspect seals, install kidney-loop filtration"),

    ("FM-70", "Wears", "Mechanical overload", "Operational", "E", "2.4",
     "0.8-1.5", "Highly variable",
     "Loads exceeding design squeeze lubricant film thinner causing asperity contact and transition from mild to severe wear with as little as 25% overload. At OCP, crusher liners wear faster in harder rock, pump wear rings open rapidly above design density, gearbox teeth pit above rated torque.",
     "Wear rate increasing above trend, operating load sustained above design, oil analysis wear metals correlated with load events",
     "CU, PU, GB, ML, CV, CO, VA", "Crusher liners, pump wear rings, mill gearboxes",
     "Load monitoring (power, current, pressure)", "Continuous", "Equipment design spec",
     "Monitor operating load", "Verify overload protection annually; torque limiter test at each shutdown", "Only sacrificial parts (shear pins, screen panels)",
     "Load consistently >100% design: investigate and reduce; wear rate >150% historical: adjust interval"),

    ("FM-71", "Wears", "Metal to metal contact", "Operational", "B", "2.4",
     "2.0-3.5", "5,000-30,000 h",
     "Direct metallic surface contact without lubricant film causes micro-welding at asperity junctions; continued motion fractures welds, transferring material per Archard's equation. At OCP, pump shaft sleeves at packing/seal zones, wire ropes on sheaves, chain pins, valve stems.",
     "Increasing shaft sleeve diameter at seal zone, wire rope flattened strands and sheave groove wear, chain elongation >2%",
     "PU, VA, CR, CV, BR, CG, GB", "Slurry pump shaft sleeves, overhead crane wire rope, chain conveyors",
     "Dimensional measurement of wear components", "3-12 mo", "OEM specification",
     "Measure shaft sleeve diameter", "Replace packing at every valve overhaul; shaft sleeves at seal change; wire rope per ISO 4309", "Acceptable for sacrificial elements (brake pads, packing, chain links)",
     "Wire rope approaching ISO 4309 discard: replace within 30d; brake disc below OEM minimum: replace immediately"),

    ("FM-72", "Wears", "Relative movement between contacting surfaces", "Operational", "B", "2.4",
     "2.0-3.5", "8,000-40,000 h",
     "Small-amplitude oscillatory motion (5-300 um) between clamped surfaces breaks oxide films, generating hard Fe2O3 debris as abrasive third body with wear coefficient 10-100x higher than sliding wear; also initiates fatigue cracks. At OCP, pump shaft bearing seats fret, coupling hubs loosen, wire ropes suffer internal strand fretting.",
     "Red-brown oxide powder at press-fit interfaces, increasing bearing clearance on shaft, vibration showing progressive looseness",
     "PU, EM, CV, CU, ML, CR, CG, FA", "Pump bearing seats, motor rotor fits, wire ropes, mill trunnion seats",
     "Vibration monitoring for progressive fit loosening", "1-4 wk", "ISO 10816, ISO 20816",
     "Monitor vibration for fit loosening", "Measure shaft at all fits during every bearing replacement; wire rope internal inspection per ISO 4309 every 6 mo", "Never for rotating shaft fits or wire ropes",
     "Shaft undersize at bearing seat: apply shaft repair (metal spray + grind); fretting cracks at shaft shoulder: replace shaft"),
]

# Field indices
FM_ID, MECHANISM, CAUSE, FREQ_BASIS, PATTERN, ISO_CODE = 0, 1, 2, 3, 4, 5
WEIBULL_BETA, WEIBULL_ETA = 6, 7
DEGRADATION, P_CONDITIONS, EQUIP_CLASSES, OCP_EQUIP = 8, 9, 10, 11
CBM_TECHNIQUE, PF_INTERVAL, REF_STD = 12, 13, 14
STRAT_CB, STRAT_FT, STRAT_RTF, KEY_THRESHOLD = 15, 16, 17, 18

# ---------------------------------------------------------------------------
# 2. STYLES
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
VALID_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GROUP_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
GROUP_FONT = Font(name="Calibri", size=11, bold=True)


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_body(ws, row, ncols, wrap_cols=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN if (wrap_cols and col in wrap_cols) else Alignment(vertical="top")


def auto_width(ws, min_width=8, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                lengths.append(max(len(line) for line in lines))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# 3. SHEET BUILDERS
# ---------------------------------------------------------------------------

def build_sheet1_72_fms(wb):
    """Sheet 1: 72 Failure Modes — main data table."""
    ws = wb.active
    ws.title = "72 Failure Modes"

    headers = [
        "FM#", "Mechanism", "Cause", "Freq. Basis", "Pattern", "ISO 14224",
        "Weibull Beta", "Weibull Eta", "Degradation Summary", "Top P-Conditions",
        "Equipment Classes", "OCP Equipment", "Primary CBM Technique",
        "P-F Interval", "Reference Standard", "Strategy (CB)", "Strategy (FT)",
        "Strategy (RTF)", "Key Threshold"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    wrap_cols = {9, 10, 11, 12, 16, 17, 18, 19}

    for i, fm in enumerate(FM_DATA, 2):
        for col, val in enumerate(fm, 1):
            ws.cell(row=i, column=col, value=val)
        style_body(ws, i, len(headers), wrap_cols)

    ws.auto_filter.ref = f"A1:S{len(FM_DATA) + 1}"
    ws.freeze_panes = "A2"

    # Column widths
    widths = {1: 7, 2: 18, 3: 30, 4: 11, 5: 8, 6: 10, 7: 12, 8: 20,
              9: 55, 10: 50, 11: 22, 12: 35, 13: 40, 14: 14, 15: 25,
              16: 45, 17: 45, 18: 40, 19: 50}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_sheet2_validation_matrix(wb):
    """Sheet 2: Validation Matrix — mechanism x cause with FM# at intersections."""
    ws = wb.create_sheet("Validation Matrix")

    # Collect all unique causes in order of first appearance
    causes_ordered = []
    seen = set()
    for fm in FM_DATA:
        c = fm[CAUSE]
        if c not in seen:
            causes_ordered.append(c)
            seen.add(c)

    # Collect mechanisms in order
    mechanisms_ordered = []
    seen_m = set()
    for fm in FM_DATA:
        m = fm[MECHANISM]
        if m not in seen_m:
            mechanisms_ordered.append(m)
            seen_m.add(m)

    # Build lookup
    lookup = {}
    for fm in FM_DATA:
        lookup[(fm[MECHANISM], fm[CAUSE])] = fm[FM_ID]

    # Headers
    ws.cell(row=1, column=1, value="Mechanism \\ Cause")
    for j, cause in enumerate(causes_ordered, 2):
        ws.cell(row=1, column=j, value=cause)
    style_header(ws, 1, len(causes_ordered) + 1)

    # Body
    for i, mech in enumerate(mechanisms_ordered, 2):
        ws.cell(row=i, column=1, value=mech)
        ws.cell(row=i, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=i, column=1).border = THIN_BORDER
        for j, cause in enumerate(causes_ordered, 2):
            cell = ws.cell(row=i, column=j)
            fm_id = lookup.get((mech, cause))
            if fm_id:
                cell.value = fm_id
                cell.fill = VALID_FILL
                cell.alignment = CENTER_ALIGN
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(causes_ordered) + 1)}{len(mechanisms_ordered) + 1}"
    ws.freeze_panes = "B2"

    # Column widths
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(causes_ordered) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 12


def build_sheet3_by_pattern(wb):
    """Sheet 3: By Pattern — grouped by Nowlan & Heap failure pattern."""
    ws = wb.create_sheet("By Pattern")

    headers = ["Pattern", "RCM Implication", "FM#", "Mechanism", "Cause"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    pattern_info = {
        "B": "Age-related (FT eligible)",
        "C": "Gradual increase (FT eligible)",
        "D": "Random + break-in (FT NOT applicable)",
        "E": "Random (FT NOT applicable)",
    }

    # Group by pattern
    by_pattern = {}
    for fm in FM_DATA:
        p = fm[PATTERN]
        by_pattern.setdefault(p, []).append(fm)

    row = 2
    for pattern in ["B", "C", "D", "E"]:
        fms = by_pattern.get(pattern, [])
        if not fms:
            continue
        # Group header row
        ws.cell(row=row, column=1, value=f"Pattern {pattern} ({len(fms)} FMs)")
        ws.cell(row=row, column=2, value=pattern_info[pattern])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.border = THIN_BORDER
        row += 1
        for fm in fms:
            ws.cell(row=row, column=1, value=pattern)
            ws.cell(row=row, column=2, value=pattern_info[pattern])
            ws.cell(row=row, column=3, value=fm[FM_ID])
            ws.cell(row=row, column=4, value=fm[MECHANISM])
            ws.cell(row=row, column=5, value=fm[CAUSE])
            style_body(ws, row, len(headers))
            row += 1

    ws.auto_filter.ref = f"A1:E{row - 1}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 35


def build_sheet4_by_freq_basis(wb):
    """Sheet 4: By Frequency Basis — Calendar vs Operational."""
    ws = wb.create_sheet("By Frequency Basis")

    headers = ["Freq. Basis", "Unit Types", "FM#", "Mechanism", "Cause", "Pattern"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    unit_types = {
        "Calendar": "Days, Weeks, Months, Years",
        "Operational": "Operating Hours, Cycles, Tonnes",
    }

    by_basis = {}
    for fm in FM_DATA:
        b = fm[FREQ_BASIS]
        by_basis.setdefault(b, []).append(fm)

    row = 2
    for basis in ["Calendar", "Operational"]:
        fms = by_basis.get(basis, [])
        # Group header
        ws.cell(row=row, column=1, value=f"{basis} ({len(fms)} FMs)")
        ws.cell(row=row, column=2, value=unit_types[basis])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.border = THIN_BORDER
        row += 1
        for fm in fms:
            ws.cell(row=row, column=1, value=basis)
            ws.cell(row=row, column=2, value=unit_types[basis])
            ws.cell(row=row, column=3, value=fm[FM_ID])
            ws.cell(row=row, column=4, value=fm[MECHANISM])
            ws.cell(row=row, column=5, value=fm[CAUSE])
            ws.cell(row=row, column=6, value=fm[PATTERN])
            style_body(ws, row, len(headers))
            row += 1

    ws.auto_filter.ref = f"A1:F{row - 1}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 10


def build_sheet5_by_iso(wb):
    """Sheet 5: By ISO 14224 — grouped by ISO failure mechanism code."""
    ws = wb.create_sheet("By ISO 14224")

    iso_descriptions = {
        "1.4": "Deformation",
        "1.5": "Looseness",
        "1.6": "Sticking",
        "2.0": "General degradation",
        "2.2": "Corrosion",
        "2.3": "Erosion",
        "2.4": "Wear",
        "2.5": "Breakage",
        "2.5/2.6": "Breakage/Fatigue",
        "2.6": "Fatigue",
        "2.7": "Overheating",
        "3.4": "Instrument drift",
        "4.1": "Short circuit",
        "4.1/4.5": "Short circuit/Insulation failure",
        "4.2": "Open circuit",
        "5.1": "Blockage/plugged",
    }

    headers = ["ISO 14224 Code", "Description", "FM#", "Mechanism", "Cause"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    by_iso = {}
    for fm in FM_DATA:
        iso = fm[ISO_CODE]
        by_iso.setdefault(iso, []).append(fm)

    row = 2
    for iso_code in sorted(by_iso.keys(), key=lambda x: float(x.split("/")[0])):
        fms = by_iso[iso_code]
        desc = iso_descriptions.get(iso_code, "")
        # Group header
        ws.cell(row=row, column=1, value=f"{iso_code} - {desc} ({len(fms)} FMs)")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.border = THIN_BORDER
        row += 1
        for fm in fms:
            ws.cell(row=row, column=1, value=iso_code)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value=fm[FM_ID])
            ws.cell(row=row, column=4, value=fm[MECHANISM])
            ws.cell(row=row, column=5, value=fm[CAUSE])
            style_body(ws, row, len(headers))
            row += 1

    ws.auto_filter.ref = f"A1:E{row - 1}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 35


def build_sheet6_cause_crossref(wb):
    """Sheet 6: Cause Cross-Reference — causes appearing in multiple mechanisms."""
    ws = wb.create_sheet("Cause Cross-Reference")

    headers = ["Cause", "# Mechanisms", "Mechanisms (FM#)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Build cause -> list of (mechanism, fm_id)
    cause_map = {}
    for fm in FM_DATA:
        cause_map.setdefault(fm[CAUSE], []).append((fm[MECHANISM], fm[FM_ID]))

    # Filter to causes with >1 mechanism and sort by count desc
    multi_causes = {c: entries for c, entries in cause_map.items() if len(entries) > 1}
    sorted_causes = sorted(multi_causes.items(), key=lambda x: -len(x[1]))

    row = 2
    for cause, entries in sorted_causes:
        ws.cell(row=row, column=1, value=cause)
        ws.cell(row=row, column=2, value=len(entries))
        detail = ", ".join(f"{m} ({fid})" for m, fid in entries)
        ws.cell(row=row, column=3, value=detail)
        style_body(ws, row, len(headers), {3})
        row += 1

    # Also add single-mechanism causes
    ws.cell(row=row, column=1)  # blank row
    row += 1
    ws.cell(row=row, column=1, value="Single-mechanism causes")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = GROUP_FILL
        cell.font = GROUP_FONT
        cell.border = THIN_BORDER
    row += 1
    single_causes = sorted(
        [(c, entries) for c, entries in cause_map.items() if len(entries) == 1],
        key=lambda x: x[0]
    )
    for cause, entries in single_causes:
        ws.cell(row=row, column=1, value=cause)
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=f"{entries[0][0]} ({entries[0][1]})")
        style_body(ws, row, len(headers), {3})
        row += 1

    ws.auto_filter.ref = f"A1:C{row - 1}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80


# ---------------------------------------------------------------------------
# 4. MAIN
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()

    print("Building Sheet 1: 72 Failure Modes...")
    build_sheet1_72_fms(wb)

    print("Building Sheet 2: Validation Matrix...")
    build_sheet2_validation_matrix(wb)

    print("Building Sheet 3: By Pattern...")
    build_sheet3_by_pattern(wb)

    print("Building Sheet 4: By Frequency Basis...")
    build_sheet4_by_freq_basis(wb)

    print("Building Sheet 5: By ISO 14224...")
    build_sheet5_by_iso(wb)

    print("Building Sheet 6: Cause Cross-Reference...")
    build_sheet6_cause_crossref(wb)

    output_path = Path(__file__).resolve().parent.parent / \
        "skills" / "00-knowledge-base" / "data-models" / "failure-modes" / "FM-MASTER-REFERENCE.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Sheet 1 rows: {wb['72 Failure Modes'].max_row - 1} data rows (expected 72)")


if __name__ == "__main__":
    main()
