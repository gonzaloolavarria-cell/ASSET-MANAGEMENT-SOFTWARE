"""
Generate Client Templates — Phase 6

Copies the 14 system templates into a client project's 5-templates/ directory,
optionally applying client branding (header color, fonts) from branding.yaml.

Usage:
    python scripts/generate_client_templates.py <client_slug> <project_slug>

Example:
    python scripts/generate_client_templates.py ocp jfc-maintenance-strategy
"""

import logging
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Regex for YAML injection prevention (reused from wizard_cli.py)
_YAML_INJECTION = re.compile(r"[{}[\]!%*&|>]")

# Template file names (numbered 01-14)
TEMPLATE_FILES = [
    "01_equipment_hierarchy.xlsx",
    "02_criticality_assessment.xlsx",
    "03_failure_modes.xlsx",
    "04_maintenance_tasks.xlsx",
    "05_work_packages.xlsx",
    "06_work_order_history.xlsx",
    "07_spare_parts_inventory.xlsx",
    "08_shutdown_calendar.xlsx",
    "09_workforce.xlsx",
    "10_field_capture.xlsx",
    "11_rca_events.xlsx",
    "12_planning_kpi_input.xlsx",
    "13_de_kpi_input.xlsx",
    "14_maintenance_strategy.xlsx",
]

# Template version metadata
TEMPLATE_VERSION = "1.0"


def _validate_slug(value: str, label: str = "slug") -> None:
    """Validate a slug (no path traversal, lowercase alphanumeric with hyphens)."""
    if not value:
        raise ValueError(f"{label} must not be empty")
    if ".." in value or "/" in value or "\\" in value:
        raise ValueError(f"{label} contains path traversal characters: {value!r}")
    pattern = re.compile(r"^[a-z0-9][a-z0-9-]*$")
    if not pattern.match(value):
        raise ValueError(
            f"{label} must match pattern [a-z0-9][a-z0-9-]*: {value!r}"
        )


def _load_branding(branding_path: Path) -> dict | None:
    """Safely load branding.yaml using yaml.safe_load."""
    if not branding_path.is_file():
        return None
    try:
        import yaml

        raw = branding_path.read_text(encoding="utf-8")
        # Sanitize YAML injection patterns
        if _YAML_INJECTION.search(raw.split("\n")[0] if raw else ""):
            logger.warning("Suspicious YAML content in %s, skipping", branding_path)
            return None
        config = yaml.safe_load(raw)
        if not isinstance(config, dict):
            logger.warning("branding.yaml at %s is not a valid dict", branding_path)
            return None
        return config
    except Exception as exc:
        logger.warning("Failed to load branding from %s: %s", branding_path, exc)
        return None


def _apply_branding(xlsx_path: Path, branding: dict) -> None:
    """Apply branding (header color, font) to an xlsx template.

    Modifies the workbook in-place. Only touches header row styling and
    the Instructions sheet footer text — never executes macros.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed, skipping branding for %s", xlsx_path)
        return

    colors = branding.get("colors", {})
    fonts_cfg = branding.get("fonts", {})
    footer_cfg = branding.get("footer", {})
    client_cfg = branding.get("client", {})

    primary_color = colors.get("primary", "").lstrip("#")
    if not primary_color or len(primary_color) != 6:
        return  # No valid color, skip

    header_fill = PatternFill(
        start_color=primary_color, end_color=primary_color, fill_type="solid"
    )
    heading_font_name = fonts_cfg.get("heading", {}).get("name", "Calibri")
    header_font = Font(name=heading_font_name, size=11, bold=True, color="FFFFFF")

    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # Apply header styling to row 1
        for cell in ws[1]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font

        # Add footer notice if Instructions sheet
        if ws.title == "Instructions" and footer_cfg.get("confidential_notice"):
            max_row = ws.max_row or 1
            ws.cell(row=max_row + 2, column=1, value=footer_cfg["confidential_notice"])

    # Set workbook properties for versioning
    wb.properties.creator = client_cfg.get("name", "VSC AMS")
    wb.properties.description = f"Template v{TEMPLATE_VERSION}"
    wb.properties.modified = datetime.now()

    wb.save(xlsx_path)


def _strip_metadata(xlsx_path: Path) -> None:
    """Remove sensitive metadata (author paths) from a template workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    wb = load_workbook(xlsx_path)
    # Clear any local path references
    if wb.properties.creator and ("Users" in wb.properties.creator or ":" in wb.properties.creator):
        wb.properties.creator = "VSC AMS"
    if wb.properties.lastModifiedBy and ("Users" in wb.properties.lastModifiedBy or ":" in wb.properties.lastModifiedBy):
        wb.properties.lastModifiedBy = "VSC AMS"
    wb.properties.description = f"Template v{TEMPLATE_VERSION}"
    wb.properties.modified = datetime.now()
    wb.save(xlsx_path)


def generate_client_templates(
    client_slug: str,
    project_slug: str,
    *,
    system_root: Path | None = None,
    client_root: Path | None = None,
) -> list[Path]:
    """Copy system templates to a client project's 5-templates/ directory.

    Args:
        client_slug: Client identifier (e.g. "ocp")
        project_slug: Project identifier (e.g. "jfc-maintenance-strategy")
        system_root: Override AMS system root (for testing)
        client_root: Override AMS client root (for testing)

    Returns:
        List of generated template paths.
    """
    _validate_slug(client_slug, "client_slug")
    _validate_slug(project_slug, "project_slug")

    # Resolve roots
    if system_root is None:
        from agents._shared.paths import get_system_root
        system_root = get_system_root()
    if client_root is None:
        from agents._shared.paths import get_client_root
        client_root = get_client_root()

    source_dir = system_root / "templates"
    target_dir = (
        client_root
        / "clients"
        / client_slug
        / "projects"
        / project_slug
        / "5-templates"
    )
    target_dir.mkdir(parents=True, exist_ok=True)

    # Resolve branding (3-level cascade)
    branding = None
    branding_candidates = [
        target_dir / "branding.yaml",  # Project-level
        client_root / "clients" / client_slug / "context" / "templates" / "branding.yaml",  # Client-level
        source_dir / "branding.yaml",  # System-level
    ]
    for bp in branding_candidates:
        branding = _load_branding(bp)
        if branding is not None:
            logger.info("Using branding from %s", bp)
            break

    # Copy templates
    generated = []
    for template_name in TEMPLATE_FILES:
        src = source_dir / template_name
        dst = target_dir / template_name
        if not src.is_file():
            logger.warning("Source template not found: %s", src)
            continue

        shutil.copy2(src, dst)

        # Strip sensitive metadata
        _strip_metadata(dst)

        # Apply branding if available
        if branding:
            _apply_branding(dst, branding)

        generated.append(dst)

    return generated


def main() -> None:
    """CLI entry point."""
    if len(sys.argv) < 3:
        print("Usage: python scripts/generate_client_templates.py <client_slug> <project_slug>")
        print("Example: python scripts/generate_client_templates.py ocp jfc-maintenance-strategy")
        sys.exit(1)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    client_slug = sys.argv[1]
    project_slug = sys.argv[2]

    print(f"Generating templates for {client_slug}/{project_slug}...")
    try:
        generated = generate_client_templates(client_slug, project_slug)
        print(f"Done! {len(generated)} templates generated:")
        for p in generated:
            print(f"  [OK] {p.name}")
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
