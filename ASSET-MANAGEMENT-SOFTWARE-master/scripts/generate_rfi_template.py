"""Generate AMS RFI Questionnaire Excel Template.

Creates an 8-sheet questionnaire + instructions sheet for collecting
client information before a maintenance strategy engagement.

Usage:
    python scripts/generate_rfi_template.py
    python scripts/generate_rfi_template.py --output custom-path.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Ensure project root is importable
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from tools.models.rfi_models import (  # noqa: E402
    AMS_TEMPLATES,
    CMMSType,
    CriticalityMethodRFI,
    DataFormat,
    Industry,
    Language,
    OrgStructure,
    SAPVersion,
    ScopeType,
    StrategyMaturity,
    WorkshopFormat,
)


# ---------------------------------------------------------------------------
# Styling constants (match existing AMS template convention)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="Calibri", size=10)
WRAP_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DEFAULT_OUTPUT = _PROJECT_ROOT / "templates" / "rfi" / "00_rfi_questionnaire.xlsx"

# Column layout for questionnaire sheets:
# A=ID  B=Field  C=Response  D=Type  E=Required  F=Help
COL_ID = 1
COL_FIELD = 2
COL_RESPONSE = 3
COL_TYPE = 4
COL_REQUIRED = 5
COL_HELP = 6
HEADER_ROW = 1
DATA_START_ROW = 2


# ---------------------------------------------------------------------------
# Sheet field definitions
# Each entry: (field_id, label, field_type, required, help_text, dropdown_values)
# ---------------------------------------------------------------------------

def _enum_values(enum_cls: type) -> list[str]:
    return [e.value for e in enum_cls]


SHEET_1_COMPANY: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("CSP-01", "Company Name", "text", True,
     "Legal name of the client company", None),
    ("CSP-02", "Industry", "dropdown", True,
     "Primary industry sector", _enum_values(Industry)),
    ("CSP-03", "Plant Name", "text", True,
     "Name of the plant or facility", None),
    ("CSP-04", "Plant Code", "text", False,
     "Internal plant code (e.g., OCP-JFC)", None),
    ("CSP-05", "Location", "text", True,
     "City, Region, Country (e.g., El Jadida, Morocco)", None),
    ("CSP-06", "Country Code (ISO)", "text", True,
     "ISO 3166-1 alpha-2 code (e.g., MA, CL, US)", None),
    ("CSP-07", "Production Capacity", "text", False,
     "Annual production capacity (e.g., 10M tonnes/year)", None),
    ("CSP-08", "Primary Language", "dropdown", True,
     "Primary working language for procedures", _enum_values(Language)),
    ("CSP-09", "Secondary Language", "dropdown", False,
     "Secondary language (if any)", _enum_values(Language)),
    ("CSP-10", "Contact Name", "text", True,
     "Name of primary project contact", None),
    ("CSP-11", "Contact Email", "text", True,
     "Email address of primary contact", None),
    ("CSP-12", "Contact Phone", "text", False,
     "Phone number of primary contact", None),
]

SHEET_2_EQUIPMENT: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("EHD-01", "Equipment List Available?", "dropdown", True,
     "Does the client have a digital equipment list?", ["Yes", "No", "Partial"]),
    ("EHD-02", "Equipment List Format", "dropdown", False,
     "Format of the equipment list", _enum_values(DataFormat)),
    ("EHD-03", "Estimated Equipment Count", "number", True,
     "Approximate number of equipment items in scope", None),
    ("EHD-04", "Hierarchy Levels", "number", False,
     "Number of hierarchy levels (1-10, typically 6)", None),
    ("EHD-05", "Naming Convention", "text", False,
     "Description of equipment naming/tagging convention", None),
    ("EHD-06", "Naming Convention Document?", "dropdown", False,
     "Is the naming convention documented?", ["Yes", "No"]),
    ("EHD-07", "BOM Available?", "dropdown", False,
     "Are Bills of Materials available?", ["Yes", "No", "Partial"]),
    ("EHD-08", "BOM Format", "dropdown", False,
     "Format of BOM data", _enum_values(DataFormat)),
    ("EHD-09", "Tag Format Example", "text", False,
     "Example of equipment tag (e.g., SAG-MILL-001)", None),
    ("EHD-10", "Functional Location Structure", "text", False,
     "SAP functional location structure description", None),
    ("EHD-11", "Equipment Master in SAP?", "dropdown", False,
     "Is equipment master data in SAP?", ["Yes", "No", "Partial"]),
]

SHEET_3_MAINTENANCE: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("MCS-01", "Strategy Maturity", "dropdown", True,
     "Current maintenance strategy maturity level", _enum_values(StrategyMaturity)),
    ("MCS-02", "CMMS Type", "dropdown", True,
     "Computerized Maintenance Management System in use", _enum_values(CMMSType)),
    ("MCS-03", "SAP Version", "dropdown", False,
     "SAP version (if SAP is the CMMS)", _enum_values(SAPVersion)),
    ("MCS-04", "Work Order History Available?", "dropdown", False,
     "Is historical work order data available?", ["Yes", "No", "Partial"]),
    ("MCS-05", "WO History Years", "number", False,
     "How many years of work order history are available?", None),
    ("MCS-06", "Failure Data Available?", "dropdown", False,
     "Is failure/breakdown data available?", ["Yes", "No", "Partial"]),
    ("MCS-07", "Failure Data Format", "dropdown", False,
     "Format of failure data", _enum_values(DataFormat)),
    ("MCS-08", "Downtime Tracking?", "dropdown", False,
     "Is equipment downtime tracked systematically?", ["Yes", "No"]),
    ("MCS-09", "Planned Maintenance Exists?", "dropdown", False,
     "Are there existing PM plans?", ["Yes", "No", "Partial"]),
    ("MCS-10", "PM Plan Format", "dropdown", False,
     "Format of existing PM plans", _enum_values(DataFormat)),
    ("MCS-11", "Prior Criticality Assessment?", "dropdown", False,
     "Has a criticality assessment been done before?", ["Yes", "No"]),
    ("MCS-12", "Criticality Method", "dropdown", False,
     "Method used for prior criticality assessment", _enum_values(CriticalityMethodRFI)),
    ("MCS-13", "Prior FMECA?", "dropdown", False,
     "Has an FMECA been done before?", ["Yes", "No"]),
    ("MCS-14", "Prior RCM Study?", "dropdown", False,
     "Has an RCM study been done before?", ["Yes", "No"]),
]

SHEET_4_ORGANIZATION: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("ORG-01", "Team Size", "number", True,
     "Total maintenance team size (including contractors)", None),
    ("ORG-02", "Organization Structure", "dropdown", False,
     "How is the maintenance team organized?", _enum_values(OrgStructure)),
    ("ORG-03", "Number of Shifts", "number", True,
     "Number of maintenance shifts (1-4)", None),
    ("ORG-04", "Contractor Maintenance?", "dropdown", False,
     "Are contractors used for maintenance?", ["Yes", "No"]),
    ("ORG-05", "Available Trades", "text", False,
     "List of maintenance trades, separated by semicolons "
     "(e.g., mechanical;electrical;instrumentation;welding)", None),
    ("ORG-06", "Dedicated Planner?", "dropdown", False,
     "Is there a dedicated maintenance planner?", ["Yes", "No"]),
    ("ORG-07", "Dedicated Reliability Engineer?", "dropdown", False,
     "Is there a dedicated reliability engineer?", ["Yes", "No"]),
    ("ORG-08", "Annual Maintenance Budget (USD)", "number", False,
     "Approximate annual maintenance budget in USD", None),
]

SHEET_5_STANDARDS: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("STD-01", "ISO Certifications", "text", False,
     "List of ISO certifications, separated by semicolons "
     "(e.g., ISO 55001;ISO 14001;ISO 45001)", None),
    ("STD-02", "Industry Regulations", "text", False,
     "Applicable industry regulations, separated by semicolons", None),
    ("STD-03", "LOTOTO Program?", "dropdown", False,
     "Is there a Lock Out/Tag Out/Try Out program?", ["Yes", "No"]),
    ("STD-04", "Procedure Documentation?", "dropdown", False,
     "Are maintenance procedures documented?", ["Yes", "No", "Partial"]),
    ("STD-05", "Procedure Language", "dropdown", True,
     "Language for maintenance procedures and work instructions",
     _enum_values(Language)),
    ("STD-06", "Safety Permits Required", "text", False,
     "Safety permits, separated by semicolons "
     "(e.g., LOTOTO;confined-space;working-at-height)", None),
    ("STD-07", "PPE Matrix Available?", "dropdown", False,
     "Is there a PPE matrix for maintenance activities?", ["Yes", "No"]),
    ("STD-08", "Environmental Permits", "text", False,
     "Environmental permits/requirements, separated by semicolons", None),
]

SHEET_6_KPI: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("KPI-01", "Current Availability (%)", "number", False,
     "Current equipment/plant availability (0-100%)", None),
    ("KPI-02", "Target Availability (%)", "number", True,
     "Target availability after strategy implementation (0-100%)", None),
    ("KPI-03", "Current MTBF (hours)", "number", False,
     "Current Mean Time Between Failures in hours", None),
    ("KPI-04", "Current MTTR (hours)", "number", False,
     "Current Mean Time To Repair in hours", None),
    ("KPI-05", "Current OEE (%)", "number", False,
     "Current Overall Equipment Effectiveness (0-100%)", None),
    ("KPI-06", "Annual Maintenance Cost (USD)", "number", False,
     "Annual maintenance expenditure in USD", None),
    ("KPI-07", "Current Planned vs Unplanned (%)", "number", False,
     "Current ratio of planned maintenance (0-100%)", None),
    ("KPI-08", "Target Planned vs Unplanned (%)", "number", False,
     "Target ratio of planned maintenance (0-100%)", None),
    ("KPI-09", "Current PM Compliance (%)", "number", False,
     "Current PM schedule compliance rate (0-100%)", None),
    ("KPI-10", "Target PM Compliance (%)", "number", False,
     "Target PM schedule compliance rate (0-100%)", None),
]

SHEET_7_SCOPE: list[tuple[str, str, str, bool, str, list[str] | None]] = [
    ("SCT-01", "Scope Type", "dropdown", True,
     "Type of scope for this engagement", _enum_values(ScopeType)),
    ("SCT-02", "Areas in Scope", "text", False,
     "List of areas/systems in scope, separated by semicolons", None),
    ("SCT-03", "Priority Equipment", "text", False,
     "Priority equipment tags, separated by semicolons "
     "(e.g., SAG-MILL-001;CRUSHER-001)", None),
    ("SCT-04", "Start Date", "date", True,
     "Project start date (YYYY-MM-DD)", None),
    ("SCT-05", "Target Completion", "date", True,
     "Target completion date (YYYY-MM-DD)", None),
    ("SCT-06", "Expected Deliverables", "text", False,
     "Expected deliverables, separated by semicolons", None),
    ("SCT-07", "Workshop Format", "dropdown", False,
     "Format for review workshops", _enum_values(WorkshopFormat)),
]

# All sheets in order (name, field definitions)
ALL_SHEETS: list[tuple[str, list[tuple[str, str, str, bool, str, list[str] | None]]]] = [
    ("1-Company Profile", SHEET_1_COMPANY),
    ("2-Equipment Data", SHEET_2_EQUIPMENT),
    ("3-Maintenance State", SHEET_3_MAINTENANCE),
    ("4-Organization", SHEET_4_ORGANIZATION),
    ("5-Standards", SHEET_5_STANDARDS),
    ("6-KPI Baseline", SHEET_6_KPI),
    ("7-Scope & Timeline", SHEET_7_SCOPE),
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, max_col: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int, options: list[str]) -> None:
    """Add dropdown validation to a range of cells."""
    formula = ",".join(options)
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "Please select from the dropdown list"
    dv.errorTitle = "Invalid input"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_length = max(max_length, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_instructions_sheet(wb: Workbook) -> None:
    """Build the general instructions sheet (first sheet)."""
    ws = wb.active
    ws.title = "Instructions"

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 60

    instructions = [
        ("AMS RFI Questionnaire", ""),
        ("", ""),
        ("Purpose:", "This questionnaire collects information about your facility, "
         "equipment, and maintenance practices to configure the AI-powered "
         "maintenance strategy development system."),
        ("", ""),
        ("How to complete:", ""),
        ("1.", "Fill in each sheet (tabs at the bottom) with the requested information."),
        ("2.", "Fields highlighted in YELLOW are REQUIRED. Green fields are OPTIONAL."),
        ("3.", "Use dropdown menus where provided. For text fields, type your response."),
        ("4.", "For list fields (e.g., trades, certifications), separate items with semicolons (;)."),
        ("5.", "Date fields should use YYYY-MM-DD format."),
        ("6.", "The last sheet (Data Availability) asks about data you can provide for each template."),
        ("", ""),
        ("Sheets:", ""),
        ("1-Company Profile", "Basic company and plant information"),
        ("2-Equipment Data", "Equipment inventory and hierarchy details"),
        ("3-Maintenance State", "Current maintenance practices and systems"),
        ("4-Organization", "Team structure and resources"),
        ("5-Standards", "Compliance and safety standards"),
        ("6-KPI Baseline", "Current and target performance metrics"),
        ("7-Scope & Timeline", "Project scope and schedule"),
        ("8-Data Availability", "Checklist of available data mapped to AMS templates"),
        ("", ""),
        ("Legend:", ""),
        ("", ""),
        ("Version:", "1.0"),
        ("Generated by:", "AMS RFI Template Generator"),
    ]

    for i, (col_a, col_b) in enumerate(instructions, start=1):
        cell_a = ws.cell(row=i, column=1, value=col_a)
        cell_b = ws.cell(row=i, column=2, value=col_b)
        cell_a.font = NORMAL_FONT
        cell_b.font = NORMAL_FONT
        cell_a.alignment = WRAP_ALIGNMENT
        cell_b.alignment = WRAP_ALIGNMENT

    # Title row
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=18, bold=True, color="1F4E79")

    # Section headers
    for row_idx in [5, 13, 23]:
        ws.cell(row=row_idx, column=1).font = SECTION_FONT

    # Legend: required/optional color samples
    legend_row = 24
    ws.cell(row=legend_row, column=1, value="REQUIRED field").fill = REQUIRED_FILL
    ws.cell(row=legend_row, column=1).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=legend_row + 1, column=1, value="OPTIONAL field").fill = OPTIONAL_FILL
    ws.cell(row=legend_row + 1, column=1).font = NORMAL_FONT


def _build_questionnaire_sheet(
    wb: Workbook,
    title: str,
    fields: list[tuple[str, str, str, bool, str, list[str] | None]],
) -> None:
    """Build a standard questionnaire sheet with ID, Field, Response, etc."""
    ws = wb.create_sheet(title=title)

    # Headers
    headers = ["ID", "Field", "Response", "Type", "Required", "Help / Instructions"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    _style_header_row(ws, HEADER_ROW, len(headers))

    # Data rows
    for i, (fid, label, ftype, required, help_text, dropdown_opts) in enumerate(fields):
        row = DATA_START_ROW + i

        ws.cell(row=row, column=COL_ID, value=fid).font = NORMAL_FONT
        ws.cell(row=row, column=COL_FIELD, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=COL_TYPE, value=ftype).font = NORMAL_FONT
        ws.cell(row=row, column=COL_REQUIRED, value="Required" if required else "Optional").font = NORMAL_FONT
        ws.cell(row=row, column=COL_HELP, value=help_text).font = NORMAL_FONT

        # Response cell styling
        response_cell = ws.cell(row=row, column=COL_RESPONSE)
        response_cell.font = NORMAL_FONT
        response_cell.alignment = WRAP_ALIGNMENT
        if required:
            response_cell.fill = REQUIRED_FILL
        else:
            response_cell.fill = OPTIONAL_FILL

        # Apply borders to all cells in row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        # Add dropdown if applicable
        if dropdown_opts:
            col_letter = get_column_letter(COL_RESPONSE)
            _add_dropdown(ws, col_letter, row, row, dropdown_opts)

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 55


def _build_data_availability_sheet(wb: Workbook) -> None:
    """Build Sheet 8: Data Availability Checklist (14 rows mapped to AMS templates)."""
    ws = wb.create_sheet(title="8-Data Availability")

    headers = [
        "Template ID", "Template Name", "Available?",
        "Format", "Quality (1-5)", "Notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    _style_header_row(ws, HEADER_ROW, len(headers))

    for i, (template_id, template_name) in enumerate(AMS_TEMPLATES):
        row = DATA_START_ROW + i

        ws.cell(row=row, column=1, value=template_id).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=template_name).font = NORMAL_FONT

        # Available? dropdown
        avail_cell = ws.cell(row=row, column=3)
        avail_cell.font = NORMAL_FONT
        avail_cell.fill = REQUIRED_FILL

        # Format dropdown
        ws.cell(row=row, column=4).font = NORMAL_FONT
        ws.cell(row=row, column=4).fill = OPTIONAL_FILL

        # Quality score
        ws.cell(row=row, column=5).font = NORMAL_FONT
        ws.cell(row=row, column=5).fill = OPTIONAL_FILL

        # Notes
        ws.cell(row=row, column=6).font = NORMAL_FONT
        ws.cell(row=row, column=6).fill = OPTIONAL_FILL

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Dropdowns for Available? column
    end_row = DATA_START_ROW + len(AMS_TEMPLATES) - 1
    _add_dropdown(ws, "C", DATA_START_ROW, end_row, ["Yes", "No", "Partial"])
    _add_dropdown(ws, "D", DATA_START_ROW, end_row, _enum_values(DataFormat))
    # Quality score dropdown (1-5)
    _add_dropdown(ws, "E", DATA_START_ROW, end_row, ["1", "2", "3", "4", "5"])

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 45


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_rfi_template(output_path: Path | None = None) -> Path:
    """Generate the complete RFI questionnaire Excel template.

    Args:
        output_path: Where to save. Defaults to templates/rfi/00_rfi_questionnaire.xlsx

    Returns:
        Path to the generated file.
    """
    if output_path is None:
        output_path = DEFAULT_OUTPUT

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 0: Instructions (uses the default active sheet)
    _build_instructions_sheet(wb)

    # Sheets 1-7: Questionnaire sheets
    for title, fields in ALL_SHEETS:
        _build_questionnaire_sheet(wb, title, fields)

    # Sheet 8: Data Availability
    _build_data_availability_sheet(wb)

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate AMS RFI Questionnaire Excel Template",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help=f"Output path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    path = generate_rfi_template(args.output)
    print(f"RFI template generated: {path}")


if __name__ == "__main__":
    main()
