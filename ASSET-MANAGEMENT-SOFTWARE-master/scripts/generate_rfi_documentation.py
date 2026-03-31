"""Generate AMS Required Documentation Excel.

Creates a structured tracker of all documents the client must provide
for the AI-Powered Asset Management and Maintenance Solution to be
fully customized. Based on the executive RFI document.

Usage:
    python scripts/generate_rfi_documentation.py
    python scripts/generate_rfi_documentation.py --output custom-path.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_PROJECT_ROOT = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------------
# Styling constants (match AMS template convention)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CRITICAL_FILL = PatternFill(start_color="FDDEDE", end_color="FDDEDE", fill_type="solid")  # light red
IMPORTANT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
DESIRABLE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # light blue
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="Calibri", size=10)
SMALL_FONT = Font(name="Calibri", size=9, color="666666")
WRAP_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

PRIORITY_FILL = {
    "CRITICAL": CRITICAL_FILL,
    "IMPORTANT": IMPORTANT_FILL,
    "DESIRABLE": DESIRABLE_FILL,
}

DEFAULT_OUTPUT = _PROJECT_ROOT / "templates" / "rfi" / "01_required_documentation.xlsx"


# ---------------------------------------------------------------------------
# Document categories and items
# Each tuple: (item_id, item_name, description, format, sap_transaction)
# Items grouped by category with priority and delivery week
# ---------------------------------------------------------------------------

CATEGORIES: list[tuple[str, str, str, str, list[tuple[str, str, str, str]]]] = [
    # (cat_id, category_name, priority, delivery_week, items)
    # items: [(item_id, item_name, description, format)]

    ("DOC-01", "Work Management Workflow Documentation", "DESIRABLE", "Week 5-6", [
        ("01.01", "AS-IS Workflow Map",
         "Current work request process: how inspectors/operators submit requests (email, verbal, SAP, etc.), "
         "current approval chain and decision points, information flow between production/maintenance/planning, "
         "average times per stage, identified friction points",
         "PowerPoint / Visio / Word"),
        ("01.02", "Work Request Standard",
         "Templates currently in use, required information fields, approval authorities by work type, "
         "priority classification criteria",
         "Word / PDF"),
        ("01.03", "Planning Process Documentation",
         "How planner receives and processes requests, information gathering steps (materials, resources, timing), "
         "communication protocols with warehouse/supervisors/production, scheduling rules and constraints",
         "Word / PDF"),
        ("01.04", "Work Management Framework & SAP Configuration Blueprint",
         "Business Process Maps for Work Management: Identification & Prioritization, Planning, Scheduling, "
         "Execution, Monitoring/Controlling/Closing. SAP PM rules and Master Data configuration",
         "PowerPoint / Visio / Word"),
    ]),

    ("DOC-02", "Equipment Hierarchy & Master Data", "CRITICAL", "Week 1-2", [
        ("02.01", "Complete SAP Hierarchical Structure",
         "Full hierarchy: Plant > Area > System > Equipment > Component. "
         "Asset Register with Equipment Make, Model, Component TAGs",
         "CSV / Excel export from SAP"),
        ("02.02", "Pilot Equipment TAGs",
         "TAGs of the 1-2 selected pilot equipment and their complete sub-hierarchy",
         "CSV / Excel"),
        ("02.03", "Criticality Classification",
         "Criticality rating for equipment (e.g., AA, A+, C, D) as currently defined",
         "CSV / Excel export from SAP"),
        ("02.04", "Physical Location Mapping",
         "Physical location identifiers for each equipment in the hierarchy",
         "CSV / Excel export from SAP"),
    ]),

    ("DOC-03", "Work Orders History", "CRITICAL", "Week 1-2", [
        ("03.01", "Work Orders Export (min. 12 months)",
         "Work order number, Equipment TAG, Type (PM01 Inspection / PM02 Preventive / PM03 Corrective), "
         "Execution date, Duration in man-hours, Equipment/Component worked on, Problem description, "
         "Materials consumed, Final status (completed/pending/rescheduled), Reason for postponement",
         "SAP export IW38/IW39"),
        ("03.02", "Work Order Notifications",
         "Associated notifications with cause codes, damage codes, and activity types",
         "SAP export IW28/IW29"),
    ]),

    ("DOC-04", "Spare Parts Catalog (BOM)", "CRITICAL", "Week 1-2", [
        ("04.01", "Bill of Materials per Pilot Equipment",
         "SAP spare part code, Spare part description, Common description, "
         "Applicable equipment, Supplier lead time, Criticality",
         "Excel / CSV from SAP"),
        ("04.02", "BOM Structure Export",
         "Full multi-level BOM structure showing parent-child relationships for pilot equipment",
         "SAP export IB01/CS12"),
    ]),

    ("DOC-05", "Current Spare Parts Inventory", "CRITICAL", "Week 1-2", [
        ("05.01", "Inventory Levels",
         "SAP code, Available quantity, Warehouse location, Minimum/safety stock levels",
         "Excel / CSV from SAP"),
        ("05.02", "Stock Movement History (12 months)",
         "Consumption history per material code — issues, receipts, adjustments",
         "SAP export MB52/MB51"),
    ]),

    ("DOC-06", "Current Maintenance Backlog", "CRITICAL", "Week 1-2", [
        ("06.01", "Pending Work Requests (last 3 months)",
         "Work request ID, Assigned priority, Creation date, Current status, "
         "Estimated resources, Reason for being pending",
         "Excel / CSV"),
        ("06.02", "Backlog Age Analysis",
         "Backlog aging breakdown (0-7 days, 7-30 days, 30-90 days, 90+ days) if available",
         "Excel / CSV"),
    ]),

    ("DOC-07", "Technical Manuals", "IMPORTANT", "Week 3-4", [
        ("07.01", "Manufacturer O&M Manual",
         "Original Equipment Manufacturer Operation & Maintenance manual for each pilot equipment",
         "PDF"),
        ("07.02", "P&IDs (Piping & Instrumentation Diagrams)",
         "Process flow diagrams and P&IDs for the pilot area/system",
         "PDF / DWG / Images"),
        ("07.03", "Technical Specifications / Datasheets",
         "Equipment datasheets with nameplate data, design parameters, operating limits",
         "PDF / Excel"),
    ]),

    ("DOC-08", "Preventive Maintenance Plans", "IMPORTANT", "Week 3-4", [
        ("08.01", "PM Plan per Pilot Equipment",
         "Frequency of each PM task, Task list (step-by-step), Associated spare parts, "
         "Estimated duration, Required specialties",
         "Excel / CSV from SAP IP10"),
        ("08.02", "PM Execution History",
         "Last execution date, Next scheduled date, Compliance rate for each PM task",
         "Excel / CSV from SAP"),
        ("08.03", "Maintenance Strategy in SAP",
         "Maintenance packages, task lists, and strategy configuration for pilot equipment",
         "SAP export IP16/IA09"),
    ]),

    ("DOC-09", "Available Workforce", "IMPORTANT", "Week 3-4", [
        ("09.01", "Technician Roster by Specialty",
         "Number of technicians per specialty: Mechanics, Electricians, Instrumentation, "
         "Other specialists (welders, riggers, etc.)",
         "Simple Excel"),
        ("09.02", "Work Shifts & Availability",
         "Shift patterns, current availability percentage, overtime rules",
         "Simple Excel"),
        ("09.03", "Contractor Resources",
         "External contractor availability, specialties, response times",
         "Simple Excel"),
    ]),

    ("DOC-10", "Shutdown Calendar", "IMPORTANT", "Week 3-4", [
        ("10.01", "Scheduled Shutdowns (next 6 months)",
         "Scheduled shutdown dates, Type (minor 8hrs / major >20hrs), "
         "Duration, Affected equipment",
         "Excel / Calendar"),
        ("10.02", "Historical Shutdown Records",
         "Past 12 months of actual shutdowns: planned vs. unplanned, duration, cause",
         "Excel / CSV"),
    ]),

    ("DOC-11", "Production Plan and Schedule", "IMPORTANT", "Week 3-4", [
        ("11.01", "Yearly Production Plan",
         "Annual production targets, throughput rates, seasonal variations",
         "Excel / Calendar"),
        ("11.02", "Monthly Production Schedule",
         "Basic monthly schedule with production windows and maintenance windows",
         "Excel / Calendar"),
        ("11.03", "Maintenance Windows & Restrictions",
         "Allowed maintenance windows, known restrictions, production vs. maintenance conflicts",
         "Excel / Calendar"),
    ]),

    ("DOC-12", "Basic FMEA (If Exists)", "DESIRABLE", "Week 5-6", [
        ("12.01", "FMEA / FMECA per Pilot Equipment",
         "Critical components, Identified failure modes, Failure effects, "
         "Recommended preventive tasks, RPN scores (if available)",
         "Excel"),
        ("12.02", "Failure History Analysis",
         "Top 10 failure modes by frequency, failure cause categorization (if available)",
         "Excel / PDF"),
    ]),

    ("DOC-13", "Reference Photographs", "DESIRABLE", "Week 5-6", [
        ("13.01", "Equipment Photographs (5-10 per equipment)",
         "Complete equipment overview, Main components, Typical spare parts, "
         "Inspection points, Nameplate",
         "JPG / PNG"),
        ("13.02", "Area / Layout Photographs",
         "General area layout, access points, safety signage",
         "JPG / PNG"),
    ]),

    ("DOC-14", "Condition Monitoring Data (If Exists)", "DESIRABLE", "Week 5-6", [
        ("14.01", "Vibration Readings",
         "Latest vibration analysis reports for rotating equipment",
         "PDF / Excel"),
        ("14.02", "Thermography Reports",
         "Recent thermography inspection reports for electrical and mechanical equipment",
         "PDF / Images"),
        ("14.03", "Active Alarms & Trends",
         "Current active alarms from DCS/SCADA, trend data for key parameters",
         "PDF / Excel / CSV"),
    ]),

    ("DOC-15", "Reference Costs", "DESIRABLE", "Week 5-6", [
        ("15.01", "Man-Hour Cost per Specialty",
         "Internal and contractor rates per trade (mechanical, electrical, instrumentation, etc.)",
         "Simple Excel"),
        ("15.02", "Average Material Costs",
         "Average cost of commonly consumed materials and spare parts",
         "Excel / CSV"),
        ("15.03", "Spares Consumption by Equipment",
         "Annual consumption cost breakdown by equipment or system",
         "Excel / CSV from SAP"),
        ("15.04", "Production Loss Cost",
         "Estimated cost per hour of downtime by area/system",
         "Simple Excel"),
    ]),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, num_cols: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int, options: list[str]) -> None:
    """Add dropdown validation to a column range."""
    formula = ",".join(options)
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "Please select from the dropdown"
    dv.errorTitle = "Invalid Value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_instructions_sheet(wb: Workbook) -> None:
    """Build the Instructions sheet."""
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "1F4E79"

    instructions = [
        ("AI-Powered Asset Management — Required Documentation", SECTION_FONT),
        ("", NORMAL_FONT),
        ("PURPOSE", Font(name="Calibri", size=11, bold=True)),
        ("This workbook lists all documents required from the client to fully customize", NORMAL_FONT),
        ("the AI-Powered Asset Management and Maintenance Solution.", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("HOW TO USE", Font(name="Calibri", size=11, bold=True)),
        ("1. Go to the 'Document Tracker' sheet — this is the main checklist", NORMAL_FONT),
        ("2. For each document item, update the STATUS column:", NORMAL_FONT),
        ("   - Pending: Not yet collected", NORMAL_FONT),
        ("   - In Progress: Being gathered by client", NORMAL_FONT),
        ("   - Received: Delivered to VSC", NORMAL_FONT),
        ("   - N/A: Not applicable for this engagement", NORMAL_FONT),
        ("3. Fill in Received Date when documents arrive", NORMAL_FONT),
        ("4. Fill in Responsible (Client) — who will provide this document", NORMAL_FONT),
        ("5. Use Quality (1-5) to rate data quality once received", NORMAL_FONT),
        ("6. Add Notes for any issues, alternatives, or clarifications", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("PRIORITY LEVELS", Font(name="Calibri", size=11, bold=True)),
        ("CRITICAL (Red)     — Week 1-2: Absolute minimum to start development", NORMAL_FONT),
        ("IMPORTANT (Yellow)  — Week 3-4: Required to complete functionalities", NORMAL_FONT),
        ("DESIRABLE (Green)   — Week 5-6: Enriches the solution", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("DELIVERY TIMELINE", Font(name="Calibri", size=11, bold=True)),
        ("Week 1-2: Equipment hierarchy, Work orders, Spare parts, Inventory, Backlog", NORMAL_FONT),
        ("Week 3-4: Technical manuals, PM plans, Workforce, Shutdowns, Production plan", NORMAL_FONT),
        ("Week 5-6: FMEA, Photographs, Condition monitoring, Costs, Workflows", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("QUALITY SCALE", Font(name="Calibri", size=11, bold=True)),
        ("1 = Unusable (major errors, missing key fields)", NORMAL_FONT),
        ("2 = Poor (many gaps, needs significant cleaning)", NORMAL_FONT),
        ("3 = Acceptable (some gaps, usable with effort)", NORMAL_FONT),
        ("4 = Good (minor gaps, mostly clean)", NORMAL_FONT),
        ("5 = Excellent (complete, clean, ready to use)", NORMAL_FONT),
    ]

    for i, (text, font) in enumerate(instructions, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 90


def _build_tracker_sheet(wb: Workbook) -> None:
    """Build the main Document Tracker sheet."""
    ws = wb.create_sheet("Document Tracker")
    ws.sheet_properties.tabColor = "C00000"

    headers = [
        "ID", "Category", "Document / Item", "Specific Content Required",
        "Expected Format", "SAP Transaction", "Priority", "Delivery Target",
        "Status", "Received Date", "Responsible (Client)", "Quality (1-5)",
        "File Name(s)", "Notes",
    ]
    for col, header in enumerate(headers, 1):
        _style_header_row(ws, 1, len(headers))
        ws.cell(row=1, column=col, value=header)

    # Freeze header row
    ws.freeze_panes = "A2"

    row = 2
    for cat_id, cat_name, priority, week, items in CATEGORIES:
        # Section header row
        ws.cell(row=row, column=1, value=cat_id).font = SECTION_FONT
        ws.cell(row=row, column=2, value=cat_name).font = SECTION_FONT
        ws.cell(row=row, column=7, value=priority).font = SECTION_FONT
        ws.cell(row=row, column=8, value=week).font = SECTION_FONT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = SECTION_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        # Item rows
        for item_id, item_name, description, fmt in items:
            full_id = f"{cat_id.replace('DOC-', '')}.{item_id.split('.')[1]}"
            ws.cell(row=row, column=1, value=item_id).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=cat_name).font = SMALL_FONT
            ws.cell(row=row, column=3, value=item_name).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=description).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=fmt).font = NORMAL_FONT
            ws.cell(row=row, column=6, value="").font = NORMAL_FONT  # SAP transaction
            ws.cell(row=row, column=7, value=priority).font = NORMAL_FONT
            ws.cell(row=row, column=8, value=week).font = NORMAL_FONT
            ws.cell(row=row, column=9, value="Pending").font = NORMAL_FONT

            # Apply priority color to status cell
            priority_fill = PRIORITY_FILL.get(priority, DESIRABLE_FILL)
            ws.cell(row=row, column=9).fill = priority_fill

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

            row += 1

    # Add SAP transactions where known
    sap_transactions = {
        "02.01": "IE01/IH01",
        "02.03": "IH06/IH08",
        "03.01": "IW38/IW39",
        "03.02": "IW28/IW29",
        "04.01": "CS12/IB01",
        "04.02": "CS12/IB01",
        "05.01": "MB52/MMBE",
        "05.02": "MB51",
        "06.01": "IW38 (open)",
        "08.01": "IP10/IP16",
        "08.02": "IP10/IP24",
        "08.03": "IP16/IA09",
    }
    for r in range(2, row):
        item_id = ws.cell(row=r, column=1).value
        if item_id in sap_transactions:
            ws.cell(row=r, column=6, value=sap_transactions[item_id]).font = SMALL_FONT

    # Dropdowns
    last_data_row = row - 1
    _add_dropdown(ws, "I", 2, last_data_row, ["Pending", "In Progress", "Received", "N/A"])
    _add_dropdown(ws, "L", 2, last_data_row, ["1", "2", "3", "4", "5"])

    # Column widths
    widths = {
        "A": 8, "B": 28, "C": 32, "D": 55, "E": 22, "F": 14,
        "G": 12, "H": 13, "I": 13, "J": 13, "K": 20, "L": 10,
        "M": 25, "N": 35,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _build_summary_sheet(wb: Workbook) -> None:
    """Build the Delivery Timeline Summary sheet."""
    ws = wb.create_sheet("Delivery Timeline")
    ws.sheet_properties.tabColor = "548235"

    headers = ["Phase", "Delivery Target", "Priority", "Categories", "# Documents",
               "What This Enables"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    phases = [
        ("Phase 1", "Week 1-2", "CRITICAL",
         "Equipment Hierarchy, Work Orders History, Spare Parts Catalog, "
         "Spare Parts Inventory, Maintenance Backlog",
         sum(len(items) for _, _, p, _, items in CATEGORIES if p == "CRITICAL"),
         "Start prototype of 3 core functionalities: Intelligent capture, "
         "Planner assistant, Backlog optimization"),
        ("Phase 2", "Week 3-4", "IMPORTANT",
         "Technical Manuals, PM Plans, Workforce, Shutdown Calendar, "
         "Production Plan & Schedule",
         sum(len(items) for _, _, p, _, items in CATEGORIES if p == "IMPORTANT"),
         "Complete functionalities: Refine algorithms, validate business logic, "
         "generate automatic weekly schedule"),
        ("Phase 3", "Week 5-6", "DESIRABLE",
         "Work Management Workflows, FMEA, Photographs, "
         "Condition Monitoring, Reference Costs",
         sum(len(items) for _, _, p, _, items in CATEGORIES if p == "DESIRABLE"),
         "Enrich solution: Predictive failure analysis, PM strategy optimization, "
         "advanced insights, optimize UX"),
    ]

    for i, (phase, target, priority, cats, count, enables) in enumerate(phases, start=2):
        ws.cell(row=i, column=1, value=phase).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=i, column=2, value=target).font = NORMAL_FONT
        ws.cell(row=i, column=3, value=priority).font = NORMAL_FONT
        ws.cell(row=i, column=4, value=cats).font = NORMAL_FONT
        ws.cell(row=i, column=5, value=count).font = NORMAL_FONT
        ws.cell(row=i, column=6, value=enables).font = NORMAL_FONT

        fill = PRIORITY_FILL.get(priority, DESIRABLE_FILL)
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).fill = fill
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = WRAP_ALIGNMENT

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 55

    # Add totals row
    total_row = 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=11, bold=True)
    total_items = sum(len(items) for _, _, _, _, items in CATEGORIES)
    ws.cell(row=total_row, column=5, value=total_items).font = Font(name="Calibri", size=11, bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).border = THIN_BORDER


def _build_delivery_format_sheet(wb: Workbook) -> None:
    """Build Preferred Delivery Format sheet."""
    ws = wb.create_sheet("Delivery Format")
    ws.sheet_properties.tabColor = "7030A0"

    # Naming convention
    ws.cell(row=1, column=1, value="Preferred Delivery Format").font = Font(
        name="Calibri", size=14, bold=True, color="1F4E79")
    ws.merge_cells("A1:D1")

    ws.cell(row=3, column=1, value="For each document:").font = Font(
        name="Calibri", size=11, bold=True)

    guidelines = [
        ("Name", "Descriptive with date, e.g.: WO_History_SAG_MILL_01_2024-2025.xlsx"),
        ("Location", "Shared folder in Drive/SharePoint organized by category"),
        ("Contact", "Person responsible for that type of information"),
        ("Notes", "Any peculiarities or special considerations"),
    ]

    for i, (field, desc) in enumerate(guidelines, start=4):
        ws.cell(row=i, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=i, column=2, value=desc).font = NORMAL_FONT

    # Folder structure
    ws.cell(row=9, column=1, value="Proposed Folder Structure:").font = Font(
        name="Calibri", size=11, bold=True)

    folders = [
        "/CLIENT_PROJECT_Data/",
        "  /01_Equipment_MasterData/",
        "  /02_Work_Orders_History/",
        "  /03_Spare_Parts/",
        "  /04_Backlog/",
        "  /05_Manuals/",
        "  /06_Maintenance_Plans/",
        "  /07_Resources/",
        "  /08_Schedules/",
        "  /09_FMEA_RCM/",
        "  /10_Photos/",
        "  /11_Condition_Monitoring/",
        "  /12_Costs/",
        "  /13_Processes_Workflows/",
    ]

    for i, folder in enumerate(folders, start=10):
        ws.cell(row=i, column=1, value=folder).font = Font(
            name="Consolas", size=10, color="333333")

    # Quality notes
    ws.cell(row=25, column=1, value="Data Quality Notes:").font = Font(
        name="Calibri", size=11, bold=True)
    quality_notes = [
        "We don't need perfection in initial data — 'good' data quickly is better than 'perfect' data late.",
        "We can iterate and improve quality progressively.",
        "Dummy/synthetic data can be generated if some elements don't exist yet.",
        "TAGs and names can be anonymized if necessary for confidentiality.",
    ]
    for i, note in enumerate(quality_notes, start=26):
        ws.cell(row=i, column=1, value=f"  - {note}").font = NORMAL_FONT

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_rfi_documentation(output_path: Path | None = None) -> Path:
    """Generate the Required Documentation Excel workbook."""
    output_path = Path(output_path) if output_path else DEFAULT_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _build_instructions_sheet(wb)
    _build_tracker_sheet(wb)
    _build_summary_sheet(wb)
    _build_delivery_format_sheet(wb)

    wb.save(str(output_path))
    print(f"Required documentation tracker generated: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate AMS Required Documentation Excel tracker"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help=f"Output path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()
    generate_rfi_documentation(args.output)


if __name__ == "__main__":
    main()
