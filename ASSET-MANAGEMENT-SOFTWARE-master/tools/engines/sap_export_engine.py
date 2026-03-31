"""
SAP Export Engine — Deterministic
Generates SAP PM upload templates from Work Packages.
Based on REF-03: 3 linked templates (Maintenance Item, Task List, Work Plan).
"""

from datetime import datetime
from pathlib import Path

from tools.models.schemas import (
    FREQ_UNIT_TRIGGER,
    MaintenanceTask,
    SAPMaintenanceItem,
    SAPMaintenancePlan,
    SAPOperation,
    SAPTaskList,
    SAPUploadPackage,
    SchedulingTrigger,
    WPConstraint,
    WorkPackage,
)

# SAP template file names (REF-03)
SAP_TEMPLATE_FILES = (
    "Maintenance Item.xlsx",
    "Task List.xlsx",
    "Work Plan.xlsx",
)

# SAP field length constants (AMS-C03 fix: no more magic numbers)
SAP_SHORT_TEXT_MAX = 72
SAP_FUNC_LOC_MAX = 40
SAP_TASK_LIST_DESC_MAX = 40

# Constraint to SAP system_condition mapping
CONSTRAINT_TO_SAP = {
    WPConstraint.ONLINE: 1,   # Running
    WPConstraint.OFFLINE: 3,  # Stopped
}

# Frequency unit mapping to SAP cycle unit
FREQ_UNIT_TO_SAP = {
    # Calendar-based (time-triggered plans)
    "DAYS": "DAY",
    "WEEKS": "WK",
    "MONTHS": "MON",
    "YEARS": "YR",
    "HOURS": "H",
    # Counter-based (usage-triggered plans)
    "HOURS_RUN": "H",
    "OPERATING_HOURS": "H",
    "TONNES": "T",
    "CYCLES": "CYC",
}


class SAPExportEngine:
    """Generates SAP PM upload package from work packages."""

    @staticmethod
    def generate_upload_package(
        work_packages: list[WorkPackage],
        plant_code: str,
        plan_description: str = "",
        tasks: dict[str, "MaintenanceTask"] | None = None,
    ) -> SAPUploadPackage:
        """
        Generate a complete SAP upload package from a list of work packages.
        Creates linked Maintenance Items, Task Lists, and a Maintenance Plan.

        Args:
            tasks: Optional dict mapping task_id → MaintenanceTask for field population.
        """
        if not work_packages:
            raise ValueError("At least one work package is required")

        tasks = tasks or {}
        maintenance_items = []
        task_lists = []

        for idx, wp in enumerate(work_packages, start=1):
            mi_ref = f"$MI{idx}"
            tl_ref = f"$TL{idx}"

            # Generate Maintenance Item
            mi = SAPMaintenanceItem(
                item_ref=mi_ref,
                description=wp.name,
                order_type="PM03",
                func_loc="",  # Must be filled from hierarchy
                main_work_center="",  # Must be filled from labour
                planner_group=1,
                task_list_ref=tl_ref,
                priority="4",  # Default planned
            )
            maintenance_items.append(mi)

            # Generate Task List with operations
            operations = []
            for task_alloc in wp.allocated_tasks:
                task = tasks.get(task_alloc.task_id)
                if task:
                    short_text = task.name[:SAP_SHORT_TEXT_MAX]
                    duration = sum(
                        lr.quantity * lr.hours_per_person for lr in task.labour_resources
                    ) or 0.5  # Fallback minimum
                    num_workers = sum(lr.quantity for lr in task.labour_resources) or 1
                    work_centre = (
                        task.labour_resources[0].specialty.value if task.labour_resources else ""
                    )
                else:
                    short_text = "Task placeholder"
                    duration = 0.5  # Default minimum
                    num_workers = 1
                    work_centre = ""

                op = SAPOperation(
                    operation_number=task_alloc.operation_number,
                    work_centre=work_centre,
                    control_key="PMIN",
                    short_text=short_text,
                    duration_hours=duration,
                    unit="H",
                    num_workers=num_workers,
                )
                operations.append(op)

            tl = SAPTaskList(
                list_ref=tl_ref,
                description=wp.name,
                func_loc="",
                system_condition=CONSTRAINT_TO_SAP.get(wp.constraint, 1),
                operations=operations,
            )
            task_lists.append(tl)

        # Determine cycle from first work package
        first_wp = work_packages[0]
        cycle_unit = FREQ_UNIT_TO_SAP.get(first_wp.frequency_unit.value, "DAY")
        trigger = FREQ_UNIT_TRIGGER.get(
            first_wp.frequency_unit, SchedulingTrigger.CALENDAR,
        )

        maintenance_plan = SAPMaintenancePlan(
            plan_id="",
            description=plan_description or f"Plan for {plant_code}",
            category="PM",
            cycle_value=int(first_wp.frequency_value),
            cycle_unit=cycle_unit,
            call_horizon_pct=50,
            scheduling_period=14,
            scheduling_unit="DAY",
            scheduling_trigger=trigger,
        )

        return SAPUploadPackage(
            plant_code=plant_code,
            maintenance_plan=maintenance_plan,
            maintenance_items=maintenance_items,
            task_lists=task_lists,
        )

    @staticmethod
    def validate_cross_references(package: SAPUploadPackage) -> list[str]:
        """Validate that all $MI/$TL cross-references are consistent."""
        errors = []

        mi_refs = {mi.item_ref for mi in package.maintenance_items}
        tl_refs = {tl.list_ref for tl in package.task_lists}

        # Every MI must reference an existing TL
        for mi in package.maintenance_items:
            if mi.task_list_ref not in tl_refs:
                errors.append(
                    f"Maintenance Item {mi.item_ref} references {mi.task_list_ref} "
                    f"which does not exist in task lists"
                )

        # Every TL should be referenced by at least one MI
        referenced_tls = {mi.task_list_ref for mi in package.maintenance_items}
        orphan_tls = tl_refs - referenced_tls
        for tl in orphan_tls:
            errors.append(f"Task List {tl} is not referenced by any Maintenance Item")

        return errors

    @staticmethod
    def validate_sap_field_lengths(package: SAPUploadPackage) -> list[str]:
        """Validate SAP field length constraints and counter-plan completeness."""
        errors = []
        for tl in package.task_lists:
            for op in tl.operations:
                if len(op.short_text) > SAP_SHORT_TEXT_MAX:
                    errors.append(
                        f"Operation {op.operation_number} in {tl.list_ref}: "
                        f"short_text exceeds {SAP_SHORT_TEXT_MAX} chars ({len(op.short_text)})"
                    )

        # Counter-based plan requires measuring point
        plan = package.maintenance_plan
        if (
            plan.scheduling_trigger == SchedulingTrigger.COUNTER
            and not plan.measuring_point
        ):
            errors.append(
                f"Counter-based plan (unit: {plan.cycle_unit}) requires "
                f"measuring_point to be filled before SAP upload."
            )

        return errors

    @staticmethod
    def resolve_sap_templates(
        client_slug: str,
        project_slug: str,
    ) -> dict[str, Path | None]:
        """Resolve SAP template file paths using the 3-level cascade.

        Returns a dict mapping template filename to its resolved Path (or None).
        Uses the template cascade: project → client → system fallback.
        """
        from agents._shared.paths import get_template_path

        return {
            name: get_template_path(name, client_slug, project_slug)
            for name in SAP_TEMPLATE_FILES
        }

    @staticmethod
    def write_to_xlsx(package: SAPUploadPackage, output_path: str | Path) -> Path:
        """Serialize a SAPUploadPackage to an .xlsx file.

        Generates 3 sheets mirroring the REF-03 SAP template structure:
          - Functional Locations  (from maintenance_items func_loc)
          - Task Lists            (operations per task list)
          - Maintenance Plans     (one row per maintenance item / plan link)

        Args:
            package: The SAP upload package to serialize.
            output_path: Destination path for the .xlsx file.

        Returns:
            The resolved output Path.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for SAP xlsx export. "
                "Install it with: pip install openpyxl"
            ) from exc

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        wb.properties.creator = "VSC AMS"
        wb.properties.description = "SAP PM Upload Package — DRAFT"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        def _style_header_row(ws, headers: list[str]) -> None:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

        # ── Sheet 1: Functional Locations ──────────────────────────────
        ws_fl = wb.active
        ws_fl.title = "Functional Locations"
        _style_header_row(ws_fl, [
            "Item Ref", "Description", "Functional Location",
            "Order Type", "Main Work Center", "Planner Group",
            "Task List Ref", "Priority",
        ])
        for mi in package.maintenance_items:
            func_loc = (mi.func_loc or "")[:SAP_FUNC_LOC_MAX]
            ws_fl.append([
                mi.item_ref,
                mi.description[:SAP_SHORT_TEXT_MAX],
                func_loc,
                mi.order_type,
                mi.main_work_center,
                mi.planner_group,
                mi.task_list_ref,
                mi.priority,
            ])

        # ── Sheet 2: Task Lists ────────────────────────────────────────
        ws_tl = wb.create_sheet("Task Lists")
        _style_header_row(ws_tl, [
            "List Ref", "Description", "Functional Location",
            "System Condition", "Op Number", "Work Centre",
            "Control Key", "Short Text", "Duration (h)", "Workers",
        ])
        for tl in package.task_lists:
            tl_desc = tl.description[:SAP_TASK_LIST_DESC_MAX]
            func_loc = (tl.func_loc or "")[:SAP_FUNC_LOC_MAX]
            if not tl.operations:
                ws_tl.append([
                    tl.list_ref, tl_desc, func_loc,
                    tl.system_condition, "", "", "", "", "", "",
                ])
            for op in tl.operations:
                ws_tl.append([
                    tl.list_ref,
                    tl_desc,
                    func_loc,
                    tl.system_condition,
                    op.operation_number,
                    op.work_centre,
                    op.control_key,
                    op.short_text[:SAP_SHORT_TEXT_MAX],
                    op.duration_hours,
                    op.num_workers,
                ])

        # ── Sheet 3: Maintenance Plans ─────────────────────────────────
        ws_mp = wb.create_sheet("Maintenance Plans")
        plan = package.maintenance_plan
        _style_header_row(ws_mp, [
            "Plan ID", "Description", "Category",
            "Cycle Value", "Cycle Unit", "Scheduling Trigger",
            "Start Date", "Item Ref", "Measuring Point",
        ])
        for mi in package.maintenance_items:
            ws_mp.append([
                plan.plan_id or "",
                plan.description[:SAP_SHORT_TEXT_MAX],
                plan.category,
                plan.cycle_value,
                plan.cycle_unit,
                plan.scheduling_trigger.value if hasattr(plan.scheduling_trigger, "value") else str(plan.scheduling_trigger),
                "",  # start_date — filled manually before SAP upload
                mi.item_ref,
                plan.measuring_point or "",
            ])

        # Auto-size columns (approximate)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(output_path)
        return output_path
