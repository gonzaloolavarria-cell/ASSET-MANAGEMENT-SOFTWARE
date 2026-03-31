"""File Parser Engine — GAP-W12.

Parses Excel (.xlsx) and CSV files into list[dict] for the import pipeline.
Deterministic — no LLM required.

Security:
  - File size limit (10 MB)
  - Row count limit (50,000)
  - Extension whitelist (.csv, .xlsx only — .xlsm rejected)
  - Read-only workbook mode (no macro execution)
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

from tools.models.schemas import FileParseError, FileParseResult

# Security constants
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 50_000
MAX_COLUMNS = 100
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
SKIP_SHEET_NAMES = {"instructions", "lookups", "rules", "validation", "lists"}

# Consecutive empty rows threshold to stop reading
EMPTY_ROW_THRESHOLD = 3


class FileParserEngine:
    """Parses Excel and CSV files into structured row data."""

    @staticmethod
    def parse_file(
        file_content: bytes,
        filename: str,
        sheet_name: str | None = None,
        max_rows: int = MAX_ROWS,
    ) -> FileParseResult:
        """Parse a file based on its extension.

        Args:
            file_content: Raw bytes of the file.
            filename: Original filename (used for extension detection).
            sheet_name: For Excel files, which sheet to parse. None = auto-detect.
            max_rows: Maximum rows to read.

        Returns:
            FileParseResult with headers, rows, and any errors.
        """
        if len(file_content) > MAX_FILE_SIZE_BYTES:
            return FileParseResult(
                success=False,
                filename=filename,
                file_type="unknown",
                errors=[FileParseError(
                    message=f"File exceeds maximum size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB",
                )],
            )

        ext = Path(filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            return FileParseResult(
                success=False,
                filename=filename,
                file_type=ext.lstrip("."),
                errors=[FileParseError(
                    message=f"Unsupported file extension '{ext}'. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
                )],
            )

        if ext == ".xlsx":
            return FileParserEngine.parse_excel(file_content, filename, sheet_name, max_rows)
        else:
            return FileParserEngine.parse_csv(file_content, filename, max_rows)

    @staticmethod
    def parse_excel(
        file_content: bytes,
        filename: str = "file.xlsx",
        sheet_name: str | None = None,
        max_rows: int = MAX_ROWS,
    ) -> FileParseResult:
        """Parse an Excel (.xlsx) file into list[dict].

        Args:
            file_content: Raw bytes of the Excel file.
            filename: Original filename.
            sheet_name: Sheet to parse. None = first data sheet (skips Instructions etc.).
            max_rows: Maximum rows to read.

        Returns:
            FileParseResult with parsed data.
        """
        errors: list[FileParseError] = []

        try:
            wb = load_workbook(
                io.BytesIO(file_content),
                data_only=True,
                read_only=True,
            )
        except Exception as e:
            return FileParseResult(
                success=False,
                filename=filename,
                file_type="xlsx",
                errors=[FileParseError(message=f"Failed to open Excel file: {e}")],
            )

        sheets_available = wb.sheetnames

        # Select sheet
        target_sheet = sheet_name
        if target_sheet is None:
            # Auto-select: first sheet whose name is NOT in SKIP_SHEET_NAMES
            for sn in sheets_available:
                if sn.lower().strip() not in SKIP_SHEET_NAMES:
                    target_sheet = sn
                    break
            if target_sheet is None and sheets_available:
                target_sheet = sheets_available[0]

        if target_sheet not in sheets_available:
            wb.close()
            return FileParseResult(
                success=False,
                filename=filename,
                file_type="xlsx",
                sheets_available=sheets_available,
                errors=[FileParseError(
                    message=f"Sheet '{target_sheet}' not found. Available: {', '.join(sheets_available)}",
                    sheet=target_sheet,
                )],
            )

        ws = wb[target_sheet]

        # Read headers from row 1
        headers: list[str] = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            wb.close()
            return FileParseResult(
                success=True,
                filename=filename,
                file_type="xlsx",
                sheets_available=sheets_available,
                sheet_parsed=target_sheet,
                total_rows=0,
            )

        for val in header_row:
            if val is None:
                break
            headers.append(str(val).strip())

        if not headers:
            wb.close()
            return FileParseResult(
                success=True,
                filename=filename,
                file_type="xlsx",
                sheets_available=sheets_available,
                sheet_parsed=target_sheet,
                total_rows=0,
            )

        if len(headers) > MAX_COLUMNS:
            headers = headers[:MAX_COLUMNS]
            errors.append(FileParseError(
                message=f"Column count exceeds {MAX_COLUMNS}, truncated",
                sheet=target_sheet,
            ))

        # Read data rows
        rows: list[dict] = []
        consecutive_empty = 0

        for row_tuple in ws.iter_rows(min_row=2, values_only=True):
            # Check max rows
            if len(rows) >= max_rows:
                errors.append(FileParseError(
                    message=f"Row limit of {max_rows} reached, remaining rows skipped",
                    sheet=target_sheet,
                ))
                break

            # Extract values for known header columns only
            values = list(row_tuple)[:len(headers)]

            # Check if row is completely empty
            if all(v is None for v in values):
                consecutive_empty += 1
                if consecutive_empty >= EMPTY_ROW_THRESHOLD:
                    break
                continue
            consecutive_empty = 0

            row_dict: dict = {}
            for col_idx, header in enumerate(headers):
                val = values[col_idx] if col_idx < len(values) else None
                row_dict[header] = _normalize_cell_value(val)

            rows.append(row_dict)

        wb.close()

        return FileParseResult(
            success=True,
            filename=filename,
            file_type="xlsx",
            sheets_available=sheets_available,
            sheet_parsed=target_sheet,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            errors=errors,
        )

    @staticmethod
    def parse_csv(
        file_content: bytes,
        filename: str = "file.csv",
        max_rows: int = MAX_ROWS,
        encoding: str | None = None,
        delimiter: str | None = None,
    ) -> FileParseResult:
        """Parse a CSV file into list[dict].

        Args:
            file_content: Raw bytes of the CSV file.
            filename: Original filename.
            max_rows: Maximum rows to read.
            encoding: Text encoding. None = auto-detect (UTF-8 → Latin-1 fallback).
            delimiter: Column delimiter. None = auto-detect via csv.Sniffer.

        Returns:
            FileParseResult with parsed data.
        """
        errors: list[FileParseError] = []

        # Detect encoding
        if encoding is None:
            encoding = FileParserEngine.detect_encoding(file_content)

        try:
            text = file_content.decode(encoding)
        except (UnicodeDecodeError, LookupError) as e:
            return FileParseResult(
                success=False,
                filename=filename,
                file_type="csv",
                errors=[FileParseError(message=f"Failed to decode file with encoding '{encoding}': {e}")],
            )

        # Strip BOM if present
        if text.startswith("\ufeff"):
            text = text[1:]

        if not text.strip():
            return FileParseResult(
                success=True,
                filename=filename,
                file_type="csv",
                total_rows=0,
            )

        # Detect delimiter
        if delimiter is None:
            delimiter = _detect_csv_delimiter(text)

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames is None:
            return FileParseResult(
                success=True,
                filename=filename,
                file_type="csv",
                total_rows=0,
            )

        headers = [h.strip() for h in reader.fieldnames if h is not None]

        if len(headers) > MAX_COLUMNS:
            headers = headers[:MAX_COLUMNS]
            errors.append(FileParseError(message=f"Column count exceeds {MAX_COLUMNS}, truncated"))

        rows: list[dict] = []
        for row_num, row in enumerate(reader, start=2):
            if len(rows) >= max_rows:
                errors.append(FileParseError(
                    message=f"Row limit of {max_rows} reached, remaining rows skipped",
                    row=row_num,
                ))
                break

            # Skip fully empty rows (only check known header columns)
            vals = [row.get(h, "") for h in headers]
            if all(not v or (isinstance(v, str) and not v.strip()) for v in vals):
                continue

            clean: dict = {}
            for h in headers:
                val = row.get(h, "")
                clean[h] = val.strip() if isinstance(val, str) else val

            rows.append(clean)

        return FileParseResult(
            success=True,
            filename=filename,
            file_type="csv",
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            errors=errors,
        )

    @staticmethod
    def list_sheets(file_content: bytes) -> list[str]:
        """List sheet names in an Excel file without parsing data."""
        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []

    @staticmethod
    def detect_encoding(file_content: bytes) -> str:
        """Detect text encoding from raw bytes.

        Strategy: try UTF-8 first (most common), then Latin-1 (SAP exports).
        """
        # Check for BOM
        if file_content[:3] == b"\xef\xbb\xbf":
            return "utf-8-sig"
        if file_content[:2] in (b"\xff\xfe", b"\xfe\xff"):
            return "utf-16"

        # Try UTF-8
        try:
            file_content[:4096].decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            pass

        # Fallback to Latin-1 (always succeeds but may misinterpret)
        return "latin-1"


def _normalize_cell_value(val: object) -> object:
    """Normalize openpyxl cell values for JSON serialization."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.date().isoformat()
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()


def _detect_csv_delimiter(text: str) -> str:
    """Auto-detect CSV delimiter using csv.Sniffer."""
    try:
        sample = text[:8192]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","
