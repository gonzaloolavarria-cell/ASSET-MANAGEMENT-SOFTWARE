"""Gantt Generator — produces Gantt data and Excel exports.

Converts WeeklyProgram work packages into structured GanttRow data
and can export to Excel with visual timeline using openpyxl.

Deterministic — no LLM required.
"""

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from tools.models.schemas import WeeklyProgram, GanttRow


# Color map by specialty
SPECIALTY_COLORS = {
    "MECHANICAL": "4472C4",
    "ELECTRICAL": "FFC000",
    "INSTRUMENTATION": "70AD47",
    "WELDING": "ED7D31",
    "GENERAL": "A5A5A5",
}

DEFAULT_COLOR = "BDD7EE"


class GanttGenerator:
    """Generates Gantt chart data and Excel exports."""

    @staticmethod
    def generate_gantt_data(program: WeeklyProgram) -> list[GanttRow]:
        """Convert WeeklyProgram work packages into Gantt rows."""
        rows: list[GanttRow] = []

        for pkg in program.work_packages:
            pkg_id = pkg.get("package_id", "")
            name = pkg.get("name", "")
            hours = pkg.get("total_duration_hours", 0.0)
            shift = pkg.get("scheduled_shift", "MORNING")
            team = pkg.get("assigned_team", [])
            items = pkg.get("grouped_items", [])

            # Determine dates
            date_str = pkg.get("scheduled_date", "")
            try:
                start = date.fromisoformat(str(date_str))
            except (ValueError, TypeError):
                start = date.today()

            # Duration in days (min 1 day)
            duration_days = max(1, int(hours / 8.0 + 0.5))
            end = start + timedelta(days=duration_days - 1)

            # Extract area from name or items
            area = ""
            if items and isinstance(items[0], str) and "-" in items[0]:
                area = "-".join(items[0].split("-")[:2])

            specialty = team[0] if team else "GENERAL"

            rows.append(GanttRow(
                package_id=pkg_id,
                name=name,
                start_date=start,
                end_date=end,
                shift=shift,
                area=area,
                specialty=specialty,
                duration_hours=hours,
            ))

        return rows

    @staticmethod
    def export_gantt_excel(gantt_rows: list[GanttRow], filepath: str) -> str:
        """Export Gantt data to Excel with schedule table and visual timeline.

        Args:
            gantt_rows: List of GanttRow objects.
            filepath: Output file path (must end in .xlsx).

        Returns:
            The filepath written.
        """
        wb = Workbook()

        # ── Sheet 1: Schedule Table ──
        ws_schedule = wb.active
        ws_schedule.title = "Schedule"

        headers = ["Package ID", "Name", "Start", "End", "Shift", "Area", "Specialty", "Hours"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws_schedule.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, gr in enumerate(gantt_rows, 2):
            values = [
                gr.package_id, gr.name,
                gr.start_date.isoformat(), gr.end_date.isoformat(),
                gr.shift, gr.area, gr.specialty, gr.duration_hours,
            ]
            for col, val in enumerate(values, 1):
                cell = ws_schedule.cell(row=row_idx, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border

        # Auto-fit column widths
        for col in range(1, len(headers) + 1):
            ws_schedule.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18

        # ── Sheet 2: Gantt Visual ──
        ws_gantt = wb.create_sheet("Gantt")

        if not gantt_rows:
            ws_gantt.cell(row=1, column=1, value="No work packages scheduled")
            wb.save(filepath)
            return filepath

        # Determine date range
        all_dates = []
        for gr in gantt_rows:
            d = gr.start_date
            while d <= gr.end_date:
                all_dates.append(d)
                d += timedelta(days=1)

        if not all_dates:
            wb.save(filepath)
            return filepath

        min_date = min(all_dates)
        max_date = max(all_dates)
        date_range = []
        d = min_date
        while d <= max_date:
            date_range.append(d)
            d += timedelta(days=1)

        # Headers: column 1 = Package Name, columns 2+ = dates
        ws_gantt.cell(row=1, column=1, value="Package").font = Font(bold=True, name="Arial", size=10)
        ws_gantt.column_dimensions["A"].width = 30
        for col_idx, dt in enumerate(date_range, 2):
            cell = ws_gantt.cell(row=1, column=col_idx, value=dt.strftime("%m/%d"))
            cell.font = Font(bold=True, name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center")
            ws_gantt.column_dimensions[cell.column_letter].width = 8

        # Rows: one per package
        for row_idx, gr in enumerate(gantt_rows, 2):
            ws_gantt.cell(row=row_idx, column=1, value=gr.name).font = Font(name="Arial", size=10)
            color = SPECIALTY_COLORS.get(gr.specialty, DEFAULT_COLOR)
            fill = PatternFill("solid", fgColor=color)

            for col_idx, dt in enumerate(date_range, 2):
                if gr.start_date <= dt <= gr.end_date:
                    cell = ws_gantt.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = thin_border

        wb.save(filepath)
        return filepath
