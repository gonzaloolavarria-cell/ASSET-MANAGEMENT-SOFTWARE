# -*- coding: utf-8 -*-
"""
=============================================================================
  DB_AM_OCP_SYNTHETIC_2026 — Gap Analysis Templates (Fase 3)
=============================================================================
  Genera plantillas identificadas en Gap Analysis vs aplicacion AMS:
    23 - Backlog Activo (IW38 ordenes abiertas)
    24 - Notificaciones SAP (IW28/IW29 avisos PM)
    25 - Documentos de Medicion (IK11/IK12 lecturas)

  Cross-references: Templates 01, 04, 06, 07, 09, 11, 15
  Blueprint: AMSA_BBP_PM_04_Rev_0
=============================================================================
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import hashlib
import json
import os
import sys
sys.stdout.reconfigure(encoding="utf-8")
random.seed(2026)

OUT = os.path.dirname(os.path.abspath(__file__))
MOCK_DIR = os.path.join(os.path.dirname(OUT), "sap_mock", "data")

# ============================================================
# STYLING (same as other scripts)
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))


def style_header(ws, headers, row=1, fill=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def save_wb(wb, name):
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"  -> Saved: {name}")
    return path


def synth_id(prefix, seq):
    h = hashlib.md5(f"SYNTH2026-{prefix}-{seq}".encode()).hexdigest()[:4].upper()
    return f"S26-{prefix}-{seq:04d}-{h}"


# ============================================================
# BLUEPRINT CONSTANTS (duplicated from generate_missing_templates.py)
# ============================================================
PLANNING_CENTER = "AN01"
PLANT_ID = "OCP-CON1"
USER_STATUS_SCHEMA = "ZPM00001"
PRIORITY_CLASS = "Z1"

PLANNING_GROUPS = {"P01": "Area seca", "P02": "Area Ripio", "P03": "Area Humeda"}
BUSINESS_AREAS = {"SEC": "Area seca", "RIP": "Area Ripio", "HUM": "Area Humeda"}
PROCESS_TO_BA = {"P01": "SEC", "P02": "RIP", "P03": "HUM"}

AREAS_FLAT = {
    "01": {"name": "Chancado", "pg": "P01"},
    "02": {"name": "Molienda", "pg": "P02"},
    "03": {"name": "Flotacion", "pg": "P03"},
    "04": {"name": "Espesado", "pg": "P03"},
    "05": {"name": "Filtrado", "pg": "P01"},
}

EQUIPMENT = [
    {"tag": "OCP-CON1-CHAN01", "desc": "Chancador Giratorio 60x113", "fl": "02-01-01-CHAN01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CHAN02", "desc": "Chancador Giratorio 60x113 StBy", "fl": "02-01-01-CHAN02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-ALIM01", "desc": "Alimentador Apron 01", "fl": "02-01-01-ALIM01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-ALIM02", "desc": "Alimentador Apron 02", "fl": "02-01-01-ALIM02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-COSE01", "desc": "Chancador Cono Secundario HP800 L1", "fl": "02-01-02-COSE01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-COSE02", "desc": "Chancador Cono Secundario HP800 L2", "fl": "02-01-02-COSE02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-COTE01", "desc": "Chancador Cono Terciario HP500 L1", "fl": "02-01-03-COTE01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-COTE02", "desc": "Chancador Cono Terciario HP500 L2", "fl": "02-01-03-COTE02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-HARV01", "desc": "Harnero Vibratorio 01", "fl": "02-01-04-HARV01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-HARV02", "desc": "Harnero Vibratorio 02", "fl": "02-01-04-HARV02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-HARV03", "desc": "Harnero Vibratorio 03", "fl": "02-01-04-HARV03", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CORR01", "desc": "Correa Transportadora CV-001", "fl": "02-01-04-CORR01", "area": "Chancado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CORR02", "desc": "Correa Transportadora CV-002", "fl": "02-01-04-CORR02", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CORR03", "desc": "Correa Transportadora CV-003", "fl": "02-01-04-CORR03", "area": "Chancado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-MSAG01", "desc": "Molino SAG 40x22 L1", "fl": "02-02-01-MSAG01", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-MSAG02", "desc": "Molino SAG 40x22 L2", "fl": "02-02-01-MSAG02", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-MBOL01", "desc": "Molino Bolas 26x40.5 L1", "fl": "02-02-02-MBOL01", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-MBOL02", "desc": "Molino Bolas 26x40.5 L2", "fl": "02-02-02-MBOL02", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-MBOL03", "desc": "Molino Bolas 26x40.5 L3", "fl": "02-02-02-MBOL03", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BOMB01", "desc": "Bomba Pulpa Descarga SAG 01", "fl": "02-02-03-BOMB01", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BOMB02", "desc": "Bomba Pulpa Descarga SAG 02", "fl": "02-02-03-BOMB02", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BOMB03", "desc": "Bomba Pulpa Descarga SAG 03 StBy", "fl": "02-02-03-BOMB03", "area": "Molienda", "pg": "P02", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-BOMB04", "desc": "Bomba Pulpa Alimentacion Bolas 01", "fl": "02-02-03-BOMB04", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BOMB05", "desc": "Bomba Pulpa Alimentacion Bolas 02", "fl": "02-02-03-BOMB05", "area": "Molienda", "pg": "P02", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-HCIC01", "desc": "Nido Hidrociclones HCI-01", "fl": "02-02-04-HCIC01", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-HCIC02", "desc": "Nido Hidrociclones HCI-02", "fl": "02-02-04-HCIC02", "area": "Molienda", "pg": "P02", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-HCIC03", "desc": "Nido Hidrociclones HCI-03 StBy", "fl": "02-02-04-HCIC03", "area": "Molienda", "pg": "P02", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CFRO01", "desc": "Celda Flotacion Rougher 300m3 C1", "fl": "02-03-01-CFRO01", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CFRO02", "desc": "Celda Flotacion Rougher 300m3 C2", "fl": "02-03-01-CFRO02", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CFRO03", "desc": "Celda Flotacion Rougher 300m3 C3", "fl": "02-03-01-CFRO03", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CFRO04", "desc": "Celda Flotacion Rougher 300m3 C4", "fl": "02-03-01-CFRO04", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CFCL01", "desc": "Celda Flotacion Cleaner 160m3 C1", "fl": "02-03-02-CFCL01", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CFCL02", "desc": "Celda Flotacion Cleaner 160m3 C2", "fl": "02-03-02-CFCL02", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CFSC01", "desc": "Celda Flotacion Scavenger 300m3 C1", "fl": "02-03-02-CFSC01", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CFSC02", "desc": "Celda Flotacion Scavenger 300m3 C2", "fl": "02-03-02-CFSC02", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-AGIT01", "desc": "Acondicionador Reactivos 01", "fl": "02-03-03-AGIT01", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-AGIT02", "desc": "Acondicionador Reactivos 02", "fl": "02-03-03-AGIT02", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "3"},
    {"tag": "OCP-CON1-DREA01", "desc": "Dosificador Reactivos Colector", "fl": "02-03-03-DREA01", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-DREA02", "desc": "Dosificador Reactivos Espumante", "fl": "02-03-03-DREA02", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-DREA03", "desc": "Dosificador Cal/pH", "fl": "02-03-03-DREA03", "area": "Flotacion", "pg": "P03", "type": "M", "abc": "3"},
    {"tag": "OCP-CON1-ESPC01", "desc": "Espesador Concentrado Hi-Rate 25m", "fl": "02-04-01-ESPC01", "area": "Espesado", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-ESPC02", "desc": "Espesador Concentrado Hi-Rate 25m StBy", "fl": "02-04-01-ESPC02", "area": "Espesado", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-ESPR01", "desc": "Espesador Relaves 45m", "fl": "02-04-02-ESPR01", "area": "Espesado", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BREL01", "desc": "Bomba Relaves 01", "fl": "02-04-02-BREL01", "area": "Espesado", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BREL02", "desc": "Bomba Relaves 02 StBy", "fl": "02-04-02-BREL02", "area": "Espesado", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-FILT01", "desc": "Filtro Prensa Concentrado FP-01", "fl": "02-05-01-FILT01", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-FILT02", "desc": "Filtro Prensa Concentrado FP-02", "fl": "02-05-01-FILT02", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-FILT03", "desc": "Filtro Prensa Concentrado FP-03 StBy", "fl": "02-05-01-FILT03", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-SECA01", "desc": "Secador Rotatorio Concentrado", "fl": "02-05-02-SECA01", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-CORR04", "desc": "Correa Transportadora CV-050", "fl": "02-05-02-CORR04", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-CORR05", "desc": "Correa Transportadora CV-051", "fl": "02-05-02-CORR05", "area": "Filtrado", "pg": "P01", "type": "M", "abc": "3"},
    {"tag": "OCP-CON1-COMP01", "desc": "Compresor Aire Planta 01", "fl": "02-05-03-COMP01", "area": "Servicios", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-COMP02", "desc": "Compresor Aire Planta 02 StBy", "fl": "02-05-03-COMP02", "area": "Servicios", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-BAGP01", "desc": "Bomba Agua Proceso 01", "fl": "02-05-03-BAGP01", "area": "Servicios", "pg": "P03", "type": "M", "abc": "1"},
    {"tag": "OCP-CON1-BAGP02", "desc": "Bomba Agua Proceso 02 StBy", "fl": "02-05-03-BAGP02", "area": "Servicios", "pg": "P03", "type": "M", "abc": "2"},
    {"tag": "OCP-CON1-ANPH01", "desc": "Analizador pH Flotacion", "fl": "02-05-03-ANPH01", "area": "Servicios", "pg": "P03", "type": "Q", "abc": "2"},
    {"tag": "OCP-CON1-FLMT01", "desc": "Flujometro Electromagnetico Pulpa", "fl": "02-05-03-FLMT01", "area": "Servicios", "pg": "P03", "type": "Q", "abc": "2"},
    {"tag": "OCP-CON1-DENS01", "desc": "Densimetro Nuclear Molienda", "fl": "02-05-03-DENS01", "area": "Servicios", "pg": "P03", "type": "Q", "abc": "1"},
    {"tag": "OCP-CON1-VIBM01", "desc": "Monitor Vibracion Molinos", "fl": "02-05-03-VIBM01", "area": "Servicios", "pg": "P03", "type": "Q", "abc": "1"},
]

# Work center mapping by planning group
PG_TO_WC = {"P01": "PASMEC01", "P02": "PARMEC01", "P03": "PAHMEC01"}
PG_TO_WC_ELE = {"P01": "PASELE01", "P02": "PARELE01", "P03": "PAHELE01"}
PG_TO_WC_INS = {"P01": "PASINS01", "P02": "PARINS01", "P03": "PAHINS01"}

SUPERVISOR_WCS = {
    "P01": "SPASMEC", "P02": "SPARELE", "P03": "SPAHMEC",
}

# Order types (from Blueprint T26)
ORDER_TYPES = {
    "PM01": {"desc": "Orden Mant. de Averia", "range_start": 1000000, "range_end": 1999999},
    "PM02": {"desc": "Orden Mant. Preventivo", "range_start": 2000000, "range_end": 2999999},
    "PM03": {"desc": "Orden de Solicitud de Mant.", "range_start": 3000000, "range_end": 3999999},
}

# Notification types (from Blueprint T20)
NOTIF_TYPES = {
    "A1": {"desc": "Aviso de mantenimiento",
           "catalogs": {"M001": "Solicitud de mantenimiento", "M002": "Averia", "M003": "Reparacion de componentes"}},
    "A2": {"desc": "Aviso predictivo e ingenieria",
           "catalogs": {"P001": "Predictivo", "P002": "Ingenieria"}},
    "A3": {"desc": "Aviso plan preventivo", "catalogs": {}},
}

PRIORITIES = {"I": "Inmediata", "A": "Alta (2-6 dias)", "M": "Media (7-14 dias)", "B": "Baja (Mayor a 14 dias)"}

# Catalog codes from template 15
DAMAGE_CODES_B = [
    "B-ROD", "B-RET", "B-EJE", "B-ACP", "B-RDC", "B-ENG",
    "B-MOT", "B-CAB", "B-BOB", "B-CON", "B-PRB", "B-SEN",
    "B-TRX", "B-VLV", "B-CIL", "B-MNG", "B-BOM", "B-IMP",
    "B-BAS", "B-SOL", "B-PER", "B-PLN",
]
SYMPTOM_CODES_C = [
    "C-VIB", "C-RUI", "C-CAL", "C-FUG", "C-DES", "C-FIS",
    "C-COR", "C-ERO", "C-BLO", "C-DBA", "C-DAL", "C-CAV",
    "C-CRT", "C-SOB", "C-AIS", "C-ARC", "C-DIS",
    "C-LEC", "C-SFL", "C-DRF", "C-DCA",
    "C-BAR", "C-CON", "C-DER", "C-OBS",
]
CAUSE_CODES_5 = [
    "5-SOC", "5-MAO", "5-ARR", "5-VEL",
    "5-FLB", "5-LBI", "5-CON", "5-MAL", "5-MTJ", "5-MRE", "5-MPM", "5-RPI",
    "5-SDI", "5-MDI", "5-DFB", "5-DEI",
    "5-AMB", "5-POL", "5-HUM", "5-TMP",
]

SPECIALTIES = ["MEC", "ELE", "INS", "LUB", "SOL"]

# Measurement types (duplicated from template 04 for cross-reference)
MEAS_TYPES = [
    ("Temperatura rodamiento LA", "TEMP_BRG_DE", "C", 1, 65, 0, 95, False),
    ("Temperatura rodamiento LOA", "TEMP_BRG_NDE", "C", 1, 65, 0, 95, False),
    ("Temperatura bobinado motor", "TEMP_MOTOR_W", "C", 1, 80, 0, 130, False),
    ("Vibracion radial lado acople", "VIB_RAD_DE", "mm/s", 2, 4.5, 0, 11.2, False),
    ("Vibracion radial lado libre", "VIB_RAD_NDE", "mm/s", 2, 4.5, 0, 11.2, False),
    ("Vibracion axial", "VIB_AXL", "mm/s", 2, 3.5, 0, 8.0, False),
    ("Presion aceite lubricacion", "PRESS_OIL_LUB", "bar", 1, 3.5, 1.5, 6.0, False),
    ("Presion hidraulica", "PRESS_HYD", "bar", 1, 200, 150, 250, False),
    ("Nivel aceite tanque", "LEVEL_OIL", "%", 0, 75, 30, 100, False),
    ("Flujo agua sello", "FLOW_SEAL_W", "L/min", 1, 15, 5, 30, False),
    ("Corriente motor", "CURR_MOTOR", "A", 1, 0, 0, 0, False),
    ("Potencia motor", "POWER_MOTOR", "kW", 1, 0, 0, 0, False),
    ("Horometro operacion", "COUNTER_HRS", "h", 0, 0, 0, 99999, True),
    ("Contador ciclos", "COUNTER_CYC", "ciclos", 0, 0, 0, 999999, True),
    ("Espesor revestimiento", "THICK_LINER", "mm", 1, 0, 0, 0, False),
    ("Temperatura proceso", "TEMP_PROC", "C", 1, 0, 0, 0, False),
    ("pH proceso", "PH_PROC", "pH", 2, 9.5, 7.0, 12.0, False),
    ("Densidad pulpa", "DENS_SLURRY", "kg/m3", 0, 1400, 1200, 1700, False),
    ("Flujo pulpa", "FLOW_SLURRY", "m3/h", 1, 0, 0, 0, False),
    ("Presion aire", "PRESS_AIR", "bar", 1, 7.0, 5.5, 8.5, False),
]

# Worker IDs (simplified from template 09 — will cross-reference by specialty)
WORKERS = {
    "MEC": [f"W-MEC-{i:03d}" for i in range(1, 31)],
    "ELE": [f"W-ELE-{i:03d}" for i in range(1, 16)],
    "INS": [f"W-INS-{i:03d}" for i in range(1, 11)],
    "LUB": [f"W-LUB-{i:03d}" for i in range(1, 6)],
    "SOL": [f"W-SOL-{i:03d}" for i in range(1, 6)],
}

# Backlog description templates
BKL_DESC = {
    "PM01": [
        "Reparar fuga aceite {eq}", "Reparar vibracion excesiva {eq}",
        "Reparar sobrecalentamiento rodamiento {eq}", "Reparar ruido anormal {eq}",
        "Reparar falla motor {eq}", "Reparar desgaste liner {eq}",
        "Reparar fuga sello mecanico {eq}", "Reparar desalineamiento {eq}",
    ],
    "PM02": [
        "Servicio preventivo {freq}W {eq}", "Inspeccion programada {eq}",
        "Lubricacion programada {eq}", "Cambio filtros {eq}",
        "Inspeccion vibraciones {eq}", "Servicio mecanico {freq}W {eq}",
    ],
    "PM03": [
        "Solicitud revision {eq}", "Solicitud calibracion {eq}",
        "Solicitud inspeccion termografica {eq}", "Solicitud alineacion {eq}",
        "Solicitud limpieza industrial {eq}", "Solicitud cambio componente {eq}",
    ],
}

# Notification description templates
NOTIF_DESC = {
    "A1": [
        "Averia: {symptom} en {eq}", "Falla: {symptom} detectada en {eq}",
        "Emergencia: {symptom} critica en {eq}", "Dano: {part} en {eq}",
    ],
    "A2": [
        "Predictivo: {symptom} tendencia en {eq}", "Ingenieria: Analisis {eq}",
        "Monitoreo: {symptom} fuera rango en {eq}", "Alerta condicion: {eq}",
    ],
    "A3": [
        "Preventivo programado: {eq}", "Plan PM: servicio {freq}W {eq}",
        "Inspeccion planificada: {eq}", "Lubricacion planificada: {eq}",
    ],
}

# Symptom descriptions for notifications
SYMPTOM_DESCS = [
    "vibracion excesiva", "ruido anormal", "sobrecalentamiento", "fuga aceite",
    "desgaste", "fisura", "corrosion", "cavitacion", "desalineamiento",
    "lectura erratica", "baja presion", "alto consumo corriente",
]
PART_DESCS = [
    "rodamiento", "sello mecanico", "impulsor", "motor electrico", "acoplamiento",
    "engranaje", "revestimiento", "valvula", "sensor", "bomba",
]


print("=" * 70)
print("  DB_AM_OCP_SYNTHETIC_2026 — Gap Templates (Fase 3)")
print("=" * 70)


# ============================================================
# 23 — BACKLOG ACTIVO (IW38 ordenes abiertas)
# 80 registros, snapshot al 2026-03-30
# ============================================================

print("\n[23] Generating 23_active_backlog.xlsx ...")
wb23 = openpyxl.Workbook()
ws23 = wb23.active
ws23.title = "Active Backlog"
h23 = [
    "backlog_id", "work_request_id", "aufnr", "auart", "auart_desc",
    "equipment_tag", "equnr", "sap_func_loc",
    "area", "planning_group", "business_area", "planning_center",
    "priority", "priority_desc", "priority_class",
    "status", "blocking_reason",
    "created_date", "age_days", "estimated_duration_hours",
    "required_specialties", "materials_ready", "shutdown_required",
    "groupable", "group_id", "work_center", "supervisor_wc",
    "description",
]
style_header(ws23, h23)

# Status distribution (total 80)
STATUS_DIST = [
    ("AWAITING_MATERIALS", 15),
    ("AWAITING_SHUTDOWN", 12),
    ("AWAITING_RESOURCES", 10),
    ("AWAITING_APPROVAL", 18),
    ("SCHEDULED", 15),
    ("IN_PROGRESS", 10),
]

# Priority distribution (total 80)
PRIORITY_DIST = [("I", 5), ("A", 20), ("M", 35), ("B", 20)]

# Order type distribution (total 80)
OT_DIST_BKL = [("PM01", 30), ("PM02", 25), ("PM03", 25)]

BLOCKING_REASONS = {
    "AWAITING_MATERIALS": [
        "Repuesto en transito ETA 2026-04-15", "Material importado en aduana",
        "Sin stock disponible, PO emitida", "Repuesto critico sin alternativa local",
    ],
    "AWAITING_SHUTDOWN": [
        "Requiere parada programada Q2-2026", "Equipo en operacion continua",
        "Parada menor proxima: 2026-04-05", "Coordinacion con produccion pendiente",
    ],
    "AWAITING_RESOURCES": [
        "Sin disponibilidad mecanicos turno A", "Grua ocupada en otra area",
        "Andamio en uso sector molienda", "Equipo intervencion disponible prox semana",
    ],
    "AWAITING_APPROVAL": [
        "Pendiente aprobacion supervisor", "Pendiente validacion planificador",
        "Requiere autorizacion gerencia", "En revision ingenieria",
    ],
    "SCHEDULED": None,
    "IN_PROGRESS": None,
}

SNAPSHOT_DATE = datetime(2026, 3, 30)

# Build flat lists for distribution
status_pool = []
for s, cnt in STATUS_DIST:
    status_pool.extend([s] * cnt)
random.shuffle(status_pool)

priority_pool = []
for p, cnt in PRIORITY_DIST:
    priority_pool.extend([p] * cnt)
random.shuffle(priority_pool)

ot_pool = []
for ot, cnt in OT_DIST_BKL:
    ot_pool.extend([ot] * cnt)
random.shuffle(ot_pool)

# Favor ABC=1 equipment (~60%)
eq_abc1 = [e for e in EQUIPMENT if e["abc"] == "1"]
eq_other = [e for e in EQUIPMENT if e["abc"] != "1"]

bkl_counter = {k: ORDER_TYPES[k]["range_start"] + 900000 for k in ORDER_TYPES}
bkl_rows = []
bkl_json = []

for i in range(80):
    eq = random.choice(eq_abc1) if random.random() < 0.60 else random.choice(eq_other)
    ot_type = ot_pool[i]
    status = status_pool[i]
    priority = priority_pool[i]

    aufnr = bkl_counter[ot_type]
    bkl_counter[ot_type] += 1

    days_ago = random.randint(5, 75)
    created = SNAPSHOT_DATE - timedelta(days=days_ago)
    age_days = days_ago

    duration = round(random.uniform(2.0, 120.0), 1)
    if ot_type == "PM02":
        duration = round(random.uniform(2.0, 24.0), 1)
    elif ot_type == "PM03":
        duration = round(random.uniform(4.0, 40.0), 1)

    n_specs = random.randint(1, 3)
    specs = ",".join(random.sample(SPECIALTIES[:3], k=n_specs))

    materials_ready = "TRUE" if status not in ("AWAITING_MATERIALS",) else "FALSE"
    shutdown_req = "TRUE" if status == "AWAITING_SHUTDOWN" or (random.random() < 0.3 and priority in ("I", "A")) else "FALSE"
    groupable = "TRUE" if random.random() < 0.4 else "FALSE"
    group_id = f"WP-GRP-{random.randint(1, 20):04d}" if groupable == "TRUE" else ""

    blocking = ""
    reasons = BLOCKING_REASONS.get(status)
    if reasons:
        blocking = random.choice(reasons)

    wc = PG_TO_WC.get(eq["pg"], "PASMEC01")
    sup_wc = SUPERVISOR_WCS.get(eq["pg"], "SPASMEC")

    desc_tpl = random.choice(BKL_DESC[ot_type])
    desc = desc_tpl.format(eq=eq["desc"][:35], freq=random.choice([4, 8, 13, 26]))[:72]

    bid = synth_id("BKL", i + 1)
    wrid = synth_id("WR", i + 1)
    equnr = 100000 + i

    row = [
        bid, wrid, aufnr, ot_type, ORDER_TYPES[ot_type]["desc"],
        eq["tag"], equnr, eq["fl"],
        eq["area"], eq["pg"], PROCESS_TO_BA.get(eq["pg"], "SEC"), PLANNING_CENTER,
        priority, PRIORITIES[priority], PRIORITY_CLASS,
        status, blocking,
        created.strftime("%Y-%m-%d"), age_days, duration,
        specs, materials_ready, shutdown_req,
        groupable, group_id, wc, sup_wc,
        desc,
    ]
    ws23.append(row)
    bkl_rows.append(row)

    bkl_json.append({
        "backlog_id": bid, "work_request_id": wrid,
        "aufnr": aufnr, "auart": ot_type,
        "equipment_tag": eq["tag"], "equnr": equnr,
        "sap_func_loc": eq["fl"], "area": eq["area"],
        "planning_group": eq["pg"], "priority": priority,
        "status": status, "blocking_reason": blocking,
        "created_date": created.strftime("%Y-%m-%d"),
        "age_days": age_days, "estimated_duration_hours": duration,
        "required_specialties": specs.split(","),
        "materials_ready": materials_ready == "TRUE",
        "shutdown_required": shutdown_req == "TRUE",
        "groupable": groupable == "TRUE",
        "work_center": wc, "description": desc,
    })

auto_width(ws23)
save_wb(wb23, "23_active_backlog.xlsx")
count_23 = 80


# ============================================================
# 24 — NOTIFICACIONES SAP (IW28/IW29)
# 220 registros: 80 abiertas + 140 cerradas
# ============================================================

print("[24] Generating 24_notifications.xlsx ...")
wb24 = openpyxl.Workbook()

# Sheet 1: Open Notifications
ws24a = wb24.active
ws24a.title = "Open Notifications"
h24 = [
    "qmnum", "qmart", "qmart_desc",
    "equipment_tag", "equnr", "sap_func_loc",
    "area", "planning_group", "business_area", "planning_center",
    "notification_catalog", "catalog_desc",
    "user_status_schema", "user_status",
    "priority", "priority_desc", "priority_class",
    "reported_by", "reported_date",
    "description", "long_text",
    "damage_code", "cause_code", "object_part_code",
    "linked_order", "system_status", "work_center",
]
style_header(ws24a, h24)

# Sheet 2: Closed Notifications
ws24b = wb24.create_sheet("Closed Notifications")
h24b = h24 + ["completed_date"]
style_header(ws24b, h24b)

# Type distribution: A1: 50%, A2: 20%, A3: 30%
NOTIF_TYPE_DIST_OPEN = [("A1", 40), ("A2", 16), ("A3", 24)]
NOTIF_TYPE_DIST_CLOSED = [("A1", 70), ("A2", 28), ("A3", 42)]

notif_serial = 5000001  # Start after Blueprint range for work orders
notif_json = {"open": [], "closed": []}


def generate_notification(idx, notif_type, is_open, linked_order_pool=None):
    global notif_serial
    eq = random.choice(eq_abc1) if random.random() < 0.55 else random.choice(eq_other)

    qmnum = notif_serial
    notif_serial += 1

    notif_cats = NOTIF_TYPES[notif_type]["catalogs"]
    notif_cat = random.choice(list(notif_cats.keys())) if notif_cats else ""
    cat_desc = notif_cats.get(notif_cat, "") if notif_cats else ""

    user_schema = USER_STATUS_SCHEMA if notif_type in ("A1", "A2") else ""
    user_status = random.choice(["APRO", "APRO", "APRO", "RECH"]) if notif_type in ("A1", "A2") else ""

    priority = random.choices(["I", "A", "M", "B"], weights=[5, 25, 45, 25])[0]

    if is_open:
        days_ago = random.randint(1, 89)
        reported = SNAPSHOT_DATE - timedelta(days=days_ago)
        sys_status = random.choice(["OSNO", "NOPR"])
    else:
        days_ago = random.randint(30, 270)
        reported = SNAPSHOT_DATE - timedelta(days=days_ago)
        sys_status = "NOCO"

    all_workers = []
    for wlist in WORKERS.values():
        all_workers.extend(wlist)
    reported_by = random.choice(all_workers)

    symptom = random.choice(SYMPTOM_DESCS)
    part = random.choice(PART_DESCS)
    desc_tpl = random.choice(NOTIF_DESC[notif_type])
    desc = desc_tpl.format(eq=eq["desc"][:30], symptom=symptom, part=part, freq=random.choice([4, 8, 13]))[:72]
    long_text = f"Detalle: {symptom} detectado en {eq['desc']}. Area: {eq['area']}. Requiere intervencion {random.choice(SPECIALTIES[:3])}."

    damage = random.choice(DAMAGE_CODES_B) if notif_type in ("A1", "A2") else ""
    cause = random.choice(CAUSE_CODES_5) if notif_type in ("A1", "A2") else ""
    obj_part = random.choice(DAMAGE_CODES_B[:10]) if notif_type == "A1" else ""

    linked = ""
    if linked_order_pool and random.random() < 0.6:
        linked = random.choice(linked_order_pool)

    wc = PG_TO_WC.get(eq["pg"], "PASMEC01")
    equnr = 200000 + idx

    row = [
        qmnum, notif_type, NOTIF_TYPES[notif_type]["desc"],
        eq["tag"], equnr, eq["fl"],
        eq["area"], eq["pg"], PROCESS_TO_BA.get(eq["pg"], "SEC"), PLANNING_CENTER,
        notif_cat, cat_desc,
        user_schema, user_status,
        priority, PRIORITIES[priority], PRIORITY_CLASS,
        reported_by, reported.strftime("%Y-%m-%d"),
        desc, long_text,
        damage, cause, obj_part,
        linked, sys_status, wc,
    ]

    json_record = {
        "qmnum": qmnum, "qmart": notif_type, "qmart_desc": NOTIF_TYPES[notif_type]["desc"],
        "equipment_tag": eq["tag"], "equnr": equnr, "sap_func_loc": eq["fl"],
        "area": eq["area"], "planning_group": eq["pg"],
        "notification_catalog": notif_cat,
        "user_status": user_status, "priority": priority,
        "reported_by": reported_by, "reported_date": reported.strftime("%Y-%m-%d"),
        "description": desc,
        "damage_code": damage, "cause_code": cause,
        "linked_order": linked, "system_status": sys_status,
        "work_center": wc,
    }

    return row, json_record, reported


# Open notifications — linked to backlog orders
open_order_pool = [r[2] for r in bkl_rows]  # aufnr from template 23

idx_global = 0
for notif_type, count in NOTIF_TYPE_DIST_OPEN:
    for _ in range(count):
        row, jrec, _ = generate_notification(idx_global, notif_type, is_open=True, linked_order_pool=open_order_pool)
        ws24a.append(row)
        notif_json["open"].append(jrec)
        idx_global += 1

# Closed notifications — linked to historical orders (template 06 range)
hist_order_pool = list(range(1000000, 1000050)) + list(range(2000000, 2000070)) + list(range(3000000, 3000025))

for notif_type, count in NOTIF_TYPE_DIST_CLOSED:
    for _ in range(count):
        row, jrec, reported = generate_notification(idx_global, notif_type, is_open=False, linked_order_pool=hist_order_pool)
        completed = reported + timedelta(days=random.randint(1, 30))
        row_closed = row + [completed.strftime("%Y-%m-%d")]
        ws24b.append(row_closed)
        jrec["completed_date"] = completed.strftime("%Y-%m-%d")
        notif_json["closed"].append(jrec)
        idx_global += 1

for ws in [ws24a, ws24b]:
    auto_width(ws)
save_wb(wb24, "24_notifications.xlsx")
count_24 = 80 + 140


# ============================================================
# 25 — DOCUMENTOS DE MEDICION (IK11/IK12)
# 250+ registros, 6 meses de lecturas
# ============================================================

print("[25] Generating 25_measurement_documents.xlsx ...")
wb25 = openpyxl.Workbook()
ws25 = wb25.active
ws25.title = "Measurement Documents"
h25 = [
    "measurement_doc_id", "measurement_point_id",
    "equipment_tag", "sap_func_loc",
    "characteristic", "characteristic_desc",
    "reading_date", "reading_time",
    "measured_value", "unit_of_measure",
    "counter_reading",
    "valuation_code", "valuation_desc",
    "lower_limit", "upper_limit", "target_value",
    "is_counter", "recorded_by", "work_center",
    "planning_group", "area",
]
style_header(ws25, h25)

# Rebuild measurement point mapping (same logic as template 04, same seed offset)
# We need to replicate the point IDs generated by generate_missing_templates.py
# Since that script uses random.seed(2026) and processes EQUIPMENT in order,
# we replicate the assignment logic to get matching point IDs

# Use a separate RNG to not disturb the main seed
mp_rng = random.Random(2026)
measurement_points = []  # list of (mp_id, eq, meas_type)

mp_serial_rebuild = 1
for eq in EQUIPMENT:
    n_points = mp_rng.randint(3, 6)
    selected = mp_rng.sample(MEAS_TYPES, k=min(n_points, len(MEAS_TYPES)))
    for mtype in selected:
        desc, char, uom, dec, target, lo, hi, is_ctr = mtype
        mp_id = f"{mp_serial_rebuild:012d}"
        mp_serial_rebuild += 1
        freq = mp_rng.choice([1, 7, 14, 30]) if not is_ctr else mp_rng.choice([1, 7])
        measurement_points.append({
            "mp_id": mp_id,
            "eq": eq,
            "char": char,
            "desc": desc,
            "uom": uom,
            "dec": dec,
            "target": target,
            "lo": lo,
            "hi": hi,
            "is_ctr": is_ctr,
            "freq_days": freq,
        })

# Generate readings over 6 months (2025-10-01 to 2026-03-30)
START_DATE = datetime(2025, 10, 1)
END_DATE = datetime(2026, 3, 30)
TOTAL_DAYS = (END_DATE - START_DATE).days

mdoc_serial = 1
mdoc_count = 0
mdoc_json = []

# Select a subset of measurement points to generate readings for
# Prioritize ABC=1 equipment (weekly) vs others (biweekly)
for mp in measurement_points:
    eq = mp["eq"]
    freq = 7 if eq["abc"] == "1" else 14

    # Skip counters for some equipment to keep count reasonable
    if mp["is_ctr"] and random.random() < 0.5:
        continue

    # Generate readings at the frequency
    current_date = START_DATE + timedelta(days=random.randint(0, freq - 1))
    counter_base = random.randint(5000, 50000) if mp["is_ctr"] else 0

    # Degradation simulation: some equipment drifts toward limits
    is_degrading = random.random() < 0.08  # 8% of points show degradation
    readings_for_point = []

    while current_date <= END_DATE:
        progress = (current_date - START_DATE).days / TOTAL_DAYS

        if mp["is_ctr"]:
            # Counter: monotonically increasing
            hours_per_day = random.uniform(18, 23)
            counter_base += int(hours_per_day * freq)
            measured_value = counter_base
        elif mp["target"] > 0 and mp["hi"] > 0:
            # Value with defined range
            if is_degrading:
                # Drift toward upper limit
                base = mp["target"] + (mp["hi"] - mp["target"]) * progress * 0.9
                noise = random.gauss(0, (mp["hi"] - mp["target"]) * 0.05)
                measured_value = round(base + noise, mp["dec"])
            else:
                # Normal operation near target
                spread = (mp["hi"] - mp["lo"]) * 0.15
                measured_value = round(random.gauss(mp["target"], spread), mp["dec"])
            measured_value = max(mp["lo"], measured_value)
        else:
            # No defined range, generate reasonable values
            if mp["char"] == "CURR_MOTOR":
                measured_value = round(random.uniform(50, 350), 1)
            elif mp["char"] == "POWER_MOTOR":
                measured_value = round(random.uniform(100, 2000), 1)
            elif mp["char"] == "THICK_LINER":
                measured_value = round(random.uniform(15, 80) - progress * 20, 1)
                measured_value = max(5, measured_value)
            elif mp["char"] == "TEMP_PROC":
                measured_value = round(random.uniform(25, 65), 1)
            elif mp["char"] == "FLOW_SLURRY":
                measured_value = round(random.uniform(200, 800), 1)
            else:
                measured_value = round(random.uniform(10, 100), mp["dec"])

        # Determine valuation
        valuation = 1  # OK
        if mp["hi"] > 0 and not mp["is_ctr"]:
            if measured_value > mp["hi"]:
                valuation = 3  # Alarm
            elif measured_value > mp["hi"] * 0.85:
                valuation = 2  # Warning
            if mp["lo"] > 0 and measured_value < mp["lo"]:
                valuation = 3  # Alarm
            elif mp["lo"] > 0 and measured_value < mp["lo"] * 1.15:
                valuation = 2  # Warning

        valuation_desc = {1: "OK", 2: "Warning", 3: "Alarm"}[valuation]

        # Reading time (aligned to 7x7 shift: 08:00-20:00)
        hour = random.randint(8, 19)
        minute = random.choice([0, 15, 30, 45])
        reading_time = f"{hour:02d}:{minute:02d}"

        # Worker
        spec_workers = WORKERS.get("INS", WORKERS["MEC"])
        if mp["char"].startswith("VIB") or mp["char"].startswith("TEMP"):
            spec_workers = WORKERS.get("INS", WORKERS["MEC"])
        elif mp["char"].startswith("PRESS") or mp["char"].startswith("FLOW"):
            spec_workers = WORKERS.get("MEC", WORKERS["MEC"])
        recorded_by = random.choice(spec_workers)

        wc = PG_TO_WC.get(eq["pg"], "PASMEC01")

        doc_id = synth_id("MDOC", mdoc_serial)
        mdoc_serial += 1

        row = [
            doc_id, mp["mp_id"],
            eq["tag"], eq["fl"],
            mp["char"], mp["desc"],
            current_date.strftime("%Y-%m-%d"), reading_time,
            measured_value, mp["uom"],
            counter_base if mp["is_ctr"] else "",
            valuation, valuation_desc,
            mp["lo"] if mp["lo"] > 0 else "", mp["hi"] if mp["hi"] > 0 else "",
            mp["target"] if mp["target"] > 0 else "",
            "TRUE" if mp["is_ctr"] else "FALSE",
            recorded_by, wc,
            eq["pg"], eq["area"],
        ]
        ws25.append(row)
        mdoc_count += 1

        mdoc_json.append({
            "measurement_doc_id": doc_id,
            "measurement_point_id": mp["mp_id"],
            "equipment_tag": eq["tag"],
            "sap_func_loc": eq["fl"],
            "characteristic": mp["char"],
            "reading_date": current_date.strftime("%Y-%m-%d"),
            "reading_time": reading_time,
            "measured_value": measured_value,
            "unit_of_measure": mp["uom"],
            "counter_reading": counter_base if mp["is_ctr"] else None,
            "valuation_code": valuation,
            "valuation_desc": valuation_desc,
            "is_counter": mp["is_ctr"],
            "recorded_by": recorded_by,
            "work_center": wc,
            "planning_group": eq["pg"],
        })

        current_date += timedelta(days=freq + random.randint(-1, 2))

    # Stop after enough readings to keep file manageable
    if mdoc_count >= 300:
        break

auto_width(ws25)
save_wb(wb25, "25_measurement_documents.xlsx")
count_25 = mdoc_count


# ============================================================
# FASE 4 CONSTANTS
# ============================================================

# Full order types including PM06/PM07 (for cost history)
ORDER_TYPES_FULL = {
    "PM01": {"desc": "Orden Mant. de Averia", "range_start": 1000000, "count": 50},
    "PM02": {"desc": "Orden Mant. Preventivo", "range_start": 2000000, "count": 70},
    "PM03": {"desc": "Orden de Solicitud de Mant.", "range_start": 3000000, "count": 25},
    "PM06": {"desc": "Orden de Inversion", "range_start": 6000000, "count": 25},
    "PM07": {"desc": "Orden de Reparacion de Componentes", "range_start": 7000000, "count": 30},
}

# Build closed order pool (cross-ref template 06 — ~65% of orders are closed/CTEC)
CLOSED_ORDERS = []
for ot, info in ORDER_TYPES_FULL.items():
    n_closed = int(info["count"] * 0.65)
    for i in range(n_closed):
        aufnr = info["range_start"] + i
        eq = EQUIPMENT[i % len(EQUIPMENT)]
        duration = round(random.uniform(2, 48) if ot != "PM07" else random.uniform(8, 120), 1)
        crew = random.randint(1, 4)
        created = SNAPSHOT_DATE - timedelta(days=random.randint(30, 365))
        actual_end = created + timedelta(days=random.randint(1, 30))
        CLOSED_ORDERS.append({
            "aufnr": aufnr, "auart": ot, "equipment_tag": eq["tag"],
            "equnr": 100000 + (i % len(EQUIPMENT)), "sap_func_loc": eq["fl"],
            "area": eq["area"], "pg": eq["pg"], "duration_hours": duration,
            "crew": crew, "created_date": created, "actual_end": actual_end,
        })

# Labor rates (from generate_synthetic_dataset.py)
LABOR_RATES = {
    "MEC": 45.0, "ELE": 48.0, "INS": 52.0, "SOL": 50.0, "LUB": 32.0,
}

# Material codes pool (cross-ref template 07)
MATERIAL_POOL = [
    {"code": f"S26-MAT-{i:04d}", "sap": f"{10000+i:012d}",
     "desc": desc, "uom": uom, "cost": cost}
    for i, (desc, uom, cost) in enumerate([
        ("Rodamiento SKF 22328 CC/W33", "EA", 1250.0),
        ("Sello mecanico John Crane 5610", "EA", 3400.0),
        ("Impulsor Warman 10/8", "EA", 8500.0),
        ("Correa Gates PowerBand 5VX1500", "EA", 185.0),
        ("Aceite Shell Omala S4 GX 320 (200L)", "DR", 1800.0),
        ("Filtro hidraulico Parker 925835", "EA", 245.0),
        ("Acoplamiento Falk 1110T", "EA", 2200.0),
        ("Liner Mn Steel Metso C160", "EA", 12500.0),
        ("Grasa Mobil SHC 220 (18kg)", "PL", 320.0),
        ("Sensor vibr Bently Nevada 330180", "EA", 1650.0),
        ("Motor WEG W22 250kW", "EA", 18500.0),
        ("Valvula Weir 8/6 Isogate", "EA", 4200.0),
        ("Malla harnero poliuretano 50x50", "EA", 680.0),
        ("Perno hex M24x100 Gr 8.8", "BX", 45.0),
        ("Empaquetadura grafito 1/2in", "MT", 28.0),
        ("Junta tórica Viton DN150", "EA", 35.0),
        ("Reductor Flender B3SH 14", "EA", 22000.0),
        ("Encoder Siemens 6FX2001", "EA", 890.0),
        ("Contactor Siemens 3RT1065", "EA", 520.0),
        ("Cable XLPE 3x95mm2", "MT", 42.0),
    ], start=1)
]

# Failure modes data (cross-ref template 03)
MECHANISMS = ["WEARS", "CORRODES", "FRACTURES", "OVERHEATS", "VIBRATES",
              "LEAKS", "SHORT_CIRCUITS", "BLOCKS", "ERODES", "FATIGUES"]
CAUSES = ["ABRASION", "CAVITATION", "CORROSION", "FATIGUE", "OVERLOAD",
           "CONTAMINATION", "MISALIGNMENT", "THERMAL_STRESS", "ELECTRICAL_SURGE",
           "LACK_LUBRICATION"]
PATTERNS = ["A_BATHTUB", "B_AGE", "C_GRADUAL", "D_INITIAL_SURGE", "E_RANDOM", "F_INFANT"]
MAINT_ITEMS = ["Rodamiento", "Impulsor", "Sello mecanico", "Eje", "Revestimiento",
               "Motor electrico", "Reductor", "Acoplamiento", "Valvula", "Correa",
               "Filtro", "Rotor", "Estator", "Malla"]

# Cost centers (cross-ref template 20)
COST_CENTERS = {
    "P01": {"MEC": "CC-PL-MEC-SEC", "ELE": "CC-PL-ELE-SEC", "INS": "CC-PL-INS-SEC"},
    "P02": {"MEC": "CC-PL-MEC-RIP", "ELE": "CC-PL-ELE-RIP", "INS": "CC-PL-INS-RIP"},
    "P03": {"MEC": "CC-PL-MEC-HUM", "ELE": "CC-PL-ELE-HUM", "INS": "CC-PL-INS-HUM"},
}

# WBS elements (cross-ref template 20)
WBS_ELEMENTS = {
    "P01": "WBS-PL-SEC-MNT-2026", "P02": "WBS-PL-RIP-MNT-2026", "P03": "WBS-PL-HUM-MNT-2026",
}

# Storage locations
STORAGE_LOCS = ["ALM-SEC-01", "ALM-RIP-01", "ALM-HUM-01", "ALM-CENTRAL"]


# ============================================================
# 26 — CONFIRMACIONES DE TIEMPO (IW41/IW42)
# 200 registros
# ============================================================

print("\n[26] Generating 26_time_confirmations.xlsx ...")
wb26 = openpyxl.Workbook()
ws26 = wb26.active
ws26.title = "Time Confirmations"
h26 = [
    "confirmation_id", "aufnr", "auart", "auart_desc",
    "equipment_tag", "sap_func_loc", "area", "planning_group",
    "operation_number", "sub_operation",
    "confirmation_type", "work_center",
    "worker_id", "worker_name", "specialty",
    "start_date", "start_time", "end_date", "end_time",
    "actual_work_hours", "break_hours", "travel_time_hours", "setup_time_hours",
    "total_duration_hours",
    "final_confirmation", "system_status_after",
    "text",
]
style_header(ws26, h26)

conf_serial = 1
conf_count = 0
conf_json = []

# Pick closed orders and generate 1-4 confirmations per order
selected_orders = random.sample(CLOSED_ORDERS, k=min(80, len(CLOSED_ORDERS)))

for order in selected_orders:
    n_ops = random.randint(1, 4)
    remaining_hours = order["duration_hours"]

    for op_idx in range(n_ops):
        op_num = (op_idx + 1) * 10  # 10, 20, 30, 40
        is_final = (op_idx == n_ops - 1)

        # Split duration across operations
        if is_final:
            work_hours = round(remaining_hours, 1)
        else:
            work_hours = round(remaining_hours * random.uniform(0.2, 0.5), 1)
            remaining_hours -= work_hours

        work_hours = max(0.5, work_hours)
        break_hrs = 1.0
        travel_hrs = round(random.uniform(0.5, 2.0), 1)
        setup_hrs = round(random.uniform(0.5, 1.5), 1)
        total = round(work_hours + break_hrs + travel_hrs + setup_hrs, 1)

        spec = random.choice(SPECIALTIES[:3])
        worker_id = random.choice(WORKERS.get(spec, WORKERS["MEC"]))
        worker_name = f"Tecnico {spec} {worker_id[-3:]}"

        wc = PG_TO_WC.get(order["pg"], "PASMEC01")
        if spec == "ELE":
            wc = PG_TO_WC_ELE.get(order["pg"], "PASELE01")
        elif spec == "INS":
            wc = PG_TO_WC_INS.get(order["pg"], "PASINS01")

        start_dt = order["actual_end"] - timedelta(days=random.randint(0, 3))
        start_h = random.randint(8, 14)
        end_h = min(start_h + int(total) + 1, 20)

        conf_id = synth_id("CONF", conf_serial)
        conf_serial += 1

        status_after = "CNF" if is_final else "PCNF"

        op_desc = random.choice([
            f"Trabajo {spec} en {order['equipment_tag'][-6:]}",
            f"Intervencion {spec} operacion {op_num:04d}",
            f"Servicio {spec} {order['auart']}",
        ])[:72]

        row = [
            conf_id, order["aufnr"], order["auart"], ORDER_TYPES_FULL[order["auart"]]["desc"],
            order["equipment_tag"], order["sap_func_loc"], order["area"], order["pg"],
            f"{op_num:04d}", "0000",
            "INDIVIDUAL", wc,
            worker_id, worker_name, spec,
            start_dt.strftime("%Y-%m-%d"), f"{start_h:02d}:00",
            start_dt.strftime("%Y-%m-%d"), f"{end_h:02d}:00",
            work_hours, break_hrs, travel_hrs, setup_hrs,
            total,
            "TRUE" if is_final else "FALSE", status_after,
            op_desc,
        ]
        ws26.append(row)
        conf_count += 1

        conf_json.append({
            "confirmation_id": conf_id, "aufnr": order["aufnr"],
            "auart": order["auart"], "equipment_tag": order["equipment_tag"],
            "operation_number": f"{op_num:04d}", "work_center": wc,
            "worker_id": worker_id, "specialty": spec,
            "start_date": start_dt.strftime("%Y-%m-%d"),
            "actual_work_hours": work_hours, "total_duration_hours": total,
            "final_confirmation": is_final, "system_status_after": status_after,
        })

auto_width(ws26)
save_wb(wb26, "26_time_confirmations.xlsx")
count_26 = conf_count


# ============================================================
# 27 — MOVIMIENTOS DE MATERIAL (MB21/MIGO/MB51)
# 200 registros: 80 reservas + 120 movimientos
# ============================================================

print("[27] Generating 27_material_movements.xlsx ...")
wb27 = openpyxl.Workbook()

# Sheet 1: Reservations
ws27a = wb27.active
ws27a.title = "Reservations"
h27a = [
    "reservation_id", "aufnr", "auart",
    "equipment_tag", "sap_func_loc", "area", "planning_group",
    "material_code", "sap_material_number", "description",
    "quantity_reserved", "unit_of_measure",
    "reservation_date", "requirement_date",
    "movement_type", "plant", "storage_location",
    "status", "cost_center",
]
style_header(ws27a, h27a)

# Sheet 2: Goods Movements
ws27b = wb27.create_sheet("Goods Movements")
h27b = [
    "movement_doc_id", "material_code", "sap_material_number", "description",
    "movement_type", "movement_type_desc", "movement_date",
    "quantity", "unit_of_measure",
    "aufnr", "cost_center", "storage_location",
    "unit_cost_usd", "total_cost_usd", "posting_date",
]
style_header(ws27b, h27b)

MOV_TYPE_DESCS = {
    "261": "Salida mercancia a orden",
    "262": "Anulacion salida mercancia",
    "101": "Entrada mercancia pedido",
    "311": "Traspaso entre almacenes",
}

res_json = []
mov_json = []

# Reservations (80)
for i in range(80):
    order = random.choice(CLOSED_ORDERS)
    mat = random.choice(MATERIAL_POOL)
    qty = random.randint(1, 10)
    res_date = order["created_date"] + timedelta(days=random.randint(0, 5))
    req_date = order["actual_end"] - timedelta(days=random.randint(0, 3))
    status = random.choices(["OPEN", "PARTIALLY_DELIVERED", "FULLY_DELIVERED"], weights=[20, 30, 50])[0]
    pg = order["pg"]
    cc = COST_CENTERS.get(pg, COST_CENTERS["P01"]).get("MEC", "CC-PL-MEC-SEC")
    sloc = random.choice(STORAGE_LOCS)

    res_id = synth_id("RES", i + 1)
    row = [
        res_id, order["aufnr"], order["auart"],
        order["equipment_tag"], order["sap_func_loc"], order["area"], pg,
        mat["code"], mat["sap"], mat["desc"],
        qty, mat["uom"],
        res_date.strftime("%Y-%m-%d"), req_date.strftime("%Y-%m-%d"),
        "261", PLANT_ID, sloc,
        status, cc,
    ]
    ws27a.append(row)
    res_json.append({
        "reservation_id": res_id, "aufnr": order["aufnr"],
        "material_code": mat["code"], "quantity_reserved": qty,
        "status": status, "storage_location": sloc,
    })

# Goods Movements (120)
# Distribution: 261=60%, 101=25%, 311=10%, 262=5%
MOV_DIST = [("261", 72), ("101", 30), ("311", 12), ("262", 6)]

mov_idx = 0
for mov_type, count in MOV_DIST:
    for _ in range(count):
        mat = random.choice(MATERIAL_POOL)
        qty = random.randint(1, 20)
        mov_date = SNAPSHOT_DATE - timedelta(days=random.randint(1, 270))

        aufnr = ""
        cc = random.choice(list(COST_CENTERS["P01"].values()))
        if mov_type in ("261", "262"):
            order = random.choice(CLOSED_ORDERS)
            aufnr = order["aufnr"]
            cc = COST_CENTERS.get(order["pg"], COST_CENTERS["P01"]).get("MEC", "CC-PL-MEC-SEC")

        sloc = random.choice(STORAGE_LOCS)
        total_cost = round(mat["cost"] * qty, 2)

        mov_id = synth_id("MOV", mov_idx + 1)
        mov_idx += 1

        row = [
            mov_id, mat["code"], mat["sap"], mat["desc"],
            mov_type, MOV_TYPE_DESCS[mov_type], mov_date.strftime("%Y-%m-%d"),
            qty, mat["uom"],
            aufnr, cc, sloc,
            mat["cost"], total_cost, mov_date.strftime("%Y-%m-%d"),
        ]
        ws27b.append(row)
        mov_json.append({
            "movement_doc_id": mov_id, "material_code": mat["code"],
            "movement_type": mov_type, "quantity": qty,
            "movement_date": mov_date.strftime("%Y-%m-%d"),
            "aufnr": aufnr, "total_cost_usd": total_cost,
        })

for ws in [ws27a, ws27b]:
    auto_width(ws)
save_wb(wb27, "27_material_movements.xlsx")
count_27 = 80 + 120


# ============================================================
# 28 — BOM DE EQUIPOS (IB01/CS01)
# 200 registros
# ============================================================

print("[28] Generating 28_equipment_bom.xlsx ...")
wb28 = openpyxl.Workbook()
ws28 = wb28.active
ws28.title = "Equipment BOM"
h28 = [
    "bom_id", "equipment_tag", "equnr", "sap_func_loc",
    "equipment_desc", "area", "planning_group",
    "bom_level", "item_number",
    "component_category", "component_category_desc",
    "material_code", "sap_material_number", "component_desc",
    "quantity", "unit_of_measure",
    "item_category", "change_frequency",
    "applicable_system", "lead_time_days",
    "unit_cost_usd", "critical_spare",
]
style_header(ws28, h28)

ITEM_CATEGORIES = ["SPARE", "WEAR", "CONSUMABLE"]
CHANGE_FREQ = ["HIGH", "MEDIUM", "LOW"]
SYSTEMS = ["LUBE", "HIDR", "TRAN", "MECA", "SELL", "ELEC", "INST", "ESTR"]

bom_serial = 1
bom_count = 0
bom_json = []

for eq in EQUIPMENT:
    n_items = random.randint(3, 5)
    mats = random.sample(MATERIAL_POOL, k=min(n_items, len(MATERIAL_POOL)))

    for item_idx, mat in enumerate(mats):
        item_num = (item_idx + 1) * 10
        comp_cat = random.choices(["L", "N", "D"], weights=[75, 20, 5])[0]
        comp_cat_desc = {"L": "Stock item", "N": "Non-stock", "D": "Document"}[comp_cat]
        qty = random.randint(1, 4) if comp_cat != "D" else 1
        item_cat = random.choice(ITEM_CATEGORIES)
        freq = random.choice(CHANGE_FREQ)
        system = random.choice(SYSTEMS)
        lead = random.randint(7, 180)
        critical = "TRUE" if item_cat == "SPARE" and eq["abc"] == "1" and random.random() < 0.5 else "FALSE"

        bom_id = synth_id("BOM", bom_serial)
        bom_serial += 1

        row = [
            bom_id, eq["tag"], 100000 + EQUIPMENT.index(eq), eq["fl"],
            eq["desc"], eq["area"], eq["pg"],
            1, f"{item_num:04d}",
            comp_cat, comp_cat_desc,
            mat["code"], mat["sap"], mat["desc"],
            qty, mat["uom"],
            item_cat, freq,
            system, lead,
            mat["cost"], critical,
        ]
        ws28.append(row)
        bom_count += 1

        bom_json.append({
            "bom_id": bom_id, "equipment_tag": eq["tag"],
            "item_number": f"{item_num:04d}",
            "material_code": mat["code"], "component_desc": mat["desc"],
            "quantity": qty, "item_category": item_cat,
            "critical_spare": critical == "TRUE",
        })

auto_width(ws28)
save_wb(wb28, "28_equipment_bom.xlsx")
count_28 = bom_count


# ============================================================
# 29 — HISTORIAL DE COSTOS (IW39/KOB1)
# 200 registros: 150 por orden + 50 resumen mensual
# ============================================================

print("[29] Generating 29_cost_history.xlsx ...")
wb29 = openpyxl.Workbook()

# Sheet 1: Order Costs
ws29a = wb29.active
ws29a.title = "Order Costs"
h29a = [
    "cost_record_id", "aufnr", "auart", "auart_desc",
    "equipment_tag", "sap_func_loc", "area", "planning_group",
    "period", "value_category", "value_category_desc",
    "cost_element", "amount_usd", "currency",
    "posting_date", "cost_center", "wbs_element",
    "settlement_rule",
]
style_header(ws29a, h29a)

VALUE_CATS = {
    "ZMANT001": {"desc": "Mano de obra interna", "element": "621000"},
    "ZMANT002": {"desc": "Materiales y repuestos", "element": "400100"},
    "ZMANT003": {"desc": "Servicios externos", "element": "473000"},
}

cost_serial = 1
cost_count = 0
cost_json = []
monthly_accum = {}  # for sheet 2

selected_cost_orders = random.sample(CLOSED_ORDERS, k=min(50, len(CLOSED_ORDERS)))

for order in selected_cost_orders:
    pg = order["pg"]
    spec = random.choice(SPECIALTIES[:3])
    rate = LABOR_RATES.get(spec, 45.0)
    cc = COST_CENTERS.get(pg, COST_CENTERS["P01"]).get(spec[:3], "CC-PL-MEC-SEC")
    wbs = WBS_ELEMENTS.get(pg, "WBS-PL-SEC-MNT-2026")
    posting = order["actual_end"]
    period = posting.strftime("%Y-%m")

    # Labour cost (ZMANT001)
    labour_amt = round(rate * order["crew"] * order["duration_hours"], 2)
    cost_id = synth_id("CST", cost_serial)
    cost_serial += 1
    ws29a.append([
        cost_id, order["aufnr"], order["auart"], ORDER_TYPES_FULL[order["auart"]]["desc"],
        order["equipment_tag"], order["sap_func_loc"], order["area"], pg,
        period, "ZMANT001", VALUE_CATS["ZMANT001"]["desc"],
        VALUE_CATS["ZMANT001"]["element"], labour_amt, "USD",
        posting.strftime("%Y-%m-%d"), cc, wbs, "FULL",
    ])
    cost_count += 1
    cost_json.append({"aufnr": order["aufnr"], "value_category": "ZMANT001", "amount_usd": labour_amt, "period": period})

    # Material cost (ZMANT002) — ~70% of orders have material costs
    if random.random() < 0.7:
        mat_amt = round(random.uniform(200, 15000), 2)
        cost_id = synth_id("CST", cost_serial)
        cost_serial += 1
        ws29a.append([
            cost_id, order["aufnr"], order["auart"], ORDER_TYPES_FULL[order["auart"]]["desc"],
            order["equipment_tag"], order["sap_func_loc"], order["area"], pg,
            period, "ZMANT002", VALUE_CATS["ZMANT002"]["desc"],
            VALUE_CATS["ZMANT002"]["element"], mat_amt, "USD",
            posting.strftime("%Y-%m-%d"), cc, wbs, "FULL",
        ])
        cost_count += 1
        cost_json.append({"aufnr": order["aufnr"], "value_category": "ZMANT002", "amount_usd": mat_amt, "period": period})
    else:
        mat_amt = 0

    # External service cost (ZMANT003) — only PM03 and some PM07
    ext_amt = 0
    if order["auart"] in ("PM03", "PM07") and random.random() < 0.5:
        ext_amt = round(random.uniform(5000, 50000), 2)
        cost_id = synth_id("CST", cost_serial)
        cost_serial += 1
        ws29a.append([
            cost_id, order["aufnr"], order["auart"], ORDER_TYPES_FULL[order["auart"]]["desc"],
            order["equipment_tag"], order["sap_func_loc"], order["area"], pg,
            period, "ZMANT003", VALUE_CATS["ZMANT003"]["desc"],
            VALUE_CATS["ZMANT003"]["element"], ext_amt, "USD",
            posting.strftime("%Y-%m-%d"), cc, wbs, "FULL",
        ])
        cost_count += 1
        cost_json.append({"aufnr": order["aufnr"], "value_category": "ZMANT003", "amount_usd": ext_amt, "period": period})

    # Accumulate for monthly summary
    key = (period, order["area"], pg)
    if key not in monthly_accum:
        monthly_accum[key] = {"orders": 0, "labour": 0, "materials": 0, "external": 0, "hours": 0}
    monthly_accum[key]["orders"] += 1
    monthly_accum[key]["labour"] += labour_amt
    monthly_accum[key]["materials"] += mat_amt
    monthly_accum[key]["external"] += ext_amt
    monthly_accum[key]["hours"] += order["duration_hours"]

# Sheet 2: Monthly Summary
ws29b = wb29.create_sheet("Monthly Summary")
h29b = [
    "period", "area", "planning_group",
    "total_orders", "total_duration_hours",
    "cost_labour_usd", "cost_materials_usd", "cost_external_usd",
    "total_cost_usd", "budget_usd",
    "variance_usd", "variance_pct",
    "maintenance_cost_per_hour",
]
style_header(ws29b, h29b)

monthly_count = 0
for (period, area, pg), vals in sorted(monthly_accum.items()):
    total_cost = vals["labour"] + vals["materials"] + vals["external"]
    budget = round(total_cost * random.uniform(0.9, 1.2), 2)
    variance = round(total_cost - budget, 2)
    var_pct = round((variance / budget) * 100, 1) if budget > 0 else 0
    cph = round(total_cost / vals["hours"], 2) if vals["hours"] > 0 else 0

    ws29b.append([
        period, area, pg,
        vals["orders"], round(vals["hours"], 1),
        round(vals["labour"], 2), round(vals["materials"], 2), round(vals["external"], 2),
        round(total_cost, 2), budget,
        variance, var_pct,
        cph,
    ])
    monthly_count += 1

for ws in [ws29a, ws29b]:
    auto_width(ws)
save_wb(wb29, "29_cost_history.xlsx")
count_29 = cost_count + monthly_count


# ============================================================
# 30 — DATOS DE CONFIABILIDAD (TTF + WEIBULL)
# 200 registros: 150 TTF + 50 Weibull
# ============================================================

print("[30] Generating 30_reliability_data.xlsx ...")
wb30 = openpyxl.Workbook()

# Sheet 1: Time-to-Failure Records
ws30a = wb30.active
ws30a.title = "Time to Failure"
h30a = [
    "ttf_record_id", "equipment_tag", "equnr", "sap_func_loc",
    "area", "planning_group", "equipment_class",
    "failure_event_id", "failure_date", "previous_failure_date",
    "time_to_failure_days", "operating_hours",
    "suspension", "failure_mechanism", "failure_cause",
    "failure_pattern", "maintainable_item",
    "repair_duration_hours",
]
style_header(ws30a, h30a)

# Group PM01 (corrective) orders by equipment for TTF calculation
pm01_by_eq = {}
for order in CLOSED_ORDERS:
    if order["auart"] == "PM01":
        tag = order["equipment_tag"]
        if tag not in pm01_by_eq:
            pm01_by_eq[tag] = []
        pm01_by_eq[tag].append(order)

# Sort each equipment's failures by date
for tag in pm01_by_eq:
    pm01_by_eq[tag].sort(key=lambda o: o["created_date"])

# Equipment classes for grouping
EQ_CLASSES = {
    "CHAN": "CHANCADOR", "ALIM": "ALIMENTADOR", "COSE": "CHANCADOR_CONO",
    "COTE": "CHANCADOR_CONO", "HARV": "HARNERO", "CORR": "CORREA_TRANSP",
    "MSAG": "MOLINO_SAG", "MBOL": "MOLINO_BOLAS", "BOMB": "BOMBA_PULPA",
    "HCIC": "HIDROCICLON", "CFRO": "CELDA_FLOTACION", "CFCL": "CELDA_FLOTACION",
    "CFSC": "CELDA_FLOTACION", "AGIT": "AGITADOR", "DREA": "DOSIFICADOR",
    "ESPC": "ESPESADOR", "ESPR": "ESPESADOR", "BREL": "BOMBA_RELAVES",
    "FILT": "FILTRO_PRENSA", "SECA": "SECADOR", "COMP": "COMPRESOR",
    "BAGP": "BOMBA_AGUA", "ANPH": "ANALIZADOR", "FLMT": "FLUJOMETRO",
    "DENS": "DENSIMETRO", "VIBM": "MONITOR_VIBRACION",
}

ttf_serial = 1
ttf_count = 0
ttf_json = []
weibull_data = {}  # eq_class -> list of TTF days

for tag, failures in pm01_by_eq.items():
    eq_info = next((e for e in EQUIPMENT if e["tag"] == tag), None)
    if not eq_info:
        continue

    eq_prefix = tag.split("-")[-1][:4]
    eq_class = EQ_CLASSES.get(eq_prefix, "GENERICO")

    for i in range(1, len(failures)):
        prev = failures[i - 1]
        curr = failures[i]
        ttf_days = (curr["created_date"] - prev["created_date"]).days
        if ttf_days <= 0:
            ttf_days = random.randint(15, 90)

        op_hours = ttf_days * random.uniform(16, 22)

        mechanism = random.choice(MECHANISMS)
        cause = random.choice(CAUSES)
        pattern = random.choice(PATTERNS)
        mi = random.choice(MAINT_ITEMS)

        ttf_id = synth_id("TTF", ttf_serial)
        ttf_serial += 1

        row = [
            ttf_id, tag, eq_info.get("equnr", 100000 + EQUIPMENT.index(eq_info)), eq_info["fl"],
            eq_info["area"], eq_info["pg"], eq_class,
            curr["aufnr"], curr["created_date"].strftime("%Y-%m-%d"),
            prev["created_date"].strftime("%Y-%m-%d"),
            ttf_days, round(op_hours, 0),
            "FALSE", mechanism, cause,
            pattern, mi,
            curr["duration_hours"],
        ]
        ws30a.append(row)
        ttf_count += 1

        ttf_json.append({
            "ttf_record_id": ttf_id, "equipment_tag": tag,
            "equipment_class": eq_class,
            "failure_date": curr["created_date"].strftime("%Y-%m-%d"),
            "time_to_failure_days": ttf_days,
            "failure_mechanism": mechanism, "failure_cause": cause,
            "failure_pattern": pattern,
        })

        if eq_class not in weibull_data:
            weibull_data[eq_class] = []
        weibull_data[eq_class].append(ttf_days)

# Add synthetic TTF records to reach 150+
# Generate TTF for equipment that only had 1 failure (supplement with synthetic intervals)
eq_abc1_tags = [e["tag"] for e in EQUIPMENT if e["abc"] == "1"]
while ttf_count < 150:
    tag = random.choice(eq_abc1_tags)
    eq_info = next((e for e in EQUIPMENT if e["tag"] == tag), None)
    if not eq_info:
        continue

    eq_prefix = tag.split("-")[-1][:4]
    eq_class = EQ_CLASSES.get(eq_prefix, "GENERICO")

    ttf_days = random.randint(20, 200)
    base_date = SNAPSHOT_DATE - timedelta(days=random.randint(30, 300))
    prev_date = base_date - timedelta(days=ttf_days)
    op_hours = ttf_days * random.uniform(16, 22)

    ttf_id = synth_id("TTF", ttf_serial)
    ttf_serial += 1

    mechanism = random.choice(MECHANISMS)
    cause = random.choice(CAUSES)
    pattern = random.choice(PATTERNS)
    mi = random.choice(MAINT_ITEMS)
    dur = round(random.uniform(2, 48), 1)

    ws30a.append([
        ttf_id, tag, 100000 + EQUIPMENT.index(eq_info), eq_info["fl"],
        eq_info["area"], eq_info["pg"], eq_class,
        "", base_date.strftime("%Y-%m-%d"), prev_date.strftime("%Y-%m-%d"),
        ttf_days, round(op_hours, 0),
        "FALSE", mechanism, cause, pattern, mi, dur,
    ])
    ttf_count += 1

    ttf_json.append({
        "ttf_record_id": ttf_id, "equipment_tag": tag,
        "equipment_class": eq_class,
        "failure_date": base_date.strftime("%Y-%m-%d"),
        "time_to_failure_days": ttf_days,
        "failure_mechanism": mechanism, "failure_cause": cause,
        "failure_pattern": pattern,
    })

    if eq_class not in weibull_data:
        weibull_data[eq_class] = []
    weibull_data[eq_class].append(ttf_days)

# Sheet 2: Weibull Parameters
ws30b = wb30.create_sheet("Weibull Parameters")
h30b = [
    "weibull_id", "equipment_class", "equipment_class_desc",
    "sample_size", "beta", "eta_days", "gamma_days",
    "r_squared", "mtbf_days",
    "failure_pattern", "failure_pattern_desc",
    "recommended_interval_days",
    "confidence_level", "status",
]
style_header(ws30b, h30b)

PATTERN_FROM_BETA = [
    (0.5, "F_INFANT", "Mortalidad infantil"),
    (1.0, "E_RANDOM", "Aleatorio"),
    (1.5, "C_GRADUAL", "Degradacion gradual"),
    (2.5, "B_AGE", "Desgaste por edad"),
    (4.0, "A_BATHTUB", "Curva banera"),
]

weibull_serial = 1
weibull_count = 0
weibull_json = []

for eq_class, ttf_list in sorted(weibull_data.items()):
    if len(ttf_list) < 3:
        # Supplement with synthetic data
        for _ in range(5 - len(ttf_list)):
            ttf_list.append(random.randint(30, 180))

    n = len(ttf_list)
    mean_ttf = sum(ttf_list) / n
    std_ttf = (sum((x - mean_ttf) ** 2 for x in ttf_list) / max(n - 1, 1)) ** 0.5

    # Estimate Weibull beta from coefficient of variation
    cv = std_ttf / mean_ttf if mean_ttf > 0 else 1.0
    if cv < 0.3:
        beta = round(random.uniform(3.0, 5.0), 2)
    elif cv < 0.6:
        beta = round(random.uniform(1.5, 3.0), 2)
    elif cv < 1.0:
        beta = round(random.uniform(0.8, 1.5), 2)
    else:
        beta = round(random.uniform(0.5, 1.0), 2)

    eta = round(mean_ttf * 1.1, 1)
    gamma = round(min(ttf_list) * 0.3, 1) if min(ttf_list) > 10 else 0
    r_sq = round(random.uniform(0.82, 0.98), 3)
    mtbf = round(mean_ttf, 1)

    # Determine pattern from beta
    pattern = "E_RANDOM"
    pattern_desc = "Aleatorio"
    for threshold, pat, desc in PATTERN_FROM_BETA:
        if beta <= threshold:
            pattern = pat
            pattern_desc = desc
            break

    rec_interval = round(eta * 0.7, 0) if beta > 1.0 else round(eta * 0.5, 0)

    wb_id = synth_id("WBL", weibull_serial)
    weibull_serial += 1

    ws30b.append([
        wb_id, eq_class, eq_class.replace("_", " ").title(),
        n, beta, eta, gamma,
        r_sq, mtbf,
        pattern, pattern_desc,
        int(rec_interval),
        0.90, "DRAFT",
    ])
    weibull_count += 1

    weibull_json.append({
        "weibull_id": wb_id, "equipment_class": eq_class,
        "sample_size": n, "beta": beta, "eta_days": eta, "gamma_days": gamma,
        "r_squared": r_sq, "mtbf_days": mtbf,
        "failure_pattern": pattern,
        "recommended_interval_days": int(rec_interval),
        "status": "DRAFT",
    })

# Supplement Weibull to reach ~50 records
EXTRA_CLASSES = [
    ("BOMBA_PULPA_SELLO", "Sello mecanico bomba pulpa"),
    ("MOLINO_LINER", "Revestimiento molino"),
    ("CORREA_POLEA", "Polea correa transportadora"),
    ("CHANCADOR_EXCENT", "Excentrica chancador"),
    ("FLOTACION_AGITADOR", "Agitador celda flotacion"),
    ("ESPESADOR_RASTRAS", "Rastras espesador"),
    ("FILTRO_PLACAS", "Placas filtro prensa"),
    ("COMPRESOR_VALV", "Valvulas compresor"),
    ("BOMBA_RODAMIENTO", "Rodamiento bomba"),
    ("MOTOR_BOBINADO", "Bobinado motor electrico"),
]

for eq_class, desc in EXTRA_CLASSES:
    if weibull_count >= 50:
        break
    n = random.randint(5, 20)
    beta = round(random.uniform(0.7, 4.5), 2)
    eta = round(random.uniform(60, 400), 1)
    gamma = round(random.uniform(0, 30), 1)
    r_sq = round(random.uniform(0.80, 0.97), 3)
    mtbf = round(eta * 0.9, 1)

    pattern = "E_RANDOM"
    pattern_desc = "Aleatorio"
    for threshold, pat, pdesc in PATTERN_FROM_BETA:
        if beta <= threshold:
            pattern = pat
            pattern_desc = pdesc
            break

    rec_interval = round(eta * 0.7, 0) if beta > 1.0 else round(eta * 0.5, 0)

    wb_id = synth_id("WBL", weibull_serial)
    weibull_serial += 1

    ws30b.append([
        wb_id, eq_class, desc,
        n, beta, eta, gamma,
        r_sq, mtbf,
        pattern, pattern_desc,
        int(rec_interval),
        0.90, "DRAFT",
    ])
    weibull_count += 1

    weibull_json.append({
        "weibull_id": wb_id, "equipment_class": eq_class,
        "sample_size": n, "beta": beta, "eta_days": eta, "gamma_days": gamma,
        "r_squared": r_sq, "mtbf_days": mtbf,
        "failure_pattern": pattern,
        "recommended_interval_days": int(rec_interval),
        "status": "DRAFT",
    })

for ws in [ws30a, ws30b]:
    auto_width(ws)
save_wb(wb30, "30_reliability_data.xlsx")
count_30 = ttf_count + weibull_count


# ============================================================
# GENERATE JSON FOR sap_mock/data/
# ============================================================

print("\n[JSON] Generating SAP mock JSON files ...")

os.makedirs(MOCK_DIR, exist_ok=True)

# Fase 3 JSONs
json_path = os.path.join(MOCK_DIR, "active_backlog.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({"snapshot_date": "2026-03-30", "total_items": len(bkl_json), "items": bkl_json}, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/active_backlog.json ({len(bkl_json)} items)")

json_path = os.path.join(MOCK_DIR, "notifications.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({
        "open_count": len(notif_json["open"]), "closed_count": len(notif_json["closed"]),
        "open": notif_json["open"], "closed": notif_json["closed"],
    }, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/notifications.json ({len(notif_json['open'])} open, {len(notif_json['closed'])} closed)")

json_path = os.path.join(MOCK_DIR, "measurement_docs.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({
        "period": {"start": "2025-10-01", "end": "2026-03-30"},
        "total_documents": len(mdoc_json), "documents": mdoc_json,
    }, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/measurement_docs.json ({len(mdoc_json)} documents)")

# Fase 4 JSONs
json_path = os.path.join(MOCK_DIR, "time_confirmations.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({"total_confirmations": len(conf_json), "confirmations": conf_json}, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/time_confirmations.json ({len(conf_json)} confirmations)")

json_path = os.path.join(MOCK_DIR, "material_movements.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({
        "reservations_count": len(res_json), "movements_count": len(mov_json),
        "reservations": res_json, "movements": mov_json,
    }, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/material_movements.json ({len(res_json)} res, {len(mov_json)} mov)")

json_path = os.path.join(MOCK_DIR, "equipment_bom.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({"total_items": len(bom_json), "bom_items": bom_json}, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/equipment_bom.json ({len(bom_json)} items)")

json_path = os.path.join(MOCK_DIR, "cost_history.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({"total_records": len(cost_json), "cost_records": cost_json}, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/cost_history.json ({len(cost_json)} records)")

json_path = os.path.join(MOCK_DIR, "reliability_data.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump({
        "ttf_count": len(ttf_json), "weibull_count": len(weibull_json),
        "time_to_failure": ttf_json, "weibull_parameters": weibull_json,
    }, f, indent=2, ensure_ascii=False)
print(f"  -> Saved: sap_mock/data/reliability_data.json ({len(ttf_json)} TTF, {len(weibull_json)} Weibull)")


# ============================================================
# SUMMARY
# ============================================================

print("\n" + "=" * 70)
print("  RESUMEN — Gap Templates Fase 3 + Fase 4")
print("=" * 70)

summary = {
    "23_active_backlog": count_23,
    "24_notifications": count_24,
    "25_measurement_documents": count_25,
    "26_time_confirmations": count_26,
    "27_material_movements": count_27,
    "28_equipment_bom": count_28,
    "29_cost_history": count_29,
    "30_reliability_data": count_30,
}

total_f3 = count_23 + count_24 + count_25
total_f4 = count_26 + count_27 + count_28 + count_29 + count_30
total = total_f3 + total_f4

print("\n  --- Fase 3 (Critico) ---")
for name in ["23_active_backlog", "24_notifications", "25_measurement_documents"]:
    cnt = summary[name]
    status = "OK >=200" if cnt >= 200 else f"OK ({cnt})"
    print(f"  {name:.<45s} {cnt:>6d}  [{status}]")

print(f"\n  --- Fase 4 (Enriquecimiento) ---")
for name in ["26_time_confirmations", "27_material_movements", "28_equipment_bom", "29_cost_history", "30_reliability_data"]:
    cnt = summary[name]
    status = "OK >=200" if cnt >= 200 else f"OK ({cnt})"
    print(f"  {name:.<45s} {cnt:>6d}  [{status}]")

print(f"\n  {'TOTAL FASE 3':.<45s} {total_f3:>6d} registros")
print(f"  {'TOTAL FASE 4':.<45s} {total_f4:>6d} registros")
print(f"  {'TOTAL NUEVAS (8 plantillas)':.<45s} {total:>6d} registros")
print(f"  {'JSON generados':.<45s}      7 archivos")
print(f"  {'TOTAL PREVIAS (22 plantillas)':.<45s}   6814 registros")
print(f"  {'GRAN TOTAL (30 plantillas)':.<45s} {total + 6814:>6d} registros")
print("=" * 70)
