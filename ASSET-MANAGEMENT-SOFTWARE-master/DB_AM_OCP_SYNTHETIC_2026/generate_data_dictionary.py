# -*- coding: utf-8 -*-
"""
=============================================================================
  DB_AM_OCP_SYNTHETIC_2026 — Data Dictionary / Glosario de Codigos
=============================================================================
  Genera un diccionario de datos unificado con TODOS los codigos usados en
  las 30 plantillas del dataset, incluyendo:
    - Codigos SAP (tipos OT, clases aviso, status, prioridades)
    - Codigos organizacionales (PG, BA, CC, WBS)
    - Codigos catalogo (B-xxx, C-xxx, 5-xxx)
    - Puestos de trabajo (PASMEC01, etc.)
    - Mapeo prioridades SAP ↔ AMS
    - Metadata por plantilla (campos, tipos, obligatoriedad)
=============================================================================
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

OUT = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# STYLING
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CAT_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MAP_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))


def style_header(ws, headers, row=1, fill=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)


print("=" * 70)
print("  DB_AM_OCP_SYNTHETIC_2026 — Data Dictionary Generator")
print("=" * 70)

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: SAP CODES (Tipos OT, Avisos, Status, Prioridades)
# ============================================================
ws1 = wb.active
ws1.title = "SAP Codes"
style_header(ws1, ["category", "code", "description_es", "description_en", "sap_transaction", "blueprint_table", "used_in_templates"])

SAP_CODES = [
    # Order Types (AUART)
    ("Tipo Orden (AUART)", "PM01", "Orden Mantenimiento de Averia", "Breakdown Maintenance Order", "IW31/IW38", "T26", "06, 23, 26, 29"),
    ("Tipo Orden (AUART)", "PM02", "Orden Mantenimiento Preventivo", "Preventive Maintenance Order", "IW31/IW38", "T26", "06, 23, 26, 29"),
    ("Tipo Orden (AUART)", "PM03", "Orden Solicitud de Mantenimiento", "Maintenance Request Order", "IW31/IW38", "T26", "06, 23, 26, 29"),
    ("Tipo Orden (AUART)", "PM06", "Orden de Inversion", "Investment Order", "IW31/IW38", "T26", "06, 29"),
    ("Tipo Orden (AUART)", "PM07", "Orden Reparacion de Componentes", "Component Repair Order", "IW31/IW38", "T26", "06, 29"),
    ("Tipo Orden (AUART)", "PM04", "NO EXISTE — Excluido por Blueprint", "DOES NOT EXIST — Excluded by Blueprint", "-", "T26", "-"),
    # PM07 Activity Classes
    ("Clase Actividad PM07", "RP1", "Reparacion Mayor", "Major Repair", "IW31", "T28", "06"),
    ("Clase Actividad PM07", "RP2", "Reparacion Menor", "Minor Repair", "IW31", "T28", "06"),
    # Notification Types (QMART)
    ("Clase Aviso (QMART)", "A1", "Aviso de mantenimiento (averia)", "Maintenance Notification (breakdown)", "IW21/IW28", "T20", "06, 24"),
    ("Clase Aviso (QMART)", "A2", "Aviso predictivo e ingenieria", "Predictive & Engineering Notification", "IW21/IW28", "T20", "06, 24"),
    ("Clase Aviso (QMART)", "A3", "Aviso plan preventivo", "Preventive Plan Notification", "IW21/IW28", "T20", "06, 24"),
    # Notification Catalogs
    ("Catalogo Aviso", "M001", "Solicitud de mantenimiento", "Maintenance Request", "QS41", "T22", "06, 24"),
    ("Catalogo Aviso", "M002", "Averia", "Breakdown", "QS41", "T22", "06, 24"),
    ("Catalogo Aviso", "M003", "Reparacion de componentes", "Component Repair", "QS41", "T22", "06, 24"),
    ("Catalogo Aviso", "P001", "Predictivo", "Predictive", "QS41", "T22", "06, 24"),
    ("Catalogo Aviso", "P002", "Ingenieria", "Engineering", "QS41", "T22", "06, 24"),
    # User Status Schema
    ("Esquema Status Usuario", "ZPM00001", "Esquema status personalizado PM", "Custom PM Status Schema", "BS02", "T25", "06, 24"),
    ("Status Usuario", "APRO", "Aprobado", "Approved", "BS02", "T25", "06, 24"),
    ("Status Usuario", "RECH", "Rechazado", "Rejected", "BS02", "T25", "06, 24"),
    # System Status
    ("Status Sistema", "ABIE", "Abierto (creado)", "Open (created)", "IW38", "-", "06"),
    ("Status Sistema", "LIBE", "Liberado para ejecucion", "Released for execution", "IW38", "-", "06"),
    ("Status Sistema", "NOTI", "Notificacion en proceso", "Notification in process", "IW38", "-", "06"),
    ("Status Sistema", "CTEC", "Terminado tecnicamente", "Technically completed", "IW38", "-", "06"),
    ("Status Sistema Aviso", "OSNO", "Aviso pendiente", "Outstanding notification", "IW28", "-", "24"),
    ("Status Sistema Aviso", "NOPR", "Aviso en proceso", "Notification in process", "IW28", "-", "24"),
    ("Status Sistema Aviso", "NOCO", "Aviso completado", "Notification completed", "IW28", "-", "24"),
    ("Status Confirmacion", "PCNF", "Parcialmente confirmado", "Partially confirmed", "IW42", "-", "26"),
    ("Status Confirmacion", "CNF", "Confirmado", "Confirmed", "IW42", "-", "26"),
    # Priority
    ("Prioridad SAP (Z1)", "I", "Inmediata", "Immediate", "IW31", "T24/T29", "06, 23, 24"),
    ("Prioridad SAP (Z1)", "A", "Alta (2-6 dias)", "High (2-6 days)", "IW31", "T24/T29", "06, 23, 24"),
    ("Prioridad SAP (Z1)", "M", "Media (7-14 dias)", "Medium (7-14 days)", "IW31", "T24/T29", "06, 23, 24"),
    ("Prioridad SAP (Z1)", "B", "Baja (mayor a 14 dias)", "Low (over 14 days)", "IW31", "T24/T29", "06, 23, 24"),
    ("Clase Prioridad", "Z1", "Clase prioridad personalizada OCP", "Custom OCP Priority Class", "SPRO", "T24/T29", "06, 23, 24"),
    # Value Categories
    ("Categoria Valor", "ZMANT001", "Mano de obra interna", "Internal Labour", "KOB1", "T30", "06, 29"),
    ("Categoria Valor", "ZMANT002", "Materiales y repuestos", "Materials & Spare Parts", "KOB1", "T30", "06, 29"),
    ("Categoria Valor", "ZMANT003", "Servicios externos", "External Services", "KOB1", "T30", "06, 29"),
    # Equipment Types
    ("Tipo Equipo SAP", "M", "Maquinas", "Machines", "IE03", "T15", "01, 04"),
    ("Tipo Equipo SAP", "Q", "Inspeccion/Medida", "Inspection/Measurement", "IE03", "T15", "01, 04"),
    # Functional Location Type
    ("Tipo Ubicacion Tecnica", "M", "Sistema tecnico estandar", "Standard Technical System", "IL03", "T14", "01"),
    # Criticality ABC
    ("Criticidad ABC", "1", "Alto", "High", "IE02", "T19", "01, 02"),
    ("Criticidad ABC", "2", "Medio", "Medium", "IE02", "T19", "01, 02"),
    ("Criticidad ABC", "3", "Bajo", "Low", "IE02", "T19", "01, 02"),
    # Route Sheet
    ("Tipo Hoja Ruta", "A", "Instruccion mantenimiento", "Maintenance Instruction", "IA05", "T34", "16"),
    ("Grupo Hoja Ruta", "PLA", "Planta", "Plant", "IA05", "T35", "16"),
    ("Grupo Hoja Ruta", "MIN", "Mina", "Mine", "IA05", "T35", "16"),
    # Maintenance Plan
    ("Tipo Plan Mant.", "PM", "Plan mantenimiento preventivo", "Preventive Maintenance Plan", "IP10", "T37", "17"),
    # Installation Status
    ("Estado Instalacion", "0", "No instalado", "Not Installed", "IE02", "T31", "22"),
    ("Estado Instalacion", "1", "Instalado", "Installed", "IE02", "T31", "22"),
    # Movement Types
    ("Tipo Movimiento", "261", "Salida mercancia a orden", "Goods Issue to Order", "MIGO", "-", "27"),
    ("Tipo Movimiento", "262", "Anulacion salida mercancia", "Reversal Goods Issue", "MIGO", "-", "27"),
    ("Tipo Movimiento", "101", "Entrada mercancia pedido", "Goods Receipt from PO", "MIGO", "-", "27"),
    ("Tipo Movimiento", "311", "Traspaso entre almacenes", "Transfer Between Warehouses", "MIGO", "-", "27"),
    # BOM Component Categories
    ("Categoria Componente BOM", "L", "Stock item (gestionado inventario)", "Stock Item (inventory managed)", "CS03", "-", "28"),
    ("Categoria Componente BOM", "N", "Non-stock (compra directa)", "Non-Stock (direct purchase)", "CS03", "-", "28"),
    ("Categoria Componente BOM", "D", "Documento (plano, manual)", "Document (drawing, manual)", "CS03", "-", "28"),
    # Backlog Status (AMS-specific)
    ("Status Backlog (AMS)", "AWAITING_MATERIALS", "Esperando materiales", "Awaiting Materials", "-", "-", "23"),
    ("Status Backlog (AMS)", "AWAITING_SHUTDOWN", "Esperando parada", "Awaiting Shutdown", "-", "-", "23"),
    ("Status Backlog (AMS)", "AWAITING_RESOURCES", "Esperando recursos", "Awaiting Resources", "-", "-", "23"),
    ("Status Backlog (AMS)", "AWAITING_APPROVAL", "Esperando aprobacion", "Awaiting Approval", "-", "-", "23"),
    ("Status Backlog (AMS)", "SCHEDULED", "Programado", "Scheduled", "-", "-", "23"),
    ("Status Backlog (AMS)", "IN_PROGRESS", "En progreso", "In Progress", "-", "-", "23"),
    # Valuation Code (Measurement)
    ("Codigo Valoracion Medicion", "1", "OK (dentro de limites)", "OK (within limits)", "IK12", "-", "25"),
    ("Codigo Valoracion Medicion", "2", "Warning (cerca de limites)", "Warning (near limits)", "IK12", "-", "25"),
    ("Codigo Valoracion Medicion", "3", "Alarm (fuera de limites)", "Alarm (out of limits)", "IK12", "-", "25"),
]

for row in SAP_CODES:
    ws1.append(list(row))
auto_width(ws1)

# ============================================================
# SHEET 2: ORGANIZATIONAL CODES (PG, BA, CC, WBS, WC)
# ============================================================
ws2 = wb.create_sheet("Organizational Codes")
style_header(ws2, ["category", "code", "description_es", "description_en", "parent", "used_in_templates"])

ORG_CODES = [
    # Planning Center
    ("Centro Planificacion", "AN01", "Centro planificacion estandar OCP", "OCP Standard Planning Center", "-", "Todos"),
    ("Planta", "OCP-CON1", "Planta Concentradora", "Concentrator Plant", "AN01", "01, 04, 22"),
    # Planning Groups
    ("Grupo Planificacion Planta", "P01", "Area Seca (Chancado + Filtrado)", "Dry Area (Crushing + Filtering)", "AN01", "06, 12, 13, 22, 23, 24"),
    ("Grupo Planificacion Planta", "P02", "Area Ripio (Molienda)", "Mid Area (Grinding)", "AN01", "06, 12, 13, 22, 23, 24"),
    ("Grupo Planificacion Planta", "P03", "Area Humeda (Flotacion + Espesado + Servicios)", "Wet Area (Flotation + Thickening + Services)", "AN01", "06, 12, 13, 22, 23, 24"),
    ("Grupo Planificacion Mina", "M01", "Perforacion", "Drilling", "AN01", "22"),
    ("Grupo Planificacion Mina", "M02", "Carguio", "Loading", "AN01", "22"),
    ("Grupo Planificacion Mina", "M03", "Transporte", "Hauling", "AN01", "22"),
    ("Grupo Planificacion Mina", "M04", "Equipos de apoyo", "Support Equipment", "AN01", "22"),
    ("Grupo Planificacion Mina", "M05", "Equipos auxiliares", "Auxiliary Equipment", "AN01", "22"),
    # Business Areas
    ("Area Empresa Planta", "SEC", "Area Seca", "Dry Area", "P01", "06, 12, 13, 22"),
    ("Area Empresa Planta", "RIP", "Area Ripio", "Mid Area", "P02", "06, 12, 13, 22"),
    ("Area Empresa Planta", "HUM", "Area Humeda", "Wet Area", "P03", "06, 12, 13, 22"),
    ("Area Empresa Mina", "PER", "Perforacion", "Drilling", "M01", "22"),
    ("Area Empresa Mina", "CAR", "Carguio", "Loading", "M02", "22"),
    ("Area Empresa Mina", "TRA", "Transporte", "Hauling", "M03", "22"),
    ("Area Empresa Mina", "APO", "Apoyo", "Support", "M04", "22"),
    ("Area Empresa Mina", "AUX", "Auxiliar", "Auxiliary", "M05", "22"),
    ("Area Empresa Mina", "TAL", "Taller", "Workshop", "M01", "22"),
    # Internal Work Centers (Planta)
    ("Puesto Trabajo INT Planta", "PASMEC01", "Planta Area Seca Mecanico 01", "Plant Dry Area Mechanic 01", "P01/SEC", "11, 06, 23"),
    ("Puesto Trabajo INT Planta", "PASELE01", "Planta Area Seca Electrico 01", "Plant Dry Area Electrician 01", "P01/SEC", "11, 06"),
    ("Puesto Trabajo INT Planta", "PASINS01", "Planta Area Seca Instrumentista 01", "Plant Dry Area Instrument Tech 01", "P01/SEC", "11, 06"),
    ("Puesto Trabajo INT Planta", "PASLUB01", "Planta Area Seca Lubricacion 01", "Plant Dry Area Lubrication 01", "P01/SEC", "11"),
    ("Puesto Trabajo INT Planta", "PARELE01", "Planta Area Ripio Electrico 01", "Plant Mid Area Electrician 01", "P02/RIP", "11, 06"),
    ("Puesto Trabajo INT Planta", "PARINS01", "Planta Area Ripio Instrumentista 01", "Plant Mid Area Instrument Tech 01", "P02/RIP", "11, 06"),
    ("Puesto Trabajo INT Planta", "PARMEC01", "Planta Area Ripio Mecanico 01", "Plant Mid Area Mechanic 01", "P02/RIP", "11, 06, 23"),
    ("Puesto Trabajo INT Planta", "PAHMEC01", "Planta Area Humeda Mecanico 01", "Plant Wet Area Mechanic 01", "P03/HUM", "11, 06, 23"),
    ("Puesto Trabajo INT Planta", "PAHELE01", "Planta Area Humeda Electrico 01", "Plant Wet Area Electrician 01", "P03/HUM", "11, 06"),
    ("Puesto Trabajo INT Planta", "PAHINS01", "Planta Area Humeda Instrumentista 01", "Plant Wet Area Instrument Tech 01", "P03/HUM", "11, 06"),
    ("Puesto Trabajo INT Planta", "PSHSIN01", "Planta Sintomatico 01", "Plant Condition Monitoring 01", "P01/SEC", "11"),
    ("Puesto Trabajo INT Planta", "PSHDCS01", "Planta DCS/Automatizacion 01", "Plant DCS/Automation 01", "P01/SEC", "11"),
    # External Work Centers
    ("Puesto Trabajo EXT", "MEXTSOL1", "Mina Externo Soldadura 1", "Mine External Welding 1", "M01", "11"),
    ("Puesto Trabajo EXT", "MEXTLAV1", "Mina Externo Lavado 1", "Mine External Washing 1", "M01", "11"),
    ("Puesto Trabajo EXT", "MEXTNEU1", "Mina Externo Neumaticos 1", "Mine External Tires 1", "M01", "11"),
    ("Puesto Trabajo EXT", "MEXTCAB1", "Mina Externo Cabina 1", "Mine External Cabin 1", "M01", "11"),
    # Supervisor Work Centers
    ("Puesto Trabajo SUP", "SPASMEC", "Supervisor Planta Area Seca Mecanico", "Supervisor Plant Dry Mechanic", "P01", "11, 06"),
    ("Puesto Trabajo SUP", "SPASELE", "Supervisor Planta Area Seca Electrico", "Supervisor Plant Dry Electric", "P01", "11"),
    ("Puesto Trabajo SUP", "SPARELE", "Supervisor Planta Area Ripio Electrico", "Supervisor Plant Mid Electric", "P02", "11"),
    ("Puesto Trabajo SUP", "SPAHMEC", "Supervisor Planta Area Humeda Mecanico", "Supervisor Plant Wet Mechanic", "P03", "11, 06"),
    # Cost Centers
    ("Centro Costo", "CC-PL-MEC-SEC", "Planta Mecanico Area Seca", "Plant Mechanic Dry Area", "P01/SEC", "11, 20, 27, 29"),
    ("Centro Costo", "CC-PL-ELE-SEC", "Planta Electrico Area Seca", "Plant Electrician Dry Area", "P01/SEC", "11, 20, 29"),
    ("Centro Costo", "CC-PL-INS-SEC", "Planta Instrumentista Area Seca", "Plant Instrument Dry Area", "P01/SEC", "11, 20"),
    ("Centro Costo", "CC-PL-MEC-RIP", "Planta Mecanico Area Ripio", "Plant Mechanic Mid Area", "P02/RIP", "11, 20, 29"),
    ("Centro Costo", "CC-PL-ELE-RIP", "Planta Electrico Area Ripio", "Plant Electrician Mid Area", "P02/RIP", "11, 20"),
    ("Centro Costo", "CC-PL-INS-RIP", "Planta Instrumentista Area Ripio", "Plant Instrument Mid Area", "P02/RIP", "11, 20"),
    ("Centro Costo", "CC-PL-MEC-HUM", "Planta Mecanico Area Humeda", "Plant Mechanic Wet Area", "P03/HUM", "11, 20, 29"),
    ("Centro Costo", "CC-PL-ELE-HUM", "Planta Electrico Area Humeda", "Plant Electrician Wet Area", "P03/HUM", "11, 20"),
    ("Centro Costo", "CC-PL-INS-HUM", "Planta Instrumentista Area Humeda", "Plant Instrument Wet Area", "P03/HUM", "11, 20"),
    # WBS Elements
    ("Elemento PEP/WBS", "WBS-PL-SEC-MNT-2026", "Mantenimiento Planta Area Seca 2026", "Maintenance Plant Dry Area 2026", "P01", "20, 29"),
    ("Elemento PEP/WBS", "WBS-PL-RIP-MNT-2026", "Mantenimiento Planta Area Ripio 2026", "Maintenance Plant Mid Area 2026", "P02", "20, 29"),
    ("Elemento PEP/WBS", "WBS-PL-HUM-MNT-2026", "Mantenimiento Planta Area Humeda 2026", "Maintenance Plant Wet Area 2026", "P03", "20, 29"),
    # Specialties
    ("Especialidad", "MEC", "Mecanico", "Mechanic", "-", "09, 11, 23, 26"),
    ("Especialidad", "ELE", "Electricista", "Electrician", "-", "09, 11, 23, 26"),
    ("Especialidad", "INS", "Instrumentista", "Instrument Technician", "-", "09, 11, 23, 26"),
    ("Especialidad", "LUB", "Lubricador", "Lubrication Technician", "-", "09, 11"),
    ("Especialidad", "SOL", "Soldador", "Welder", "-", "09, 11, 23"),
    ("Especialidad", "SIN", "Tecnico Sintomatico", "Condition Monitoring Tech", "-", "11"),
    ("Especialidad", "DCS", "Tecnico DCS/Automatizacion", "DCS/Automation Tech", "-", "11"),
    # Storage Locations
    ("Ubicacion Almacen", "ALM-SEC-01", "Almacen Area Seca 01", "Dry Area Warehouse 01", "SEC", "27"),
    ("Ubicacion Almacen", "ALM-RIP-01", "Almacen Area Ripio 01", "Mid Area Warehouse 01", "RIP", "27"),
    ("Ubicacion Almacen", "ALM-HUM-01", "Almacen Area Humeda 01", "Wet Area Warehouse 01", "HUM", "27"),
    ("Ubicacion Almacen", "ALM-CENTRAL", "Almacen Central Planta", "Central Plant Warehouse", "-", "27"),
]

for row in ORG_CODES:
    ws2.append(list(row))
auto_width(ws2)

# ============================================================
# SHEET 3: CATALOG CODES (B-xxx, C-xxx, 5-xxx from template 15)
# ============================================================
ws3 = wb.create_sheet("Catalog Codes")
style_header(ws3, ["catalog_type", "catalog_type_desc", "group_code", "group_desc", "code", "description_es", "applicable_to"])

CATALOG_CODES = [
    # Type D (Order Determination)
    ("D", "Determinacion Orden", "D-M001", "Solicitud mantenimiento", "D-M001", "Solicitud de mantenimiento", "A1"),
    ("D", "Determinacion Orden", "D-M002", "Averia", "D-M002", "Averia", "A1"),
    ("D", "Determinacion Orden", "D-M003", "Reparacion componentes", "D-M003", "Reparacion de componentes", "A1"),
    ("D", "Determinacion Orden", "D-P001", "Predictivo", "D-P001", "Predictivo", "A2"),
    ("D", "Determinacion Orden", "D-P002", "Ingenieria", "D-P002", "Ingenieria", "A2"),
    # Type B (Object Parts) - Key mechanical
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-ROD", "Rodamiento", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-RET", "Retenedor/Sello", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-EJE", "Eje", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-ACP", "Acoplamiento", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-RDC", "Reductor", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-ENG", "Engranaje", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-IMP", "Impulsor", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-BOM", "Bomba (cuerpo)", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-MEC", "Mecanico", "B-VLV", "Valvula", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-ELE", "Electrico", "B-MOT", "Motor electrico", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-ELE", "Electrico", "B-CAB", "Cableado", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-ELE", "Electrico", "B-BOB", "Bobinado", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-ELE", "Electrico", "B-CON", "Contactor/Proteccion", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-INS", "Instrumentacion", "B-SEN", "Sensor", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-INS", "Instrumentacion", "B-TRX", "Transmisor", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-INS", "Instrumentacion", "B-PRB", "Probe/Sonda", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-EST", "Estructural", "B-BAS", "Bastidor", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-EST", "Estructural", "B-SOL", "Soldadura", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-EST", "Estructural", "B-PER", "Pernos/Sujeciones", "A1,A2,A3"),
    ("B", "Parte del Objeto", "B-EST", "Estructural", "B-PLN", "Planchas desgaste", "A1,A2,A3"),
    # Type C (Symptoms) - Key entries
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-VIB", "Vibracion excesiva", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-RUI", "Ruido anormal", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-CAL", "Sobrecalentamiento", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-FUG", "Fuga aceite/fluido", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-DES", "Desgaste", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-FIS", "Fisura/Fractura", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-BLO", "Bloqueo/Atasco", "A1,A2"),
    ("C", "Sintomas/Dano", "C-MEC", "Mecanico", "C-DAL", "Desalineamiento", "A1,A2"),
    ("C", "Sintomas/Dano", "C-ELE", "Electrico", "C-CRT", "Cortocircuito", "A1,A2"),
    ("C", "Sintomas/Dano", "C-ELE", "Electrico", "C-SOB", "Sobrecarga", "A1,A2"),
    ("C", "Sintomas/Dano", "C-ELE", "Electrico", "C-AIS", "Falla aislacion", "A1,A2"),
    ("C", "Sintomas/Dano", "C-INS", "Instrumentacion", "C-LEC", "Lectura erratica", "A1,A2"),
    ("C", "Sintomas/Dano", "C-INS", "Instrumentacion", "C-DCA", "Descalibracion", "A1,A2"),
    # Type 5 (Causes) - Key entries
    ("5", "Causas", "5-OPE", "Operacionales", "5-SOC", "Sobrecarga operacional", "A1,A2"),
    ("5", "Causas", "5-OPE", "Operacionales", "5-MAO", "Mala operacion", "A1,A2"),
    ("5", "Causas", "5-MNT", "Mantenimiento", "5-FLB", "Falta lubricacion", "A1,A2"),
    ("5", "Causas", "5-MNT", "Mantenimiento", "5-MAL", "Mala alineacion", "A1,A2"),
    ("5", "Causas", "5-MNT", "Mantenimiento", "5-MPM", "Preventivo omitido", "A1,A2"),
    ("5", "Causas", "5-MNT", "Mantenimiento", "5-RPI", "Repuesto inadecuado", "A1,A2"),
    ("5", "Causas", "5-DIS", "Diseno", "5-SDI", "Subdimensionamiento", "A1,A2"),
    ("5", "Causas", "5-DIS", "Diseno", "5-MDI", "Material inadecuado", "A1,A2"),
    ("5", "Causas", "5-EXT", "Externas", "5-AMB", "Condicion ambiental", "A1,A2"),
    ("5", "Causas", "5-EXT", "Externas", "5-POL", "Polvo/Contaminacion externa", "A1,A2"),
]

for row in CATALOG_CODES:
    ws3.append(list(row))
auto_width(ws3)

# ============================================================
# SHEET 4: PRIORITY MAPPING SAP ↔ AMS
# ============================================================
ws4 = wb.create_sheet("Priority Mapping")
style_header(ws4, ["sap_code", "sap_description", "ams_code", "ams_description", "response_time", "priority_class", "notes"], fill=MAP_FILL)

PRIORITY_MAP = [
    ("I", "Inmediata", "1_EMERGENCY", "Emergency", "< 24 horas", "Z1", "Riesgo seguridad o parada planta. Atencion inmediata."),
    ("A", "Alta (2-6 dias)", "2_URGENT", "Urgent", "2-6 dias", "Z1", "Degradacion en progreso. Programar en proxima ventana."),
    ("M", "Media (7-14 dias)", "3_NORMAL", "Normal", "7-14 dias", "Z1", "Condicion requiere atencion planificada."),
    ("B", "Baja (>14 dias)", "4_PLANNED", "Planned", "> 14 dias", "Z1", "Puede esperar proxima parada programada."),
]

for row in PRIORITY_MAP:
    ws4.append(list(row))
auto_width(ws4)

# ============================================================
# SHEET 5: TEMPLATE INDEX (Metadata por plantilla)
# ============================================================
ws5 = wb.create_sheet("Template Index")
style_header(ws5, ["template_num", "filename", "description_es", "records", "sheets", "sap_transaction", "fase", "key_fields", "cross_references"], fill=CAT_FILL)

TEMPLATES = [
    ("01", "01_equipment_hierarchy.xlsx", "Jerarquia de Equipos (6 niveles + BOM)", "894", "3", "IE03/IL03", "1", "equipment_tag, equnr, tplnr, abc_criticality", "-"),
    ("02", "02_criticality_assessment.xlsx", "Evaluacion de Criticidad ABC 1/2/3", "465", "1", "IE02", "1", "equipment_tag, criticality_score, criticality_rank", "01"),
    ("03", "03_failure_modes.xlsx", "Modos de Falla FMECA", "274", "1", "-", "1", "equipment_tag, mechanism, cause, rpn_total", "01"),
    ("04", "04_measurement_points.xlsx", "Puntos de Medida (config CINI_PM_004)", "255", "1", "IK01", "2", "measurement_point_id, characteristic, unit_of_measure, limits", "01"),
    ("05", "05_work_packages.xlsx", "Paquetes de Trabajo", "235", "1", "IA05", "1", "work_package_id, equipment_tag, frequency, constraint", "01, 14"),
    ("06", "06_work_order_history.xlsx", "Historial OTs (PM01-PM07) con cierre", "200", "1", "IW38/IW39", "1", "aufnr, auart, equipment_tag, system_status, cause/solution/closure", "01"),
    ("07", "07_spare_parts_inventory.xlsx", "Inventario Repuestos (VED/FSN/ABC)", "201", "1", "MM60", "1", "material_code, description, quantity_on_hand, ved_class", "01"),
    ("08", "08_shutdown_calendar.xlsx", "Calendario Paradas", "221", "1", "-", "1", "shutdown_name, type, planned_start, planned_end", "-"),
    ("09", "09_workforce.xlsx", "Fuerza Laboral (turno 7x7)", "200", "1", "-", "1", "worker_id, name, specialty, shift, planning_group", "11"),
    ("10", "10_field_capture.xlsx", "Capturas de Campo", "220", "1", "-", "1", "capture_id, technician_id, equipment_tag, raw_text", "01, 09"),
    ("11", "11_work_centers.xlsx", "Maestro Puestos de Trabajo (CINI_PM_001)", "44", "4", "CR01", "2", "work_center_code, type, specialty, planning_group, cost_center", "20"),
    ("12", "12_planning_kpi_input.xlsx", "KPIs Planificacion (semanal x PG)", "208", "1", "-", "1", "planning_group, period, wo_planned, wo_completed", "-"),
    ("13", "13_de_kpi_input.xlsx", "KPIs Eliminacion Defectos (mensual x PG)", "204", "1", "-", "1", "planning_group, period, events_reported, events_closed", "-"),
    ("14", "14_maintenance_strategy.xlsx", "Estrategias Mantenimiento RCM", "284", "1", "-", "1", "equipment_tag, strategy_type, frequency, tactics_type", "01, 03"),
    ("15", "15_catalog_profiles.xlsx", "Catalogos y Perfiles (D/B/C/5)", "296", "3", "QS41", "2", "catalog_type, code, description, applicable_notification_type", "-"),
    ("16", "16_route_sheets.xlsx", "Hojas de Ruta Detalladas (CINI_PM_005/006)", "1015", "1", "IA05", "2", "route_sheet_id, equipment_tag, operation_number, work_center", "01, 11"),
    ("17", "17_maintenance_plans.xlsx", "Planes Mantenimiento Formales (CINI_PM_007)", "818", "1", "IP10/IP41", "2", "plan_id, equipment_tag, frequency, strategy", "01, 16"),
    ("18", "18_dms_maf_documents.xlsx", "Documentos MAF/DMS (CINI_PM_008)", "271", "1", "CV01N", "2", "document_id, equipment_tag, document_type, revision", "01"),
    ("19", "19_classification.xlsx", "Clasificacion SAP (CINI_PM_011)", "215", "1", "CL02", "2", "class_code, characteristic, equipment_tag", "01"),
    ("20", "20_financial_assignments.xlsx", "Asignaciones Financieras (CO/PS/FI-AA)", "142", "4", "KS01/CJ01/AS01", "2", "cost_center, wbs_element, fixed_asset, activity_type", "11"),
    ("21", "21_configuration_points.xlsx", "Puntos Configuracion (CONF_PM_001-058)", "58", "1", "SPRO", "2", "config_point_id, description, value", "-"),
    ("22", "22_org_structure_config.xlsx", "Estructura Organizacional", "94", "5", "SPRO", "2", "structure_indicator, number_ranges, installation_status, company_areas, planning_groups", "-"),
    ("23", "23_active_backlog.xlsx", "Backlog Activo (ordenes abiertas IW38)", "80", "1", "IW38", "3", "backlog_id, aufnr, auart, status, priority, equipment_tag", "01, 06"),
    ("24", "24_notifications.xlsx", "Notificaciones SAP (avisos IW28/IW29)", "220", "2", "IW28/IW29", "3", "qmnum, qmart, equipment_tag, damage_code, cause_code, system_status", "01, 06, 15, 23"),
    ("25", "25_measurement_documents.xlsx", "Documentos de Medicion (lecturas IK11)", "307", "1", "IK11/IK12", "3", "measurement_point_id, equipment_tag, measured_value, valuation_code", "01, 04, 09"),
    ("26", "26_time_confirmations.xlsx", "Confirmaciones de Tiempo (IW41/IW42)", "208", "1", "IW41/IW42", "4", "confirmation_id, aufnr, worker_id, actual_work_hours", "06, 09, 11"),
    ("27", "27_material_movements.xlsx", "Movimientos de Material (MB21/MIGO)", "200", "2", "MB21/MIGO/MB51", "4", "material_code, movement_type, quantity, aufnr", "06, 07, 20"),
    ("28", "28_equipment_bom.xlsx", "BOM de Equipos (IB01/CS01)", "246", "1", "IB01/CS01", "4", "equipment_tag, material_code, quantity, item_category, critical_spare", "01, 07"),
    ("29", "29_cost_history.xlsx", "Historial de Costos (IW39/KOB1)", "123", "2", "IW39/KOB1", "4", "aufnr, value_category, amount_usd, cost_center, wbs_element", "06, 20"),
    ("30", "30_reliability_data.xlsx", "Datos de Confiabilidad (TTF + Weibull)", "178", "2", "-", "4", "equipment_tag, time_to_failure_days, beta, eta, failure_pattern", "01, 03, 06"),
]

for row in TEMPLATES:
    ws5.append(list(row))
auto_width(ws5)

# ============================================================
# SHEET 6: FIELD GLOSSARY (Column header definitions)
# ============================================================
ws6 = wb.create_sheet("Field Glossary")
style_header(ws6, ["field_name", "sap_field", "description_es", "description_en", "data_type", "example", "used_in_templates"])

FIELD_GLOSSARY = [
    # --- SAP Equipment & Hierarchy Fields ---
    ("equipment_tag", "EQUNR (alias)", "Tag tecnico del equipo (formato OCP)", "Technical equipment tag (OCP format)", "String", "OCP-CON1-MSAG01", "01-30 (todos)"),
    ("equnr", "EQUNR", "Numero de equipo SAP (12 digitos)", "SAP Equipment Number (12 digits)", "String", "000000100001", "01, 02, 06, 23, 24, 26, 28"),
    ("eqktx", "EQKTX", "Texto corto / descripcion del equipo", "Equipment Short Text / Description", "String(40)", "Molino SAG 40x22 L1", "01, 02"),
    ("eqart", "EQART", "Tipo de equipo SAP (M=Maquina, Q=Inspeccion)", "SAP Equipment Type (M=Machine, Q=Inspection)", "String(1)", "M", "01"),
    ("eqart_desc", "-", "Descripcion del tipo de equipo", "Equipment type description", "String", "Maquinas", "01"),
    ("sap_func_loc", "TPLNR", "Ubicacion tecnica SAP (jerarquia)", "SAP Functional Location (hierarchy)", "String(40)", "02-01-01-CHAN01", "01-30 (todos)"),
    ("tplnr", "TPLNR", "Ubicacion tecnica SAP (nombre original SAP)", "SAP Technical Place Number", "String(40)", "02-01-01-CHAN01", "06"),
    ("fl_type", "FLTYP", "Tipo de ubicacion tecnica (M=estandar)", "Functional Location Type (M=standard)", "String(1)", "M", "01"),
    ("pltxt", "PLTXT", "Texto descriptivo de la ubicacion tecnica", "Functional Location Description Text", "String(40)", "Chancado Primario", "01"),
    ("level", "-", "Nivel jerarquico (1=Planta, 2=Area, 3=Sistema, 4=Equipo)", "Hierarchy Level (1=Plant to 4=Equipment)", "Integer", "4", "01"),
    ("parent_fl", "TPLMA", "Ubicacion tecnica padre (jerarquia superior)", "Parent Functional Location", "String(40)", "02-01-01", "01"),
    ("herst", "HERST", "Fabricante del equipo", "Equipment Manufacturer", "String(30)", "Metso", "01"),
    ("model", "-", "Modelo del equipo", "Equipment Model", "String", "C160", "01"),
    ("power_kw", "-", "Potencia nominal del equipo (kW)", "Rated Power (kW)", "Decimal", "5500", "01"),
    ("weight_kg", "-", "Peso del equipo (kg)", "Equipment Weight (kg)", "Decimal", "120000", "01"),
    ("abckz", "ABCKZ", "Indicador criticidad ABC (1=Alto, 2=Medio, 3=Bajo)", "ABC Criticality Indicator (1=High, 2=Med, 3=Low)", "String(1)", "1", "01, 02"),
    ("abckz_desc", "-", "Descripcion del nivel de criticidad", "Criticality level description", "String", "Alto", "01, 02"),
    ("stat", "STAT", "Estado del equipo (ACTIVO, INACTIVO)", "Equipment Status", "String", "ACTIVO", "01"),
    ("install_date", "INBDT", "Fecha de instalacion del equipo", "Equipment Installation Date", "Date", "2020-03-15", "01"),
    ("serial_number", "SERNR", "Numero de serie del equipo", "Equipment Serial Number", "String(30)", "SN-2020-001234", "01"),
    # --- Work Order Fields ---
    ("aufnr", "AUFNR", "Numero de orden de trabajo SAP", "SAP Work Order Number", "Integer", "1000001", "06, 23, 26, 27, 29"),
    ("auart", "AUART", "Tipo de orden (PM01/PM02/PM03/PM06/PM07)", "Order Type", "String(4)", "PM01", "06, 23, 26, 27, 29"),
    ("auart_desc", "-", "Descripcion del tipo de orden", "Order Type Description", "String", "Orden Mant. de Averia", "06, 23, 26"),
    ("erdat", "ERDAT", "Fecha de creacion de la orden", "Order Created Date", "Date", "2026-01-15", "06"),
    ("gstrp", "GSTRP", "Fecha inicio planificada (basica)", "Basic Start Date (planned)", "DateTime", "2026-01-20 08:00", "06"),
    ("gltrp", "GLTRP", "Fecha fin planificada (basica)", "Basic End Date (planned)", "DateTime", "2026-01-20 16:00", "06"),
    ("iedd", "IEDD", "Fecha inicio real de la orden", "Actual Start Date", "DateTime", "2026-01-20 09:00", "06"),
    ("iedt", "IEDT", "Fecha fin real de la orden", "Actual End Date", "DateTime", "2026-01-20 15:30", "06"),
    ("duration_hours", "-", "Duracion total de la intervencion (horas)", "Total intervention duration (hours)", "Decimal", "6.5", "06"),
    ("system_status", "STAT", "Status sistema SAP de la orden (ABIE/LIBE/NOTI/CTEC)", "SAP System Status of Order", "String", "CTEC", "06, 24"),
    ("description", "KTEXT", "Texto corto de la orden/aviso (max 72 chars SAP)", "Short Text (max 72 chars SAP limit)", "String(72)", "Reparar fuga aceite Sist Lubricacion", "06, 23"),
    ("failure_found", "-", "Hallazgo al intervenir (que se encontro)", "Finding during intervention (what was found)", "String", "Rodamiento colapsado", "06"),
    ("cause_description", "-", "Descripcion de la causa raiz", "Root cause description", "String", "Desgaste normal por horas operacion", "06"),
    ("solution_description", "-", "Accion correctiva ejecutada", "Corrective action taken", "String", "Se reemplazo componente danado", "06"),
    ("closure_comments", "-", "Notas de cierre y recomendaciones", "Closure notes and recommendations", "String", "Equipo entregado operativo", "06"),
    ("crew_size", "-", "Cantidad de trabajadores asignados", "Number of workers assigned", "Integer", "3", "06"),
    ("specialty", "-", "Especialidad del puesto trabajo (MEC/ELE/INS/LUB/SOL)", "Work Center specialty", "String(3)", "MEC", "06, 26"),
    ("synth_trace_id", "-", "ID trazabilidad datos sinteticos (interno)", "Synthetic data trace ID (internal)", "String", "S26-OT-0001-A1B2", "06"),
    # --- Priority Fields ---
    ("priokx", "PRIOKX", "Codigo prioridad SAP clase Z1 (I/A/M/B)", "SAP Priority Code class Z1", "String(1)", "A", "06"),
    ("priokx_desc", "-", "Descripcion de la prioridad", "Priority description", "String", "Alta (2-6 dias)", "06"),
    ("priority_class", "PRIOK", "Clase de prioridad SAP", "SAP Priority Class", "String(2)", "Z1", "06, 23, 24"),
    ("priority", "-", "Codigo prioridad (I/A/M/B o status backlog)", "Priority code", "String", "A", "23, 24"),
    ("priority_desc", "-", "Descripcion de la prioridad", "Priority description", "String", "Alta (2-6 dias)", "23, 24"),
    # --- Notification Fields ---
    ("ilart", "ILART", "Tipo de aviso/notificacion SAP (A1/A2/A3)", "SAP Notification Type (linked to order)", "String(2)", "A1", "06"),
    ("ilart_desc", "-", "Descripcion del tipo de aviso", "Notification Type description", "String", "Aviso de mantenimiento", "06"),
    ("qmnum", "QMNUM", "Numero de notificacion/aviso SAP", "SAP Notification Number", "Integer", "5000001", "06, 24"),
    ("qmart", "QMART", "Clase de aviso SAP (A1/A2/A3)", "SAP Notification Class", "String(2)", "A1", "24"),
    ("qmart_desc", "-", "Descripcion de la clase de aviso", "Notification Class description", "String", "Aviso de mantenimiento", "24"),
    ("notification_catalog", "-", "Catalogo de aviso asignado (M001-M003, P001-P002)", "Assigned notification catalog", "String", "M001", "06, 24"),
    ("notification_status_schema", "-", "Esquema status usuario del aviso (ZPM00001)", "Notification user status schema", "String", "ZPM00001", "06"),
    ("notif_user_status", "-", "Status usuario del aviso (APRO/RECH)", "Notification user status", "String(4)", "APRO", "06"),
    ("user_status_schema", "-", "Esquema status usuario (ZPM00001)", "User status schema", "String", "ZPM00001", "24"),
    ("user_status", "-", "Status usuario (APRO=Aprobado, RECH=Rechazado)", "User status", "String(4)", "APRO", "24"),
    ("reported_by", "-", "ID del trabajador que reporto", "Reporting worker ID", "String", "W-MEC-001", "24"),
    ("reported_date", "-", "Fecha en que se reporto el aviso", "Notification reported date", "Date", "2026-02-15", "24"),
    ("long_text", "-", "Texto largo del aviso (descripcion extendida)", "Notification long text (extended description)", "String", "Detalle: vibracion excesiva...", "24"),
    ("damage_code", "-", "Codigo de dano (catalogo B de template 15)", "Damage code (catalog B from template 15)", "String", "B-ROD", "24"),
    ("cause_code", "-", "Codigo de causa (catalogo 5 de template 15)", "Cause code (catalog 5 from template 15)", "String", "5-FLB", "24"),
    ("object_part_code", "-", "Codigo parte objeto (catalogo B)", "Object part code (catalog B)", "String", "B-MOT", "24"),
    ("linked_order", "-", "Numero de orden vinculada al aviso", "Linked order number", "Integer", "1900001", "24"),
    ("completed_date", "-", "Fecha completado (avisos cerrados)", "Completion date (closed notifications)", "Date", "2026-03-01", "24"),
    # --- Work Center & Organization Fields ---
    ("arbpl", "ARBPL", "Puesto de trabajo SAP (8 caracteres)", "SAP Work Center (8 characters)", "String(8)", "PASMEC01", "06"),
    ("supervisor_wc", "-", "Puesto de trabajo del supervisor", "Supervisor Work Center", "String(8)", "SPASMEC", "06, 23"),
    ("work_center", "ARBPL", "Puesto de trabajo responsable", "Responsible Work Center", "String(8)", "PASMEC01", "23-27"),
    ("planning_group", "INGRP", "Grupo de planificacion (P01/P02/P03)", "Planning Group", "String(3)", "P01", "06, 12, 13, 23-30"),
    ("planning_center", "IWERK", "Centro de planificacion SAP", "SAP Planning Center", "String(4)", "AN01", "06, 23, 24"),
    ("business_area", "GSBER", "Area de empresa (SEC/RIP/HUM)", "Business Area", "String(3)", "SEC", "06, 23, 24"),
    ("area", "-", "Nombre del area de proceso", "Process area name", "String", "Chancado", "01-30 (todos)"),
    ("subprocess", "-", "Nombre del subproceso", "Sub-process name", "String", "Chancado Primario", "06"),
    # --- Cost Fields ---
    ("cost_labour_usd", "-", "Costo mano de obra interna (USD)", "Internal labour cost (USD)", "Decimal", "2340.00", "06"),
    ("cost_materials_usd", "-", "Costo materiales y repuestos (USD)", "Materials & spare parts cost (USD)", "Decimal", "5600.00", "06"),
    ("cost_external_usd", "-", "Costo servicios externos (USD)", "External services cost (USD)", "Decimal", "12000.00", "06"),
    ("value_cat_labour", "-", "Categoria valor mano obra (ZMANT001)", "Labour value category", "String", "ZMANT001", "06"),
    ("value_cat_materials", "-", "Categoria valor materiales (ZMANT002)", "Materials value category", "String", "ZMANT002", "06"),
    ("value_cat_external", "-", "Categoria valor servicios ext. (ZMANT003)", "External services value category", "String", "ZMANT003", "06"),
    ("value_category", "-", "Categoria de valor (ZMANT001/002/003)", "Value category", "String", "ZMANT001", "29"),
    ("amount_usd", "-", "Monto en dolares", "Amount in USD", "Decimal", "3500.00", "29"),
    ("cost_center", "-", "Centro de costo SAP", "SAP Cost Center", "String", "CC-PL-MEC-SEC", "27, 29"),
    ("wbs_element", "-", "Elemento PEP/WBS del proyecto", "WBS / PEP Element", "String", "WBS-PL-SEC-MNT-2026", "29"),
    ("cost_element", "-", "Elemento de costo SAP (cuenta CO)", "SAP Cost Element (CO account)", "String", "621000", "29"),
    ("budget_usd", "-", "Presupuesto asignado (USD)", "Assigned budget (USD)", "Decimal", "45000.00", "29"),
    ("variance_usd", "-", "Varianza costo vs presupuesto (USD)", "Cost vs budget variance (USD)", "Decimal", "-2300.00", "29"),
    ("variance_pct", "-", "Varianza porcentual (%)", "Variance percentage (%)", "Decimal", "-5.1", "29"),
    ("unit_cost_usd", "-", "Costo unitario del material (USD)", "Unit cost of material (USD)", "Decimal", "1250.00", "27, 28"),
    ("total_cost_usd", "-", "Costo total (cantidad x unitario)", "Total cost (qty x unit cost)", "Decimal", "3750.00", "27"),
    # --- Material & BOM Fields ---
    ("material_code", "MATNR (alias)", "Codigo material sintetico (S26-MAT-NNNN)", "Synthetic material code", "String", "S26-MAT-0001", "07, 27, 28"),
    ("sap_material_number", "MATNR", "Numero de material SAP (12 digitos)", "SAP Material Number (12 digits)", "String(12)", "000000010001", "27, 28"),
    ("component_desc", "-", "Descripcion del componente/material", "Component/Material description", "String", "Rodamiento SKF 22328", "28"),
    ("quantity", "-", "Cantidad (en la unidad de medida)", "Quantity (in unit of measure)", "Decimal", "2", "27, 28"),
    ("unit_of_measure", "MEINS", "Unidad de medida (EA/KG/L/M/DR/PL/BX/MT)", "Unit of Measure", "String(3)", "EA", "07, 25, 27, 28"),
    ("item_number", "-", "Numero posicion BOM (incrementos de 10)", "BOM Item Number (increments of 10)", "String(4)", "0010", "28"),
    ("item_category", "-", "Categoria del item (SPARE/WEAR/CONSUMABLE)", "Item category", "String", "SPARE", "28"),
    ("critical_spare", "-", "Es repuesto critico (TRUE/FALSE)", "Is critical spare (TRUE/FALSE)", "Boolean", "TRUE", "28"),
    ("component_category", "-", "Categoria componente BOM (L=stock/N=non-stock/D=doc)", "BOM Component Category", "String(1)", "L", "28"),
    ("movement_type", "BWART", "Tipo de movimiento SAP (261/262/101/311)", "SAP Movement Type", "String(3)", "261", "27"),
    ("storage_location", "LGORT", "Ubicacion almacenamiento", "Storage Location", "String", "ALM-CENTRAL", "27"),
    ("reservation_id", "-", "Numero de reserva de material", "Material reservation number", "String", "S26-RES-0001", "27"),
    ("movement_doc_id", "-", "Numero documento movimiento", "Movement document number", "String", "S26-MOV-0001", "27"),
    ("lead_time_days", "-", "Tiempo de entrega (dias)", "Lead time (days)", "Integer", "45", "07, 28"),
    # --- Measurement Fields ---
    ("measurement_point_id", "POINT", "ID punto de medida SAP (12 digitos)", "SAP Measurement Point ID (12 digits)", "String(12)", "000000000001", "04, 25"),
    ("measurement_doc_id", "-", "ID documento de medicion", "Measurement document ID", "String", "S26-MDOC-0001", "25"),
    ("characteristic", "ATNAM", "Caracteristica medida (TEMP_BRG_DE, VIB_RAD_DE...)", "Measured characteristic", "String", "TEMP_BRG_DE", "04, 25"),
    ("characteristic_desc", "-", "Descripcion de la caracteristica medida", "Characteristic description", "String", "Temperatura rodamiento LA", "25"),
    ("reading_date", "-", "Fecha de la lectura/medicion", "Reading/measurement date", "Date", "2026-02-15", "25"),
    ("reading_time", "-", "Hora de la lectura (turno 08:00-20:00)", "Reading time (shift 08:00-20:00)", "Time", "10:30", "25"),
    ("measured_value", "-", "Valor medido", "Measured value", "Decimal", "72.5", "25"),
    ("counter_reading", "-", "Lectura del contador (horometro, ciclos)", "Counter reading (hours, cycles)", "Integer", "45230", "25"),
    ("valuation_code", "-", "Codigo valoracion (1=OK, 2=Warning, 3=Alarm)", "Valuation code (1=OK, 2=Warning, 3=Alarm)", "Integer", "1", "25"),
    ("lower_limit", "-", "Limite inferior de medicion", "Lower measurement limit", "Decimal", "0", "04, 25"),
    ("upper_limit", "-", "Limite superior de medicion", "Upper measurement limit", "Decimal", "95", "04, 25"),
    ("target_value", "-", "Valor objetivo de medicion", "Target measurement value", "Decimal", "65", "04, 25"),
    ("is_counter", "-", "Es tipo contador (TRUE/FALSE)", "Is counter type (TRUE/FALSE)", "Boolean", "FALSE", "04, 25"),
    # --- Time Confirmation Fields ---
    ("confirmation_id", "-", "ID de confirmacion de tiempo", "Time confirmation ID", "String", "S26-CONF-0001", "26"),
    ("operation_number", "VORNR", "Numero de operacion (0010, 0020, 0030...)", "Operation Number (increments of 10)", "String(4)", "0010", "26"),
    ("sub_operation", "-", "Sub-operacion (0000 por defecto)", "Sub-operation (0000 default)", "String(4)", "0000", "26"),
    ("confirmation_type", "-", "Tipo confirmacion (INDIVIDUAL/COLLECTIVE)", "Confirmation type", "String", "INDIVIDUAL", "26"),
    ("worker_id", "-", "ID del trabajador que ejecuto", "Executing worker ID", "String", "W-MEC-001", "09, 25, 26"),
    ("worker_name", "-", "Nombre del trabajador", "Worker name", "String", "Tecnico MEC 001", "26"),
    ("actual_work_hours", "-", "Horas netas trabajadas", "Actual net work hours", "Decimal", "4.5", "26"),
    ("break_hours", "-", "Horas de descanso (1h turno 7x7)", "Break hours (1h per 7x7 shift)", "Decimal", "1.0", "26"),
    ("travel_time_hours", "-", "Tiempo de traslado al equipo (horas)", "Travel time to equipment (hours)", "Decimal", "0.5", "26"),
    ("setup_time_hours", "-", "Tiempo de preparacion (horas)", "Setup time (hours)", "Decimal", "1.0", "26"),
    ("total_duration_hours", "-", "Duracion total (trabajo+descanso+traslado+prep)", "Total duration (work+break+travel+setup)", "Decimal", "7.0", "26"),
    ("final_confirmation", "-", "Es ultima confirmacion de la orden (TRUE/FALSE)", "Is final confirmation for order (TRUE/FALSE)", "Boolean", "TRUE", "26"),
    ("system_status_after", "-", "Status sistema despues de confirmacion (PCNF/CNF)", "System status after confirmation", "String(4)", "CNF", "26"),
    # --- Reliability / Weibull Fields ---
    ("ttf_record_id", "-", "ID registro time-to-failure", "Time-to-failure record ID", "String", "S26-TTF-0001", "30"),
    ("failure_event_id", "-", "ID del evento de falla (ref OT PM01)", "Failure event ID (ref PM01 order)", "Integer", "1000005", "30"),
    ("failure_date", "-", "Fecha del evento de falla", "Failure event date", "Date", "2026-02-10", "30"),
    ("previous_failure_date", "-", "Fecha falla previa en mismo equipo", "Previous failure date on same equipment", "Date", "2025-11-15", "30"),
    ("time_to_failure_days", "-", "Dias entre fallas (TTF)", "Days between failures (TTF)", "Integer", "87", "30"),
    ("operating_hours", "-", "Horas operacion entre fallas", "Operating hours between failures", "Integer", "1740", "30"),
    ("suspension", "-", "Dato censurado (TRUE=aun funcionando)", "Censored data (TRUE=still running)", "Boolean", "FALSE", "30"),
    ("failure_mechanism", "-", "Mecanismo de falla (18 valores validos)", "Failure mechanism (18 valid values)", "String", "WEARS", "03, 30"),
    ("failure_cause", "-", "Causa de falla (44 valores validos)", "Failure cause (44 valid values)", "String", "ABRASION", "03, 30"),
    ("failure_pattern", "-", "Patron falla Nowlan&Heap (A-F)", "Failure pattern Nowlan&Heap (A-F)", "String", "B_AGE", "30"),
    ("maintainable_item", "-", "Item mantenible que fallo", "Maintainable item that failed", "String", "Rodamiento", "30"),
    ("repair_duration_hours", "-", "Duracion de la reparacion (horas)", "Repair duration (hours)", "Decimal", "12.5", "30"),
    ("beta", "-", "Parametro forma Weibull (>1=desgaste, <1=infant, =1=aleatorio)", "Weibull shape parameter", "Decimal", "2.1", "30"),
    ("eta_days", "-", "Parametro escala Weibull (vida caracteristica, dias)", "Weibull scale parameter (characteristic life, days)", "Decimal", "180.5", "30"),
    ("gamma_days", "-", "Parametro ubicacion Weibull (periodo libre falla, dias)", "Weibull location parameter (failure-free period, days)", "Decimal", "10.0", "30"),
    ("r_squared", "-", "Bondad de ajuste del modelo Weibull", "Weibull model goodness of fit", "Decimal", "0.94", "30"),
    ("mtbf_days", "-", "Tiempo medio entre fallas (dias)", "Mean Time Between Failures (days)", "Decimal", "162.5", "30"),
    ("recommended_interval_days", "-", "Intervalo optimo recomendado (dias)", "Recommended optimal interval (days)", "Integer", "126", "30"),
    ("equipment_class", "-", "Clase de equipo para agrupacion (ej: BOMBA_PULPA)", "Equipment class for grouping", "String", "BOMBA_PULPA", "30"),
    ("sample_size", "-", "Tamano de muestra para calculo Weibull", "Sample size for Weibull calculation", "Integer", "12", "30"),
    ("confidence_level", "-", "Nivel de confianza del calculo", "Calculation confidence level", "Decimal", "0.90", "30"),
    # --- Backlog-specific Fields ---
    ("backlog_id", "-", "ID del item de backlog", "Backlog item ID", "String", "S26-BKL-0001", "23"),
    ("work_request_id", "-", "ID de la solicitud de trabajo vinculada", "Linked work request ID", "String", "S26-WR-0001", "23"),
    ("status", "-", "Estado del item (backlog, reserva, etc.)", "Item status", "String", "AWAITING_MATERIALS", "23, 27"),
    ("blocking_reason", "-", "Razon de bloqueo del item backlog", "Backlog item blocking reason", "String", "Repuesto en transito", "23"),
    ("created_date", "-", "Fecha de creacion del registro", "Record creation date", "Date", "2026-02-01", "23"),
    ("age_days", "-", "Antiguedad del item en dias", "Item age in days", "Integer", "45", "23"),
    ("estimated_duration_hours", "-", "Duracion estimada de la intervencion (horas)", "Estimated intervention duration (hours)", "Decimal", "8.0", "23"),
    ("required_specialties", "-", "Especialidades requeridas (CSV: MEC,ELE,INS)", "Required specialties (CSV)", "String", "MEC,ELE", "23"),
    ("materials_ready", "-", "Materiales disponibles (TRUE/FALSE)", "Materials available (TRUE/FALSE)", "Boolean", "TRUE", "23"),
    ("shutdown_required", "-", "Requiere parada de equipo (TRUE/FALSE)", "Equipment shutdown required (TRUE/FALSE)", "Boolean", "FALSE", "23"),
    ("groupable", "-", "Agrupable en paquete de trabajo (TRUE/FALSE)", "Groupable in work package (TRUE/FALSE)", "Boolean", "TRUE", "23"),
    ("group_id", "-", "ID del grupo/paquete de trabajo", "Work package group ID", "String", "WP-GRP-0005", "23"),
    # --- Generic / Common Fields ---
    ("period", "-", "Periodo contable (YYYY-MM)", "Accounting period", "String", "2026-02", "12, 13, 29"),
    ("recorded_by", "-", "ID del tecnico que registro la lectura", "Recording technician ID", "String", "W-INS-003", "25"),
    ("posting_date", "-", "Fecha de contabilizacion", "Posting date", "Date", "2026-02-15", "27, 29"),
    ("settlement_rule", "-", "Regla de liquidacion (FULL/PERIODIC)", "Settlement rule", "String", "FULL", "29"),
    ("pm07_activity_class", "-", "Clase actividad PM07 (RP1=Mayor, RP2=Menor)", "PM07 activity class (RP1=Major, RP2=Minor)", "String(3)", "RP1", "06"),
    ("pm07_activity_desc", "-", "Descripcion clase actividad PM07", "PM07 activity class description", "String", "Reparacion Mayor", "06"),
    ("bom_id", "-", "ID de lista de materiales", "Bill of Materials ID", "String", "S26-BOM-0001", "28"),
    ("bom_level", "-", "Nivel de BOM (1=superior)", "BOM level (1=top level)", "Integer", "1", "28"),
    ("change_frequency", "-", "Frecuencia de cambio (HIGH/MEDIUM/LOW)", "Change frequency", "String", "MEDIUM", "28"),
    ("applicable_system", "-", "Sistema aplicable (LUBE/HIDR/TRAN/MECA/SELL/ELEC)", "Applicable system", "String", "MECA", "28"),
    ("currency", "-", "Moneda (USD)", "Currency", "String(3)", "USD", "29"),
    ("text", "-", "Texto libre / notas de la operacion", "Free text / operation notes", "String", "Trabajo MEC en MSAG01", "26"),
]

for row in FIELD_GLOSSARY:
    ws6.append(list(row))
auto_width(ws6)

# ============================================================
# SAVE
# ============================================================
path = os.path.join(OUT, "00_data_dictionary.xlsx")
wb.save(path)
print(f"\n  -> Saved: 00_data_dictionary.xlsx")

total_codes = len(SAP_CODES) + len(ORG_CODES) + len(CATALOG_CODES) + len(PRIORITY_MAP) + len(TEMPLATES) + len(FIELD_GLOSSARY)
print(f"     Sheets: 6 (SAP Codes, Organizational, Catalogs, Priority Mapping, Template Index, Field Glossary)")
print(f"     Total entries: {total_codes} ({len(FIELD_GLOSSARY)} field definitions)")
print("=" * 70)
