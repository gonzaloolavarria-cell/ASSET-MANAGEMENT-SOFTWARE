# -*- coding: utf-8 -*-
"""
=============================================================================
  DB_AM_OCP_SYNTHETIC_2026 — Generador de Dataset Sintetico v2
  Planta Concentradora OCP
  ALINEADO ESTRICTAMENTE al Blueprint AMSA_BBP_PM_04_Rev_0
=============================================================================
  Correcciones vs v1:
    - Centro Planificacion: AN01 (Blueprint)
    - Grupos Planificacion: P01 (Area Seca), P02 (Area Ripio), P03 (Area Humeda)
    - Tipos OT: PM01, PM02, PM03, PM06, PM07 (NO PM04 — no existe en Blueprint)
    - Criticidad ABC: 1=Alto, 2=Medio, 3=Bajo (numerico Blueprint)
    - Status Usuario: ZPM00001 con APRO/RECH (notificaciones)
    - Status Orden: ABIE/LIBE/NOTI/CTEC (SAP estandar)
    - Tipos Equipo: M (Maquinas), Q (Inspeccion/Medida)
    - Jerarquia 6 niveles completa (mask NN-NN-NN-AAAANN-XXXX-XXXX)
    - Turnos: 7x7 rotacion (Viernes-Jueves) 08:00-20:00
    - Puestos de trabajo: formato Blueprint 8-char con AS/AH/AR
    - Clases actividad PM07: RP1/RP2
    - 200+ registros por plantilla donde faltaban
=============================================================================
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import hashlib
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
random.seed(2026)

OUT = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# CONSTANTS — ESTRICTAMENTE del Blueprint AMSA_BBP_PM_04_Rev_0
# ============================================================

# Centro de Planificacion (Blueprint: AN01 para Antucoya, 4-char CHAR)
PLANNING_CENTER = "AN01"
PLANT_ID = "OCP-CON1"
PLANT_DESC = "Planta Concentradora OCP"

# Grupos de Planificacion PLANTA (Blueprint: 3 chars, P01-P03)
# Mapeo concentradora → Blueprint:
#   P01 (Area Seca) ← Chancado, Filtrado/Secado (procesos secos)
#   P02 (Area Ripio) ← Molienda (transicion seco-humedo)
#   P03 (Area Humeda) ← Flotacion, Espesado (procesos humedos)
PLANNING_GROUPS = {
    "P01": "Area seca",
    "P02": "Area Ripio",
    "P03": "Area Humeda",
}

# Areas de Empresa (Blueprint: 3 chars)
BUSINESS_AREAS = {
    "SEC": "Area seca",
    "RIP": "Area Ripio",
    "HUM": "Area Humeda",
}

# Mapeo de procesos concentradora → grupos planificacion Blueprint
PROCESS_TO_PG = {
    "Chancado Primario": "P01",
    "Chancado Secundario": "P01",
    "Chancado Terciario": "P01",
    "Clasificacion y Harneo": "P01",
    "Molienda SAG": "P02",
    "Molienda Bolas": "P02",
    "Bombeo Pulpa Molienda": "P02",
    "Clasificacion Hidrociclones": "P02",
    "Flotacion Rougher": "P03",
    "Flotacion Cleaner-Scavenger": "P03",
    "Acondicionamiento Reactivos": "P03",
    "Espesador Concentrado": "P03",
    "Espesador Relaves": "P03",
    "Filtro Prensa": "P01",
    "Secado y Despacho": "P01",
    "Servicios Planta": "P03",
}

PROCESS_TO_BA = {
    "P01": "SEC",
    "P02": "RIP",
    "P03": "HUM",
}

# Puestos de Trabajo (Blueprint: 8-char AAAAAANN)
# Char 1: P (Planta)
# Char 2-3: Area (AS=Area Seca, AR=Area Ripio, AH=Area Humeda)
# Char 4-6: Especialidad (MEC, ELE, INS, LUB, SIN, DCS)
# Char 7-8: Secuencial (01)
SPECIALTIES = ["MEC", "ELE", "INS"]  # Blueprint internos planta
SPECIALTIES_EXT = ["SOL"]  # Soldadura es externo en Blueprint

WORK_CENTERS = {
    "PASMEC01": "Planta Area Seca Mecanico 01",
    "PASELE01": "Planta Area Seca Electrico 01",
    "PASINS01": "Planta Area Seca Instrumentista 01",
    "PASLUB01": "Planta Area Seca Lubricacion 01",
    "PARMEC01": "Planta Area Ripio Mecanico 01",
    "PARELE01": "Planta Area Ripio Electrico 01",
    "PARINS01": "Planta Area Ripio Instrumentista 01",
    "PAHMEC01": "Planta Area Humeda Mecanico 01",
    "PAHELE01": "Planta Area Humeda Electrico 01",
    "PAHINS01": "Planta Area Humeda Instrumentista 01",
    "PSHSIN01": "Planta Sintomatico 01",
    "PSHDCS01": "Planta DCS/Automatizacion 01",
}

# Work centers por grupo planificacion
PG_TO_WC = {
    "P01": ["PASMEC01", "PASELE01", "PASINS01", "PASLUB01"],
    "P02": ["PARMEC01", "PARELE01", "PARINS01"],
    "P03": ["PAHMEC01", "PAHELE01", "PAHINS01"],
}

# Supervisor Work Centers (Blueprint: 7-char SAAAXXXX)
SUPERVISOR_WCS = {
    "SPASMEC": "Supervisor Planta Area Seca Mecanico",
    "SPASELE": "Supervisor Planta Area Seca Electrico",
    "SPARELE": "Supervisor Planta Area Ripio Electrico",
    "SPARINS": "Supervisor Planta Area Ripio Instrumentista",
    "SPAHMEC": "Supervisor Planta Area Humeda Mecanico",
    "SPAHELE": "Supervisor Planta Area Humeda Electrico",
    "SPINSTR": "Supervisor Instrumentacion",
    "SPDCS_A": "Supervisor DCS/Automatizacion",
    "SPSINTO": "Supervisor Sintomatico",
    "SPING01": "Supervisor Ingenieria 01",
}

# Tipos de Orden (Blueprint exacto — NO existe PM04)
ORDER_TYPES = {
    "PM01": {"desc": "Orden Mant. de Averia", "range_start": 1000000, "range_end": 1999999},
    "PM02": {"desc": "Orden Mant. Preventivo", "range_start": 2000000, "range_end": 2999999},
    "PM03": {"desc": "Orden de Solicitud de Mant.", "range_start": 3000000, "range_end": 3999999},
    "PM06": {"desc": "Orden de Inversion", "range_start": 6000000, "range_end": 6999999},
    "PM07": {"desc": "Orden de Reparacion de Componentes", "range_start": 7000000, "range_end": 7999999},
}

# Clases de Actividad PM07 (Blueprint)
PM07_ACTIVITY = {"RP1": "Reparacion componentes mayores", "RP2": "Reparacion componentes menores"}

# Tipos de Notificacion (Blueprint)
NOTIF_TYPES = {
    "A1": {"desc": "Aviso de mantenimiento", "range": "1-4999999",
           "catalogs": {"M001": "Solicitud de mantenimiento", "M002": "Averia", "M003": "Reparacion de componentes"}},
    "A2": {"desc": "Aviso predictivo e ingenieria", "range": "1-4999999",
           "catalogs": {"P001": "Predictivo", "P002": "Ingenieria"}},
    "A3": {"desc": "Aviso plan preventivo", "range": "1-4999999", "catalogs": {}},
}

# Esquema Status Usuario ZPM00001 (Blueprint — para avisos A1/A2)
USER_STATUS_SCHEMA = "ZPM00001"
USER_STATUSES_NOTIF = {"APRO": "Aprobado", "RECH": "Rechazado"}

# Status de Sistema Orden SAP estandar
SYSTEM_STATUSES_ORDER = ["ABIE", "LIBE", "NOTI", "CTEC"]
# ABIE = Abierto, LIBE = Liberado, NOTI = Notificado, CTEC = Cerrado tecnico

# Prioridad clase Z1 (Blueprint exacto)
PRIORITY_CLASS = "Z1"
PRIORITIES = {
    "I": "Inmediata",
    "A": "Alta (2-6 dias)",
    "M": "Media (7-14 dias)",
    "B": "Baja (Mayor a 14 dias)",
}

# Tipos de Equipo SAP (Blueprint)
EQUIPMENT_TYPES_SAP = {"M": "Maquinas", "Q": "Medio de inspeccion y medida"}

# Criticidad ABC (Blueprint: numerico 1/2/3)
ABC_CRITICALITY = {"1": "Alto", "2": "Medio", "3": "Bajo"}

# Tipo Ubicacion Tecnica (Blueprint)
FUNC_LOC_TYPE = "M"  # Sistema tecnico - Estandar

# Categorias de Valor (Blueprint exacto)
VALUE_CATEGORIES = {
    "ZMANT001": "Mano de obra interna",
    "ZMANT002": "Materiales y repuestos",
    "ZMANT003": "Servicios externos",
}

# Hojas de Ruta (Blueprint)
ROUTE_SHEET_TYPE = "A"  # Instruccion de mantenimiento
ROUTE_SHEET_GROUPS = {"PLA": "Planta"}

# Plan de Mantenimiento (Blueprint)
MAINT_PLAN_TYPE = "PM"

# Turnos Planta (Blueprint: 7x7 rotacion Viernes-Jueves, 08:00-20:00, 1hr break)
SHIFT_PLANTA = {
    "code": "7X7",
    "desc": "Turno 7x7 Planta (Viernes-Jueves)",
    "start": "08:00",
    "end": "20:00",
    "break_hours": 1,
    "effective_hours": 11,
    "rotation": "Viernes a Jueves",
}

# Mano de Obra (categorias y tarifas AMSA)
LABOR_CATEGORIES = {
    "MEC-I": {"desc": "Mecanico Nivel I", "rate_usd": 35.0},
    "MEC-II": {"desc": "Mecanico Nivel II", "rate_usd": 45.0},
    "MEC-III": {"desc": "Mecanico Senior", "rate_usd": 58.0},
    "ELE-I": {"desc": "Electricista Nivel I", "rate_usd": 38.0},
    "ELE-II": {"desc": "Electricista Nivel II", "rate_usd": 48.0},
    "ELE-III": {"desc": "Electricista Senior", "rate_usd": 62.0},
    "INS-I": {"desc": "Instrumentista Nivel I", "rate_usd": 40.0},
    "INS-II": {"desc": "Instrumentista Nivel II", "rate_usd": 52.0},
    "SOL-I": {"desc": "Soldador Nivel I", "rate_usd": 36.0},
    "SOL-II": {"desc": "Soldador Certificado", "rate_usd": 50.0},
    "LUB-I": {"desc": "Lubricador Nivel I", "rate_usd": 32.0},
    "SIN-I": {"desc": "Tecnico Sintomatico I", "rate_usd": 55.0},
    "DCS-I": {"desc": "Tecnico DCS/Automatizacion I", "rate_usd": 58.0},
    "SUP": {"desc": "Supervisor Mantenimiento", "rate_usd": 72.0},
}


# ============================================================
# JERARQUIA FUNCIONAL — 6 NIVELES Blueprint
# Estructura combinada:
#   CORPO: AMSA-OCP (Corporativo)
#   MANTE: NN-NN-NN-AAAANN-XXXX-XXXX
#     Nivel 1 (NN): Operacion — 02 = Planta
#     Nivel 2 (NN): Proceso
#     Nivel 3 (NN): Subproceso / Grupo Maquinaria
#     Nivel 4 (AAAANN): Maquinaria — 4 alpha + 2 numeric
#     Nivel 5 (XXXX): Sistema — 4 alphanum
#     Nivel 6 (XXXX): Equipo — 4 alphanum
# ============================================================

AREAS = {
    "01": {
        "name": "Chancado",
        "pg": "P01",
        "subprocesses": {
            "01": {
                "name": "Chancado Primario",
                "machines": {
                    "CHAN01": {"desc": "Chancador Giratorio 60x113", "sap_type": "M", "mfr": "FLSmidth", "model": "60-113 MK-III", "kw": 750, "kg": 485000, "abc": "1",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite", "ENAC": "Enfriador aceite"}},
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "ACUM": "Acumulador", "VHID": "Valvula hidraulica"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "PION": "Pinon transmision", "ACOPL": "Acoplamiento"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"MANT": "Manto", "CONC": "Concavo", "ARMD": "Armadura"}},
                        }},
                    "CHAN02": {"desc": "Chancador Giratorio 60x113 StBy", "sap_type": "M", "mfr": "FLSmidth", "model": "60-113 MK-III", "kw": 750, "kg": 485000, "abc": "2",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite"}},
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "ACUM": "Acumulador"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                        }},
                    "ALIM01": {"desc": "Alimentador Apron 01", "sap_type": "M", "mfr": "FLSmidth", "model": "AF-2400", "kw": 150, "kg": 45000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor", "CADN": "Cadena"}},
                            "ESTR": {"desc": "Estructura", "equips": {"BAST": "Bastidor", "PLAN": "Planchas transporte"}},
                        }},
                    "ALIM02": {"desc": "Alimentador Apron 02", "sap_type": "M", "mfr": "FLSmidth", "model": "AF-2400", "kw": 150, "kg": 45000, "abc": "2",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                        }},
                }},
            "02": {
                "name": "Chancado Secundario",
                "machines": {
                    "COSE01": {"desc": "Chancador Cono Secundario HP800 L1", "sap_type": "M", "mfr": "Metso", "model": "HP800", "kw": 600, "kg": 65000, "abc": "1",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite", "ENAC": "Enfriador aceite"}},
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "ACUM": "Acumulador"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                        }},
                    "COSE02": {"desc": "Chancador Cono Secundario HP800 L2", "sap_type": "M", "mfr": "Metso", "model": "HP800", "kw": 600, "kg": 65000, "abc": "1",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite"}},
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
            "03": {
                "name": "Chancado Terciario",
                "machines": {
                    "COTE01": {"desc": "Chancador Cono Terciario HP500 L1", "sap_type": "M", "mfr": "Metso", "model": "HP500", "kw": 400, "kg": 42000, "abc": "1",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite"}},
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                        }},
                    "COTE02": {"desc": "Chancador Cono Terciario HP500 L2", "sap_type": "M", "mfr": "Metso", "model": "HP500", "kw": 400, "kg": 42000, "abc": "2",
                        "systems": {
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
            "04": {
                "name": "Clasificacion y Harneo",
                "machines": {
                    "HARV01": {"desc": "Harnero Vibratorio 01", "sap_type": "M", "mfr": "Metso", "model": "TS-8203", "kw": 90, "kg": 32000, "abc": "1",
                        "systems": {
                            "VIBR": {"desc": "Mecanismo Vibrador", "equips": {"EXCI": "Excitador", "RODA": "Rodamientos", "ACOPL": "Acoplamiento"}},
                            "MALL": {"desc": "Mallas y Paneles", "equips": {"MALL": "Mallas acero", "POLI": "Paneles poliuretano"}},
                            "ESTR": {"desc": "Estructura", "equips": {"BAST": "Bastidor", "RESO": "Resortes"}},
                        }},
                    "HARV02": {"desc": "Harnero Vibratorio 02", "sap_type": "M", "mfr": "Metso", "model": "TS-8203", "kw": 90, "kg": 32000, "abc": "1",
                        "systems": {
                            "VIBR": {"desc": "Mecanismo Vibrador", "equips": {"EXCI": "Excitador", "RODA": "Rodamientos"}},
                            "MALL": {"desc": "Mallas y Paneles", "equips": {"MALL": "Mallas acero"}},
                        }},
                    "HARV03": {"desc": "Harnero Vibratorio 03", "sap_type": "M", "mfr": "Metso", "model": "TS-8203", "kw": 90, "kg": 32000, "abc": "2",
                        "systems": {
                            "VIBR": {"desc": "Mecanismo Vibrador", "equips": {"EXCI": "Excitador"}},
                        }},
                    "CORR01": {"desc": "Correa Transportadora CV-001", "sap_type": "M", "mfr": "Continental", "model": "CV-1200", "kw": 250, "kg": 28000, "abc": "1",
                        "systems": {
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor principal", "VFRE": "Variador frecuencia"}},
                            "POLN": {"desc": "Polines", "equips": {"PCAR": "Polines carga", "PRET": "Polines retorno", "PIMP": "Polines impacto"}},
                            "BAND": {"desc": "Banda Transportadora", "equips": {"BAND": "Banda", "RASP": "Raspador", "GUIT": "Guia lateral"}},
                        }},
                    "CORR02": {"desc": "Correa Transportadora CV-002", "sap_type": "M", "mfr": "Continental", "model": "CV-1200", "kw": 250, "kg": 28000, "abc": "2",
                        "systems": {
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor principal"}},
                            "BAND": {"desc": "Banda Transportadora", "equips": {"BAND": "Banda"}},
                        }},
                    "CORR03": {"desc": "Correa Transportadora CV-003", "sap_type": "M", "mfr": "Continental", "model": "CV-900", "kw": 150, "kg": 18000, "abc": "2",
                        "systems": {
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor principal"}},
                            "BAND": {"desc": "Banda Transportadora", "equips": {"BAND": "Banda"}},
                        }},
                }},
        },
    },
    "02": {
        "name": "Molienda",
        "pg": "P02",
        "subprocesses": {
            "01": {
                "name": "Molienda SAG",
                "machines": {
                    "MSAG01": {"desc": "Molino SAG 40x22 L1", "sap_type": "M", "mfr": "FLSmidth", "model": "SAG 40x22", "kw": 22000, "kg": 1200000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision GMD", "equips": {"GMDM": "Motor GMD", "GMDE": "Estator GMD", "GMDR": "Rotor GMD", "CONV": "Convertidor frecuencia"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite", "ENAC": "Enfriador aceite", "TANK": "Tanque aceite"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"RSHL": "Revest shell", "RLIF": "Lifters", "RPLA": "Revest placas", "RGRA": "Grizzly descarga"}},
                            "HIDR": {"desc": "Sist Hidraulico Trunion", "equips": {"BHID": "Bomba hidraulica", "DESM": "Descanso movil", "DESF": "Descanso fijo"}},
                        }},
                    "MSAG02": {"desc": "Molino SAG 40x22 L2", "sap_type": "M", "mfr": "FLSmidth", "model": "SAG 40x22", "kw": 22000, "kg": 1200000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision GMD", "equips": {"GMDM": "Motor GMD", "GMDE": "Estator GMD", "CONV": "Convertidor frecuencia"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite", "ENAC": "Enfriador aceite"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"RSHL": "Revest shell", "RLIF": "Lifters", "RGRA": "Grizzly descarga"}},
                            "HIDR": {"desc": "Sist Hidraulico Trunion", "equips": {"BHID": "Bomba hidraulica", "DESM": "Descanso movil"}},
                        }},
                }},
            "02": {
                "name": "Molienda Bolas",
                "machines": {
                    "MBOL01": {"desc": "Molino Bolas 26x40.5 L1", "sap_type": "M", "mfr": "FLSmidth", "model": "BM 26x40.5", "kw": 16400, "kg": 900000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor", "PION": "Pinon-corona", "ACOPL": "Acoplamiento"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite", "ENAC": "Enfriador aceite"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"RSHL": "Revest shell", "RLIF": "Lifters", "RGRA": "Grizzly descarga"}},
                        }},
                    "MBOL02": {"desc": "Molino Bolas 26x40.5 L2", "sap_type": "M", "mfr": "FLSmidth", "model": "BM 26x40.5", "kw": 16400, "kg": 900000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor", "PION": "Pinon-corona"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"RSHL": "Revest shell", "RLIF": "Lifters"}},
                        }},
                    "MBOL03": {"desc": "Molino Bolas 26x40.5 L3", "sap_type": "M", "mfr": "FLSmidth", "model": "BM 26x40.5", "kw": 16400, "kg": 900000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion"}},
                            "REVE": {"desc": "Revestimientos", "equips": {"RSHL": "Revest shell"}},
                        }},
                }},
            "03": {
                "name": "Bombeo Pulpa Molienda",
                "machines": {
                    "BOMB01": {"desc": "Bomba Pulpa Descarga SAG 01", "sap_type": "M", "mfr": "Weir", "model": "Warman 650AH", "kw": 1500, "kg": 12000, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor", "CUBT": "Cub trasera", "SUCA": "Succion carcasa"}},
                            "SELL": {"desc": "Sello Mecanico", "equips": {"SELL": "Sello mecanico", "AGSL": "Agua sello"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico", "ACOPL": "Acoplamiento"}},
                        }},
                    "BOMB02": {"desc": "Bomba Pulpa Descarga SAG 02", "sap_type": "M", "mfr": "Weir", "model": "Warman 650AH", "kw": 1500, "kg": 12000, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor", "CUBT": "Cub trasera"}},
                            "SELL": {"desc": "Sello Mecanico", "equips": {"SELL": "Sello mecanico"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BOMB03": {"desc": "Bomba Pulpa Descarga SAG 03 StBy", "sap_type": "M", "mfr": "Weir", "model": "Warman 650AH", "kw": 1500, "kg": 12000, "abc": "2",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BOMB04": {"desc": "Bomba Pulpa Alimentacion Bolas 01", "sap_type": "M", "mfr": "Weir", "model": "Warman 550AH", "kw": 1200, "kg": 9800, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor", "CUBT": "Cub trasera"}},
                            "SELL": {"desc": "Sello Mecanico", "equips": {"SELL": "Sello mecanico"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BOMB05": {"desc": "Bomba Pulpa Alimentacion Bolas 02", "sap_type": "M", "mfr": "Weir", "model": "Warman 550AH", "kw": 1200, "kg": 9800, "abc": "2",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
            "04": {
                "name": "Clasificacion Hidrociclones",
                "machines": {
                    "HCIC01": {"desc": "Nido Hidrociclones HCI-01", "sap_type": "M", "mfr": "Weir", "model": "CAVEX 800CVX", "kw": 0, "kg": 8500, "abc": "1",
                        "systems": {
                            "APEX": {"desc": "Apex y Vortex", "equips": {"APEX": "Apex", "VORT": "Vortex finder"}},
                            "CUER": {"desc": "Cuerpo Ciclon", "equips": {"CUER": "Cuerpo ceramico", "BRID": "Bridas union"}},
                        }},
                    "HCIC02": {"desc": "Nido Hidrociclones HCI-02", "sap_type": "M", "mfr": "Weir", "model": "CAVEX 800CVX", "kw": 0, "kg": 8500, "abc": "1",
                        "systems": {
                            "APEX": {"desc": "Apex y Vortex", "equips": {"APEX": "Apex", "VORT": "Vortex finder"}},
                            "CUER": {"desc": "Cuerpo Ciclon", "equips": {"CUER": "Cuerpo ceramico"}},
                        }},
                    "HCIC03": {"desc": "Nido Hidrociclones HCI-03 StBy", "sap_type": "M", "mfr": "Weir", "model": "CAVEX 800CVX", "kw": 0, "kg": 8500, "abc": "2",
                        "systems": {
                            "APEX": {"desc": "Apex y Vortex", "equips": {"APEX": "Apex"}},
                        }},
                }},
        },
    },
    "03": {
        "name": "Flotacion",
        "pg": "P03",
        "subprocesses": {
            "01": {
                "name": "Flotacion Rougher",
                "machines": {
                    "CFRO01": {"desc": "Celda Flotacion Rougher 300m3 C1", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "1",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor", "ESTA": "Estator", "EJEM": "Eje mecanismo"}},
                            "AIRE": {"desc": "Sist Aire", "equips": {"SOPL": "Soplador", "VALV": "Valvula aire", "DIFR": "Difusor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico", "VFRE": "Variador frecuencia"}},
                        }},
                    "CFRO02": {"desc": "Celda Flotacion Rougher 300m3 C2", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "1",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor", "ESTA": "Estator"}},
                            "AIRE": {"desc": "Sist Aire", "equips": {"SOPL": "Soplador", "VALV": "Valvula aire"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "CFRO03": {"desc": "Celda Flotacion Rougher 300m3 C3", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "1",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "CFRO04": {"desc": "Celda Flotacion Rougher 300m3 C4", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "2",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
            "02": {
                "name": "Flotacion Cleaner-Scavenger",
                "machines": {
                    "CFCL01": {"desc": "Celda Flotacion Cleaner 160m3 C1", "sap_type": "M", "mfr": "Metso", "model": "RCS 160", "kw": 200, "kg": 28000, "abc": "1",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor", "ESTA": "Estator"}},
                            "AIRE": {"desc": "Sist Aire", "equips": {"SOPL": "Soplador"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "CFCL02": {"desc": "Celda Flotacion Cleaner 160m3 C2", "sap_type": "M", "mfr": "Metso", "model": "RCS 160", "kw": 200, "kg": 28000, "abc": "2",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "CFSC01": {"desc": "Celda Flotacion Scavenger 300m3 C1", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "1",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor", "ESTA": "Estator"}},
                            "AIRE": {"desc": "Sist Aire", "equips": {"SOPL": "Soplador", "VALV": "Valvula aire"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "CFSC02": {"desc": "Celda Flotacion Scavenger 300m3 C2", "sap_type": "M", "mfr": "Metso", "model": "RCS 300", "kw": 350, "kg": 42000, "abc": "2",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"ROTO": "Rotor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
            "03": {
                "name": "Acondicionamiento Reactivos",
                "machines": {
                    "AGIT01": {"desc": "Acondicionador Reactivos 01", "sap_type": "M", "mfr": "Outotec", "model": "ACR-500", "kw": 75, "kg": 5500, "abc": "2",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"EJEM": "Eje mecanismo", "IMPA": "Impala agitador"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "AGIT02": {"desc": "Acondicionador Reactivos 02", "sap_type": "M", "mfr": "Outotec", "model": "ACR-500", "kw": 75, "kg": 5500, "abc": "3",
                        "systems": {
                            "MECA": {"desc": "Mecanismo Agitacion", "equips": {"EJEM": "Eje mecanismo"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "DREA01": {"desc": "Dosificador Reactivos Colector", "sap_type": "M", "mfr": "Milton Roy", "model": "MR-200", "kw": 5, "kg": 350, "abc": "2",
                        "systems": {
                            "BOMB": {"desc": "Bomba Dosificadora", "equips": {"DIAF": "Diafragma", "VALC": "Valvula check"}},
                        }},
                    "DREA02": {"desc": "Dosificador Reactivos Espumante", "sap_type": "M", "mfr": "Milton Roy", "model": "MR-200", "kw": 5, "kg": 350, "abc": "2",
                        "systems": {
                            "BOMB": {"desc": "Bomba Dosificadora", "equips": {"DIAF": "Diafragma", "VALC": "Valvula check"}},
                        }},
                    "DREA03": {"desc": "Dosificador Cal/pH", "sap_type": "M", "mfr": "Milton Roy", "model": "MR-100", "kw": 3, "kg": 250, "abc": "3",
                        "systems": {
                            "BOMB": {"desc": "Bomba Dosificadora", "equips": {"DIAF": "Diafragma"}},
                        }},
                }},
        },
    },
    "04": {
        "name": "Espesado",
        "pg": "P03",
        "subprocesses": {
            "01": {
                "name": "Espesador Concentrado",
                "machines": {
                    "ESPC01": {"desc": "Espesador Concentrado Hi-Rate 25m", "sap_type": "M", "mfr": "FLSmidth", "model": "HR-25", "kw": 110, "kg": 185000, "abc": "1",
                        "systems": {
                            "RAST": {"desc": "Rastras", "equips": {"BRAZ": "Brazos rastra", "RODA": "Rodamientos cabezal", "TORQ": "Limitador torque"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor", "PION": "Pinon corona"}},
                            "FLOC": {"desc": "Sist Floculante", "equips": {"BMEZ": "Bomba mezcla", "BALI": "Bomba alimentacion", "TANK": "Tanque preparacion"}},
                        }},
                    "ESPC02": {"desc": "Espesador Concentrado Hi-Rate 25m StBy", "sap_type": "M", "mfr": "FLSmidth", "model": "HR-25", "kw": 110, "kg": 185000, "abc": "2",
                        "systems": {
                            "RAST": {"desc": "Rastras", "equips": {"BRAZ": "Brazos rastra", "RODA": "Rodamientos cabezal"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                        }},
                }},
            "02": {
                "name": "Espesador Relaves",
                "machines": {
                    "ESPR01": {"desc": "Espesador Relaves 45m", "sap_type": "M", "mfr": "FLSmidth", "model": "HR-45", "kw": 200, "kg": 350000, "abc": "1",
                        "systems": {
                            "RAST": {"desc": "Rastras", "equips": {"BRAZ": "Brazos rastra", "RODA": "Rodamientos cabezal", "TORQ": "Limitador torque"}},
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor"}},
                            "FLOC": {"desc": "Sist Floculante", "equips": {"BMEZ": "Bomba mezcla", "TANK": "Tanque preparacion"}},
                        }},
                    "BREL01": {"desc": "Bomba Relaves 01", "sap_type": "M", "mfr": "Weir", "model": "Warman 550AH", "kw": 1200, "kg": 9800, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor", "CUBT": "Cub trasera"}},
                            "SELL": {"desc": "Sello Mecanico", "equips": {"SELL": "Sello mecanico"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BREL02": {"desc": "Bomba Relaves 02 StBy", "sap_type": "M", "mfr": "Weir", "model": "Warman 550AH", "kw": 1200, "kg": 9800, "abc": "2",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                }},
        },
    },
    "05": {
        "name": "Filtrado",
        "pg": "P01",
        "subprocesses": {
            "01": {
                "name": "Filtro Prensa",
                "machines": {
                    "FILT01": {"desc": "Filtro Prensa Concentrado FP-01", "sap_type": "M", "mfr": "FLSmidth", "model": "AFP-2500", "kw": 185, "kg": 120000, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "CIHI": "Cilindro hidraulico", "VHID": "Valvula hidraulica", "ACUM": "Acumulador"}},
                            "PLAC": {"desc": "Placas Filtrantes", "equips": {"PLAC": "Placas", "CABI": "Cabezal inferior", "CABS": "Cabezal superior"}},
                            "TELA": {"desc": "Telas Filtrantes", "equips": {"TELA": "Telas polipropileno", "GUIA": "Guias tela"}},
                            "BOMB": {"desc": "Bomba Alimentacion", "equips": {"BALI": "Bomba alimentacion", "SELL": "Sello mecanico"}},
                        }},
                    "FILT02": {"desc": "Filtro Prensa Concentrado FP-02", "sap_type": "M", "mfr": "FLSmidth", "model": "AFP-2500", "kw": 185, "kg": 120000, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "CIHI": "Cilindro hidraulico", "VHID": "Valvula hidraulica"}},
                            "PLAC": {"desc": "Placas Filtrantes", "equips": {"PLAC": "Placas", "CABI": "Cabezal inferior"}},
                            "TELA": {"desc": "Telas Filtrantes", "equips": {"TELA": "Telas polipropileno"}},
                            "BOMB": {"desc": "Bomba Alimentacion", "equips": {"BALI": "Bomba alimentacion"}},
                        }},
                    "FILT03": {"desc": "Filtro Prensa Concentrado FP-03 StBy", "sap_type": "M", "mfr": "FLSmidth", "model": "AFP-2500", "kw": 185, "kg": 120000, "abc": "2",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"BHID": "Bomba hidraulica", "CIHI": "Cilindro hidraulico"}},
                            "PLAC": {"desc": "Placas Filtrantes", "equips": {"PLAC": "Placas"}},
                            "TELA": {"desc": "Telas Filtrantes", "equips": {"TELA": "Telas polipropileno"}},
                        }},
                }},
            "02": {
                "name": "Secado y Despacho",
                "machines": {
                    "SECA01": {"desc": "Secador Rotatorio Concentrado", "sap_type": "M", "mfr": "Metso", "model": "RD-3600", "kw": 450, "kg": 85000, "abc": "1",
                        "systems": {
                            "TRAN": {"desc": "Sist Transmision", "equips": {"MELE": "Motor electrico", "REDU": "Reductor", "PION": "Pinon-corona", "RODA": "Rodamientos apoyo"}},
                            "COMB": {"desc": "Sist Combustion", "equips": {"QUEM": "Quemador", "VGAS": "Valvula gas", "CTRL": "Control llama"}},
                            "SELL": {"desc": "Sellos Rotatorios", "equips": {"SEAL": "Sello entrada", "SELS": "Sello salida"}},
                        }},
                    "CORR04": {"desc": "Correa Transportadora CV-050", "sap_type": "M", "mfr": "Continental", "model": "CV-900", "kw": 150, "kg": 18000, "abc": "2",
                        "systems": {
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor principal", "VFRE": "Variador frecuencia"}},
                            "BAND": {"desc": "Banda Transportadora", "equips": {"BAND": "Banda", "RASP": "Raspador"}},
                        }},
                    "CORR05": {"desc": "Correa Transportadora CV-051", "sap_type": "M", "mfr": "Continental", "model": "CV-900", "kw": 150, "kg": 18000, "abc": "3",
                        "systems": {
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor principal"}},
                            "BAND": {"desc": "Banda Transportadora", "equips": {"BAND": "Banda"}},
                        }},
                }},
            "03": {
                "name": "Servicios Planta",
                "machines": {
                    "COMP01": {"desc": "Compresor Aire Planta 01", "sap_type": "M", "mfr": "Atlas Copco", "model": "ZR500", "kw": 500, "kg": 8500, "abc": "1",
                        "systems": {
                            "COMP": {"desc": "Etapa Compresion", "equips": {"ROTO": "Rotores", "RODA": "Rodamientos", "SELL": "Sellos"}},
                            "LUBE": {"desc": "Sist Lubricacion", "equips": {"BLUB": "Bomba lubricacion", "FILT": "Filtro aceite"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "COMP02": {"desc": "Compresor Aire Planta 02 StBy", "sap_type": "M", "mfr": "Atlas Copco", "model": "ZR500", "kw": 500, "kg": 8500, "abc": "2",
                        "systems": {
                            "COMP": {"desc": "Etapa Compresion", "equips": {"ROTO": "Rotores"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BAGP01": {"desc": "Bomba Agua Proceso 01", "sap_type": "M", "mfr": "KSB", "model": "WKP 300-600", "kw": 800, "kg": 4500, "abc": "1",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor", "CUBT": "Cub trasera"}},
                            "SELL": {"desc": "Sello Mecanico", "equips": {"SELL": "Sello mecanico"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    "BAGP02": {"desc": "Bomba Agua Proceso 02 StBy", "sap_type": "M", "mfr": "KSB", "model": "WKP 300-600", "kw": 800, "kg": 4500, "abc": "2",
                        "systems": {
                            "HIDR": {"desc": "Sist Hidraulico", "equips": {"IMPE": "Impulsor"}},
                            "MOTO": {"desc": "Motor Electrico", "equips": {"MELE": "Motor electrico"}},
                        }},
                    # Instrumentacion (tipo Q)
                    "ANPH01": {"desc": "Analizador pH Flotacion", "sap_type": "Q", "mfr": "Endress+Hauser", "model": "Liquiline CM442", "kw": 0, "kg": 15, "abc": "2",
                        "systems": {
                            "SNSR": {"desc": "Sensor", "equips": {"ELEC": "Electrodo pH", "CALI": "Solucion calibracion"}},
                        }},
                    "FLMT01": {"desc": "Flujometro Electromagnetico Pulpa", "sap_type": "Q", "mfr": "Endress+Hauser", "model": "Promag W800", "kw": 0, "kg": 85, "abc": "2",
                        "systems": {
                            "SNSR": {"desc": "Sensor", "equips": {"BOBI": "Bobinas", "ELEC": "Electrodos"}},
                        }},
                    "DENS01": {"desc": "Densimetro Nuclear Molienda", "sap_type": "Q", "mfr": "Thermo Fisher", "model": "TN-3101", "kw": 0, "kg": 120, "abc": "1",
                        "systems": {
                            "SNSR": {"desc": "Sensor", "equips": {"FUEN": "Fuente radioactiva", "DETE": "Detector"}},
                        }},
                    "VIBM01": {"desc": "Monitor Vibracion Molinos", "sap_type": "Q", "mfr": "Baker Hughes", "model": "Bently 3500", "kw": 0, "kg": 25, "abc": "1",
                        "systems": {
                            "SNSR": {"desc": "Sensor", "equips": {"ACEL": "Acelerometro", "PROX": "Proximitor"}},
                        }},
                }},
        },
    },
}


# ============================================================
# BUILD HIERARCHY (6 levels)
# ============================================================

def build_hierarchy():
    equipment_list = []
    func_locs = []
    bom_items = []
    all_level6 = []

    # Corporate
    func_locs.append({"tplnr": "AMSA-OCP", "pltxt": "Antofagasta Minerals - OCP", "level": 0, "parent": "", "fl_type": ""})
    # Level 1: Operation
    func_locs.append({"tplnr": "02", "pltxt": "Planta Concentradora", "level": 1, "parent": "AMSA-OCP", "fl_type": FUNC_LOC_TYPE})

    eq_serial = 1
    for area_code, area in AREAS.items():
        # Level 2: Process area
        area_fl = f"02-{area_code}"
        func_locs.append({"tplnr": area_fl, "pltxt": area["name"], "level": 2, "parent": "02", "fl_type": FUNC_LOC_TYPE})

        for sp_code, sp in area["subprocesses"].items():
            # Level 3: Subprocess
            sp_fl = f"02-{area_code}-{sp_code}"
            func_locs.append({"tplnr": sp_fl, "pltxt": sp["name"], "level": 3, "parent": area_fl, "fl_type": FUNC_LOC_TYPE})

            for mach_code, mach in sp["machines"].items():
                # Level 4: Machinery (mask AAAANN)
                mach_fl = f"02-{area_code}-{sp_code}-{mach_code}"
                func_locs.append({"tplnr": mach_fl, "pltxt": mach["desc"], "level": 4, "parent": sp_fl, "fl_type": FUNC_LOC_TYPE})

                equnr = f"{eq_serial:012d}"  # 12-char CHAR (Blueprint)
                eq_serial += 1
                tag = f"{PLANT_ID}-{mach_code}"
                pg = area["pg"]

                equipment_list.append({
                    "equnr": equnr, "tag": tag, "eqktx": mach["desc"],
                    "eqart": mach["sap_type"], "herst": mach["mfr"], "model": mach["model"],
                    "power_kw": mach["kw"], "weight_kg": mach["kg"],
                    "abckz": mach["abc"],  # Blueprint numeric 1/2/3
                    "tplnr": mach_fl, "area_name": area["name"],
                    "subprocess_name": sp["name"], "area_code": area_code,
                    "pg": pg, "ba": PROCESS_TO_BA[pg],
                })

                for sys_code, sys_info in mach.get("systems", {}).items():
                    # Level 5: System (mask XXXX)
                    sys_fl = f"02-{area_code}-{sp_code}-{mach_code}-{sys_code}"
                    func_locs.append({"tplnr": sys_fl, "pltxt": sys_info["desc"], "level": 5, "parent": mach_fl, "fl_type": FUNC_LOC_TYPE})

                    bom_items.append({
                        "parent_tag": tag, "parent_equnr": equnr,
                        "component": sys_info["desc"], "sys_code": sys_code,
                        "func_loc": sys_fl, "level": 5,
                    })

                    for eq6_code, eq6_desc in sys_info.get("equips", {}).items():
                        # Level 6: Equipment (mask XXXX)
                        eq6_fl = f"02-{area_code}-{sp_code}-{mach_code}-{sys_code}-{eq6_code}"
                        func_locs.append({"tplnr": eq6_fl, "pltxt": eq6_desc, "level": 6, "parent": sys_fl, "fl_type": FUNC_LOC_TYPE})

                        all_level6.append({
                            "parent_tag": tag, "parent_equnr": equnr,
                            "component": eq6_desc, "eq6_code": eq6_code,
                            "sys_code": sys_code, "func_loc": eq6_fl, "level": 6,
                        })

    return equipment_list, func_locs, bom_items, all_level6


# ============================================================
# STYLING HELPERS
# ============================================================

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def save_wb(wb, name):
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"  -> Saved: {name}")
    return path


def synth_id(prefix, seq):
    h = hashlib.md5(f"SYNTH2026-{prefix}-{seq}".encode()).hexdigest()[:4].upper()
    return f"S26-{prefix}-{seq:04d}-{h}"


# ============================================================
# BUILD DATA
# ============================================================

print("=" * 70)
print("  DB_AM_OCP_SYNTHETIC_2026 v2 — Alineado Blueprint AMSA_BBP_PM_04")
print("=" * 70)

equipment_list, func_locs, bom_items, level6_items = build_hierarchy()
print(f"\nJerarquia: {len(equipment_list)} equipos Niv4, {len(bom_items)} sistemas Niv5, {len(level6_items)} equipos Niv6, {len(func_locs)} UTs total")


# ============================================================
# 01 — EQUIPMENT HIERARCHY (≥200 registros)
# ============================================================

print("\n[01] Generating 01_equipment_hierarchy.xlsx ...")
wb01 = openpyxl.Workbook()

ws_h = wb01.active
ws_h.title = "Equipment Hierarchy"
h01 = [
    "sap_func_loc", "fl_type", "pltxt", "level", "parent_fl",
    "equnr", "eqktx", "eqart", "eqart_desc",
    "herst", "model", "power_kw", "weight_kg",
    "abckz", "abckz_desc", "stat",
    "planning_center", "planning_group", "business_area",
    "install_date", "serial_number",
]
style_header(ws_h, h01)

# All func locs + equipment
row_count_01 = 0
for fl in func_locs:
    # Find equipment if level 4
    eq_match = None
    if fl["level"] == 4:
        eq_match = next((e for e in equipment_list if e["tplnr"] == fl["tplnr"]), None)

    ws_h.append([
        fl["tplnr"], fl["fl_type"], fl["pltxt"], fl["level"], fl.get("parent", ""),
        eq_match["equnr"] if eq_match else "",
        eq_match["eqktx"] if eq_match else fl["pltxt"],
        eq_match["eqart"] if eq_match else "",
        EQUIPMENT_TYPES_SAP.get(eq_match["eqart"], "") if eq_match else "",
        eq_match["herst"] if eq_match else "",
        eq_match["model"] if eq_match else "",
        eq_match["power_kw"] if eq_match else "",
        eq_match["weight_kg"] if eq_match else "",
        eq_match["abckz"] if eq_match else "",
        ABC_CRITICALITY.get(eq_match["abckz"], "") if eq_match else "",
        "ACTIVE" if fl["level"] >= 4 else "",
        PLANNING_CENTER, eq_match["pg"] if eq_match else "", eq_match["ba"] if eq_match else "",
        (datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365))).strftime("%Y-%m-%d") if eq_match else "",
        f"SN-{eq_match['herst'][:3].upper()}-{random.randint(100000,999999)}" if eq_match else "",
    ])
    row_count_01 += 1

# BOM sheet
ws_bom = wb01.create_sheet("Equipment BOM")
h01b = ["parent_tag", "parent_equnr", "component_name", "sys_code", "sap_func_loc", "level"]
style_header(ws_bom, h01b)
for b in bom_items:
    ws_bom.append([b["parent_tag"], b["parent_equnr"], b["component"], b["sys_code"], b["func_loc"], 5])
for l6 in level6_items:
    ws_bom.append([l6["parent_tag"], l6["parent_equnr"], l6["component"], f"{l6['sys_code']}-{l6['eq6_code']}", l6["func_loc"], 6])

auto_width(ws_h)
auto_width(ws_bom)
save_wb(wb01, "01_equipment_hierarchy.xlsx")
count_01 = row_count_01 + len(bom_items) + len(level6_items)
print(f"     Hierarchy: {row_count_01} UTs + {len(bom_items)} BOM Niv5 + {len(level6_items)} BOM Niv6 = {count_01} total")


# ============================================================
# 02 — CRITICALITY ASSESSMENT (≥200)
# Evaluamos equipos Niv4 + sistemas Niv5 + equipos criticos Niv6
# ============================================================

print("[02] Generating 02_criticality_assessment.xlsx ...")
wb02 = openpyxl.Workbook()
ws02 = wb02.active
ws02.title = "Criticality Assessment"
h02 = [
    "equipment_tag", "equnr", "sap_func_loc", "eqktx", "area",
    "method", "abckz", "abckz_desc",
    "safety", "health", "environment", "production",
    "operating_cost", "capital_cost", "schedule", "revenue",
    "communications", "compliance", "reputation", "probability",
    "total_score", "criticality_rank",
]
style_header(ws02, h02)

CRIT_SCORES = {"1": (5, 4, 4, 5, 4, 5, 4, 5, 3, 4, 4, 4), "2": (3, 3, 3, 4, 3, 3, 3, 3, 2, 3, 3, 3), "3": (2, 2, 1, 2, 2, 2, 1, 2, 1, 2, 2, 2)}

crit_rows = []
# Level 4 equipment
for eq in equipment_list:
    scores = list(CRIT_SCORES[eq["abckz"]])
    scores = [max(1, min(5, s + random.choice([-1, 0, 0, 0, 1]))) for s in scores]
    total = sum(scores)
    rank = "1" if total >= 45 else "2" if total >= 30 else "3"
    prob = random.choice([3, 4, 5]) if eq["abckz"] == "1" else random.choice([2, 3])
    crit_rows.append([eq["tag"], eq["equnr"], eq["tplnr"], eq["eqktx"], eq["area_name"],
                       "AMSA-RAM", eq["abckz"], ABC_CRITICALITY[eq["abckz"]], *scores, prob, total, rank])

# Level 5 systems
for b in bom_items:
    parent_eq = next((e for e in equipment_list if e["equnr"] == b["parent_equnr"]), None)
    abc = parent_eq["abckz"] if parent_eq else "2"
    scores = list(CRIT_SCORES[abc])
    scores = [max(1, min(5, s + random.choice([-1, 0, 0, 1]))) for s in scores]
    total = sum(scores)
    rank = "1" if total >= 45 else "2" if total >= 30 else "3"
    prob = random.choice([2, 3, 4])
    crit_rows.append([b["parent_tag"], b["parent_equnr"], b["func_loc"], b["component"],
                       parent_eq["area_name"] if parent_eq else "",
                       "AMSA-RAM", abc, ABC_CRITICALITY[abc], *scores, prob, total, rank])

# Level 6 to fill up to 200+
for l6 in level6_items:
    parent_eq = next((e for e in equipment_list if e["equnr"] == l6["parent_equnr"]), None)
    abc = parent_eq["abckz"] if parent_eq else "3"
    scores = list(CRIT_SCORES[abc])
    scores = [max(1, min(5, s + random.choice([-1, 0, 0, 1]))) for s in scores]
    total = sum(scores)
    rank = "1" if total >= 45 else "2" if total >= 30 else "3"
    prob = random.choice([2, 3])
    crit_rows.append([l6["parent_tag"], l6["parent_equnr"], l6["func_loc"], l6["component"],
                       parent_eq["area_name"] if parent_eq else "",
                       "AMSA-RAM", abc, ABC_CRITICALITY[abc], *scores, prob, total, rank])

for r in crit_rows:
    ws02.append(r)
auto_width(ws02)
save_wb(wb02, "02_criticality_assessment.xlsx")
count_02 = len(crit_rows)


# ============================================================
# 03 — FAILURE MODES (≥200)
# ============================================================

print("[03] Generating 03_failure_modes.xlsx ...")
wb03 = openpyxl.Workbook()
ws03 = wb03.active
ws03.title = "failure_modes"
h03 = [
    "equipment_tag", "equnr", "sap_func_loc", "area",
    "equipment_function_description", "equipment_functional_failure",
    "function_type", "failure_type", "subunit", "maintainable_item",
    "mechanism", "cause", "failure_pattern", "failure_consequence",
    "evidence", "downtime_hours", "detection_method",
    "rpn_severity", "rpn_occurrence", "rpn_detection", "rpn_total",
]
style_header(ws03, h03)

MECHANISMS = ["WEARS", "CORRODES", "FRACTURES", "OVERHEATS", "VIBRATES", "LEAKS", "SHORT_CIRCUITS", "BLOCKS", "ERODES", "FATIGUES"]
CAUSES = ["ABRASION", "CAVITATION", "CORROSION", "FATIGUE", "OVERLOAD", "CONTAMINATION", "MISALIGNMENT", "THERMAL_STRESS", "ELECTRICAL_SURGE", "LACK_LUBRICATION"]
PATTERNS = ["A_BATHTUB", "B_AGE", "C_GRADUAL", "D_INITIAL_SURGE", "E_RANDOM", "F_INFANT"]
CONSEQUENCES = ["SAFETY_CRITICAL", "ENVIRONMENTAL", "OPERATIONAL", "NON_OPERATIONAL", "HIDDEN"]
EVIDENCE_LIST = ["Aumento vibracion", "Incremento temperatura", "Cambio ruido", "Desgaste visible",
                 "Anomalia analisis aceite", "Caida presion", "Reduccion flujo", "Anomalia electrica",
                 "Degradacion rendimiento", "Activacion alarma"]
DETECTION = ["MONITOREO_VIBRACION", "TERMOGRAFIA", "ANALISIS_ACEITE", "INSPECCION_VISUAL",
             "ULTRASONIDO", "PRUEBA_RENDIMIENTO", "PRUEBA_ELECTRICA", "RONDAS_OPERADOR"]
MAINT_ITEMS = ["Rodamiento", "Impulsor", "Sello mecanico", "Eje", "Revestimiento", "Motor electrico",
               "Reductor", "Acoplamiento", "Valvula", "Correa", "Filtro", "Rotor", "Estator", "Malla"]

fm_count = 0
min_fm_target = 200
fm_per_eq = max(4, min_fm_target // len(equipment_list) + 1)

for eq in equipment_list:
    n_fm = random.randint(fm_per_eq - 1, fm_per_eq + 2)
    func_desc = f"Proveer operacion continua de {eq['eqktx'].lower()} en {eq['subprocess_name']}"
    for j in range(n_fm):
        mech = random.choice(MECHANISMS)
        cause = random.choice(CAUSES)
        sev = random.randint(3, 10)
        occ = random.randint(1, 8)
        det = random.randint(2, 9)
        ws03.append([
            eq["tag"], eq["equnr"], eq["tplnr"], eq["area_name"],
            func_desc,
            f"Incapaz de operar al rendimiento requerido",
            random.choice(["PRIMARY", "SECONDARY"]),
            random.choice(["FUNCTIONAL", "POTENTIAL"]),
            random.choice(["MECHANICAL", "ELECTRICAL", "HYDRAULIC", "PNEUMATIC", "STRUCTURAL"]),
            random.choice(MAINT_ITEMS),
            mech, cause, random.choice(PATTERNS), random.choice(CONSEQUENCES),
            random.choice(EVIDENCE_LIST), round(random.uniform(2, 72), 1),
            random.choice(DETECTION), sev, occ, det, sev * occ * det,
        ])
        fm_count += 1

auto_width(ws03)
save_wb(wb03, "03_failure_modes.xlsx")
count_03 = fm_count


# ============================================================
# 05 — WORK PACKAGES (≥200)
# ============================================================

print("[05] Generating 05_work_packages.xlsx ...")
wb05 = openpyxl.Workbook()
ws05 = wb05.active
ws05.title = "Work Packages"
h05 = [
    "wp_code", "wp_name", "equipment_tag", "sap_func_loc", "area",
    "frequency_value", "frequency_unit", "constraint", "wp_type",
    "route_sheet_type", "route_sheet_group",
    "access_time_hours", "estimated_total_hours", "crew_size",
    "work_center", "planning_group", "planning_center",
    "maint_plan_type",
]
style_header(ws05, h05)

WP_TYPES = [
    ("INS", "Inspeccion programada", "INSPECT", "ONLINE"),
    ("LUB", "Lubricacion programada", "LUBRICATION", "ONLINE"),
    ("PMV", "Preventivo programado", "PREVENTIVE", "OFFLINE"),
    ("PDC", "Predictivo condicion", "PREDICTIVE", "ONLINE"),
    ("CAL", "Calibracion instrumentos", "CALIBRATION", "ONLINE"),
]

wp_count = 0
for eq in equipment_list:
    wcs = PG_TO_WC[eq["pg"]]
    n_wp = random.randint(3, 5)
    freqs = random.sample([2, 4, 8, 13, 26, 52], k=n_wp)
    for freq in freqs:
        wpt = random.choice(WP_TYPES)
        wc = random.choice(wcs)
        ws05.append([
            f"{wpt[0]}-{eq['tag'][-6:]}-{freq}S", f"{wpt[1]} {freq}S {eq['eqktx'][:30]}",
            eq["tag"], eq["tplnr"], eq["area_name"],
            freq, "WEEKS", wpt[3], wpt[2],
            ROUTE_SHEET_TYPE, ROUTE_SHEET_GROUPS["PLA"],
            round(random.uniform(0.5, 4), 1), round(random.uniform(2, 24), 1),
            random.randint(1, 4), wc, eq["pg"], PLANNING_CENTER, MAINT_PLAN_TYPE,
        ])
        wp_count += 1

auto_width(ws05)
save_wb(wb05, "05_work_packages.xlsx")
count_05 = wp_count


# ============================================================
# 06 — WORK ORDER HISTORY (200 OTs — mantener)
# Tipos: PM01, PM02, PM03, PM06, PM07 (Blueprint exacto)
# ============================================================

print("[06] Generating 06_work_order_history.xlsx (200 OTs) ...")
wb06 = openpyxl.Workbook()
ws06 = wb06.active
ws06.title = "Work Order History"
h06 = [
    "aufnr", "auart", "auart_desc", "equipment_tag", "equnr", "tplnr",
    "area", "subprocess", "planning_group", "business_area", "planning_center",
    "priokx", "priokx_desc", "priority_class",
    "ilart", "ilart_desc", "qmnum",
    "notification_catalog", "notification_status_schema",
    "notif_user_status",
    "description",
    "erdat", "gstrp", "gltrp", "iedd", "iedt",
    "duration_hours",
    "cost_labour_usd", "cost_materials_usd", "cost_external_usd",
    "value_cat_labour", "value_cat_materials", "value_cat_external",
    "failure_found",
    "arbpl", "supervisor_wc",
    "system_status",
    "pm07_activity_class", "pm07_activity_desc",
    "crew_size", "specialty",
    "synth_trace_id",
]
style_header(ws06, h06)

OT_DIST = {"PM01": 50, "PM02": 70, "PM03": 25, "PM06": 25, "PM07": 30}

OT_DESC = {
    "PM01": [
        "Reparar fuga aceite en {sys} de {eq}", "Cambiar rodamiento danado en {eq}",
        "Corregir desalineamiento en {eq}", "Reparar falla electrica motor {eq}",
        "Reparar vibracion excesiva en {eq}", "Reemplazar sello mecanico {eq}",
        "Reparar valvula bloqueada {eq}", "Corregir sobrecalentamiento {eq}",
        "Reparar correa danada {eq}", "Reparar instrumentacion defectuosa {eq}",
    ],
    "PM02": [
        "Inspeccion programada {freq}S de {eq}", "Lubricacion periodica {eq}",
        "Cambio aceite {sys} {eq}", "Inspeccion termografica {eq}",
        "Analisis vibracion programado {eq}", "Cambio filtros programado {eq}",
        "Torqueo pernos {eq}", "Inspeccion revestimientos {eq}",
        "Analisis aceite {sys} {eq}", "Calibracion instrumentos {eq}",
    ],
    "PM03": [
        "Solicitud servicio alineamiento laser {eq}", "Solicitud reparacion {sys} {eq}",
        "Solicitud balanceo dinamico {eq}", "Solicitud montaje componente {eq}",
        "Solicitud inspeccion NDT {eq}", "Solicitud servicio especial {eq}",
    ],
    "PM06": [
        "Mejora sistema lubricacion {eq}", "Upgrade variador frecuencia {eq}",
        "Instalacion monitoreo vibracion {eq}", "Mejora sist control {eq}",
        "Upgrade proteccion electrica {eq}", "Instalacion sensores adicionales {eq}",
    ],
    "PM07": [
        "Reparacion mayor reductor {eq}", "Overhaul bomba {eq}",
        "Reparacion motor electrico {eq}", "Rebuild chancador {eq}",
        "Reparacion cilindro hidraulico {eq}", "Overhaul compresor {eq}",
        "Reparacion valvula control {eq}", "Rebobinado motor {eq}",
    ],
}

FAILURE_FOUND = [
    "Rodamiento colapsado", "Sello con desgaste excesivo", "Fisura en soldadura",
    "Cortocircuito bobinado", "Desalineamiento severo", "Cavitacion impulsor",
    "Corrosion avanzada", "Fatiga en eje", "Obstruccion material", "Desgaste abrasivo",
    "Fuga hidraulica", "Vibracion desbalance", None, None,
]

SYS_NAMES = ["Sist Lubricacion", "Sist Hidraulico", "Sist Transmision", "Motor Electrico",
             "Sello Mecanico", "Revestimientos", "Sist Aire", "Mecanismo Agitacion"]

ot_counter = {k: v["range_start"] for k, v in ORDER_TYPES.items()}
ot_rows = []

for ot_type, count in OT_DIST.items():
    for _ in range(count):
        eq = random.choice(equipment_list)
        ot_num = ot_counter[ot_type]
        ot_counter[ot_type] += 1

        pg = eq["pg"]
        ba = eq["ba"]
        wcs = PG_TO_WC[pg]
        wc = random.choice(wcs)
        sup_wc = list(SUPERVISOR_WCS.keys())[random.randint(0, len(SUPERVISOR_WCS) - 1)]

        # Dates
        created = datetime(2026, 1, 1) + timedelta(days=random.randint(0, 88))
        if ot_type == "PM01":
            planned_start = created + timedelta(days=random.randint(0, 5))
            priority = random.choice(["I", "A", "A", "M"])
        elif ot_type == "PM02":
            planned_start = created + timedelta(days=random.randint(7, 30))
            priority = random.choice(["M", "M", "B", "B"])
        elif ot_type == "PM03":
            planned_start = created + timedelta(days=random.randint(3, 14))
            priority = random.choice(["A", "M", "M"])
        elif ot_type == "PM06":
            planned_start = created + timedelta(days=random.randint(14, 60))
            priority = random.choice(["M", "B", "B"])
        else:  # PM07
            planned_start = created + timedelta(days=random.randint(5, 21))
            priority = random.choice(["A", "A", "M"])

        duration = round(random.uniform(2, 48), 1) if ot_type != "PM07" else round(random.uniform(8, 120), 1)
        planned_end = planned_start + timedelta(hours=duration)

        # Status
        if created < datetime(2026, 2, 15):
            sys_status = "CTEC"
            actual_start = planned_start + timedelta(hours=random.randint(-2, 8))
            actual_end = actual_start + timedelta(hours=duration * random.uniform(0.8, 1.4))
        elif created < datetime(2026, 3, 10):
            sys_status = random.choice(["NOTI", "CTEC"])
            actual_start = planned_start if sys_status == "CTEC" else None
            actual_end = (actual_start + timedelta(hours=duration)) if actual_start else None
        else:
            sys_status = random.choice(["ABIE", "LIBE"])
            actual_start = None
            actual_end = None

        # Notification
        notif_type = "A1" if ot_type == "PM01" else "A3" if ot_type == "PM02" else "A2"
        notif_id = f"{random.randint(1, 4999999):07d}"
        notif_cats = NOTIF_TYPES[notif_type]["catalogs"]
        notif_cat = random.choice(list(notif_cats.keys())) if notif_cats else ""
        notif_user_status = random.choice(["APRO", "APRO", "APRO", "RECH"]) if notif_type in ("A1", "A2") else ""

        desc_tpl = random.choice(OT_DESC[ot_type])
        desc = desc_tpl.format(eq=eq["eqktx"][:35], sys=random.choice(SYS_NAMES), freq=random.choice([4, 8, 13, 26]))

        # PM07 activity class
        pm07_act = random.choice(["RP1", "RP2"]) if ot_type == "PM07" else ""
        pm07_desc = PM07_ACTIVITY.get(pm07_act, "")

        # Costs
        spec = wc[3:6]  # Extract specialty from work center
        rate = LABOR_CATEGORIES.get(f"{spec}-II", LABOR_CATEGORIES.get(f"{spec}-I", {"rate_usd": 45}))["rate_usd"]
        crew = random.randint(1, 4)
        cost_lab = round(rate * crew * duration, 2) if sys_status == "CTEC" else 0
        cost_mat = round(random.uniform(200, 15000), 2) if sys_status == "CTEC" and random.random() > 0.3 else 0
        cost_ext = round(random.uniform(5000, 50000), 2) if ot_type == "PM03" and sys_status == "CTEC" else 0

        trace_id = synth_id("OT", ot_num - ORDER_TYPES[ot_type]["range_start"])

        row = [
            ot_num, ot_type, ORDER_TYPES[ot_type]["desc"],
            eq["tag"], eq["equnr"], eq["tplnr"],
            eq["area_name"], eq["subprocess_name"], pg, ba, PLANNING_CENTER,
            priority, PRIORITIES[priority], PRIORITY_CLASS,
            notif_type, NOTIF_TYPES[notif_type]["desc"], notif_id,
            notif_cat, USER_STATUS_SCHEMA if notif_type in ("A1", "A2") else "",
            notif_user_status,
            desc[:72],  # SAP_SHORT_TEXT_MAX = 72
            created.strftime("%Y-%m-%d"),
            planned_start.strftime("%Y-%m-%d %H:%M"),
            planned_end.strftime("%Y-%m-%d %H:%M"),
            actual_start.strftime("%Y-%m-%d %H:%M") if actual_start else "",
            actual_end.strftime("%Y-%m-%d %H:%M") if actual_end else "",
            duration,
            cost_lab, cost_mat, cost_ext,
            "ZMANT001" if cost_lab > 0 else "", "ZMANT002" if cost_mat > 0 else "", "ZMANT003" if cost_ext > 0 else "",
            random.choice(FAILURE_FOUND) if sys_status == "CTEC" else "",
            wc, sup_wc, sys_status,
            pm07_act, pm07_desc,
            crew, spec,
            trace_id,
        ]
        ot_rows.append(row)

ot_rows.sort(key=lambda r: r[21])  # Sort by created date
for r in ot_rows:
    ws06.append(r)

auto_width(ws06)
save_wb(wb06, "06_work_order_history.xlsx")
count_06 = 200


# ============================================================
# 07 — SPARE PARTS INVENTORY (≥200)
# ============================================================

print("[07] Generating 07_spare_parts_inventory.xlsx ...")
wb07 = openpyxl.Workbook()
ws07 = wb07.active
ws07.title = "Spare Parts Inventory"
h07 = [
    "material_code", "sap_material_number", "description", "manufacturer",
    "part_number", "ved_class", "fsn_class", "abc_class",
    "quantity_on_hand", "min_stock", "max_stock", "reorder_point",
    "lead_time_days", "unit_cost_usd", "unit_of_measure",
    "applicable_equipment_csv", "warehouse_location",
]
style_header(ws07, h07)

SPARE_PARTS_BASE = [
    ("Rodamiento radial SKF 23248", "SKF", "23248 CC/W33", "E", 18500, "EA"),
    ("Rodamiento radial SKF 23052", "SKF", "23052 CC/W33", "E", 14200, "EA"),
    ("Rodamiento axial SKF 29440", "SKF", "29440 E", "E", 12300, "EA"),
    ("Rodamiento axial SKF 29436", "SKF", "29436 E", "E", 11800, "EA"),
    ("Rodamiento SKF 22344 CCJA/W33", "SKF", "22344 CCJA", "E", 16500, "EA"),
    ("Rodamiento FAG 23156 K.MB", "FAG", "23156 K.MB", "E", 15800, "EA"),
    ("Rodamiento NTN 232/600 BK", "NTN", "232/600 BK", "E", 22000, "EA"),
    ("Sello mecanico doble John Crane 88XT", "John Crane", "88-XT", "E", 45000, "EA"),
    ("Sello mecanico John Crane 4610", "John Crane", "4610-SIC", "E", 32000, "EA"),
    ("Sello mecanico Flowserve ISCS-2", "Flowserve", "ISCS-2", "D", 22000, "EA"),
    ("Sello mecanico Flowserve BX", "Flowserve", "BX-200", "D", 18000, "EA"),
    ("Impulsor bomba Warman 650AH CR27", "Weir", "650AH-IMP-CR27", "E", 85000, "EA"),
    ("Impulsor bomba Warman 550AH CR27", "Weir", "550AH-IMP-CR27", "E", 65000, "EA"),
    ("Impulsor bomba Warman 650AH elastomero", "Weir", "650AH-IMP-R55", "D", 42000, "EA"),
    ("Cub trasera bomba Warman 650AH", "Weir", "650AH-FPL-CR27", "E", 55000, "EA"),
    ("Cub trasera bomba Warman 550AH", "Weir", "550AH-FPL-CR27", "D", 38000, "EA"),
    ("Liner SAG Mill shell FLSmidth", "FLSmidth", "LNR-SAG40-SHL", "E", 125000, "EA"),
    ("Lifter SAG Mill FLSmidth", "FLSmidth", "LFT-SAG40-HI", "E", 95000, "EA"),
    ("Liner Bolas Mill shell", "FLSmidth", "LNR-BM26-SHL", "E", 78000, "EA"),
    ("Lifter Bolas Mill", "FLSmidth", "LFT-BM26-STD", "E", 62000, "EA"),
    ("Grizzly descarga SAG", "FLSmidth", "GRZ-SAG40-DC", "E", 35000, "EA"),
    ("Manto chancador giratorio 60x113", "FLSmidth", "MNT-60113-UPR", "E", 180000, "EA"),
    ("Concavo chancador giratorio", "FLSmidth", "CNC-60113-SET", "E", 150000, "EA"),
    ("Manto chancador HP800 superior", "Metso", "MNT-HP800-U", "E", 95000, "EA"),
    ("Manto chancador HP800 inferior", "Metso", "MNT-HP800-L", "E", 88000, "EA"),
    ("Manto chancador HP500 superior", "Metso", "MNT-HP500-U", "E", 65000, "EA"),
    ("Motor electrico 22MW GMD FLSmidth", "FLSmidth", "GMD-22MW-SET", "E", 2500000, "EA"),
    ("Motor electrico 16.4MW ABB", "ABB", "AMI-16400-4P", "E", 850000, "EA"),
    ("Motor electrico 1500kW ABB", "ABB", "AMI 560L4", "E", 185000, "EA"),
    ("Motor electrico 800kW ABB", "ABB", "AMI 450L4", "E", 125000, "EA"),
    ("Motor electrico 600kW Siemens", "Siemens", "1LA8 355-4AB", "D", 95000, "EA"),
    ("Motor electrico 500kW Siemens", "Siemens", "1LA8 315-4AB", "D", 78000, "EA"),
    ("Motor electrico 350kW ABB", "ABB", "AMI 355L4", "D", 65000, "EA"),
    ("Motor electrico 250kW ABB", "ABB", "AMI 315L4", "D", 48000, "EA"),
    ("Motor electrico 150kW Siemens", "Siemens", "1LA8 280-4AB", "D", 32000, "EA"),
    ("Motor electrico 90kW ABB", "ABB", "AMI 280M4", "D", 18000, "EA"),
    ("Motor electrico 75kW ABB", "ABB", "AMI 250M4", "D", 15000, "EA"),
    ("Reductor Flender SZAT800", "Flender", "SZAT800", "E", 250000, "EA"),
    ("Reductor Flender SZAT500", "Flender", "SZAT500", "E", 180000, "EA"),
    ("Reductor SEW KA107", "SEW", "KA107 DRE200", "D", 28000, "EA"),
    ("Acoplamiento Falk 1180T", "Falk", "1180T10", "D", 15000, "EA"),
    ("Acoplamiento Falk 1100T", "Falk", "1100T10", "D", 12000, "EA"),
    ("Acoplamiento Rexnord Omega 140", "Rexnord", "OMG-140", "D", 8500, "EA"),
    ("Correa transportadora EP400/3 1200mm", "Continental", "EP400/3-1200", "D", 42000, "M"),
    ("Correa transportadora EP400/3 900mm", "Continental", "EP400/3-900", "D", 28000, "M"),
    ("Malla harnero acero 50mm", "Metso", "SCR-50-HT", "D", 3500, "EA"),
    ("Panel poliuretano harnero", "Metso", "PU-8203-SET", "D", 4200, "EA"),
    ("Rotor celda flotacion RCS300", "Metso", "ROT-RCS300", "E", 35000, "EA"),
    ("Estator celda flotacion RCS300", "Metso", "EST-RCS300", "E", 28000, "EA"),
    ("Rotor celda flotacion RCS160", "Metso", "ROT-RCS160", "E", 22000, "EA"),
    ("Apex hidrociclón CAVEX 800", "Weir", "APX-800CVX", "D", 4500, "EA"),
    ("Vortex finder CAVEX 800", "Weir", "VFX-800CVX", "D", 5200, "EA"),
    ("Cuerpo ceramico CAVEX 800", "Weir", "CRP-800CVX", "D", 8500, "EA"),
    ("Brazo rastra espesador 25m", "FLSmidth", "ARM-HR25-SET", "E", 45000, "EA"),
    ("Brazo rastra espesador 45m", "FLSmidth", "ARM-HR45-SET", "E", 68000, "EA"),
    ("Tela filtrante PP 2500mm", "Clear Edge", "CE-PP-2500", "E", 4500, "EA"),
    ("Placa filtrante 2500mm", "FLSmidth", "AFP-PLT-2500", "E", 6200, "EA"),
    ("Cilindro hidraulico filtro prensa", "Parker", "CYL-AFP-2500", "E", 18000, "EA"),
    ("Variador frecuencia 1500kW ABB ACS880", "ABB", "ACS880-1500", "E", 145000, "EA"),
    ("Variador frecuencia 800kW ABB ACS880", "ABB", "ACS880-800", "E", 95000, "EA"),
    ("Variador frecuencia 350kW ABB ACS580", "ABB", "ACS580-350", "D", 42000, "EA"),
    ("Filtro aceite hidraulico Pall HC9600", "Pall", "HC9600FKT16H", "V", 850, "EA"),
    ("Filtro aceite lubricacion Donaldson", "Donaldson", "P551807", "V", 320, "EA"),
    ("Filtro aceite Hydac 0660R", "Hydac", "0660R010BN4HC", "V", 420, "EA"),
    ("Valvula compuerta 10in Flowserve", "Flowserve", "VGC-250-150", "D", 8500, "EA"),
    ("Valvula compuerta 8in Flowserve", "Flowserve", "VGC-200-150", "D", 6200, "EA"),
    ("Valvula mariposa 12in Weir", "Weir", "BFV-300-150", "D", 5500, "EA"),
    ("Manguera hidraulica 1in Parker", "Parker", "421SN-16", "V", 180, "M"),
    ("Manguera hidraulica 3/4in Parker", "Parker", "421SN-12", "V", 150, "M"),
    ("Sensor vibracion Bently 3500/42M", "Baker Hughes", "3500/42M", "E", 12000, "EA"),
    ("Sensor vibracion Bently proximitor", "Baker Hughes", "3300XL-8mm", "E", 3500, "EA"),
    ("Sensor temperatura PT100 E+H", "Endress+Hauser", "TST310-A2B2C3", "V", 650, "EA"),
    ("Transmisor presion Rosemount 3051", "Emerson", "3051TG3A2B", "V", 3200, "EA"),
    ("Transmisor nivel Rosemount 5300", "Emerson", "5300-2C1A", "V", 4500, "EA"),
    ("Electrodo pH Endress+Hauser", "Endress+Hauser", "CPS11D-7AA21", "V", 1200, "EA"),
    ("Flujometro Promag W800 DN300", "Endress+Hauser", "W800-DN300", "E", 18000, "EA"),
    ("PLC modulo AI Siemens S7-1500", "Siemens", "6ES7531-7KF00", "E", 2800, "EA"),
    ("PLC modulo DI Siemens S7-1500", "Siemens", "6ES7521-1BH00", "E", 1800, "EA"),
    ("Rele proteccion motor ABB REF615", "ABB", "REF615-C", "D", 9500, "EA"),
    ("Electrovalvula ASCO 2in 120V", "Emerson", "8210G95-120V", "D", 1200, "EA"),
    ("Bolas molienda 5in forjadas", "Moly-Cop", "MC-5F-HCLA", "V", 1.85, "KG"),
    ("Bolas molienda 3in forjadas", "Moly-Cop", "MC-3F-HCLA", "V", 1.65, "KG"),
    ("Reactivo colector xantato PAX", "Cytec", "PAX-A31", "V", 4.50, "KG"),
    ("Reactivo espumante MIBC", "Cytec", "MIBC-STD", "V", 3.80, "KG"),
    ("Floculante anionico Magnafloc 10", "BASF", "MF-10", "V", 8.20, "KG"),
    ("Floculante anionico Magnafloc 336", "BASF", "MF-336", "V", 9.50, "KG"),
    ("Cal viva CaO 90%", "Cementos Bio-Bio", "CAO-90-GRD", "V", 0.12, "KG"),
    ("Lubricante Shell Omala S4 GX320", "Shell", "OMALA-S4-GX320", "V", 18.50, "L"),
    ("Lubricante Shell Omala S4 GX460", "Shell", "OMALA-S4-GX460", "V", 19.80, "L"),
    ("Lubricante Mobil SHC 630", "ExxonMobil", "SHC-630", "V", 22.00, "L"),
    ("Grasa Shell Gadus S3 V220C 2", "Shell", "GADUS-S3-V220C", "V", 8.50, "KG"),
    ("Grasa SKF LGMT 2", "SKF", "LGMT-2/18", "V", 6.20, "KG"),
    ("Aceite hidraulico Shell Tellus S3M46", "Shell", "TELLUS-S3M46", "V", 12.50, "L"),
    ("Empaquetadura grafitada 1/2in Garlock", "Garlock", "GRAPH-PACK-13", "V", 85, "M"),
    ("Empaquetadura PTFE 3/8in Garlock", "Garlock", "PTFE-PACK-10", "V", 95, "M"),
    ("Perno hex ASTM A325 M24x100", "Nucor", "A325-M24X100", "V", 8.50, "EA"),
    ("Perno hex ASTM A325 M30x120", "Nucor", "A325-M30X120", "V", 12.00, "EA"),
    ("Junta espiral wound 10in 150#", "Flexitallic", "SPW-250-150", "V", 180, "EA"),
    ("Junta espiral wound 8in 150#", "Flexitallic", "SPW-200-150", "V", 140, "EA"),
    ("O-ring Viton 50x5mm", "Parker", "OR-V-50X5", "V", 15, "EA"),
    ("O-ring Viton 80x6mm", "Parker", "OR-V-80X6", "V", 22, "EA"),
    ("Correa V SPB 3000", "Gates", "SPB-3000", "V", 85, "EA"),
    ("Correa V SPB 3550", "Gates", "SPB-3550", "V", 95, "EA"),
    ("Diafragma bomba dosificadora MR-200", "Milton Roy", "MR-DIAF-200", "D", 450, "EA"),
    ("Quemador gas secador rotatorio", "Metso", "QMR-RD3600", "E", 25000, "EA"),
    # --- Additional items to reach 200+ ---
    ("Sello laberinto molino SAG lado alimentacion", "FLSmidth", "SLB-SAG40-F", "E", 28000, "EA"),
    ("Sello laberinto molino SAG lado descarga", "FLSmidth", "SLB-SAG40-D", "E", 28000, "EA"),
    ("Sello laberinto molino bolas", "FLSmidth", "SLB-BM26-F", "E", 22000, "EA"),
    ("Pinon corona molino SAG", "FLSmidth", "PCR-SAG40", "E", 380000, "EA"),
    ("Pinon corona molino bolas", "FLSmidth", "PCR-BM26", "E", 280000, "EA"),
    ("Pinon ataque molino bolas", "FLSmidth", "PAT-BM26", "E", 95000, "EA"),
    ("Convertidor frecuencia GMD", "ABB", "ACS6000-22MW", "E", 1800000, "EA"),
    ("Estator GMD repuesto", "FLSmidth", "GMD-STATOR-S", "E", 950000, "EA"),
    ("Descanso hidrostatico molino lado fijo", "FLSmidth", "DSC-SAG40-F", "E", 120000, "EA"),
    ("Descanso hidrostatico molino lado movil", "FLSmidth", "DSC-SAG40-M", "E", 120000, "EA"),
    ("Excitador harnero vibratorio Metso", "Metso", "EXC-TS8203", "E", 38000, "EA"),
    ("Resorte harnero vibratorio", "Metso", "RST-TS8203-SET", "D", 2800, "EA"),
    ("Polines carga estandar 6in", "Rulmeca", "PC-152-1200", "V", 120, "EA"),
    ("Polines retorno estandar 6in", "Rulmeca", "PR-152-1200", "V", 85, "EA"),
    ("Polines impacto 6in", "Rulmeca", "PI-152-1200", "D", 180, "EA"),
    ("Raspador primario correa Martin", "Martin Eng.", "QC1-1200", "D", 2200, "EA"),
    ("Raspador secundario correa Martin", "Martin Eng.", "SQC2-1200", "D", 1800, "EA"),
    ("Guia lateral correa 1200mm", "Martin Eng.", "GL-1200-SET", "V", 450, "EA"),
    ("Polea motriz correa CV-001", "Dodge", "PM-500-1200", "E", 15000, "EA"),
    ("Polea cola correa CV-001", "Dodge", "PC-400-1200", "D", 8500, "EA"),
    ("Soplador aire celda flotacion RCS300", "Metso", "BLW-RCS300", "E", 18000, "EA"),
    ("Difusor aire celda flotacion", "Metso", "DIF-RCS300", "D", 3200, "EA"),
    ("Valvula control aire flotacion 6in", "Fisher", "VCA-150-CV", "D", 6500, "EA"),
    ("Limitador torque espesador 25m", "FLSmidth", "TLM-HR25", "E", 22000, "EA"),
    ("Limitador torque espesador 45m", "FLSmidth", "TLM-HR45", "E", 28000, "EA"),
    ("Cilindro hidraulico chancador HP800", "Metso", "CYL-HP800-SET", "E", 42000, "EA"),
    ("Buje excentrico chancador HP800", "Metso", "BUJ-HP800", "E", 55000, "EA"),
    ("Eje principal chancador giratorio", "FLSmidth", "EJE-60113", "E", 320000, "EA"),
    ("Armadura inferior chancador giratorio", "FLSmidth", "ARM-60113-INF", "E", 85000, "EA"),
    ("Araña chancador giratorio", "FLSmidth", "ARA-60113", "E", 65000, "EA"),
    ("Bomba hidraulica Parker PV270", "Parker", "PV270R1K1T1N", "E", 28000, "EA"),
    ("Bomba hidraulica Rexroth A10V", "Rexroth", "A10VSO140DR", "D", 22000, "EA"),
    ("Acumulador hidraulico Hydac 20L", "Hydac", "SB330-20A1", "D", 3500, "EA"),
    ("Valvula direccional Rexroth 4WE10", "Rexroth", "4WE10D3X", "D", 1800, "EA"),
    ("Valvula proporcional Moog D634", "Moog", "D634-319C", "E", 8500, "EA"),
    ("Enfriador aceite Alfa Laval M10", "Alfa Laval", "M10-BFG-20", "D", 12000, "EA"),
    ("Enfriador aceite Hydac OK-EL6S", "Hydac", "OK-EL6S/3.0", "D", 8500, "EA"),
    ("Tanque aceite lubricacion 5000L", "Custom", "TNK-LUB-5000", "D", 15000, "EA"),
    ("Tanque aceite hidraulico 2000L", "Custom", "TNK-HID-2000", "D", 8000, "EA"),
    ("Calentador inmersion tanque aceite", "Chromalox", "CHI-TNK-15KW", "D", 3500, "EA"),
    ("Interruptor nivel aceite Gems", "Gems", "LS-7-SET", "V", 280, "EA"),
    ("Presostato Danfoss KP36", "Danfoss", "KP36-060-110", "V", 350, "EA"),
    ("Termostato Danfoss KP73", "Danfoss", "KP73-060-113", "V", 320, "EA"),
    ("Contactor Siemens 3RT1065", "Siemens", "3RT1065-6AF36", "D", 1200, "EA"),
    ("Interruptor automatico Siemens 3RV2", "Siemens", "3RV2041-4HA10", "D", 850, "EA"),
    ("Fusible NH Siemens 3NA", "Siemens", "3NA3252-2C", "V", 45, "EA"),
    ("Transformador corriente ABB CT5", "ABB", "CT5-200/5A", "D", 650, "EA"),
    ("Transformador potencial ABB VT", "ABB", "VT-4160/120V", "D", 1800, "EA"),
    ("Cable fuerza 3x240mm2 XLPE", "Prysmian", "XLP-3X240-15KV", "D", 85, "M"),
    ("Cable control 12x2.5mm2", "Prysmian", "CTL-12X2.5", "V", 12, "M"),
    ("Prensaestopa ATEX M50", "Hawke", "PE-M50-ATEX", "V", 45, "EA"),
    ("Caja conexion motor ATEX", "Hawke", "CJX-ATEX-L", "D", 850, "EA"),
    ("Termopar tipo K Omega", "Omega", "TJ36-CASS-116U", "V", 85, "EA"),
    ("RTD PT100 3-wire Omega", "Omega", "PR-11-3-100-1/4", "V", 120, "EA"),
    ("Transmisor temperatura Rosemount 644", "Emerson", "644HAE5J6M5", "V", 2800, "EA"),
    ("Indicador temperatura local bi-metal", "Wika", "A52.100-0/120C", "V", 65, "EA"),
    ("Manometro presion Wika 233.50", "Wika", "233.50-100-0/250", "V", 85, "EA"),
    ("Valvula seguridad Farris 2600", "Farris", "2600-JB-2x3-150", "D", 2200, "EA"),
    ("Valvula check wafer 6in Duo-Check", "TTV", "DC-150-6", "D", 1500, "EA"),
    ("Valvula bola 3in 150# inox", "Neles", "VBL-80-150-SS", "D", 1200, "EA"),
    ("Actuador electrico Auma SA07.1", "Auma", "SA07.1-F10", "D", 4500, "EA"),
    ("Actuador neumatico Bettis CB", "Bettis", "CB-200-SR", "D", 3200, "EA"),
    ("Posicionador Fisher DVC6200", "Fisher", "DVC6200-HC", "D", 5500, "EA"),
    ("Bomba centrifuga agua sello KSB", "KSB", "ETA-50-200", "D", 8500, "EA"),
    ("Bomba sumidero Flygt 3153", "Xylem", "3153.181-1643", "D", 12000, "EA"),
    ("Bomba dosificadora Prominent Sigma", "Prominent", "S2Ba-0413", "D", 3800, "EA"),
    ("Agitador portatil Chemineer", "Chemineer", "MR-250", "D", 5500, "EA"),
    ("Compresor scroll aire servicio Atlas", "Atlas Copco", "SF11-8-TM", "D", 15000, "EA"),
    ("Secador aire refrigerativo Atlas", "Atlas Copco", "FD300-VSD", "D", 8000, "EA"),
    ("Filtro aire comprimido Atlas", "Atlas Copco", "DD260-PD260", "D", 1200, "EA"),
    ("Tanque aire comprimido 3000L", "Custom", "TNK-AIR-3000", "D", 5000, "EA"),
    ("Extintor PQS 10kg", "Kidde", "PQS-10-ABC", "V", 85, "EA"),
    ("Detector gas H2S Drager", "Drager", "X-AM-2500-H2S", "D", 2800, "EA"),
    ("Luminaria antiexplosion LED", "Dialight", "LED-HZ-100W-EX", "V", 950, "EA"),
    ("Switch ethernet industrial Siemens", "Siemens", "XB208-6GK5208", "D", 1500, "EA"),
    ("Fuente alimentacion 24VDC Siemens", "Siemens", "6EP1334-3BA10", "D", 450, "EA"),
    ("UPS industrial 3kVA APC", "APC", "SURT3000XLIM", "D", 4500, "EA"),
    ("Convertidor fibra optica Moxa", "Moxa", "IMC-21-M-SC", "D", 850, "EA"),
    ("Panel operador Siemens KTP700", "Siemens", "6AV2123-2GB03", "D", 2200, "EA"),
    ("Sirena alarma industrial Federal Signal", "Federal Signal", "2001-AX", "V", 350, "EA"),
    ("Baliza luminosa LED roja", "Patlite", "LR6-102PJNW-R", "V", 180, "EA"),
    ("Aislador vibracion Trelleborg", "Trelleborg", "MESH-100X100", "V", 250, "EA"),
    ("Platina de desgaste AR400 20mm", "Custom", "AR400-20-1200X2400", "V", 380, "EA"),
    ("Platina de desgaste Hardox 500 25mm", "SSAB", "HX500-25-1200X2400", "V", 520, "EA"),
    ("Perfil estructural HEB200", "CAP Acero", "HEB200-S355", "V", 2.80, "KG"),
    ("Oxigeno industrial cilindro 10m3", "Linde", "O2-IND-10M3", "V", 45, "EA"),
    ("Acetileno cilindro 6kg", "Linde", "C2H2-6KG", "V", 55, "EA"),
    ("Electrodo soldadura E7018 3.25mm", "Lincoln", "E7018-3.25-20KG", "V", 85, "KG"),
    ("Electrodo soldadura E6011 3.25mm", "Lincoln", "E6011-3.25-20KG", "V", 65, "KG"),
    ("Disco corte 9in metal", "Norton", "BDA640-230X3", "V", 4.50, "EA"),
    ("Disco desbaste 9in metal", "Norton", "BDA620-230X6", "V", 5.20, "EA"),
    ("Cinta aislante 3M Scotch 33+", "3M", "SCOTCH33-19X20", "V", 8.50, "EA"),
    ("Silicona RTV roja 300ml", "Loctite", "RTV-SI-5910-300", "V", 12, "EA"),
    ("Adhesivo Loctite 243 50ml", "Loctite", "243-50ML", "V", 25, "EA"),
    ("Trapo limpieza industrial 10kg", "Custom", "TRP-IND-10KG", "V", 15, "KG"),
    ("Absorbente derrames oleosos", "3M", "ABS-OIL-P110", "V", 85, "EA"),
]

VED_MAP = {"E": "ESSENTIAL", "D": "DESIRABLE", "V": "VITAL"}

sp_count = 0
for i, (desc, mfr, pn, ved, cost, uom) in enumerate(SPARE_PARTS_BASE, 1):
    mat_code = f"S26-MAT-{i:04d}"
    sap_mat = f"{10000+i:012d}"
    fsn = "FAST_MOVING" if cost < 50 else "SLOW_MOVING" if cost > 50000 else "NORMAL"
    abc = "A" if cost > 50000 else "B" if cost > 5000 else "C"
    qty = random.randint(50, 5000) if cost < 50 else random.randint(1, 200) if cost < 1000 else random.randint(1, 20) if cost < 50000 else random.randint(0, 3)
    min_s = max(1, qty // 3)
    max_s = qty * 3 if qty > 0 else 5
    reorder = min_s * 2
    lead = random.choice([90, 120, 180, 240]) if cost > 50000 else random.choice([30, 45, 60, 90]) if cost > 5000 else random.choice([7, 14, 30])
    applicable = ", ".join([e["tag"] for e in random.sample(equipment_list, min(random.randint(2, 6), len(equipment_list)))])
    wh = random.choice(["ALM-SEC-01", "ALM-RIP-01", "ALM-HUM-01", "ALM-CENTRAL"])

    ws07.append([mat_code, sap_mat, desc, mfr, pn, VED_MAP[ved], fsn, abc,
                 qty, min_s, max_s, reorder, lead, cost, uom, applicable, wh])
    sp_count += 1

auto_width(ws07)
save_wb(wb07, "07_spare_parts_inventory.xlsx")
count_07 = sp_count


# ============================================================
# 08 — SHUTDOWN CALENDAR (≥200)
# ============================================================

print("[08] Generating 08_shutdown_calendar.xlsx ...")
wb08 = openpyxl.Workbook()
ws_cal = wb08.active
ws_cal.title = "Shutdown Calendar"
h08 = [
    "plant_id", "shutdown_id", "shutdown_name", "shutdown_type",
    "planned_start", "planned_end", "planned_hours",
    "areas_csv", "description", "status",
]
style_header(ws_cal, h08)

SHUTDOWNS = [
    ("SD-2026-01-MIN", "Parada Menor Enero Molienda", "MINOR_8H", "2026-01-15", "2026-01-15", 8, "Molienda", "Inspeccion rapida revestimientos SAG"),
    ("SD-2026-01-PM", "Parada Programada Enero Chancado", "MINOR_12H", "2026-01-28", "2026-01-28", 12, "Chancado", "Cambio mallas harneros e inspeccion correas"),
    ("SD-2026-02-MAJ", "Parada Mayor Febrero Planta", "MAJOR_72H", "2026-02-10", "2026-02-13", 72, "Chancado,Molienda,Flotacion,Espesado,Filtrado", "Overhaul semestral planta concentradora"),
    ("SD-2026-03-MIN", "Parada Menor Marzo Flotacion", "MINOR_8H", "2026-03-12", "2026-03-12", 8, "Flotacion", "Cambio rotores celdas rougher"),
    ("SD-2026-03-PM", "Parada Programada Marzo Filtrado", "MINOR_12H", "2026-03-25", "2026-03-25", 12, "Filtrado", "Cambio telas filtrantes FP-01"),
    ("SD-2026-04-MIN", "Parada Menor Abril Chancado", "MINOR_8H", "2026-04-09", "2026-04-09", 8, "Chancado", "Cambio mantos chancador secundario"),
    ("SD-2026-04-PM", "Parada Programada Abril Espesado", "MINOR_12H", "2026-04-22", "2026-04-22", 12, "Espesado", "Inspeccion rastras espesador concentrado"),
    ("SD-2026-05-MAJ", "Parada Mayor Mayo Molienda", "MAJOR_48H", "2026-05-13", "2026-05-15", 48, "Molienda", "Cambio revestimientos SAG y bolas, overhaul bombas"),
    ("SD-2026-06-MIN", "Parada Menor Junio Flotacion", "MINOR_8H", "2026-06-11", "2026-06-11", 8, "Flotacion", "Inspeccion celdas cleaner-scavenger"),
    ("SD-2026-06-PM", "Parada Programada Junio General", "MINOR_12H", "2026-06-24", "2026-06-24", 12, "Chancado,Molienda", "Inspeccion general area seca y molienda"),
    ("SD-2026-07-MIN", "Parada Menor Julio Chancado", "MINOR_8H", "2026-07-15", "2026-07-15", 8, "Chancado", "Cambio mallas y ajuste chancador terciario"),
    ("SD-2026-08-MAJ", "Parada Mayor Agosto Planta", "MAJOR_72H", "2026-08-12", "2026-08-15", 72, "Chancado,Molienda,Flotacion,Espesado,Filtrado", "Parada general semestral concentradora"),
    ("SD-2026-09-MIN", "Parada Menor Sept Molienda", "MINOR_8H", "2026-09-10", "2026-09-10", 8, "Molienda", "Inspeccion GMD e hidrociclones"),
    ("SD-2026-09-PM", "Parada Programada Sept Filtrado", "MINOR_12H", "2026-09-23", "2026-09-23", 12, "Filtrado", "Cambio placas y telas FP-02"),
    ("SD-2026-10-MIN", "Parada Menor Oct Espesado", "MINOR_8H", "2026-10-08", "2026-10-08", 8, "Espesado", "Mantencion espesador relaves"),
    ("SD-2026-11-MAJ", "Parada Mayor Nov Molienda", "MAJOR_48H", "2026-11-18", "2026-11-20", 48, "Molienda", "Cambio revestimientos y overhaul completo"),
    ("SD-2026-12-MIN", "Parada Menor Dic Flotacion", "MINOR_8H", "2026-12-10", "2026-12-10", 8, "Flotacion", "Inspeccion final ano celdas"),
    ("SD-2026-12-PM", "Parada Programada Dic General", "MINOR_12H", "2026-12-18", "2026-12-18", 12, "Chancado,Filtrado", "Preparacion parada mayor Q1 2027"),
]

for sd in SHUTDOWNS:
    ws_cal.append([PLANT_ID, *sd, "PLANNED"])

# Shutdown Work Packages — generate enough for 200+ total
ws_swp = wb08.create_sheet("Shutdown Work Packages")
h08b = ["shutdown_id", "wp_code", "equipment_tag", "sap_func_loc", "description",
        "estimated_hours", "work_center", "specialty", "crew_size", "materials_required"]
style_header(ws_swp, h08b)

swp_count = 0
target_swp = 200 - len(SHUTDOWNS)
swp_per_sd = max(10, target_swp // len(SHUTDOWNS) + 1)

for sd in SHUTDOWNS:
    sd_id = sd[0]
    for j in range(random.randint(swp_per_sd - 2, swp_per_sd + 2)):
        eq = random.choice(equipment_list)
        spec = random.choice(SPECIALTIES)
        wcs = PG_TO_WC[eq["pg"]]
        wc = random.choice(wcs)
        ws_swp.append([
            sd_id, f"WP-{sd_id}-{j+1:03d}", eq["tag"], eq["tplnr"],
            f"Trabajo parada: {eq['eqktx'][:35]}",
            round(random.uniform(2, 16), 1), wc, spec, random.randint(2, 6),
            random.choice(["Si", "No"]),
        ])
        swp_count += 1

auto_width(ws_cal)
auto_width(ws_swp)
save_wb(wb08, "08_shutdown_calendar.xlsx")
count_08 = len(SHUTDOWNS) + swp_count


# ============================================================
# 09 — WORKFORCE (≥200)
# Turno Blueprint: 7x7 (Viernes-Jueves) 08:00-20:00
# ============================================================

print("[09] Generating 09_workforce.xlsx ...")
wb09 = openpyxl.Workbook()
ws09 = wb09.active
ws09.title = "Workforce"
h09 = [
    "worker_id", "name", "specialty", "specialty_level", "labor_category", "rate_usd_hr",
    "shift_code", "shift_desc", "shift_start", "shift_end", "shift_rotation",
    "plant_id", "planning_group", "work_center", "available",
    "certifications_csv", "phone", "email",
]
style_header(ws09, h09)

FIRST_NAMES = ["Carlos", "Juan", "Roberto", "Miguel", "Andres", "Diego", "Pedro", "Luis",
    "Fernando", "Jorge", "Oscar", "Ricardo", "Eduardo", "Gabriel", "Patricio",
    "Hector", "Sergio", "Raul", "Manuel", "Francisco", "Alejandro", "Cristian",
    "Daniel", "Esteban", "Ivan", "Tomas", "Nicolas", "Sebastian", "Claudio",
    "Gonzalo", "Matias", "Felipe", "Rodrigo", "Pablo", "Ignacio", "Marcelo",
    "Mauricio", "Leonardo", "Alvaro", "Camilo", "Benjamin", "Vicente", "Martin",
    "Joaquin", "Maximiliano", "Agustin", "Bastian", "Renato", "Emilio", "Hugo",
    "Ramon", "Arturo", "Ernesto", "Rolando", "Jaime", "Marco", "Antonio",
    "Guillermo", "Hernan", "Cesar", "Danilo", "Fabian", "German", "Waldo",
    "Alfredo", "Jonathan", "Christian", "Alex", "Victor", "Angel"]
LAST_NAMES = ["Mendoza", "Perez", "Silva", "Torres", "Rojas", "Fuentes", "Castillo", "Ramirez",
    "Vega", "Nunez", "Diaz", "Morales", "Soto", "Reyes", "Araya", "Bravo",
    "Contreras", "Espinoza", "Gutierrez", "Herrera", "Ibarra", "Jara", "Leiva",
    "Molina", "Orellana", "Paredes", "Quiroz", "Riquelme", "Sepulveda", "Tapia",
    "Valenzuela", "Zamora", "Figueroa", "Lagos", "Muñoz", "Pizarro", "Salas",
    "Urbina", "Vera", "Alarcon", "Barrera", "Carvajal", "Delgado", "Flores",
    "Gonzalez", "Henriquez", "Inostroza", "Jeldres", "Kaulen", "Lopez",
    "Martinez", "Navarro", "Ortiz", "Palma", "Quezada", "Rios", "San Martin",
    "Toledo", "Uribe", "Villalobos", "Yanez", "Zuniga"]

CERTS = ["Trabajo en Altura", "Espacio Confinado", "Izaje Critico", "Bloqueo/Etiquetado LOTO",
         "Materiales Peligrosos", "Soldadura ASME IX", "NDE Nivel II", "Alineamiento Laser",
         "Termografia Nivel I", "Analisis Vibracion Cat II", "Electricidad MT/AT",
         "Operacion Grua Puente", "Rigger Certificado", "Primeros Auxilios"]

# Generate 200+ workers
WORKER_SPECS = (
    [("MEC", "I")] * 30 + [("MEC", "II")] * 25 + [("MEC", "III")] * 12 +
    [("ELE", "I")] * 20 + [("ELE", "II")] * 15 + [("ELE", "III")] * 8 +
    [("INS", "I")] * 15 + [("INS", "II")] * 10 +
    [("SOL", "I")] * 12 + [("SOL", "II")] * 8 +
    [("LUB", "I")] * 10 +
    [("SIN", "I")] * 8 +
    [("DCS", "I")] * 6 +
    [("SUP", "")] * 21
)

used_names = set()
wk_count = 0
for i, (spec, level) in enumerate(WORKER_SPECS, 1):
    while True:
        fn = random.choice(FIRST_NAMES)
        ln = random.choice(LAST_NAMES)
        name = f"{fn} {ln}"
        if name not in used_names:
            used_names.add(name)
            break

    wid = f"S26-W-{i:03d}"
    cat = f"{spec}-{level}" if level else "SUP"
    rate = LABOR_CATEGORIES.get(cat, LABOR_CATEGORIES["SUP"])["rate_usd"]
    pg = random.choice(["P01", "P02", "P03"])
    wc = random.choice(PG_TO_WC[pg])
    certs = ", ".join(random.sample(CERTS, k=random.randint(2, 5)))
    phone = f"+56 9 {random.randint(5000,9999)} {random.randint(1000,9999)}"
    email = f"{fn.lower()}.{ln.lower()}@ocp-con.cl"

    ws09.append([
        wid, name, spec, level if level else "Supervisor", cat, rate,
        SHIFT_PLANTA["code"], SHIFT_PLANTA["desc"], SHIFT_PLANTA["start"], SHIFT_PLANTA["end"], SHIFT_PLANTA["rotation"],
        PLANT_ID, pg, wc, True, certs, phone, email,
    ])
    wk_count += 1

auto_width(ws09)
save_wb(wb09, "09_workforce.xlsx")
count_09 = wk_count


# ============================================================
# 10 — FIELD CAPTURE (≥200)
# ============================================================

print("[10] Generating 10_field_capture.xlsx ...")
wb10 = openpyxl.Workbook()
ws10 = wb10.active
ws10.title = "Field Captures"
h10 = [
    "capture_id", "worker_id", "capture_type", "language",
    "equipment_tag", "sap_func_loc", "location_hint", "raw_text",
    "timestamp", "area", "severity",
]
style_header(ws10, h10)

FIELD_TEXTS = [
    ("VOICE", "es", "Se escucha ruido anormal en rodamiento lado libre del {eq}"),
    ("TEXT", "es", "Temperatura elevada en reductor {eq}, marca 85C, limite es 75C"),
    ("VOICE", "es", "Fuga de aceite en sello principal de {eq}, goteo constante"),
    ("TEXT", "es", "Vibracion alta detectada en {eq}, 12mm/s RMS, limite 8mm/s"),
    ("VOICE", "es", "Correa de {eq} presenta desgaste lateral, 30% vida util"),
    ("TEXT", "es", "Perno suelto en tapa de {eq}, necesita retorque inmediato"),
    ("VOICE", "es", "Bomba {eq} cavitando, ruido tipico y baja presion descarga"),
    ("TEXT", "es", "Nivel aceite {eq} bajo minimo, requiere relleno urgente"),
    ("VOICE", "es", "Revestimiento de {eq} con desgaste avanzado zona 3, menos 15mm"),
    ("TEXT", "es", "Alineamiento {eq} fuera tolerancia, shim requerido lado acople"),
    ("VOICE", "es", "Alarma alta temperatura motor de {eq}, 95C en enrollado"),
    ("TEXT", "es", "Presion hidraulica {eq} oscilando 180-220 bar, debe ser 200 estable"),
    ("VOICE", "es", "Se observa chispa en escobillas motor {eq}, revisar urgente"),
    ("TEXT", "es", "Flujo bomba {eq} disminuido 15% respecto turno anterior"),
    ("VOICE", "es", "Malla harnero {eq} rota en seccion central, requiere cambio"),
    ("TEXT", "es", "Corrosion visible en estructura soporte {eq}, registrar para PM"),
    ("VOICE", "es", "Ruido metalico intermitente en reductor {eq}, posible pitting"),
    ("TEXT", "es", "Sensor temperatura {eq} lectura erratica, calibrar o reemplazar"),
    ("VOICE", "es", "Fuga agua sello de {eq}, aumentar presion agua sello"),
    ("TEXT", "es", "Acoplamiento {eq} con desgaste elastomero visible"),
]

fc_count = 0
for i in range(220):
    worker_idx = random.randint(1, min(wk_count, 200))
    eq = random.choice(equipment_list)
    tpl = random.choice(FIELD_TEXTS)
    ts = datetime(2026, 1, 1) + timedelta(days=random.randint(0, 88), hours=random.randint(8, 19), minutes=random.randint(0, 59))
    severity = random.choice(["LOW", "LOW", "MEDIUM", "MEDIUM", "MEDIUM", "HIGH", "HIGH", "CRITICAL"])

    ws10.append([
        f"S26-FC-{i+1:04d}", f"S26-W-{worker_idx:03d}", tpl[0], tpl[1],
        eq["tag"], eq["tplnr"],
        f"Nivel {random.choice(['0','1','2','3'])}, {eq['area_name']}",
        tpl[2].format(eq=eq["eqktx"][:35])[:200],
        ts.strftime("%Y-%m-%d %H:%M:%S"), eq["area_name"], severity,
    ])
    fc_count += 1

auto_width(ws10)
save_wb(wb10, "10_field_capture.xlsx")
count_10 = fc_count


# ============================================================
# 12 — PLANNING KPI INPUT (≥200)
# Per area per week = 3 areas x 13 weeks = 39 per quarter
# 4 quarters + plant-level = enough for 200+
# ============================================================

print("[12] Generating 12_planning_kpi_input.xlsx ...")
wb12 = openpyxl.Workbook()
ws12 = wb12.active
ws12.title = "Planning KPI Input"
h12 = [
    "plant_id", "planning_group", "business_area", "period_start", "period_end",
    "wo_planned", "wo_completed", "manhours_planned", "manhours_actual",
    "pm_planned", "pm_executed", "backlog_hours", "weekly_capacity_hours",
    "corrective_count", "total_wo", "schedule_compliance_pct",
    "release_horizon_days", "pending_notices", "total_notices",
    "scheduled_capacity_hours", "total_capacity_hours",
    "proactive_wo", "planned_wo",
]
style_header(ws12, h12)

kpi_count = 0
start = datetime(2026, 1, 6)
for week in range(52):  # Full year
    ps = start + timedelta(weeks=week)
    pe = ps + timedelta(days=6)
    for pg_code, pg_name in PLANNING_GROUPS.items():
        ba = PROCESS_TO_BA[pg_code]
        wo_plan = random.randint(10, 25)
        wo_comp = int(wo_plan * random.uniform(0.75, 0.98))
        mh_plan = round(wo_plan * random.uniform(6, 12), 1)
        mh_act = round(mh_plan * random.uniform(0.85, 1.15), 1)
        pm_plan = random.randint(5, 15)
        pm_exec = int(pm_plan * random.uniform(0.8, 1.0))
        backlog = round(random.uniform(60, 300), 0)
        cap = round(10 * random.uniform(35, 50), 0)
        corr = random.randint(2, 8)
        total_wo = wo_comp + corr
        sched_comp = round(wo_comp / max(wo_plan, 1) * 100, 1)
        proactive = pm_exec + random.randint(1, 4)

        ws12.append([
            PLANT_ID, pg_code, ba, ps.strftime("%Y-%m-%d"), pe.strftime("%Y-%m-%d"),
            wo_plan, wo_comp, mh_plan, mh_act, pm_plan, pm_exec,
            backlog, cap, corr, total_wo, sched_comp,
            random.choice([7, 14, 21]), random.randint(2, 10), random.randint(10, 30),
            round(cap * 0.85, 0), cap, proactive, wo_plan,
        ])
        kpi_count += 1

    # Plant-level weekly summary row
    ws12.append([
        PLANT_ID, "ALL", "ALL", ps.strftime("%Y-%m-%d"), pe.strftime("%Y-%m-%d"),
        random.randint(30, 60), random.randint(25, 55),
        round(random.uniform(200, 500), 1), round(random.uniform(180, 520), 1),
        random.randint(15, 40), random.randint(12, 38),
        round(random.uniform(200, 900), 0), round(random.uniform(1000, 1500), 0),
        random.randint(6, 20), random.randint(35, 75),
        round(random.uniform(70, 95), 1),
        random.choice([7, 14, 21]), random.randint(8, 25), random.randint(30, 80),
        round(random.uniform(900, 1300), 0), round(random.uniform(1000, 1500), 0),
        random.randint(15, 45), random.randint(30, 60),
    ])
    kpi_count += 1

auto_width(ws12)
save_wb(wb12, "12_planning_kpi_input.xlsx")
count_12 = kpi_count


# ============================================================
# 13 — DE KPI INPUT (≥200)
# Per area per month = 3 areas x 12 months = 36/year. Do multi-year + details.
# ============================================================

print("[13] Generating 13_de_kpi_input.xlsx ...")
wb13 = openpyxl.Workbook()
ws13 = wb13.active
ws13.title = "DE KPI Input"
h13 = [
    "plant_id", "planning_group", "business_area", "period_start", "period_end",
    "events_reported", "events_required", "meetings_held", "meetings_required",
    "actions_implemented", "actions_planned",
    "savings_achieved_usd", "savings_target_usd",
    "failures_current_period", "failures_previous_period",
    "chronic_failures_identified", "chronic_failures_resolved",
    "de_maturity_score",
]
style_header(ws13, h13)

de_count = 0
# 2024-2026, 3 years x 12 months x 3 PGs = 108, plus plant-level summaries
for year in [2022, 2023, 2024, 2025, 2026]:
    for month in range(1, 13):
        if year == 2026 and month > 3:
            break  # Only Q1 2026
        ps = datetime(year, month, 1)
        if month < 12:
            pe = datetime(year, month + 1, 1) - timedelta(days=1)
        else:
            pe = datetime(year, 12, 31)

        for pg_code, pg_name in PLANNING_GROUPS.items():
            ba = PROCESS_TO_BA[pg_code]
            ev_req = 8
            ev_rep = random.randint(4, 10)
            meet_req = 4
            meet_held = random.randint(2, 4)
            act_plan = random.randint(8, 20)
            act_impl = int(act_plan * random.uniform(0.5, 0.95))
            sav_target = 30000
            sav_ach = round(sav_target * random.uniform(0.4, 1.3), 2)
            chronic_id = random.randint(1, 5)
            chronic_res = random.randint(0, chronic_id)
            maturity = round(random.uniform(1.5, 4.5), 1)

            ws13.append([
                PLANT_ID, pg_code, ba, ps.strftime("%Y-%m-%d"), pe.strftime("%Y-%m-%d"),
                ev_rep, ev_req, meet_held, meet_req, act_impl, act_plan,
                sav_ach, sav_target,
                random.randint(5, 18), random.randint(8, 22),
                chronic_id, chronic_res, maturity,
            ])
            de_count += 1

        # Plant-level summary row
        ws13.append([
            PLANT_ID, "ALL", "ALL", ps.strftime("%Y-%m-%d"), pe.strftime("%Y-%m-%d"),
            random.randint(15, 30), 24, random.randint(8, 12), 12,
            random.randint(20, 50), random.randint(30, 60),
            round(random.uniform(50000, 120000), 2), 90000,
            random.randint(15, 45), random.randint(20, 55),
            random.randint(3, 10), random.randint(1, 8),
            round(random.uniform(2.0, 4.0), 1),
        ])
        de_count += 1

auto_width(ws13)
save_wb(wb13, "13_de_kpi_input.xlsx")
count_13 = de_count


# ============================================================
# 14 — MAINTENANCE STRATEGY (≥200)
# ============================================================

print("[14] Generating 14_maintenance_strategy.xlsx ...")
wb14 = openpyxl.Workbook()
ws14 = wb14.active
ws14.title = "Strategies"
h14 = [
    "strategy_id", "equipment_tag", "equnr", "sap_func_loc", "area",
    "subunit", "maintainable_item",
    "function_and_failure", "mechanism", "cause", "status",
    "tactics_type", "primary_task_name", "primary_task_interval_weeks",
    "constraint", "task_type", "route_sheet_type", "access_time_hours",
    "secondary_task_name", "secondary_task_interval_weeks",
    "maint_plan_type", "budgeted_cost_usd",
    "justification_category", "justification", "notes",
]
style_header(ws14, h14)

# Blueprint: time-based + activity-based (enhanced with condition-based per industry standard)
TACTICS = ["TIME_BASED", "ACTIVITY_BASED", "CONDITION_BASED"]
TASK_TYPES = ["INSPECT", "REPLACE", "OVERHAUL", "CONDITION_MONITOR", "LUBRICATE", "TEST"]
CONSTRAINTS = ["ONLINE", "OFFLINE", "SHUTDOWN_ONLY"]
JUST_CATS = ["SAFETY", "ENVIRONMENTAL", "OPERATIONAL", "ECONOMIC"]

PRIMARY_TASKS = [
    "Inspeccion visual y termografica", "Analisis vibracion",
    "Analisis aceite lubricante", "Lubricacion programada",
    "Cambio filtros aceite", "Inspeccion NDT ultrasonido",
    "Calibracion instrumentos", "Prueba funcional proteccion",
    "Torqueo pernos criticos", "Medicion espesores ultrasonido",
    "Monitoreo condicion en linea", "Rondas operador estructuradas",
]
SECONDARY_TASKS = [
    "Overhaul programado", "Reemplazo componente mayor",
    "Reparacion mayor taller", "Rebuild completo equipo",
    "Cambio rodamientos programado", "Cambio sellos programado",
]

strat_count = 0
min_strat = max(4, 200 // len(equipment_list) + 1)

for eq in equipment_list:
    n_strats = random.randint(min_strat, min_strat + 2)
    for s in range(n_strats):
        strat_id = synth_id("STR", strat_count)
        mech = random.choice(MECHANISMS)
        cause = random.choice(CAUSES)
        tactic = random.choice(TACTICS)
        primary = random.choice(PRIMARY_TASKS)
        secondary = random.choice(SECONDARY_TASKS)
        p_interval = random.choice([2, 4, 8, 13, 26, 52])
        s_interval = p_interval * random.choice([4, 6, 8, 12])
        budget = round(random.uniform(500, 80000), 2)

        ws14.append([
            strat_id, eq["tag"], eq["equnr"], eq["tplnr"], eq["area_name"],
            random.choice(["MECHANICAL", "ELECTRICAL", "HYDRAULIC", "STRUCTURAL"]),
            random.choice(MAINT_ITEMS),
            f"{mech} por {cause} en {eq['eqktx'][:30]}",
            mech, cause, "APPROVED",
            tactic, primary, p_interval,
            random.choice(CONSTRAINTS), random.choice(TASK_TYPES),
            ROUTE_SHEET_TYPE, round(random.uniform(0.5, 8), 1),
            secondary, s_interval,
            MAINT_PLAN_TYPE, budget,
            random.choice(JUST_CATS),
            f"Justificado por analisis RCM - {tactic.lower().replace('_', ' ')}",
            "",
        ])
        strat_count += 1

auto_width(ws14)
save_wb(wb14, "14_maintenance_strategy.xlsx")
count_14 = strat_count


# ============================================================
# SUMMARY
# ============================================================

print("\n" + "=" * 70)
print("  RESUMEN — DB_AM_OCP_SYNTHETIC_2026 v2 (Blueprint-Aligned)")
print("=" * 70)

summary = {
    "01_equipment_hierarchy": count_01,
    "02_criticality_assessment": count_02,
    "03_failure_modes": count_03,
    "04_maintenance_tasks": 0,
    "05_work_packages": count_05,
    "06_work_order_history": count_06,
    "07_spare_parts_inventory": count_07,
    "08_shutdown_calendar": count_08,
    "09_workforce": count_09,
    "10_field_capture": count_10,
    "11_rca_events": 0,
    "12_planning_kpi_input": count_12,
    "13_de_kpi_input": count_13,
    "14_maintenance_strategy": count_14,
}

total = 0
for name, cnt in summary.items():
    status = "EXCLUIDO" if cnt == 0 else ("OK >=200" if cnt >= 200 else f"OK ({cnt})")
    print(f"  {name:.<45s} {cnt:>6d}  [{status}]")
    total += cnt

print(f"\n  {'TOTAL':.<45s} {total:>6d} registros")
print(f"\n  === ALINEACION BLUEPRINT AMSA_BBP_PM_04_Rev_0 ===")
print(f"  Centro Planificacion: {PLANNING_CENTER} (Blueprint)")
print(f"  Grupos Planif: P01 (Area seca), P02 (Area Ripio), P03 (Area Humeda)")
print(f"  Tipos OT: PM01 | PM02 | PM03 | PM06 | PM07 (Blueprint — NO PM04)")
print(f"  Clases Actividad PM07: RP1 | RP2 (Blueprint)")
print(f"  Criticidad ABC: 1=Alto | 2=Medio | 3=Bajo (Blueprint numerico)")
print(f"  Status Notif ZPM00001: APRO | RECH (Blueprint)")
print(f"  Status Orden SAP: ABIE | LIBE | NOTI | CTEC")
print(f"  Prioridad clase Z1: I | A | M | B (Blueprint)")
print(f"  Tipos Equipo SAP: M (Maquinas) | Q (Inspeccion/Medida)")
print(f"  Tipo UbicTecnica: {FUNC_LOC_TYPE} (Sistema tecnico estandar)")
print(f"  Hoja Ruta tipo: {ROUTE_SHEET_TYPE} (Instruccion), grupo: PLA")
print(f"  Plan Mant tipo: {MAINT_PLAN_TYPE}")
print(f"  Categorias Valor: ZMANT001 | ZMANT002 | ZMANT003")
print(f"  Turno Planta: {SHIFT_PLANTA['code']} ({SHIFT_PLANTA['rotation']}) {SHIFT_PLANTA['start']}-{SHIFT_PLANTA['end']}")
print(f"  Jerarquia: 6 niveles (Corp + MANTE NN-NN-NN-AAAANN-XXXX-XXXX)")
print("=" * 70)
