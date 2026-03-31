"""Tests for RFI Excel processor."""

from datetime import date
from pathlib import Path

import pytest
import yaml

from scripts.generate_rfi_template import generate_rfi_template
from scripts.process_ams_rfi import (
    _append_or_create,
    _safe_bool,
    _safe_date,
    _safe_enum,
    _safe_float,
    _safe_int,
    _safe_str,
    _split_list,
    build_project_yaml,
    generate_data_availability_report,
    generate_followup,
    generate_global_requirements,
    generate_kpi_appendix,
    generate_maintenance_strategy_requirements,
    generate_scope_assessment,
    generate_standards_appendix,
    generate_work_planning_requirements,
    process_rfi,
)
from tools.models.rfi_models import (
    AMS_TEMPLATES,
    CMMSType,
    CompanySiteProfile,
    CriticalityMethodRFI,
    DataAvailabilityChecklist,
    DataAvailabilityItem,
    DataFormat,
    EquipmentHierarchyData,
    Industry,
    KPIBaselineTargets,
    Language,
    MaintenanceCurrentState,
    OrganizationResources,
    RFISubmission,
    SAPVersion,
    ScopeTimeline,
    ScopeType,
    StandardsCompliance,
    StrategyMaturity,
)
from tools.validators.quality_validator import ValidationResult


# ---------------------------------------------------------------------------
# Safe cell readers
# ---------------------------------------------------------------------------

class TestSafeStr:
    def test_none(self):
        assert _safe_str(None) == ""

    def test_string(self):
        assert _safe_str("hello") == "hello"

    def test_number(self):
        assert _safe_str(42) == "42"

    def test_whitespace(self):
        assert _safe_str("  test  ") == "test"


class TestSafeInt:
    def test_none(self):
        assert _safe_int(None) is None

    def test_valid_int(self):
        assert _safe_int(42) == 42

    def test_float_truncation(self):
        assert _safe_int(42.7) == 42

    def test_string_int(self):
        assert _safe_int("100") == 100

    def test_invalid(self):
        assert _safe_int("abc") is None

    def test_default(self):
        assert _safe_int("abc", default=0) == 0


class TestSafeFloat:
    def test_none(self):
        assert _safe_float(None) is None

    def test_valid(self):
        assert _safe_float(3.14) == 3.14

    def test_string_float(self):
        assert _safe_float("95.5") == 95.5

    def test_invalid(self):
        assert _safe_float("abc") is None


class TestSafeBool:
    def test_none(self):
        assert _safe_bool(None) is None

    def test_yes(self):
        assert _safe_bool("Yes") is True

    def test_no(self):
        assert _safe_bool("No") is False

    def test_true(self):
        assert _safe_bool("True") is True

    def test_partial(self):
        assert _safe_bool("Partial") is True

    def test_case_insensitive(self):
        assert _safe_bool("YES") is True
        assert _safe_bool("no") is False


class TestSafeDate:
    def test_none(self):
        assert _safe_date(None) is None

    def test_date_object(self):
        d = date(2026, 3, 5)
        assert _safe_date(d) == d

    def test_string(self):
        assert _safe_date("2026-03-05") == date(2026, 3, 5)

    def test_invalid(self):
        assert _safe_date("not-a-date") is None


class TestSafeEnum:
    def test_valid(self):
        assert _safe_enum("mining", Industry) == Industry.MINING

    def test_case_insensitive(self):
        assert _safe_enum("MINING", Industry) == Industry.MINING

    def test_none(self):
        assert _safe_enum(None, Industry) is None

    def test_invalid(self):
        assert _safe_enum("nonexistent", Industry) is None


class TestSplitList:
    def test_none(self):
        assert _split_list(None) == []

    def test_semicolon(self):
        assert _split_list("a;b;c") == ["a", "b", "c"]

    def test_whitespace(self):
        assert _split_list(" a ; b ; c ") == ["a", "b", "c"]

    def test_empty_items(self):
        assert _split_list("a;;b") == ["a", "b"]


# ---------------------------------------------------------------------------
# Output generators
# ---------------------------------------------------------------------------

def _make_complete_submission() -> RFISubmission:
    """Helper to create a complete submission for testing."""
    return RFISubmission(
        company_site=CompanySiteProfile(
            company_name="OCP Group",
            industry=Industry.MINING,
            plant_name="Jorf Fertilizer Complex",
            plant_code="OCP-JFC",
            location="El Jadida, Morocco",
            country="MA",
            primary_language=Language.FR,
            contact_name="John Doe",
            contact_email="john@ocp.ma",
        ),
        equipment_hierarchy=EquipmentHierarchyData(
            equipment_list_available=True,
            equipment_list_format=DataFormat.SAP_EXPORT,
            estimated_equipment_count=500,
            hierarchy_levels=6,
            naming_convention="AREA-SYSTEM-EQUIP-NNN",
            tag_format_example="SAG-MILL-001",
            bom_available=True,
            bom_format=DataFormat.SAP_EXPORT,
        ),
        maintenance_state=MaintenanceCurrentState(
            strategy_maturity=StrategyMaturity.DEVELOPING,
            cmms_type=CMMSType.SAP_PM,
            sap_version=SAPVersion.S4HANA,
            wo_history_available=True,
            wo_history_years=5,
            failure_data_available=True,
            failure_data_format=DataFormat.SAP_EXPORT,
            planned_maintenance_exists=True,
            pm_plan_format=DataFormat.SAP_EXPORT,
            prior_criticality_assessment=True,
            criticality_method=CriticalityMethodRFI.R8_4LEVEL,
        ),
        organization=OrganizationResources(
            team_size=50,
            org_structure="area-based",
            shifts=3,
            trades=["mechanical", "electrical", "instrumentation", "welding"],
        ),
        standards=StandardsCompliance(
            procedure_language=Language.FR,
            iso_certifications=["ISO 55001", "ISO 14001", "ISO 45001"],
            safety_permits_required=["LOTOTO", "confined-space", "working-at-height"],
            lototo_program=True,
        ),
        kpi_baseline=KPIBaselineTargets(
            current_availability=85.0,
            target_availability=95.0,
        ),
        scope_timeline=ScopeTimeline(
            scope_type=ScopeType.SYSTEM,
            start_date=date(2026, 2, 1),
            target_completion=date(2026, 6, 30),
            areas_in_scope=["Phosphate Processing"],
            priority_equipment=["SAG-MILL-001"],
        ),
        data_availability=DataAvailabilityChecklist(
            items=[
                DataAvailabilityItem(
                    template_id=tid, template_name=name,
                    available=(i < 7),
                    data_format=DataFormat.EXCEL if i < 7 else None,
                    quality_score=3 if i < 7 else None,
                )
                for i, (tid, name) in enumerate(AMS_TEMPLATES)
            ],
        ),
    )


class TestBuildProjectYaml:
    def test_structure(self):
        sub = _make_complete_submission()
        result = build_project_yaml(sub)

        # Top-level keys
        assert "project" in result
        assert "client" in result
        assert "scope" in result
        assert "maintenance_context" in result
        assert "organization" in result
        assert "standards" in result
        assert "kpi_baseline" in result

    def test_client_data(self):
        sub = _make_complete_submission()
        result = build_project_yaml(sub)

        assert result["client"]["name"] == "OCP Group"
        assert result["client"]["industry"] == "mining"
        assert result["client"]["country"] == "MA"
        assert result["client"]["language"] == "fr"

    def test_scope_data(self):
        sub = _make_complete_submission()
        result = build_project_yaml(sub)

        assert result["scope"]["type"] == "system"
        assert result["scope"]["plant"]["name"] == "Jorf Fertilizer Complex"
        assert result["scope"]["plant"]["code"] == "OCP-JFC"
        assert result["scope"]["priority_equipment"] == ["SAG-MILL-001"]

    def test_maintenance_context(self):
        sub = _make_complete_submission()
        result = build_project_yaml(sub)

        mc = result["maintenance_context"]
        assert mc["cmms"] == "sap-pm"
        assert mc["sap_version"] == "S/4HANA"
        assert mc["availability_target"] == 0.95
        assert mc["existing_data"]["equipment_list"] is True

    def test_yaml_roundtrip(self):
        sub = _make_complete_submission()
        result = build_project_yaml(sub)
        yaml_str = yaml.dump(result, default_flow_style=False, allow_unicode=True)
        parsed = yaml.safe_load(yaml_str)
        assert parsed["client"]["name"] == "OCP Group"

    def test_empty_submission(self):
        sub = RFISubmission()
        result = build_project_yaml(sub)
        assert result["client"]["name"] is None
        assert result["kpi_baseline"]["availability"] is None


class TestDataAvailabilityReport:
    def test_contains_table(self):
        sub = _make_complete_submission()
        report = generate_data_availability_report(sub)
        assert "| Template |" in report
        assert "Equipment Hierarchy" in report

    def test_summary_counts(self):
        sub = _make_complete_submission()
        report = generate_data_availability_report(sub)
        assert "Data available" in report


class TestScopeAssessment:
    def test_contains_scope(self):
        sub = _make_complete_submission()
        report = generate_scope_assessment(sub)
        assert "Scope Assessment" in report
        assert "500" in report  # equipment count
        assert "High" in report  # complexity for 500 items

    def test_recommended_milestone_with_criticality(self):
        sub = _make_complete_submission()
        report = generate_scope_assessment(sub)
        assert "Milestone 2" in report  # has criticality but also FMECA

    def test_recommended_milestone_without_data(self):
        sub = RFISubmission()
        report = generate_scope_assessment(sub)
        assert "Milestone 1" in report


class TestGlobalRequirements:
    def test_contains_language(self):
        sub = _make_complete_submission()
        report = generate_global_requirements(sub)
        assert "fr" in report
        assert "work instructions MUST be written in" in report.lower() or "Work instructions MUST" in report

    def test_contains_naming(self):
        sub = _make_complete_submission()
        report = generate_global_requirements(sub)
        assert "SAG-MILL-001" in report

    def test_contains_standards(self):
        sub = _make_complete_submission()
        report = generate_global_requirements(sub)
        assert "ISO 55001" in report

    def test_contains_cmms(self):
        sub = _make_complete_submission()
        report = generate_global_requirements(sub)
        assert "sap-pm" in report


class TestFollowup:
    def test_none_when_no_issues(self):
        result = generate_followup(RFISubmission(), [])
        assert result is None

    def test_generated_with_errors(self):
        errors = [
            ValidationResult("RFI-001", "ERROR", "Company name is required"),
        ]
        result = generate_followup(RFISubmission(), errors)
        assert result is not None
        assert "RFI-001" in result
        assert "MUST resolve" in result

    def test_generated_with_warnings(self):
        warnings = [
            ValidationResult("RFI-011", "WARNING", "WO years not specified"),
        ]
        result = generate_followup(RFISubmission(), warnings)
        assert result is not None
        assert "RFI-011" in result


# ---------------------------------------------------------------------------
# E2E: generate template -> fill -> process
# ---------------------------------------------------------------------------

class TestE2EProcess:
    def test_process_empty_template(self, tmp_path, monkeypatch):
        """Process an unfilled template (all responses empty)."""
        # Generate template
        template_path = tmp_path / "rfi.xlsx"
        generate_rfi_template(template_path)

        # Set up mock client root
        client_root = tmp_path / "client-data"
        client_root.mkdir()
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))

        # Process (dry run)
        result = process_rfi(
            template_path,
            client_slug="test-client",
            project_slug="test-project",
            dry_run=True,
        )

        assert result["completeness_score"] == 0.0
        assert len(result["files"]) >= 4  # project.yaml + 3 reports

    def test_process_generates_valid_yaml(self, tmp_path, monkeypatch):
        """Process should generate parseable YAML."""
        template_path = tmp_path / "rfi.xlsx"
        generate_rfi_template(template_path)

        # Fill in some data
        from openpyxl import load_workbook
        wb = load_workbook(str(template_path))
        ws = wb["1-Company Profile"]
        ws.cell(row=2, column=3, value="Test Corp")  # CSP-01: Company Name
        ws.cell(row=3, column=3, value="mining")       # CSP-02: Industry
        ws.cell(row=4, column=3, value="Test Plant")   # CSP-03: Plant Name
        ws.cell(row=7, column=3, value="US")            # CSP-06: Country
        ws.cell(row=9, column=3, value="en")            # CSP-08: Language

        ws2 = wb["7-Scope & Timeline"]
        ws2.cell(row=2, column=3, value="system")      # SCT-01: Scope Type
        ws2.cell(row=5, column=3, value="2026-01-01")  # SCT-04: Start Date
        ws2.cell(row=6, column=3, value="2026-12-31")  # SCT-05: Target Completion

        filled_path = tmp_path / "filled_rfi.xlsx"
        wb.save(str(filled_path))

        # Set up mock client root
        client_root = tmp_path / "client-data"
        client_root.mkdir()
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))

        result = process_rfi(
            filled_path,
            client_slug="test-client",
            project_slug="test-project",
            scaffold=True,
        )

        # Check project.yaml was written and is valid YAML
        project_root = client_root / "clients" / "test-client" / "projects" / "test-project"
        project_yaml = project_root / "project.yaml"
        assert project_yaml.exists()

        parsed = yaml.safe_load(project_yaml.read_text(encoding="utf-8"))
        assert parsed["client"]["name"] == "Test Corp"
        assert parsed["scope"]["type"] == "system"

    def test_file_not_found(self, tmp_path, monkeypatch):
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(tmp_path))
        with pytest.raises(FileNotFoundError):
            process_rfi(
                tmp_path / "nonexistent.xlsx",
                client_slug="test",
                project_slug="test",
            )

    def test_wrong_extension(self, tmp_path, monkeypatch):
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(tmp_path))
        bad_file = tmp_path / "test.csv"
        bad_file.write_text("test")
        with pytest.raises(ValueError, match=".xlsx"):
            process_rfi(bad_file, client_slug="test", project_slug="test")


# ---------------------------------------------------------------------------
# Phase 9: Memory-seeding generators
# ---------------------------------------------------------------------------

class TestMaintenanceStrategyRequirements:
    def test_contains_current_state(self):
        sub = _make_complete_submission()
        content = generate_maintenance_strategy_requirements(sub)
        assert "Maintenance Strategy Requirements" in content
        assert "Current State" in content
        assert "developing" in content  # strategy_maturity
        assert "sap-pm" in content      # cmms_type

    def test_contains_scope_priorities(self):
        sub = _make_complete_submission()
        content = generate_maintenance_strategy_requirements(sub)
        assert "Scope & Priorities" in content
        assert "system" in content  # scope_type
        assert "Phosphate Processing" in content  # areas_in_scope
        assert "SAG-MILL-001" in content  # priority_equipment

    def test_wo_history_fields(self):
        sub = _make_complete_submission()
        content = generate_maintenance_strategy_requirements(sub)
        assert "WO history available: Yes" in content
        assert "WO history years: 5" in content

    def test_dates_included(self):
        sub = _make_complete_submission()
        content = generate_maintenance_strategy_requirements(sub)
        assert "2026-02-01" in content
        assert "2026-06-30" in content

    def test_empty_submission(self):
        sub = RFISubmission()
        content = generate_maintenance_strategy_requirements(sub)
        assert "Not specified" in content
        assert "WO history available: No" in content


class TestWorkPlanningRequirements:
    def test_contains_org_data(self):
        sub = _make_complete_submission()
        content = generate_work_planning_requirements(sub)
        assert "Work Planning Requirements" in content
        assert "Team size: 50" in content
        assert "Shifts: 3" in content

    def test_trades_listed(self):
        sub = _make_complete_submission()
        content = generate_work_planning_requirements(sub)
        assert "mechanical" in content
        assert "electrical" in content
        assert "instrumentation" in content

    def test_empty_submission(self):
        sub = RFISubmission()
        content = generate_work_planning_requirements(sub)
        assert "Not specified" in content
        assert "Contractor maintenance: No" in content

    def test_budget_formatted(self):
        sub = _make_complete_submission()
        sub.organization.maintenance_budget_usd = 1_500_000.0
        content = generate_work_planning_requirements(sub)
        assert "$1,500,000" in content


class TestStandardsAppendix:
    def test_contains_hse_section(self):
        sub = _make_complete_submission()
        content = generate_standards_appendix(sub)
        assert "HSE & Environmental" in content
        assert "Sheet 5" in content

    def test_safety_permits(self):
        sub = _make_complete_submission()
        content = generate_standards_appendix(sub)
        assert "LOTOTO" in content
        assert "confined-space" in content

    def test_empty_submission(self):
        sub = RFISubmission()
        content = generate_standards_appendix(sub)
        assert "PPE matrix documented: No" in content


class TestKPIAppendix:
    def test_contains_kpi_section(self):
        sub = _make_complete_submission()
        content = generate_kpi_appendix(sub)
        assert "KPI Targets" in content
        assert "Sheet 6" in content

    def test_target_availability(self):
        sub = _make_complete_submission()
        content = generate_kpi_appendix(sub)
        assert "95.0%" in content

    def test_empty_submission(self):
        sub = RFISubmission()
        content = generate_kpi_appendix(sub)
        assert "Not specified" in content

    def test_full_kpi_data(self):
        sub = _make_complete_submission()
        sub.kpi_baseline.target_planned_vs_unplanned = 80.0
        sub.kpi_baseline.target_pm_compliance = 90.0
        sub.kpi_baseline.annual_maintenance_cost_usd = 2_000_000.0
        content = generate_kpi_appendix(sub)
        assert "80.0%" in content
        assert "90.0%" in content
        assert "$2,000,000" in content


# ---------------------------------------------------------------------------
# Phase 9: Append-or-create helper
# ---------------------------------------------------------------------------

class TestAppendOrCreate:
    def test_creates_new_file(self, tmp_path):
        target = tmp_path / "sub" / "test.md"
        _append_or_create(target, "Hello World")
        assert target.exists()
        assert target.read_text(encoding="utf-8") == "Hello World"

    def test_appends_to_existing(self, tmp_path):
        target = tmp_path / "test.md"
        target.write_text("Original content", encoding="utf-8")
        _append_or_create(target, "New section")
        text = target.read_text(encoding="utf-8")
        assert "Original content" in text
        assert "New section" in text

    def test_idempotent_no_duplicate(self, tmp_path):
        target = tmp_path / "test.md"
        _append_or_create(target, "Section A")
        _append_or_create(target, "Section A")  # second call
        text = target.read_text(encoding="utf-8")
        assert text.count("Section A") == 1

    def test_dry_run_no_write(self, tmp_path):
        target = tmp_path / "test.md"
        _append_or_create(target, "Content", dry_run=True)
        assert not target.exists()

    def test_creates_parent_dirs(self, tmp_path):
        target = tmp_path / "deep" / "nested" / "dir" / "file.md"
        _append_or_create(target, "Content")
        assert target.exists()


# ---------------------------------------------------------------------------
# Phase 9: E2E memory-seeding in process_rfi
# ---------------------------------------------------------------------------

class TestMemorySeedingE2E:
    def test_process_creates_memory_files(self, tmp_path, monkeypatch):
        """process_rfi should create maintenance-strategy and work-planning memory files."""
        from scripts.generate_rfi_template import generate_rfi_template

        template_path = tmp_path / "rfi.xlsx"
        generate_rfi_template(template_path)

        client_root = tmp_path / "client-data"
        client_root.mkdir()
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))

        result = process_rfi(
            template_path,
            client_slug="test-client",
            project_slug="test-project",
            scaffold=True,
        )

        memory_dir = (
            client_root / "clients" / "test-client" / "projects" / "test-project" / "3-memory"
        )

        # Check maintenance-strategy/requirements.md was created
        maint_strat = memory_dir / "maintenance-strategy" / "requirements.md"
        assert maint_strat.exists(), "maintenance-strategy/requirements.md not created"
        assert "Maintenance Strategy Requirements" in maint_strat.read_text(encoding="utf-8")

        # Check work-planning/requirements.md was created
        work_plan = memory_dir / "work-planning" / "requirements.md"
        assert work_plan.exists(), "work-planning/requirements.md not created"
        assert "Work Planning Requirements" in work_plan.read_text(encoding="utf-8")

        # Check global-requirements.md has appendices
        global_req = memory_dir / "global-requirements.md"
        assert global_req.exists()
        global_text = global_req.read_text(encoding="utf-8")
        assert "HSE & Environmental" in global_text
        assert "KPI Targets" in global_text

    def test_memory_files_in_result_dict(self, tmp_path, monkeypatch):
        """Result dict should include memory-seeded file paths."""
        from scripts.generate_rfi_template import generate_rfi_template

        template_path = tmp_path / "rfi.xlsx"
        generate_rfi_template(template_path)

        client_root = tmp_path / "client-data"
        client_root.mkdir()
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))

        result = process_rfi(
            template_path,
            client_slug="test-client",
            project_slug="test-project",
            scaffold=True,
        )

        file_paths = list(result["files"].keys())
        has_maint = any("maintenance-strategy" in p for p in file_paths)
        has_work = any("work-planning" in p for p in file_paths)
        assert has_maint, "maintenance-strategy not in result files"
        assert has_work, "work-planning not in result files"

    def test_reprocess_no_duplication(self, tmp_path, monkeypatch):
        """Running process_rfi twice should not duplicate appendix content."""
        from scripts.generate_rfi_template import generate_rfi_template

        template_path = tmp_path / "rfi.xlsx"
        generate_rfi_template(template_path)

        client_root = tmp_path / "client-data"
        client_root.mkdir()
        monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))

        kwargs = dict(
            client_slug="test-client",
            project_slug="test-project",
            scaffold=True,
        )
        process_rfi(template_path, **kwargs)
        process_rfi(template_path, **kwargs)  # second run

        memory_dir = (
            client_root / "clients" / "test-client" / "projects" / "test-project" / "3-memory"
        )
        global_text = (memory_dir / "global-requirements.md").read_text(encoding="utf-8")
        assert global_text.count("HSE & Environmental") == 1
        assert global_text.count("KPI Targets") == 1
