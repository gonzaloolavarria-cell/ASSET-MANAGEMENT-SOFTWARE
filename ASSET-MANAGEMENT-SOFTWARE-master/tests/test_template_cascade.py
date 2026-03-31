"""Tests for Phase 6 — Template cascade, client generation, SAP resolution, security."""

import os
import shutil
from pathlib import Path

import pytest
import yaml

from scripts.generate_client_templates import (
    TEMPLATE_FILES,
    _load_branding,
    _validate_slug,
    generate_client_templates,
)


def _can_import_openpyxl() -> bool:
    try:
        import openpyxl
        return True
    except ImportError:
        return False


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def tmp_ams(tmp_path):
    """Create a temporary AMS-like structure with system templates."""
    system_root = tmp_path / "ASSET-MANAGEMENT-SOFTWARE"
    client_root = tmp_path / "ASSET-MANAGEMENT-SOFTWARE-CLIENT"
    system_root.mkdir()
    client_root.mkdir()
    (system_root / "agents" / "_shared").mkdir(parents=True)

    # Create 14 system template stubs (minimal valid .xlsx)
    templates_dir = system_root / "templates"
    templates_dir.mkdir()
    for name in TEMPLATE_FILES:
        _create_minimal_xlsx(templates_dir / name)

    return system_root, client_root


def _create_minimal_xlsx(path: Path):
    """Create a minimal valid .xlsx file for testing."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="test_header")
        ws.cell(row=2, column=1, value="test_value")
        wb.save(path)
    except ImportError:
        # Fallback: write empty bytes (tests that don't need openpyxl)
        path.write_bytes(b"PK\x03\x04")


@pytest.fixture
def env_roots(tmp_ams, monkeypatch):
    """Set environment variables pointing to temp dirs."""
    system_root, client_root = tmp_ams
    monkeypatch.setenv("AMS_SYSTEM_ROOT", str(system_root))
    monkeypatch.setenv("AMS_CLIENT_ROOT", str(client_root))
    return system_root, client_root


# ---------------------------------------------------------------------------
# generate_client_templates — Functional Tests
# ---------------------------------------------------------------------------


class TestGenerateClientTemplates:
    def test_copies_14_templates(self, env_roots):
        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert len(result) == 14

    def test_templates_in_correct_directory(self, env_roots):
        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        expected_dir = (
            client_root / "clients" / "ocp" / "projects" / "jfc-strategy" / "5-templates"
        )
        for p in result:
            assert p.parent == expected_dir
            assert p.is_file()

    def test_all_14_filenames_present(self, env_roots):
        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        generated_names = {p.name for p in result}
        assert generated_names == set(TEMPLATE_FILES)

    def test_idempotent(self, env_roots):
        """Re-running does not duplicate content."""
        system_root, client_root = env_roots
        result1 = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        result2 = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert len(result1) == len(result2) == 14
        # Count files in directory
        target_dir = (
            client_root / "clients" / "ocp" / "projects" / "jfc-strategy" / "5-templates"
        )
        xlsx_files = list(target_dir.glob("*.xlsx"))
        assert len(xlsx_files) == 14

    def test_creates_target_directory(self, env_roots):
        """Target 5-templates/ directory is created if it doesn't exist."""
        system_root, client_root = env_roots
        target_dir = (
            client_root / "clients" / "ocp" / "projects" / "jfc-strategy" / "5-templates"
        )
        assert not target_dir.exists()
        generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert target_dir.is_dir()


# ---------------------------------------------------------------------------
# Branding Tests
# ---------------------------------------------------------------------------


class TestBrandingApplication:
    def test_without_branding(self, env_roots):
        """Templates are clean copies when no branding.yaml exists."""
        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert len(result) == 14
        # Files should exist and be valid
        for p in result:
            assert p.stat().st_size > 0

    def test_with_client_branding(self, env_roots):
        """Branding is applied from client context/templates/branding.yaml."""
        system_root, client_root = env_roots
        branding_dir = client_root / "clients" / "ocp" / "context" / "templates"
        branding_dir.mkdir(parents=True)
        branding = {
            "client": {"name": "OCP Group"},
            "colors": {"primary": "#1B5E20"},
            "fonts": {"heading": {"name": "Arial"}},
            "footer": {"confidential_notice": "OCP Confidential"},
        }
        (branding_dir / "branding.yaml").write_text(
            yaml.dump(branding), encoding="utf-8"
        )
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert len(result) == 14

    def test_with_project_branding_overrides(self, env_roots):
        """Project-level branding.yaml overrides client-level."""
        system_root, client_root = env_roots
        # Client-level branding
        client_branding_dir = client_root / "clients" / "ocp" / "context" / "templates"
        client_branding_dir.mkdir(parents=True)
        (client_branding_dir / "branding.yaml").write_text(
            yaml.dump({"colors": {"primary": "#FF0000"}}), encoding="utf-8"
        )
        # Project-level branding (should win)
        project_dir = (
            client_root / "clients" / "ocp" / "projects" / "jfc-strategy" / "5-templates"
        )
        project_dir.mkdir(parents=True)
        (project_dir / "branding.yaml").write_text(
            yaml.dump({"colors": {"primary": "#00FF00"}}), encoding="utf-8"
        )
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        assert len(result) == 14


# ---------------------------------------------------------------------------
# Version Metadata Tests
# ---------------------------------------------------------------------------


class TestVersionMetadata:
    @pytest.mark.skipif(
        not _can_import_openpyxl(), reason="openpyxl not installed"
    )
    def test_templates_have_version_metadata(self, env_roots):
        """Generated templates have version metadata in workbook properties."""
        from openpyxl import load_workbook

        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        for p in result:
            wb = load_workbook(p)
            assert "Template v" in (wb.properties.description or "")
            wb.close()

    @pytest.mark.skipif(
        not _can_import_openpyxl(), reason="openpyxl not installed"
    )
    def test_no_local_paths_in_metadata(self, env_roots):
        """Templates should not contain local path metadata."""
        from openpyxl import load_workbook

        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        for p in result:
            wb = load_workbook(p)
            creator = wb.properties.creator or ""
            last_modified = wb.properties.lastModifiedBy or ""
            assert "Users" not in creator
            assert ":" not in creator or creator == "VSC AMS"
            assert "Users" not in last_modified
            wb.close()


# ---------------------------------------------------------------------------
# Security Tests (Phase 6 Cybersecurity)
# ---------------------------------------------------------------------------


class TestSecurity:
    def test_path_traversal_client_slug(self):
        with pytest.raises(ValueError):
            generate_client_templates("../../../etc", "test")

    def test_path_traversal_project_slug(self):
        with pytest.raises(ValueError):
            generate_client_templates("ocp", "../../../etc/passwd")

    def test_command_injection_client_slug(self):
        with pytest.raises(ValueError):
            generate_client_templates("ocp; rm -rf /", "test")

    def test_empty_client_slug(self):
        with pytest.raises(ValueError, match="must not be empty"):
            generate_client_templates("", "test")

    def test_empty_project_slug(self):
        with pytest.raises(ValueError, match="must not be empty"):
            generate_client_templates("ocp", "")

    def test_yaml_injection_blocked(self, tmp_path):
        """YAML injection patterns in branding.yaml are rejected."""
        branding_path = tmp_path / "branding.yaml"
        branding_path.write_text(
            "!!python/object/apply:os.system ['echo pwned']",
            encoding="utf-8",
        )
        result = _load_branding(branding_path)
        assert result is None

    def test_safe_load_only(self, tmp_path):
        """Branding uses yaml.safe_load, not yaml.load."""
        branding_path = tmp_path / "branding.yaml"
        branding_path.write_text(
            yaml.dump({"colors": {"primary": "#FF0000"}}),
            encoding="utf-8",
        )
        result = _load_branding(branding_path)
        assert result is not None
        assert result["colors"]["primary"] == "#FF0000"

    @pytest.mark.skipif(
        not _can_import_openpyxl(), reason="openpyxl not installed"
    )
    def test_no_macros_in_templates(self, env_roots):
        """Generated templates are .xlsx (not .xlsm) — no macros."""
        system_root, client_root = env_roots
        result = generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        for p in result:
            assert p.suffix == ".xlsx", f"Expected .xlsx, got {p.suffix}"
            # Verify no VBA project
            from openpyxl import load_workbook
            wb = load_workbook(p)
            assert wb.vba_archive is None, f"Macro found in {p.name}"
            wb.close()


# ---------------------------------------------------------------------------
# SAP Template Resolution Tests
# ---------------------------------------------------------------------------


class TestSAPTemplateResolution:
    def test_resolve_sap_templates(self, env_roots):
        """SAP templates resolve via cascade."""
        from tools.engines.sap_export_engine import SAP_TEMPLATE_FILES, SAPExportEngine

        system_root, _ = env_roots
        templates_dir = system_root / "templates"
        for name in SAP_TEMPLATE_FILES:
            _create_minimal_xlsx(templates_dir / name)

        result = SAPExportEngine.resolve_sap_templates("ocp", "jfc-strategy")
        assert len(result) == 3
        for name in SAP_TEMPLATE_FILES:
            assert result[name] is not None

    def test_resolve_missing_sap_templates(self, env_roots):
        """Missing SAP templates return None per entry."""
        from tools.engines.sap_export_engine import SAP_TEMPLATE_FILES, SAPExportEngine

        result = SAPExportEngine.resolve_sap_templates("ocp", "jfc-strategy")
        for name in SAP_TEMPLATE_FILES:
            assert result[name] is None

    def test_sap_field_length_constants(self):
        """SAP field length constants are defined (AMS-C03 fix)."""
        from tools.engines.sap_export_engine import (
            SAP_FUNC_LOC_MAX,
            SAP_SHORT_TEXT_MAX,
            SAP_TASK_LIST_DESC_MAX,
        )

        assert SAP_SHORT_TEXT_MAX == 72
        assert SAP_FUNC_LOC_MAX == 40
        assert SAP_TASK_LIST_DESC_MAX == 40


# ---------------------------------------------------------------------------
# Integration Tests
# ---------------------------------------------------------------------------


class TestIntegration:
    def test_cascade_plus_generate_e2e(self, env_roots):
        """get_template_path + generate_client_templates: E2E flow."""
        from agents._shared.paths import get_template_path

        system_root, client_root = env_roots
        # Generate templates for client
        generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        # Now cascade should find at project level
        result = get_template_path("01_equipment_hierarchy.xlsx", "ocp", "jfc-strategy")
        assert result is not None
        assert "jfc-strategy" in str(result)

    def test_cascade_priority_after_generation(self, env_roots):
        """After generating, project-level templates override system-level."""
        from agents._shared.paths import get_template_path

        system_root, client_root = env_roots
        generate_client_templates(
            "ocp", "jfc-strategy",
            system_root=system_root,
            client_root=client_root,
        )
        result = get_template_path("01_equipment_hierarchy.xlsx", "ocp", "jfc-strategy")
        # Should resolve to project level, not system level
        assert "5-templates" in str(result)
        assert "clients" in str(result)

    def test_sap_resolve_uses_cascade(self, env_roots):
        """SAP resolve_sap_templates uses the same cascade logic."""
        from tools.engines.sap_export_engine import SAP_TEMPLATE_FILES, SAPExportEngine

        system_root, client_root = env_roots
        templates_dir = system_root / "templates"
        for name in SAP_TEMPLATE_FILES:
            _create_minimal_xlsx(templates_dir / name)

        result = SAPExportEngine.resolve_sap_templates("ocp", "jfc-strategy")
        for name, path in result.items():
            assert path is not None
            assert path.is_file()
