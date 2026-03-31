"""Tests for Data Import Engine — Phase 6 + GAP-W12."""

import pytest

from tools.engines.data_import_engine import DataImportEngine
from tools.models.schemas import ImportSource


class TestValidateHierarchyData:

    def test_valid_rows(self):
        rows = [
            {"equipment_id": "EQ-1", "description": "Pump A", "equipment_type": "ROTATING"},
            {"equipment_id": "EQ-2", "description": "Motor B", "equipment_type": "ELECTRICAL"},
        ]
        result = DataImportEngine.validate_hierarchy_data(rows)
        assert result.total_rows == 2
        assert result.valid_rows == 2
        assert result.error_rows == 0

    def test_missing_required_column(self):
        rows = [{"equipment_id": "EQ-1", "description": "Pump A"}]  # missing equipment_type
        result = DataImportEngine.validate_hierarchy_data(rows)
        assert result.error_rows == 1
        assert len(result.errors) == 1
        assert result.errors[0].column == "equipment_type"

    def test_empty_value(self):
        rows = [{"equipment_id": "", "description": "Pump A", "equipment_type": "ROTATING"}]
        result = DataImportEngine.validate_hierarchy_data(rows)
        assert result.error_rows == 1

    def test_empty_rows(self):
        result = DataImportEngine.validate_hierarchy_data([])
        assert result.total_rows == 0
        assert result.valid_rows == 0

    def test_with_column_mapping(self):
        rows = [{"asset_id": "EQ-1", "desc": "Pump A", "type": "ROTATING"}]
        mapping = {"asset_id": "equipment_id", "desc": "description", "type": "equipment_type"}
        result = DataImportEngine.validate_hierarchy_data(rows, column_mapping=mapping)
        assert result.valid_rows == 1


class TestValidateFailureHistory:

    def test_valid_rows(self):
        rows = [
            {"equipment_id": "EQ-1", "failure_date": "2025-01-15", "failure_mode": "BEARING_FAILURE"},
        ]
        result = DataImportEngine.validate_failure_history(rows)
        assert result.valid_rows == 1

    def test_invalid_date(self):
        rows = [
            {"equipment_id": "EQ-1", "failure_date": "not-a-date", "failure_mode": "BEARING_FAILURE"},
        ]
        result = DataImportEngine.validate_failure_history(rows)
        assert result.error_rows == 1
        assert any("date format" in e.message.lower() for e in result.errors)

    def test_missing_equipment_id(self):
        rows = [{"failure_date": "2025-01-15", "failure_mode": "BEARING_FAILURE"}]
        result = DataImportEngine.validate_failure_history(rows)
        assert result.error_rows == 1


class TestValidateMaintenancePlan:

    def test_valid_rows(self):
        rows = [
            {"equipment_id": "EQ-1", "task_description": "Oil change", "frequency": "MONTHLY"},
        ]
        result = DataImportEngine.validate_maintenance_plan(rows)
        assert result.valid_rows == 1

    def test_missing_frequency(self):
        rows = [{"equipment_id": "EQ-1", "task_description": "Oil change"}]
        result = DataImportEngine.validate_maintenance_plan(rows)
        assert result.error_rows == 1


class TestDetectColumnMapping:

    def test_exact_match(self):
        headers = ["equipment_id", "description", "equipment_type"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.EQUIPMENT_HIERARCHY)
        assert mapping.confidence == 1.0
        assert len(mapping.mapping) == 3

    def test_alias_match(self):
        headers = ["asset_id", "name", "category"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.EQUIPMENT_HIERARCHY)
        assert mapping.confidence == 1.0
        assert mapping.mapping["asset_id"] == "equipment_id"

    def test_partial_match(self):
        headers = ["equipment_id", "random_col"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.EQUIPMENT_HIERARCHY)
        assert mapping.confidence < 1.0

    def test_no_match(self):
        headers = ["foo", "bar", "baz"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.EQUIPMENT_HIERARCHY)
        assert mapping.confidence == 0.0


class TestSummarizeImport:

    def test_summary_all_valid(self):
        rows = [
            {"equipment_id": "EQ-1", "description": "Pump", "equipment_type": "ROTATING"},
            {"equipment_id": "EQ-2", "description": "Motor", "equipment_type": "ELECTRICAL"},
        ]
        result = DataImportEngine.validate_hierarchy_data(rows)
        summary = DataImportEngine.summarize_import(result)
        assert summary.total_rows == 2
        assert summary.valid_pct == 100.0

    def test_summary_with_errors(self):
        rows = [
            {"equipment_id": "EQ-1", "description": "Pump", "equipment_type": "ROTATING"},
            {"equipment_id": "", "description": "Bad", "equipment_type": ""},
        ]
        result = DataImportEngine.validate_hierarchy_data(rows)
        summary = DataImportEngine.summarize_import(result)
        assert summary.valid_pct == 50.0
        assert len(summary.error_summary) > 0


# ── Generic validate_data entry point ────────────────────────────


class TestValidateDataGeneric:

    def test_dispatches_to_hierarchy(self):
        rows = [{"equipment_id": "EQ-1", "description": "Pump", "equipment_type": "ROTATING"}]
        result = DataImportEngine.validate_data(rows, ImportSource.EQUIPMENT_HIERARCHY)
        assert result.valid_rows == 1
        assert result.source == ImportSource.EQUIPMENT_HIERARCHY

    def test_dispatches_to_failure_history(self):
        rows = [{"equipment_id": "EQ-1", "failure_date": "2025-01-15", "failure_mode": "BEARING"}]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_HISTORY)
        assert result.valid_rows == 1

    def test_empty_rows(self):
        result = DataImportEngine.validate_data([], ImportSource.EQUIPMENT_HIERARCHY)
        assert result.total_rows == 0


# ── Criticality Assessment ───────────────────────────────────────


class TestValidateCriticalityAssessment:

    def test_valid_rows(self):
        rows = [{"equipment_tag": "SAG-001", "method": "FULL_MATRIX", "safety": 3, "production": 4}]
        result = DataImportEngine.validate_data(rows, ImportSource.CRITICALITY_ASSESSMENT)
        assert result.valid_rows == 1

    def test_invalid_method(self):
        rows = [{"equipment_tag": "SAG-001", "method": "INVALID"}]
        result = DataImportEngine.validate_data(rows, ImportSource.CRITICALITY_ASSESSMENT)
        assert result.error_rows == 1
        assert any("method" in e.column for e in result.errors)

    def test_score_out_of_range(self):
        rows = [{"equipment_tag": "SAG-001", "method": "FULL_MATRIX", "safety": 6}]
        result = DataImportEngine.validate_data(rows, ImportSource.CRITICALITY_ASSESSMENT)
        assert result.error_rows == 1
        assert any("safety" in e.column for e in result.errors)

    def test_score_zero_invalid(self):
        rows = [{"equipment_tag": "SAG-001", "method": "SIMPLIFIED", "safety": 0}]
        result = DataImportEngine.validate_data(rows, ImportSource.CRITICALITY_ASSESSMENT)
        assert result.error_rows == 1

    def test_missing_equipment_tag(self):
        rows = [{"method": "FULL_MATRIX"}]
        result = DataImportEngine.validate_data(rows, ImportSource.CRITICALITY_ASSESSMENT)
        assert result.error_rows == 1
        assert any("equipment_tag" in e.column for e in result.errors)


# ── Failure Modes (72-combo) ─────────────────────────────────────


class TestValidateFailureModes:

    def test_valid_72_combo(self):
        rows = [{
            "equipment_tag": "SAG-001",
            "function_description": "Rotate drum",
            "mechanism": "BLOCKS",
            "cause": "CONTAMINATION",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_MODES)
        assert result.valid_rows == 1

    def test_invalid_mechanism(self):
        rows = [{
            "equipment_tag": "SAG-001",
            "function_description": "Rotate drum",
            "mechanism": "EXPLODES",
            "cause": "CONTAMINATION",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_MODES)
        assert result.error_rows == 1
        assert any("mechanism" in e.column for e in result.errors)

    def test_invalid_cause(self):
        rows = [{
            "equipment_tag": "SAG-001",
            "function_description": "Rotate drum",
            "mechanism": "BLOCKS",
            "cause": "MAGIC",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_MODES)
        assert result.error_rows == 1
        assert any("cause" in e.column for e in result.errors)

    def test_invalid_combination(self):
        # Both valid individually, but NOT a valid pair
        rows = [{
            "equipment_tag": "SAG-001",
            "function_description": "Rotate drum",
            "mechanism": "ARCS",
            "cause": "CONTAMINATION",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_MODES)
        assert result.error_rows == 1
        assert any("combination" in e.message.lower() for e in result.errors)

    def test_missing_required(self):
        rows = [{"equipment_tag": "SAG-001", "function_description": "Rotate drum"}]
        result = DataImportEngine.validate_data(rows, ImportSource.FAILURE_MODES)
        assert result.error_rows == 1


# ── Maintenance Tasks ────────────────────────────────────────────


class TestValidateMaintenanceTasks:

    def test_valid_rows(self):
        rows = [{"task_id": "T-001", "task_name": "Oil change", "task_type": "LUBRICATE", "constraint": "ONLINE"}]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_TASKS)
        assert result.valid_rows == 1

    def test_invalid_task_type(self):
        rows = [{"task_id": "T-001", "task_name": "X", "task_type": "DESTROY", "constraint": "ONLINE"}]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_TASKS)
        assert result.error_rows == 1
        assert any("task_type" in e.column for e in result.errors)

    def test_invalid_constraint(self):
        rows = [{"task_id": "T-001", "task_name": "X", "task_type": "INSPECT", "constraint": "MAYBE"}]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_TASKS)
        assert result.error_rows == 1
        assert any("constraint" in e.column for e in result.errors)


# ── Work Order History ───────────────────────────────────────────


class TestValidateWorkOrderHistory:

    def test_valid_rows(self):
        rows = [{
            "wo_id": "WO-001", "order_type": "PM01",
            "equipment_tag": "SAG-001", "status": "COMPLETED",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.WORK_ORDER_HISTORY)
        assert result.valid_rows == 1

    def test_invalid_order_type(self):
        rows = [{"wo_id": "WO-001", "order_type": "PM99", "equipment_tag": "X", "status": "COMPLETED"}]
        result = DataImportEngine.validate_data(rows, ImportSource.WORK_ORDER_HISTORY)
        assert result.error_rows == 1

    def test_invalid_status(self):
        rows = [{"wo_id": "WO-001", "order_type": "PM01", "equipment_tag": "X", "status": "UNKNOWN"}]
        result = DataImportEngine.validate_data(rows, ImportSource.WORK_ORDER_HISTORY)
        assert result.error_rows == 1

    def test_optional_date_validated(self):
        rows = [{
            "wo_id": "WO-001", "order_type": "PM01",
            "equipment_tag": "X", "status": "COMPLETED",
            "created_date": "not-a-date",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.WORK_ORDER_HISTORY)
        assert result.error_rows == 1
        assert any("date" in e.message.lower() for e in result.errors)

    def test_valid_optional_dates(self):
        rows = [{
            "wo_id": "WO-001", "order_type": "PM01",
            "equipment_tag": "X", "status": "COMPLETED",
            "created_date": "2025-03-01", "planned_start": "2025-03-05",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.WORK_ORDER_HISTORY)
        assert result.valid_rows == 1


# ── Spare Parts Inventory ────────────────────────────────────────


class TestValidateSpareParts:

    def test_valid_rows(self):
        rows = [{"material_code": "MAT-001", "description": "Bearing", "ved_class": "VITAL", "quantity_on_hand": 10}]
        result = DataImportEngine.validate_data(rows, ImportSource.SPARE_PARTS_INVENTORY)
        assert result.valid_rows == 1

    def test_invalid_ved_class(self):
        rows = [{"material_code": "MAT-001", "description": "Bearing", "ved_class": "LUXURY", "quantity_on_hand": 10}]
        result = DataImportEngine.validate_data(rows, ImportSource.SPARE_PARTS_INVENTORY)
        assert result.error_rows == 1

    def test_negative_quantity(self):
        rows = [{"material_code": "MAT-001", "description": "Bearing", "ved_class": "VITAL", "quantity_on_hand": -5}]
        result = DataImportEngine.validate_data(rows, ImportSource.SPARE_PARTS_INVENTORY)
        assert result.error_rows == 1
        assert any("quantity" in e.message.lower() for e in result.errors)

    def test_non_numeric_quantity(self):
        rows = [{"material_code": "MAT-001", "description": "Bearing", "ved_class": "VITAL", "quantity_on_hand": "abc"}]
        result = DataImportEngine.validate_data(rows, ImportSource.SPARE_PARTS_INVENTORY)
        assert result.error_rows == 1

    def test_zero_quantity_valid(self):
        rows = [{"material_code": "MAT-001", "description": "Bearing", "ved_class": "ESSENTIAL", "quantity_on_hand": 0}]
        result = DataImportEngine.validate_data(rows, ImportSource.SPARE_PARTS_INVENTORY)
        assert result.valid_rows == 1


# ── Shutdown Calendar ────────────────────────────────────────────


class TestValidateShutdownCalendar:

    def test_valid_rows(self):
        rows = [{
            "plant_id": "OCP-JFC", "shutdown_name": "Annual",
            "shutdown_type": "MAJOR_20H_PLUS", "planned_start": "2025-06-01",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.SHUTDOWN_CALENDAR)
        assert result.valid_rows == 1

    def test_invalid_shutdown_type(self):
        rows = [{"plant_id": "P1", "shutdown_name": "X", "shutdown_type": "HUGE", "planned_start": "2025-06-01"}]
        result = DataImportEngine.validate_data(rows, ImportSource.SHUTDOWN_CALENDAR)
        assert result.error_rows == 1

    def test_invalid_start_date(self):
        rows = [{"plant_id": "P1", "shutdown_name": "X", "shutdown_type": "MINOR_8H", "planned_start": "bad-date"}]
        result = DataImportEngine.validate_data(rows, ImportSource.SHUTDOWN_CALENDAR)
        assert result.error_rows == 1

    def test_start_after_end(self):
        rows = [{
            "plant_id": "P1", "shutdown_name": "X",
            "shutdown_type": "MINOR_8H",
            "planned_start": "2025-06-10", "planned_end": "2025-06-01",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.SHUTDOWN_CALENDAR)
        assert result.error_rows == 1
        assert any("after" in e.message.lower() for e in result.errors)

    def test_start_before_end_valid(self):
        rows = [{
            "plant_id": "P1", "shutdown_name": "X",
            "shutdown_type": "MINOR_8H",
            "planned_start": "2025-06-01", "planned_end": "2025-06-10",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.SHUTDOWN_CALENDAR)
        assert result.valid_rows == 1


# ── Workforce ────────────────────────────────────────────────────


class TestValidateWorkforce:

    def test_valid_rows(self):
        rows = [{"worker_id": "W-001", "name": "Ali Bouzidi", "specialty": "FITTER", "shift": "MORNING"}]
        result = DataImportEngine.validate_data(rows, ImportSource.WORKFORCE)
        assert result.valid_rows == 1

    def test_invalid_specialty(self):
        rows = [{"worker_id": "W-001", "name": "Ali", "specialty": "WIZARD", "shift": "MORNING"}]
        result = DataImportEngine.validate_data(rows, ImportSource.WORKFORCE)
        assert result.error_rows == 1

    def test_invalid_shift(self):
        rows = [{"worker_id": "W-001", "name": "Ali", "specialty": "FITTER", "shift": "MIDNIGHT"}]
        result = DataImportEngine.validate_data(rows, ImportSource.WORKFORCE)
        assert result.error_rows == 1


# ── Field Capture ────────────────────────────────────────────────


class TestValidateFieldCapture:

    def test_valid_rows(self):
        rows = [{"technician_id": "T-01", "capture_type": "VOICE", "raw_text": "Bearing noise", "timestamp": "2025-03-01"}]
        result = DataImportEngine.validate_data(rows, ImportSource.FIELD_CAPTURE)
        assert result.valid_rows == 1

    def test_invalid_capture_type(self):
        rows = [{"technician_id": "T-01", "capture_type": "TELEPATHY", "raw_text": "X", "timestamp": "2025-03-01"}]
        result = DataImportEngine.validate_data(rows, ImportSource.FIELD_CAPTURE)
        assert result.error_rows == 1


# ── RCA Events ───────────────────────────────────────────────────


class TestValidateRCAEvents:

    def test_valid_rows(self):
        rows = [{"event_description": "Bearing failure on SAG", "plant_id": "OCP-JFC", "level": "2"}]
        result = DataImportEngine.validate_data(rows, ImportSource.RCA_EVENTS)
        assert result.valid_rows == 1

    def test_invalid_level(self):
        rows = [{"event_description": "X", "plant_id": "P1", "level": "5"}]
        result = DataImportEngine.validate_data(rows, ImportSource.RCA_EVENTS)
        assert result.error_rows == 1


# ── Planning KPI ─────────────────────────────────────────────────


class TestValidatePlanningKPI:

    def test_valid_rows(self):
        rows = [{
            "plant_id": "OCP-JFC", "period_start": "2025-01-01",
            "period_end": "2025-01-31", "wo_planned": 100, "wo_completed": 85,
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.PLANNING_KPI)
        assert result.valid_rows == 1

    def test_invalid_date(self):
        rows = [{
            "plant_id": "P1", "period_start": "not-a-date",
            "period_end": "2025-01-31", "wo_planned": 10, "wo_completed": 5,
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.PLANNING_KPI)
        assert result.error_rows == 1


# ── DE KPI ───────────────────────────────────────────────────────


class TestValidateDEKPI:

    def test_valid_rows(self):
        rows = [{
            "plant_id": "OCP-JFC", "period_start": "2025-01-01",
            "period_end": "2025-01-31", "events_reported": 20, "events_required": 25,
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.DE_KPI)
        assert result.valid_rows == 1

    def test_invalid_date(self):
        rows = [{
            "plant_id": "P1", "period_start": "2025-01-01",
            "period_end": "bad", "events_reported": 10, "events_required": 15,
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.DE_KPI)
        assert result.error_rows == 1


# ── Maintenance Strategy ─────────────────────────────────────────


class TestValidateMaintenanceStrategy:

    def test_valid_rows(self):
        rows = [{
            "strategy_id": "S-001", "equipment_tag": "SAG-001",
            "mechanism": "BLOCKS", "cause": "CONTAMINATION",
            "tactics_type": "CONDITION_BASED",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_STRATEGY)
        assert result.valid_rows == 1

    def test_invalid_72_combo(self):
        rows = [{
            "strategy_id": "S-001", "equipment_tag": "SAG-001",
            "mechanism": "ARCS", "cause": "CONTAMINATION",
            "tactics_type": "CONDITION_BASED",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_STRATEGY)
        assert result.error_rows == 1

    def test_invalid_tactics_type(self):
        rows = [{
            "strategy_id": "S-001", "equipment_tag": "SAG-001",
            "mechanism": "BLOCKS", "cause": "CONTAMINATION",
            "tactics_type": "PRAY",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_STRATEGY)
        assert result.error_rows == 1

    def test_missing_strategy_id(self):
        rows = [{
            "equipment_tag": "SAG-001",
            "mechanism": "BLOCKS", "cause": "CONTAMINATION",
            "tactics_type": "CONDITION_BASED",
        }]
        result = DataImportEngine.validate_data(rows, ImportSource.MAINTENANCE_STRATEGY)
        assert result.error_rows == 1


# ── Column mapping for extended types ────────────────────────────


class TestDetectColumnMappingExtended:

    def test_criticality_mapping(self):
        headers = ["tag", "crit_method"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.CRITICALITY_ASSESSMENT)
        assert mapping.mapping.get("tag") == "equipment_tag"
        assert mapping.mapping.get("crit_method") == "method"
        assert mapping.confidence == 1.0

    def test_spare_parts_alias(self):
        headers = ["part_number", "description", "ved", "stock"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.SPARE_PARTS_INVENTORY)
        assert mapping.confidence == 1.0
        assert mapping.mapping.get("part_number") == "material_code"
        assert mapping.mapping.get("stock") == "quantity_on_hand"

    def test_workforce_alias(self):
        headers = ["employee_id", "full_name", "trade", "shift_type"]
        mapping = DataImportEngine.detect_column_mapping(headers, ImportSource.WORKFORCE)
        assert mapping.confidence == 1.0


# ── parse_and_validate integration ───────────────────────────────


class TestParseAndValidate:

    def test_valid_xlsx(self):
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="equipment_id")
        ws.cell(row=1, column=2, value="description")
        ws.cell(row=1, column=3, value="equipment_type")
        ws.cell(row=2, column=1, value="EQ-001")
        ws.cell(row=2, column=2, value="SAG Mill")
        ws.cell(row=2, column=3, value="MILL")
        buf = io.BytesIO()
        wb.save(buf)

        result = DataImportEngine.parse_and_validate(
            buf.getvalue(), "test.xlsx", ImportSource.EQUIPMENT_HIERARCHY,
        )
        assert result.valid_rows == 1
        assert result.source == ImportSource.EQUIPMENT_HIERARCHY

    def test_parse_failure_propagates(self):
        result = DataImportEngine.parse_and_validate(
            b"not-a-valid-file", "test.xlsx", ImportSource.EQUIPMENT_HIERARCHY,
        )
        assert result.valid_rows == 0
        assert len(result.errors) > 0
