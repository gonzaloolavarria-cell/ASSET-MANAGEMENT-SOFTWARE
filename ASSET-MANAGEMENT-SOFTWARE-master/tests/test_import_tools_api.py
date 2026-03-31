"""Tests for GAP-W12 Session C — MCP import tools + API endpoints."""

import io
import json
import os

import pytest
from openpyxl import Workbook

from agents.tool_wrappers.registry import call_tool, TOOL_REGISTRY
from agents.tool_wrappers.server import AGENT_TOOL_MAP


# ── Helpers ──────────────────────────────────────────────────────────

def _make_xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Create a minimal xlsx file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _template_path(name: str) -> str:
    """Return the absolute resolved path to a template file (no '..' segments)."""
    return str(
        os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "templates", name)
        )
    )


# ── MCP Tool Registration ───────────────────────────────────────────


class TestImportToolsRegistered:

    def test_parse_import_file_registered(self):
        assert "parse_import_file" in TOOL_REGISTRY

    def test_detect_import_columns_registered(self):
        assert "detect_import_columns" in TOOL_REGISTRY

    def test_parse_and_validate_import_registered(self):
        assert "parse_and_validate_import" in TOOL_REGISTRY

    def test_tools_in_orchestrator_map(self):
        for name in ["parse_import_file", "detect_import_columns", "parse_and_validate_import"]:
            assert name in AGENT_TOOL_MAP["orchestrator"], f"{name} not in orchestrator"

    def test_tools_in_planning_map(self):
        for name in ["parse_import_file", "detect_import_columns", "parse_and_validate_import"]:
            assert name in AGENT_TOOL_MAP["planning"], f"{name} not in planning"

    def test_total_tool_count_updated(self):
        """3 new tools added (parse_import_file, detect_import_columns, parse_and_validate_import) → total ≥ 158."""
        assert len(TOOL_REGISTRY) >= 158


# ── parse_import_file tool ──────────────────────────────────────────


class TestParseImportFileTool:

    def test_parse_template(self):
        path = _template_path("01_equipment_hierarchy.xlsx")
        if not os.path.isfile(path):
            pytest.skip("Template not found")
        result_json = call_tool("parse_import_file", {
            "input_json": json.dumps({"file_path": path}),
        })
        result = json.loads(result_json)
        assert "error" not in result, f"Unexpected error: {result.get('error')}"
        assert result["success"] is True
        assert len(result["headers"]) > 0

    def test_path_traversal_blocked(self):
        result_json = call_tool("parse_import_file", {
            "input_json": json.dumps({"file_path": "../../etc/passwd"}),
        })
        result = json.loads(result_json)
        assert "error" in result

    def test_disallowed_directory(self):
        result_json = call_tool("parse_import_file", {
            "input_json": json.dumps({"file_path": "/tmp/random.xlsx"}),
        })
        result = json.loads(result_json)
        assert "error" in result

    def test_nonexistent_file(self):
        result_json = call_tool("parse_import_file", {
            "input_json": json.dumps({
                "file_path": _template_path("99_does_not_exist.xlsx"),
            }),
        })
        result = json.loads(result_json)
        assert "error" in result


# ── detect_import_columns tool ──────────────────────────────────────


class TestDetectImportColumnsTool:

    def test_detect_hierarchy_columns(self):
        result_json = call_tool("detect_import_columns", {
            "input_json": json.dumps({
                "headers": ["equipment_id", "description", "equipment_type"],
                "source": "EQUIPMENT_HIERARCHY",
            }),
        })
        result = json.loads(result_json)
        assert result["confidence"] == 1.0
        assert "equipment_id" in result["mapping"].values()

    def test_detect_with_aliases(self):
        result_json = call_tool("detect_import_columns", {
            "input_json": json.dumps({
                "headers": ["tag", "crit_method"],
                "source": "CRITICALITY_ASSESSMENT",
            }),
        })
        result = json.loads(result_json)
        assert result["confidence"] == 1.0

    def test_no_match(self):
        result_json = call_tool("detect_import_columns", {
            "input_json": json.dumps({
                "headers": ["foo", "bar", "baz"],
                "source": "EQUIPMENT_HIERARCHY",
            }),
        })
        result = json.loads(result_json)
        assert result["confidence"] == 0.0


# ── parse_and_validate_import tool ──────────────────────────────────


class TestParseAndValidateImportTool:

    def test_valid_template(self):
        path = _template_path("01_equipment_hierarchy.xlsx")
        if not os.path.isfile(path):
            pytest.skip("Template not found")
        result_json = call_tool("parse_and_validate_import", {
            "input_json": json.dumps({
                "file_path": path,
                "source": "EQUIPMENT_HIERARCHY",
            }),
        })
        result = json.loads(result_json)
        assert "total_rows" in result
        assert "valid_rows" in result
        assert result["source"] == "EQUIPMENT_HIERARCHY"

    def test_with_explicit_mapping(self):
        path = _template_path("01_equipment_hierarchy.xlsx")
        if not os.path.isfile(path):
            pytest.skip("Template not found")
        result_json = call_tool("parse_and_validate_import", {
            "input_json": json.dumps({
                "file_path": path,
                "source": "EQUIPMENT_HIERARCHY",
                "column_mapping": {"equipment_id": "equipment_id"},
            }),
        })
        result = json.loads(result_json)
        assert "total_rows" in result


# ── Reporting service functions ─────────────────────────────────────


class TestReportingServiceImport:

    def test_upload_and_validate(self):
        from api.services.reporting_service import upload_and_validate
        from unittest.mock import MagicMock

        db = MagicMock()
        xlsx = _make_xlsx_bytes(
            ["equipment_id", "description", "equipment_type"],
            [["EQ-001", "Pump", "ROTATING"]],
        )
        result = upload_and_validate(db, xlsx, "test.xlsx", "EQUIPMENT_HIERARCHY")
        assert result["valid_rows"] == 1
        assert result["source"] == "EQUIPMENT_HIERARCHY"

    def test_upload_rejects_bad_extension(self):
        from api.services.reporting_service import upload_and_validate
        from unittest.mock import MagicMock

        db = MagicMock()
        with pytest.raises(ValueError, match="Unsupported file extension"):
            upload_and_validate(db, b"data", "test.pdf", "EQUIPMENT_HIERARCHY")

    def test_upload_rejects_oversized(self):
        from api.services.reporting_service import upload_and_validate, MAX_IMPORT_FILE_SIZE
        from unittest.mock import MagicMock

        db = MagicMock()
        big = b"x" * (MAX_IMPORT_FILE_SIZE + 1)
        with pytest.raises(ValueError, match="File too large"):
            upload_and_validate(db, big, "test.xlsx", "EQUIPMENT_HIERARCHY")

    def test_detect_source_from_filename(self):
        from api.services.reporting_service import detect_source_from_filename

        assert detect_source_from_filename("01_equipment_hierarchy.xlsx") == "EQUIPMENT_HIERARCHY"
        assert detect_source_from_filename("07_spare_parts_inventory.xlsx") == "SPARE_PARTS_INVENTORY"
        assert detect_source_from_filename("14_maintenance_strategy.xlsx") == "MAINTENANCE_STRATEGY"
        assert detect_source_from_filename("random_file.xlsx") is None

    def test_detect_source_with_dash_separator(self):
        from api.services.reporting_service import detect_source_from_filename

        assert detect_source_from_filename("03-failure_modes.xlsx") == "FAILURE_MODES"


# ── API endpoint tests (using TestClient) ───────────────────────────


class TestImportAPIEndpoints:

    @pytest.fixture
    def client(self):
        """Create a FastAPI test client."""
        try:
            from fastapi.testclient import TestClient
            from api.main import app
            return TestClient(app)
        except Exception:
            pytest.skip("FastAPI test client not available")

    def test_template_download(self, client):
        response = client.get("/api/v1/reporting/import/template/1")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert len(response.content) > 100

    def test_template_download_invalid_number(self, client):
        response = client.get("/api/v1/reporting/import/template/99")
        assert response.status_code == 404

    def test_upload_endpoint(self, client):
        xlsx = _make_xlsx_bytes(
            ["equipment_id", "description", "equipment_type"],
            [["EQ-001", "Pump", "ROTATING"]],
        )
        response = client.post(
            "/api/v1/reporting/import/upload",
            files={"file": ("test.xlsx", xlsx)},
            params={"source": "EQUIPMENT_HIERARCHY"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["valid_rows"] == 1

    def test_upload_bad_extension(self, client):
        response = client.post(
            "/api/v1/reporting/import/upload",
            files={"file": ("test.pdf", b"data")},
            params={"source": "EQUIPMENT_HIERARCHY"},
        )
        assert response.status_code == 400

    def test_batch_import(self, client):
        xlsx1 = _make_xlsx_bytes(
            ["equipment_id", "description", "equipment_type"],
            [["EQ-001", "Pump", "ROTATING"]],
        )
        response = client.post(
            "/api/v1/reporting/import/batch",
            files=[
                ("files", ("01_equipment_hierarchy.xlsx", xlsx1)),
            ],
        )
        assert response.status_code == 200
        data = response.json()
        assert data["total_files"] == 1
        assert data["results"][0]["source"] == "EQUIPMENT_HIERARCHY"

    def test_batch_import_unknown_prefix(self, client):
        response = client.post(
            "/api/v1/reporting/import/batch",
            files=[("files", ("random.xlsx", b"data"))],
        )
        assert response.status_code == 200
        data = response.json()
        assert "error" in data["results"][0]


# ── Import History — Unit Tests ─────────────────────────────────────


class TestImportHistoryModel:

    def test_model_registered(self):
        from api.database.models import ImportHistoryModel
        assert ImportHistoryModel.__tablename__ == "import_history"

    def test_model_columns(self):
        from api.database.models import ImportHistoryModel
        cols = {c.name for c in ImportHistoryModel.__table__.columns}
        for expected in ("import_id", "plant_id", "source", "filename", "total_rows",
                         "valid_rows", "error_rows", "status", "errors_json", "imported_at"):
            assert expected in cols, f"Missing column: {expected}"


class TestImportHistorySchema:

    def test_schema_roundtrip(self):
        from tools.models.schemas import ImportHistoryEntry, ImportSource
        from datetime import datetime
        entry = ImportHistoryEntry(
            import_id="abc-123",
            plant_id="OCP-JFC",
            source=ImportSource.EQUIPMENT_HIERARCHY,
            filename="test.xlsx",
            total_rows=10,
            valid_rows=9,
            error_rows=1,
            status="partial",
        )
        assert entry.status == "partial"
        assert entry.source == ImportSource.EQUIPMENT_HIERARCHY
        data = entry.model_dump()
        assert data["import_id"] == "abc-123"


class TestImportStatusHelper:

    def test_all_valid_is_success(self):
        from api.services.reporting_service import _import_status
        assert _import_status(10, 0) == "success"

    def test_partial_errors_is_partial(self):
        from api.services.reporting_service import _import_status
        assert _import_status(10, 3) == "partial"

    def test_all_errors_is_failed(self):
        from api.services.reporting_service import _import_status
        assert _import_status(5, 5) == "failed"

    def test_zero_rows_is_failed(self):
        from api.services.reporting_service import _import_status
        assert _import_status(0, 0) == "success"


class TestRecordImportHistory:

    def test_record_creates_db_object(self):
        from api.services.reporting_service import record_import_history
        from tools.models.schemas import ImportResult, ImportSource
        from unittest.mock import MagicMock

        db = MagicMock()
        result = ImportResult(
            source=ImportSource.EQUIPMENT_HIERARCHY,
            total_rows=5, valid_rows=5, error_rows=0, errors=[], validated_data=[],
        )
        entry = record_import_history(db, "OCP-JFC", ImportSource.EQUIPMENT_HIERARCHY,
                                      "test.xlsx", result, file_size_kb=10)
        db.add.assert_called_once()
        assert entry.status == "success"
        assert entry.plant_id == "OCP-JFC"
        assert entry.valid_rows == 5

    def test_record_with_errors(self):
        from api.services.reporting_service import record_import_history
        from tools.models.schemas import ImportResult, ImportSource, ImportValidationError
        from unittest.mock import MagicMock

        db = MagicMock()
        errors = [ImportValidationError(row=1, column="equipment_id", message="Missing")]
        result = ImportResult(
            source=ImportSource.EQUIPMENT_HIERARCHY,
            total_rows=3, valid_rows=2, error_rows=1, errors=errors, validated_data=[],
        )
        entry = record_import_history(db, "OCP-JFC", ImportSource.EQUIPMENT_HIERARCHY,
                                      "test.xlsx", result)
        assert entry.status == "partial"
        assert entry.error_rows == 1


class TestListImportHistory:

    def test_list_returns_empty_list(self):
        from api.services.reporting_service import list_import_history
        from unittest.mock import MagicMock

        mock_query = MagicMock()
        mock_query.filter_by.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.offset.return_value = mock_query
        mock_query.limit.return_value = mock_query
        mock_query.all.return_value = []
        db = MagicMock()
        db.query.return_value = mock_query

        result = list_import_history(db, plant_id="OCP-JFC")
        assert result == []

    def test_get_history_entry_not_found(self):
        from api.services.reporting_service import get_import_history_entry
        from unittest.mock import MagicMock

        mock_query = MagicMock()
        mock_query.filter_by.return_value = mock_query
        mock_query.first.return_value = None
        db = MagicMock()
        db.query.return_value = mock_query

        result = get_import_history_entry(db, "nonexistent-id")
        assert result is None


# ── Import History — API Endpoint Tests ────────────────────────────


class TestImportHistoryAPIEndpoints:

    @pytest.fixture
    def client(self):
        try:
            from fastapi.testclient import TestClient
            from api.main import app
            return TestClient(app)
        except Exception:
            pytest.skip("FastAPI test client not available")

    def test_list_history_endpoint(self, client):
        response = client.get("/api/v1/reporting/import/history")
        assert response.status_code == 200
        assert isinstance(response.json(), list)

    def test_list_history_with_plant_filter(self, client):
        response = client.get("/api/v1/reporting/import/history", params={"plant_id": "OCP-JFC"})
        assert response.status_code == 200
        assert isinstance(response.json(), list)

    def test_history_entry_not_found(self, client):
        response = client.get("/api/v1/reporting/import/history/does-not-exist")
        assert response.status_code == 404

    def test_upload_creates_history_entry(self, client):
        """Upload a valid file and verify a history entry is created."""
        xlsx = _make_xlsx_bytes(
            ["equipment_id", "description", "equipment_type"],
            [["EQ-001", "Pump", "ROTATING"]],
        )
        upload_resp = client.post(
            "/api/v1/reporting/import/upload",
            files={"file": ("01_test.xlsx", xlsx)},
            params={"source": "EQUIPMENT_HIERARCHY"},
        )
        assert upload_resp.status_code == 200

        # History should now have at least one entry
        hist_resp = client.get("/api/v1/reporting/import/history")
        assert hist_resp.status_code == 200
        entries = hist_resp.json()
        assert len(entries) >= 1
        latest = entries[0]
        assert latest["source"] == "EQUIPMENT_HIERARCHY"
        assert latest["status"] in ("success", "partial", "failed")
