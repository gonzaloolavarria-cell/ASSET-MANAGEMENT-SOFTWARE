"""Tests for File Parser Engine — GAP-W12."""

import csv
import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools.engines.file_parser_engine import (
    ALLOWED_EXTENSIONS,
    MAX_FILE_SIZE_BYTES,
    FileParserEngine,
)

TEMPLATE_DIR = Path(__file__).parent.parent / "templates"


def _make_xlsx_bytes(
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Data",
    extra_sheets: list[str] | None = None,
) -> bytes:
    """Helper: create a minimal .xlsx file in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    if extra_sheets:
        for name in extra_sheets:
            wb.create_sheet(title=name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv_bytes(
    headers: list[str],
    rows: list[list],
    delimiter: str = ",",
    encoding: str = "utf-8",
) -> bytes:
    """Helper: create CSV bytes."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode(encoding)


# ── Excel Tests ───────────────────────────────────────────────────


class TestParseExcel:

    def test_parse_simple_xlsx(self):
        content = _make_xlsx_bytes(
            ["equipment_id", "description", "equipment_type"],
            [["EQ-001", "SAG Mill", "MILL"], ["EQ-002", "Ball Mill", "MILL"]],
        )
        result = FileParserEngine.parse_file(content, "test.xlsx")
        assert result.success is True
        assert result.file_type == "xlsx"
        assert result.total_rows == 2
        assert len(result.headers) == 3
        assert result.rows[0]["equipment_id"] == "EQ-001"
        assert result.rows[1]["description"] == "Ball Mill"

    def test_parse_specific_sheet(self):
        content = _make_xlsx_bytes(
            ["a", "b"],
            [["1", "2"]],
            sheet_name="Tasks",
            extra_sheets=["Instructions", "Other"],
        )
        result = FileParserEngine.parse_excel(content, "test.xlsx", sheet_name="Tasks")
        assert result.success is True
        assert result.sheet_parsed == "Tasks"
        assert result.total_rows == 1

    def test_instructions_sheet_skipped(self):
        """When no sheet_name given, skip 'Instructions' and pick first data sheet."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Instructions"
        ws1.cell(row=1, column=1, value="Instructions here")
        ws2 = wb.create_sheet("Data")
        ws2.cell(row=1, column=1, value="col_a")
        ws2.cell(row=2, column=1, value="val_1")
        buf = io.BytesIO()
        wb.save(buf)

        result = FileParserEngine.parse_excel(buf.getvalue(), "test.xlsx")
        assert result.success is True
        assert result.sheet_parsed == "Data"
        assert result.total_rows == 1

    def test_numeric_cells_preserved(self):
        content = _make_xlsx_bytes(
            ["name", "power_kw", "weight_kg"],
            [["Pump", 250.5, 1200]],
        )
        result = FileParserEngine.parse_file(content, "test.xlsx")
        assert result.rows[0]["power_kw"] == 250.5
        assert result.rows[0]["weight_kg"] == 1200

    def test_date_cells_to_iso(self):
        from datetime import datetime
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="date_col")
        ws.cell(row=2, column=1, value=datetime(2025, 6, 15))
        buf = io.BytesIO()
        wb.save(buf)

        result = FileParserEngine.parse_excel(buf.getvalue())
        assert result.rows[0]["date_col"] == "2025-06-15"

    def test_empty_rows_trimmed(self):
        """3 consecutive empty rows should stop reading."""
        content = _make_xlsx_bytes(
            ["a", "b"],
            [["1", "2"], [None, None], [None, None], [None, None], ["3", "4"]],
        )
        result = FileParserEngine.parse_file(content, "test.xlsx")
        # Should stop after 3 empty rows → only first row
        assert result.total_rows == 1

    def test_max_rows_enforced(self):
        rows = [["val"] for _ in range(100)]
        content = _make_xlsx_bytes(["col"], rows)
        result = FileParserEngine.parse_excel(content, "test.xlsx", max_rows=10)
        assert result.total_rows == 10
        assert any("limit" in e.message.lower() for e in result.errors)

    def test_oversized_file_rejected(self):
        # Create content larger than MAX_FILE_SIZE_BYTES (fake it)
        content = b"x" * (MAX_FILE_SIZE_BYTES + 1)
        result = FileParserEngine.parse_file(content, "big.xlsx")
        assert result.success is False
        assert any("size" in e.message.lower() for e in result.errors)

    def test_nonexistent_sheet(self):
        content = _make_xlsx_bytes(["a"], [["1"]])
        result = FileParserEngine.parse_excel(content, "test.xlsx", sheet_name="NoSuchSheet")
        assert result.success is False
        assert any("not found" in e.message.lower() for e in result.errors)

    def test_empty_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        buf = io.BytesIO()
        wb.save(buf)

        result = FileParserEngine.parse_excel(buf.getvalue())
        assert result.success is True
        assert result.total_rows == 0

    def test_list_sheets(self):
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        buf = io.BytesIO()
        wb.save(buf)

        sheets = FileParserEngine.list_sheets(buf.getvalue())
        assert sheets == ["Sheet1", "Sheet2", "Sheet3"]

    def test_sheets_available_in_result(self):
        content = _make_xlsx_bytes(
            ["a"], [["1"]], sheet_name="Data", extra_sheets=["Instructions"],
        )
        result = FileParserEngine.parse_file(content, "test.xlsx")
        assert "Data" in result.sheets_available
        assert "Instructions" in result.sheets_available


# ── CSV Tests ─────────────────────────────────────────────────────


class TestParseCsv:

    def test_parse_simple_csv(self):
        content = _make_csv_bytes(
            ["equipment_id", "description"],
            [["EQ-001", "SAG Mill"], ["EQ-002", "Ball Mill"]],
        )
        result = FileParserEngine.parse_file(content, "test.csv")
        assert result.success is True
        assert result.file_type == "csv"
        assert result.total_rows == 2
        assert result.rows[0]["equipment_id"] == "EQ-001"

    def test_auto_detect_semicolon(self):
        content = _make_csv_bytes(
            ["a", "b", "c"],
            [["1", "2", "3"]],
            delimiter=";",
        )
        result = FileParserEngine.parse_csv(content, "test.csv")
        assert result.success is True
        assert result.total_rows == 1
        assert result.rows[0]["a"] == "1"

    def test_encoding_latin1(self):
        content = _make_csv_bytes(
            ["name", "desc"],
            [["Válvula", "Descripción con acentos"]],
            encoding="latin-1",
        )
        result = FileParserEngine.parse_csv(content, "test.csv")
        assert result.success is True
        assert "Válvula" in result.rows[0]["name"]

    def test_utf8_with_bom(self):
        raw = b"\xef\xbb\xbfname,value\ntest,123\n"
        result = FileParserEngine.parse_csv(raw, "test.csv")
        assert result.success is True
        assert result.total_rows == 1
        assert "name" in result.headers

    def test_empty_csv(self):
        result = FileParserEngine.parse_csv(b"", "empty.csv")
        assert result.success is True
        assert result.total_rows == 0

    def test_csv_max_rows(self):
        rows = [[str(i)] for i in range(100)]
        content = _make_csv_bytes(["val"], rows)
        result = FileParserEngine.parse_csv(content, "big.csv", max_rows=5)
        assert result.total_rows == 5

    def test_skip_empty_csv_rows(self):
        text = "a,b\n1,2\n,,\n3,4\n"
        result = FileParserEngine.parse_csv(text.encode(), "test.csv")
        assert result.total_rows == 2  # empty row skipped


# ── Dispatch Tests ────────────────────────────────────────────────


class TestParseFile:

    def test_dispatch_xlsx(self):
        content = _make_xlsx_bytes(["a"], [["1"]])
        result = FileParserEngine.parse_file(content, "data.xlsx")
        assert result.file_type == "xlsx"

    def test_dispatch_csv(self):
        content = _make_csv_bytes(["a"], [["1"]])
        result = FileParserEngine.parse_file(content, "data.csv")
        assert result.file_type == "csv"

    def test_reject_unknown_extension(self):
        result = FileParserEngine.parse_file(b"data", "file.json")
        assert result.success is False
        assert any("unsupported" in e.message.lower() for e in result.errors)

    def test_reject_xlsm(self):
        result = FileParserEngine.parse_file(b"data", "file.xlsm")
        assert result.success is False


# ── Encoding Detection ────────────────────────────────────────────


class TestDetectEncoding:

    def test_utf8_default(self):
        assert FileParserEngine.detect_encoding(b"hello world") == "utf-8"

    def test_utf8_bom(self):
        assert FileParserEngine.detect_encoding(b"\xef\xbb\xbfhello") == "utf-8-sig"

    def test_latin1_fallback(self):
        # Bytes that are invalid UTF-8 but valid Latin-1
        content = b"caf\xe9 cr\xe8me"
        assert FileParserEngine.detect_encoding(content) == "latin-1"


# ── Template Integration Tests ────────────────────────────────────


class TestParseTemplates:
    """Test parsing actual template files from templates/ directory."""

    @pytest.mark.skipif(
        not (TEMPLATE_DIR / "01_equipment_hierarchy.xlsx").exists(),
        reason="Template files not generated",
    )
    def test_parse_equipment_hierarchy_template(self):
        path = TEMPLATE_DIR / "01_equipment_hierarchy.xlsx"
        content = path.read_bytes()
        result = FileParserEngine.parse_file(content, path.name)
        assert result.success is True
        assert result.total_rows >= 3  # Templates have 3-5 example rows
        assert "equipment_tag" in result.headers

    @pytest.mark.skipif(
        not (TEMPLATE_DIR / "06_work_order_history.xlsx").exists(),
        reason="Template files not generated",
    )
    def test_parse_work_order_template(self):
        path = TEMPLATE_DIR / "06_work_order_history.xlsx"
        content = path.read_bytes()
        result = FileParserEngine.parse_file(content, path.name)
        assert result.success is True
        assert result.total_rows >= 3
        assert "wo_id" in result.headers

    @pytest.mark.skipif(
        not (TEMPLATE_DIR / "07_spare_parts_inventory.xlsx").exists(),
        reason="Template files not generated",
    )
    def test_parse_spare_parts_template(self):
        path = TEMPLATE_DIR / "07_spare_parts_inventory.xlsx"
        content = path.read_bytes()
        result = FileParserEngine.parse_file(content, path.name)
        assert result.success is True
        assert "material_code" in result.headers
