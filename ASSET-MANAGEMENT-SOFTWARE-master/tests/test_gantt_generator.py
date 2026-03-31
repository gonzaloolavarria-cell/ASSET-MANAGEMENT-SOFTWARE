"""Tests for Gantt Generator — Phase 4B."""

import os
import tempfile
from datetime import date

from tools.processors.gantt_generator import GanttGenerator
from tools.engines.scheduling_engine import SchedulingEngine
from tools.models.schemas import (
    BacklogWorkPackage, ShiftType, MaterialsReadyStatus,
    WeeklyProgram, WeeklyProgramStatus, GanttRow,
)


def _make_program(n=3):
    """Create a test program with n packages."""
    pkgs = []
    for i in range(n):
        pkgs.append(BacklogWorkPackage(
            package_id=f"WP-G-{i+1:03d}",
            name=f"Gantt Test {i+1}",
            grouped_items=[f"BRY-SAG-ML-001-ITEM-{i+1}"],
            reason_for_grouping="Gantt test",
            scheduled_date=date(2025, 6, 5 + i),
            scheduled_shift=ShiftType.MORNING if i % 2 == 0 else ShiftType.AFTERNOON,
            total_duration_hours=4.0 + i * 2,
            assigned_team=["MECHANICAL"] if i % 2 == 0 else ["ELECTRICAL"],
            materials_status=MaterialsReadyStatus.READY,
        ))
    return SchedulingEngine.create_weekly_program("P1", 23, 2025, pkgs)


class TestGenerateGanttData:

    def test_correct_row_count(self):
        program = _make_program(3)
        rows = GanttGenerator.generate_gantt_data(program)
        assert len(rows) == 3

    def test_row_has_dates(self):
        program = _make_program(1)
        rows = GanttGenerator.generate_gantt_data(program)
        assert rows[0].start_date is not None
        assert rows[0].end_date is not None
        assert rows[0].end_date >= rows[0].start_date

    def test_row_has_specialty(self):
        program = _make_program(1)
        rows = GanttGenerator.generate_gantt_data(program)
        assert rows[0].specialty in ("MECHANICAL", "ELECTRICAL", "INSTRUMENTATION", "GENERAL")

    def test_row_duration_matches(self):
        program = _make_program(1)
        rows = GanttGenerator.generate_gantt_data(program)
        assert rows[0].duration_hours == 4.0

    def test_empty_program(self):
        program = _make_program(0)
        rows = GanttGenerator.generate_gantt_data(program)
        assert len(rows) == 0


class TestExportGanttExcel:

    def test_creates_file(self):
        program = _make_program(2)
        rows = GanttGenerator.generate_gantt_data(program)
        filepath = os.path.join(tempfile.gettempdir(), "test_gantt.xlsx")
        result = GanttGenerator.export_gantt_excel(rows, filepath)
        assert result == filepath
        assert os.path.exists(filepath)
        os.remove(filepath)

    def test_has_two_sheets(self):
        from openpyxl import load_workbook
        program = _make_program(2)
        rows = GanttGenerator.generate_gantt_data(program)
        filepath = os.path.join(tempfile.gettempdir(), "test_gantt_sheets.xlsx")
        GanttGenerator.export_gantt_excel(rows, filepath)
        wb = load_workbook(filepath)
        assert "Schedule" in wb.sheetnames
        assert "Gantt" in wb.sheetnames
        wb.close()
        os.remove(filepath)

    def test_schedule_sheet_has_headers(self):
        from openpyxl import load_workbook
        program = _make_program(1)
        rows = GanttGenerator.generate_gantt_data(program)
        filepath = os.path.join(tempfile.gettempdir(), "test_gantt_headers.xlsx")
        GanttGenerator.export_gantt_excel(rows, filepath)
        wb = load_workbook(filepath)
        ws = wb["Schedule"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert "Package ID" in headers
        assert "Hours" in headers
        wb.close()
        os.remove(filepath)

    def test_empty_rows_no_error(self):
        filepath = os.path.join(tempfile.gettempdir(), "test_gantt_empty.xlsx")
        result = GanttGenerator.export_gantt_excel([], filepath)
        assert os.path.exists(result)
        os.remove(filepath)
