"""Tests for Excel template generation (D1).

Updated to match current template structure after migration:
- Sheet names: some changed (e.g., "Failure Modes" → "failure_modes")
- Instructions sheets: removed from most templates
- Multi-resource sheets: consolidated (T04 now has only "Tasks" sheet)
- Data validations: removed from most templates
- New sheets: T01 has "Equipment BOM", T05 has "WP Task Details",
  T08 has "Shutdown Work Packages"
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook

TEMPLATES_DIR = Path(__file__).parent.parent / "templates"


# ── Helper ────────────────────────────────────────────────────────
def _load(filename: str):
    path = TEMPLATES_DIR / filename
    assert path.exists(), f"Template not found: {path}"
    return load_workbook(path)


def _get_col_index(ws, col_name: str) -> int:
    """Find column index by header name."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == col_name:
            return c
    raise ValueError(f"Column '{col_name}' not found in {ws.title}")


# ══════════════════════════════════════════════════════════════════
# File Existence
# ══════════════════════════════════════════════════════════════════

EXPECTED_FILES = [
    "01_equipment_hierarchy.xlsx",
    "02_criticality_assessment.xlsx",
    "03_failure_modes.xlsx",
    "04_maintenance_tasks.xlsx",
    "05_work_packages.xlsx",
    "06_work_order_history.xlsx",
    "07_spare_parts_inventory.xlsx",
    "08_shutdown_calendar.xlsx",
    "09_workforce.xlsx",
    "10_field_capture.xlsx",
    "11_rca_events.xlsx",
    "12_planning_kpi_input.xlsx",
    "13_de_kpi_input.xlsx",
    "14_maintenance_strategy.xlsx",
]


class TestTemplateExistence:
    """All 14 templates should exist."""

    @pytest.mark.parametrize("filename", EXPECTED_FILES)
    def test_template_exists(self, filename):
        assert (TEMPLATES_DIR / filename).exists()

    def test_all_14_templates(self):
        # Filter out Excel temporary lock files (~$*.xlsx)
        xlsx_files = sorted(
            f.name for f in TEMPLATES_DIR.glob("*.xlsx")
            if not f.name.startswith("~$")
        )
        assert len(xlsx_files) == 14


# ══════════════════════════════════════════════════════════════════
# Sheet Structure
# ══════════════════════════════════════════════════════════════════

class TestSheetStructure:
    """Each template should have correct sheets."""

    def test_01_has_hierarchy_and_bom(self):
        wb = _load("01_equipment_hierarchy.xlsx")
        assert "Equipment Hierarchy" in wb.sheetnames
        assert "Equipment BOM" in wb.sheetnames

    def test_02_has_criticality_assessment(self):
        wb = _load("02_criticality_assessment.xlsx")
        assert "Criticality Assessment" in wb.sheetnames

    def test_03_has_failure_modes(self):
        wb = _load("03_failure_modes.xlsx")
        assert "failure_modes" in wb.sheetnames

    def test_04_has_tasks(self):
        wb = _load("04_maintenance_tasks.xlsx")
        assert "Tasks" in wb.sheetnames

    def test_05_has_work_packages(self):
        wb = _load("05_work_packages.xlsx")
        assert "Work Packages" in wb.sheetnames
        assert "WP Task Details" in wb.sheetnames

    def test_14_has_strategies(self):
        wb = _load("14_maintenance_strategy.xlsx")
        assert "Strategies" in wb.sheetnames


# ══════════════════════════════════════════════════════════════════
# Headers
# ══════════════════════════════════════════════════════════════════

class TestHeaders:
    """Verify key columns exist in each template."""

    def test_01_hierarchy_headers(self):
        wb = _load("01_equipment_hierarchy.xlsx")
        ws = wb["Equipment Hierarchy"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "plant_id" in headers
        assert "equipment_tag" in headers
        assert "criticality" in headers

    def test_02_criticality_headers(self):
        wb = _load("02_criticality_assessment.xlsx")
        ws = wb["Criticality Assessment"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "equipment_tag" in headers
        assert "safety" in headers
        assert "probability" in headers
        # tag + method + 11 categories + probability = 14
        assert len(headers) == 14

    def test_03_failure_mode_headers(self):
        wb = _load("03_failure_modes.xlsx")
        ws = wb["failure_modes"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "mechanism" in headers
        assert "cause" in headers
        assert "failure_consequence" in headers

    def test_04_task_headers(self):
        wb = _load("04_maintenance_tasks.xlsx")
        ws = wb["Tasks"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "task_id" in headers
        assert "task_name" in headers
        assert "task_type" in headers

    def test_07_spare_parts_headers(self):
        wb = _load("07_spare_parts_inventory.xlsx")
        ws = wb["Spare Parts Inventory"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "ved_class" in headers
        assert "fsn_class" in headers
        assert "abc_class" in headers

    def test_14_strategy_headers(self):
        """Template 14 strategy must have FM + tactics + primary/secondary task fields."""
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Identity
        assert "strategy_id" in headers
        assert "equipment_tag" in headers
        assert "maintainable_item" in headers
        # Failure mode
        assert "what" in headers
        assert "mechanism" in headers
        assert "cause" in headers
        # Strategy decision
        assert "status" in headers
        assert "tactics_type" in headers
        # Primary task
        assert "primary_task_id" in headers
        assert "primary_task_interval" in headers
        assert "primary_task_acceptable_limits" in headers
        assert "primary_task_conditional_comments" in headers
        # Secondary task
        assert "secondary_task_id" in headers
        assert "secondary_task_constraint" in headers
        # Budget
        assert "budgeted_as" in headers


# ══════════════════════════════════════════════════════════════════
# Example Data
# ══════════════════════════════════════════════════════════════════

class TestExampleData:
    """Each template should have example data rows."""

    @pytest.mark.parametrize("filename,sheet_name,min_rows", [
        ("01_equipment_hierarchy.xlsx", "Equipment Hierarchy", 5),
        ("02_criticality_assessment.xlsx", "Criticality Assessment", 5),
        ("03_failure_modes.xlsx", "failure_modes", 5),
        ("04_maintenance_tasks.xlsx", "Tasks", 5),
        ("05_work_packages.xlsx", "Work Packages", 4),
        ("06_work_order_history.xlsx", "Work Order History", 5),
        ("07_spare_parts_inventory.xlsx", "Spare Parts Inventory", 5),
        ("08_shutdown_calendar.xlsx", "Shutdown Calendar", 5),
        ("09_workforce.xlsx", "Workforce", 5),
        ("10_field_capture.xlsx", "Field Captures", 5),
        ("11_rca_events.xlsx", "RCA Events", 5),
        ("12_planning_kpi_input.xlsx", "Planning KPI Input", 3),
        ("13_de_kpi_input.xlsx", "DE KPI Input", 3),
        ("14_maintenance_strategy.xlsx", "Strategies", 5),
    ])
    def test_has_example_rows(self, filename, sheet_name, min_rows):
        wb = _load(filename)
        ws = wb[sheet_name]
        data_rows = ws.max_row - 1  # exclude header
        assert data_rows >= min_rows, f"{filename}/{sheet_name}: expected >= {min_rows} rows, got {data_rows}"


# ══════════════════════════════════════════════════════════════════
# Styling
# ══════════════════════════════════════════════════════════════════

class TestStyling:
    """Headers should have OCP green styling."""

    def test_header_fill_color(self):
        wb = _load("01_equipment_hierarchy.xlsx")
        ws = wb["Equipment Hierarchy"]
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "001B5E20" or cell.fill.start_color.rgb == "FF1B5E20"

    def test_header_font_white_bold(self):
        wb = _load("01_equipment_hierarchy.xlsx")
        ws = wb["Equipment Hierarchy"]
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb in ("00FFFFFF", "FFFFFFFF")

    def test_frozen_panes(self):
        wb = _load("01_equipment_hierarchy.xlsx")
        ws = wb["Equipment Hierarchy"]
        assert ws.freeze_panes == "A2"


# ══════════════════════════════════════════════════════════════════
# Strategy Integrity (Template 14)
# ══════════════════════════════════════════════════════════════════

class TestStrategyIntegrity:
    """Template 14 strategy should follow correct task assignment rules."""

    def test_no_duplicate_strategy_ids(self):
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        sid_col = _get_col_index(ws, "strategy_id")
        ids = []
        for row in range(2, ws.max_row + 1):
            sid = ws.cell(row=row, column=sid_col).value
            if sid:
                ids.append(sid)
        assert len(ids) == len(set(ids)), "Duplicate strategy_ids"

    def test_cb_has_both_tasks(self):
        """Condition-Based strategies should have both primary and secondary tasks."""
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        type_col = _get_col_index(ws, "tactics_type")
        pt_col = _get_col_index(ws, "primary_task_id")
        st_col = _get_col_index(ws, "secondary_task_id")
        status_col = _get_col_index(ws, "status")

        for row in range(2, ws.max_row + 1):
            ttype = ws.cell(row=row, column=type_col).value
            status = ws.cell(row=row, column=status_col).value
            if ttype == "CONDITION_BASED" and status == "RECOMMENDED":
                pt = ws.cell(row=row, column=pt_col).value
                st = ws.cell(row=row, column=st_col).value
                assert pt is not None, f"Row {row}: CB strategy missing primary_task_id"
                assert st is not None, f"Row {row}: CB strategy missing secondary_task_id"

    def test_rtf_has_no_primary_task(self):
        """Run-to-Failure strategies should NOT have a primary task."""
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        type_col = _get_col_index(ws, "tactics_type")
        pt_col = _get_col_index(ws, "primary_task_id")
        st_col = _get_col_index(ws, "secondary_task_id")

        for row in range(2, ws.max_row + 1):
            ttype = ws.cell(row=row, column=type_col).value
            if ttype == "RUN_TO_FAILURE":
                pt = ws.cell(row=row, column=pt_col).value
                st = ws.cell(row=row, column=st_col).value
                assert pt is None, f"Row {row}: RTF strategy should not have primary_task_id"
                assert st is not None, f"Row {row}: RTF strategy must have secondary_task_id"

    def test_ft_has_primary_task(self):
        """Fixed-Time strategies should have primary task."""
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        type_col = _get_col_index(ws, "tactics_type")
        pt_col = _get_col_index(ws, "primary_task_id")

        for row in range(2, ws.max_row + 1):
            ttype = ws.cell(row=row, column=type_col).value
            if ttype == "FIXED_TIME":
                pt = ws.cell(row=row, column=pt_col).value
                assert pt is not None, f"Row {row}: FT strategy missing primary_task_id"

    def test_has_recommended_example(self):
        """Should include at least one RECOMMENDED strategy example."""
        wb = _load("14_maintenance_strategy.xlsx")
        ws = wb["Strategies"]
        status_col = _get_col_index(ws, "status")
        statuses = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=status_col).value
            if val:
                statuses.add(val)
        assert "RECOMMENDED" in statuses
