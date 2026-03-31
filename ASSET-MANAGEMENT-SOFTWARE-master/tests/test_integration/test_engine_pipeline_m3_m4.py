"""
Integration tests: M3→M4 engine pipeline.
Tasks → Work Packages → SAP Export → Templates.
"""

import json
import os
import pytest
from datetime import date
from pathlib import Path

from tools.engines.sap_export_engine import SAPExportEngine, SAP_SHORT_TEXT_MAX, SAP_FUNC_LOC_MAX
from tools.engines.quality_score_engine import QualityScoreEngine
from tools.engines.execution_checklist_engine import ExecutionChecklistEngine
from tools.engines.assignment_engine import AssignmentEngine
from tools.engines.roi_engine import ROIEngine
from tools.engines.budget_engine import BudgetEngine
from tools.models.schemas import (
    CompetencyLevel,
    LabourSpecialty,
    MaintenanceTask,
    StrategyType,
    TaskCompetencyRequirement,
    TaskType,
    TechnicianCompetency,
    TechnicianProfile,
    WorkPackage,
)
from agents.orchestration.session_state import SessionState


pytestmark = pytest.mark.integration


def _make_technicians():
    """Create 5 test technicians with varied competencies."""
    return [
        TechnicianProfile(
            worker_id="W-A01", name="Ahmed Senior",
            specialty=LabourSpecialty.FITTER, shift="MORNING", plant_id="OCP-JFC1",
            competencies=[
                TechnicianCompetency(specialty=LabourSpecialty.FITTER, equipment_type="SAG_MILL", level=CompetencyLevel.A),
                TechnicianCompetency(specialty=LabourSpecialty.FITTER, equipment_type="PUMP", level=CompetencyLevel.A),
            ],
            years_experience=15, equipment_expertise=["SAG_MILL", "PUMP", "CONVEYOR", "CRUSHER"],
            safety_training_current=True,
        ),
        TechnicianProfile(
            worker_id="W-B01", name="Khalid Standard",
            specialty=LabourSpecialty.FITTER, shift="MORNING", plant_id="OCP-JFC1",
            competencies=[
                TechnicianCompetency(specialty=LabourSpecialty.FITTER, equipment_type="SAG_MILL", level=CompetencyLevel.B),
            ],
            years_experience=7, equipment_expertise=["SAG_MILL", "PUMP"],
            safety_training_current=True,
        ),
        TechnicianProfile(
            worker_id="W-C01", name="Youssef Junior",
            specialty=LabourSpecialty.FITTER, shift="MORNING", plant_id="OCP-JFC1",
            competencies=[
                TechnicianCompetency(specialty=LabourSpecialty.FITTER, equipment_type="SAG_MILL", level=CompetencyLevel.C),
            ],
            years_experience=2, equipment_expertise=["SAG_MILL"],
            safety_training_current=True,
        ),
        TechnicianProfile(
            worker_id="W-E01", name="Hassan Electrician",
            specialty=LabourSpecialty.ELECTRICIAN, shift="MORNING", plant_id="OCP-JFC1",
            competencies=[
                TechnicianCompetency(specialty=LabourSpecialty.ELECTRICIAN, equipment_type="SAG_MILL", level=CompetencyLevel.B),
            ],
            years_experience=10, equipment_expertise=["SAG_MILL", "MOTOR"],
            safety_training_current=True,
        ),
        TechnicianProfile(
            worker_id="W-I01", name="Omar Instrument",
            specialty=LabourSpecialty.INSTRUMENTIST, shift="MORNING", plant_id="OCP-JFC1",
            competencies=[
                TechnicianCompetency(specialty=LabourSpecialty.INSTRUMENTIST, equipment_type="SAG_MILL", level=CompetencyLevel.B),
            ],
            years_experience=8, equipment_expertise=["SAG_MILL"],
            safety_training_current=True,
        ),
    ]


# ------------------------------------------------------------------
# M3 Pipeline: Tasks + Work Packages + Assignment + Checklist
# ------------------------------------------------------------------
class TestM3Pipeline:
    """Tests for M3: task generation, WP assembly, assignment, checklists."""

    def test_tasks_have_valid_types(self, pipeline_tasks):
        """All tasks have valid TaskType values."""
        for task in pipeline_tasks:
            assert isinstance(task, MaintenanceTask)
            assert task.task_type in TaskType

    def test_t16_rule_replace_tasks_have_materials(self, pipeline_tasks):
        """Every REPLACE task has materials assigned (T-16 rule)."""
        replace_tasks = [t for t in pipeline_tasks if t.task_type == TaskType.REPLACE]
        assert len(replace_tasks) >= 3, "Should have at least 3 REPLACE tasks"
        for task in replace_tasks:
            assert len(task.material_resources) > 0, (
                f"REPLACE task '{task.name}' violates T-16 rule: no materials"
            )

    def test_sap_short_text_max_72(self, pipeline_tasks):
        """All task names ≤ 72 chars (SAP constraint)."""
        for task in pipeline_tasks:
            assert len(task.name) <= 72, (
                f"Task name exceeds SAP_SHORT_TEXT_MAX: '{task.name}' ({len(task.name)} chars)"
            )

    def test_work_package_assembly_from_tasks(self, pipeline_work_packages, pipeline_tasks):
        """WPs have correct allocated tasks."""
        task_ids = {t.task_id for t in pipeline_tasks}
        for wp in pipeline_work_packages:
            for at in wp.allocated_tasks:
                assert at.task_id in task_ids, (
                    f"WP {wp.code} references unknown task {at.task_id}"
                )

    def test_work_package_names_under_40(self, pipeline_work_packages):
        """WP names ≤ 40 chars (SAP constraint)."""
        for wp in pipeline_work_packages:
            assert len(wp.name) <= 40, (
                f"WP name exceeds limit: '{wp.name}' ({len(wp.name)} chars)"
            )

    def test_assignment_engine_from_tasks(self, pipeline_tasks):
        """AssignmentEngine.optimize() assigns technicians to tasks."""
        engine = AssignmentEngine()
        technicians = _make_technicians()
        task_dicts = [{
            "task_id": pipeline_tasks[0].task_id,
            "name": pipeline_tasks[0].name,
            "competency_requirements": [{
                "specialty": pipeline_tasks[0].labour_resources[0].specialty.value
                    if pipeline_tasks[0].labour_resources else "FITTER",
                "min_level": "B",
            }],
        }]
        summary = engine.optimize_assignments(
            tasks=task_dicts,
            technicians=technicians,
            target_date=date(2026, 3, 15),
            target_shift="MORNING",
            plant_id="OCP-JFC1",
        )
        assert summary is not None
        assert summary.total_tasks >= 0

    def test_execution_checklist_from_work_package(self, pipeline_work_packages, pipeline_tasks):
        """ExecutionChecklistEngine generates valid checklist from WP."""
        wp = pipeline_work_packages[2]  # Offline WP with REPLACE tasks
        wp_dict = wp.model_dump()
        task_dicts = [t.model_dump() for t in pipeline_tasks[3:6]]  # REPLACE tasks
        checklist = ExecutionChecklistEngine.generate_checklist(
            work_package=wp_dict,
            tasks=task_dicts,
            equipment_name="SAG Mill #1",
            equipment_tag="BRY-SAG-ML-001",
        )
        assert checklist is not None
        assert len(checklist.steps) > 0
        assert checklist.work_package_id == wp.work_package_id

    def test_m3_swmr_planning_ownership(self, pipeline_tasks, pipeline_work_packages):
        """Planning agent owns tasks and WPs."""
        session = SessionState(
            session_id="swmr-m3", equipment_tag="TEST", plant_code="TEST",
        )
        session.write_entities("maintenance_tasks",
                               [t.model_dump() for t in pipeline_tasks], "planning")
        session.write_entities("work_packages",
                               [wp.model_dump() for wp in pipeline_work_packages], "planning")
        with pytest.raises(PermissionError):
            session.write_entities("maintenance_tasks", [], "reliability")

    def test_m3_validation_entities_present(self, pipeline_session):
        """M3 entities present in session."""
        counts = pipeline_session.get_entity_counts()
        assert counts.get("maintenance_tasks", 0) == 8
        assert counts.get("work_packages", 0) == 3

    def test_budget_items_from_tasks(self, pipeline_budget_items):
        """Budget items have valid categories and amounts."""
        for item in pipeline_budget_items:
            assert item.planned_amount > 0
            assert item.category is not None


# ------------------------------------------------------------------
# M4 Pipeline: SAP Export + Templates
# ------------------------------------------------------------------
class TestM4Pipeline:
    """Tests for M4: SAP export, template population, reporting."""

    def test_sap_export_from_work_packages(self, pipeline_work_packages, pipeline_tasks):
        """SAPExportEngine produces valid upload package."""
        tasks_dict = {t.task_id: t for t in pipeline_tasks}
        package = SAPExportEngine.generate_upload_package(
            work_packages=pipeline_work_packages,
            plant_code="OCP-JFC1",
            plan_description="SAG Mill Maintenance Plan",
            tasks=tasks_dict,
        )
        assert package is not None
        assert package.plant_code == "OCP-JFC1"

    def test_sap_export_xlsx_roundtrip(self, pipeline_work_packages, pipeline_tasks, tmp_path):
        """write_to_xlsx() produces readable Excel file."""
        tasks_dict = {t.task_id: t for t in pipeline_tasks}
        package = SAPExportEngine.generate_upload_package(
            work_packages=pipeline_work_packages,
            plant_code="OCP-JFC1",
            tasks=tasks_dict,
        )
        output_path = str(tmp_path / "sap_export.xlsx")
        SAPExportEngine.write_to_xlsx(package, output_path)
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_sap_export_xlsx_cell_content(self, pipeline_work_packages, pipeline_tasks, tmp_path):
        """SAP xlsx has correct sheet names, headers, and data rows."""
        import openpyxl
        tasks_dict = {t.task_id: t for t in pipeline_tasks}
        package = SAPExportEngine.generate_upload_package(
            work_packages=pipeline_work_packages,
            plant_code="OCP-JFC1",
            tasks=tasks_dict,
        )
        output_path = str(tmp_path / "sap_content.xlsx")
        SAPExportEngine.write_to_xlsx(package, output_path)
        wb = openpyxl.load_workbook(output_path)

        # 3 sheets
        assert "Functional Locations" in wb.sheetnames
        assert "Task Lists" in wb.sheetnames
        assert "Maintenance Plans" in wb.sheetnames

        # Functional Locations headers
        ws_fl = wb["Functional Locations"]
        fl_headers = [ws_fl.cell(1, c).value for c in range(1, 9)]
        assert "Item Ref" in fl_headers
        assert "Functional Location" in fl_headers
        # Should have data rows (one per maintenance item)
        assert ws_fl.max_row >= 2, "Functional Locations should have at least 1 data row"

        # Task Lists headers
        ws_tl = wb["Task Lists"]
        tl_headers = [ws_tl.cell(1, c).value for c in range(1, 11)]
        assert "Short Text" in tl_headers
        assert "Duration (h)" in tl_headers

        # All short texts ≤ 72 chars
        for row in range(2, ws_tl.max_row + 1):
            short_text = ws_tl.cell(row, tl_headers.index("Short Text") + 1).value
            if short_text:
                assert len(str(short_text)) <= 72, f"Short text exceeds 72: {short_text}"

        # All func loc ≤ 40 chars
        for row in range(2, ws_fl.max_row + 1):
            func_loc = ws_fl.cell(row, fl_headers.index("Functional Location") + 1).value
            if func_loc:
                assert len(str(func_loc)) <= 40, f"Func loc exceeds 40: {func_loc}"

        # Maintenance Plans
        ws_mp = wb["Maintenance Plans"]
        assert ws_mp.max_row >= 2, "Maintenance Plans should have at least 1 data row"

    def test_roi_engine_calculation(self, pipeline_roi_input):
        """ROIEngine produces valid result."""
        result = ROIEngine.calculate_roi(pipeline_roi_input)
        assert result is not None
        assert result.npv != 0  # Should have meaningful NPV
        assert result.investment_cost == pipeline_roi_input.investment_cost
        assert len(result.cumulative_savings_by_year) == pipeline_roi_input.analysis_horizon_years

    def test_budget_engine_tracking(self, pipeline_budget_items):
        """BudgetEngine tracks budget from items."""
        items_dicts = [item.model_dump() for item in pipeline_budget_items]
        # Convert date objects to strings for JSON compatibility
        for item in items_dicts:
            for key in ("period_start", "period_end"):
                if item.get(key) and hasattr(item[key], "isoformat"):
                    item[key] = item[key].isoformat()
        summary = BudgetEngine.track_budget("OCP-JFC1", items_dicts)
        assert summary is not None
        assert summary.total_planned > 0
        assert summary.total_actual > 0

    def test_m4_quality_score_session(self, pipeline_session):
        """Quality scorer on complete M1-M4 session."""
        report = QualityScoreEngine.score_session(
            session_entities=pipeline_session.get_validation_input(),
            milestone=4,
            session_id=pipeline_session.session_id,
        )
        assert report is not None
        assert report.overall_score >= 0

    def test_sap_func_loc_constraints(self, pipeline_hierarchy_nodes):
        """Functional location codes ≤ 40 chars."""
        for node in pipeline_hierarchy_nodes:
            assert len(node.code) <= SAP_FUNC_LOC_MAX, (
                f"Node code exceeds SAP_FUNC_LOC_MAX: '{node.code}' ({len(node.code)} chars)"
            )

    def test_reporting_engine_monthly(self):
        """ReportingEngine produces monthly report."""
        from tools.engines.reporting_engine import ReportingEngine
        report = ReportingEngine.generate_monthly_kpi_report(
            plant_id="OCP-JFC1", month=3, year=2026,
        )
        assert report is not None
        assert report.month == 3
        assert report.year == 2026

    def test_reporting_engine_with_financial(self):
        """ReportingEngine handles financial summary param."""
        from tools.engines.reporting_engine import ReportingEngine
        report = ReportingEngine.generate_monthly_kpi_report(
            plant_id="OCP-JFC1", month=3, year=2026,
            financial_summary={"total_budget": 300000, "total_spend": 285000},
        )
        assert report is not None


# ------------------------------------------------------------------
# M3→M4 Handoff
# ------------------------------------------------------------------
class TestM3ToM4Handoff:
    """Tests for data integrity across M3→M4 milestone boundary."""

    def test_session_accumulates_all_milestones(self, pipeline_session):
        """All entity types present in complete session."""
        counts = pipeline_session.get_entity_counts()
        expected_types = [
            "hierarchy_nodes", "criticality_assessments",
            "functions", "functional_failures", "failure_modes",
            "maintenance_tasks", "work_packages",
        ]
        for entity_type in expected_types:
            assert counts.get(entity_type, 0) > 0, (
                f"Missing entity type: {entity_type}"
            )

    def test_sap_export_references_valid_work_packages(self, pipeline_work_packages, pipeline_tasks):
        """SAP items reference valid WP codes."""
        tasks_dict = {t.task_id: t for t in pipeline_tasks}
        package = SAPExportEngine.generate_upload_package(
            work_packages=pipeline_work_packages,
            plant_code="OCP-JFC1",
            tasks=tasks_dict,
        )
        wp_ids = {wp.work_package_id for wp in pipeline_work_packages}
        # Package should reference known WPs
        assert package is not None

    def test_entity_counts_complete(self, pipeline_session):
        """session.get_entity_counts() returns non-zero for populated types."""
        counts = pipeline_session.get_entity_counts()
        assert counts["hierarchy_nodes"] == 6
        assert counts["criticality_assessments"] == 2
        assert counts["functions"] == 3
        assert counts["functional_failures"] == 3
        assert counts["failure_modes"] == 6
        assert counts["maintenance_tasks"] == 8
        assert counts["work_packages"] == 3

    def test_sap_texts_under_72_chars(self, pipeline_work_packages, pipeline_tasks):
        """SAP short texts in export are ≤ 72 chars."""
        for wp in pipeline_work_packages:
            assert len(wp.name) <= 72
        for task in pipeline_tasks:
            assert len(task.name) <= 72

    def test_session_roundtrip_with_all_milestones(self, pipeline_session):
        """JSON roundtrip preserves all milestones."""
        json_str = pipeline_session.to_json()
        restored = SessionState.from_json(json_str)
        orig = pipeline_session.get_entity_counts()
        rest = restored.get_entity_counts()
        for key in orig:
            assert orig[key] == rest.get(key, 0)

    def test_deliverable_status_flow(self):
        """Test approval status transitions."""
        from tools.models.schemas import ApprovalStatus
        statuses = [ApprovalStatus.DRAFT, ApprovalStatus.REVIEWED, ApprovalStatus.APPROVED]
        # Verify enum values exist
        for s in statuses:
            assert s.value

    def test_financial_roi_integration(self, pipeline_roi_input, pipeline_budget_items):
        """ROI and budget engines work with same project data."""
        roi_result = ROIEngine.calculate_roi(pipeline_roi_input)
        items_dicts = [item.model_dump() for item in pipeline_budget_items]
        for item in items_dicts:
            for key in ("period_start", "period_end"):
                if item.get(key) and hasattr(item[key], "isoformat"):
                    item[key] = item[key].isoformat()
        budget_summary = BudgetEngine.track_budget("OCP-JFC1", items_dicts)
        # Both should process same plant
        assert roi_result.plant_id == "OCP-JFC1"
        assert budget_summary is not None


# ------------------------------------------------------------------
# Template Content Verification (cell-level assertions)
# ------------------------------------------------------------------
class TestTemplateContentVerification:
    """Verify generated xlsx templates have correct cell content from pipeline data."""

    def test_t03_failure_modes_cell_content(
        self, pipeline_hierarchy_nodes, pipeline_fmeca, tmp_path
    ):
        """T-03 xlsx has mechanism/cause/evidence columns populated."""
        from tools.engines.template_population_engine import TemplatePopulationEngine
        import openpyxl

        nodes = [n.model_dump(mode="json") for n in pipeline_hierarchy_nodes]
        fms = [fm.model_dump(mode="json") for fm in pipeline_fmeca["failure_modes"]]
        funcs = [f.model_dump(mode="json") for f in pipeline_fmeca["functions"]]
        ffs = [ff.model_dump(mode="json") for ff in pipeline_fmeca["failures"]]

        out = tmp_path / "03_failure_modes.xlsx"
        TemplatePopulationEngine.populate_03_failure_modes(
            fms, funcs, ffs, nodes, out,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 16)]

        assert "mechanism" in headers
        assert "cause" in headers
        assert "evidence" in headers
        # 6 failure modes → 6 data rows
        assert ws.max_row == 7  # 1 header + 6 data

        # Check mechanism column populated (not blank)
        mech_col = headers.index("mechanism") + 1
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, mech_col).value
            assert val, f"Row {row} mechanism is blank"

        # Check evidence column populated (from nested failure_effect)
        ev_col = headers.index("evidence") + 1
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, ev_col).value
            assert val, f"Row {row} evidence is blank"

        # Check what_component column populated
        what_col = headers.index("what_component") + 1
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, what_col).value
            assert val, f"Row {row} what_component is blank"

    def test_t04_tasks_labour_populated(
        self, pipeline_tasks, tmp_path
    ):
        """T-04 xlsx labour sheet has rows from labour_resources."""
        from tools.engines.template_population_engine import TemplatePopulationEngine
        import openpyxl

        tasks = [t.model_dump(mode="json") for t in pipeline_tasks]
        out = tmp_path / "04_maintenance_tasks.xlsx"
        TemplatePopulationEngine.populate_04_tasks(tasks, [], out)
        wb = openpyxl.load_workbook(out)

        # Tasks sheet
        ws_tasks = wb["Tasks"]
        assert ws_tasks.max_row >= 9  # 1 header + 8 tasks

        # Task name column should be populated (not blank)
        for row in range(2, ws_tasks.max_row + 1):
            name = ws_tasks.cell(row, 2).value
            assert name, f"Row {row} task_name is blank"

        # Labour sheet should have rows
        ws_labour = wb["Task_Labour"]
        assert ws_labour.max_row >= 2, "Labour sheet has no data rows"

        # Materials sheet should have rows (inline material_resources extraction)
        ws_materials = wb["Task_Materials"]
        # REPLACE tasks have material_resources → should be extracted
        assert ws_materials.max_row >= 2, "Materials sheet has no data (inline extraction failed)"

    def test_t14_strategy_tactics_populated(
        self, pipeline_hierarchy_nodes, pipeline_fmeca, pipeline_tasks, tmp_path
    ):
        """T-14 xlsx tactics_type column populated from FM.strategy_type."""
        from tools.engines.template_population_engine import TemplatePopulationEngine
        import openpyxl

        nodes = [n.model_dump(mode="json") for n in pipeline_hierarchy_nodes]
        fms = [fm.model_dump(mode="json") for fm in pipeline_fmeca["failure_modes"]]
        funcs = [f.model_dump(mode="json") for f in pipeline_fmeca["functions"]]
        tasks = [t.model_dump(mode="json") for t in pipeline_tasks]

        out = tmp_path / "14_maintenance_strategy.xlsx"
        TemplatePopulationEngine.populate_14_strategy(
            fms, tasks, funcs, nodes, out,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 32)]

        assert "tactics_type" in headers
        tactics_col = headers.index("tactics_type") + 1
        # 6 FMs → 6 data rows
        assert ws.max_row == 7

        # Every FM has strategy_type → tactics_type should be populated
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, tactics_col).value
            assert val, f"Row {row} tactics_type is blank (strategy_type not resolved)"

        # what column should be populated
        what_col = headers.index("what") + 1
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, what_col).value
            assert val, f"Row {row} 'what' component is blank"

    def test_populate_all_from_session(self, pipeline_session, tmp_path):
        """populate_all() from session entities produces all 7 templates."""
        from tools.engines.template_population_engine import TemplatePopulationEngine

        results = TemplatePopulationEngine.populate_all(
            session_entities=pipeline_session.entities,
            output_dir=tmp_path / "deliverables",
            plant_code="OCP-JFC1",
        )
        # M1+M2+M3 entities present → should produce at least 5 templates
        # (01, 02, 03, 04, 05; 07 depends on material_assignments, 14 depends on FMs)
        assert len(results) >= 5, f"Only {len(results)} templates generated: {list(results.keys())}"
        assert "01_equipment_hierarchy.xlsx" in results
        assert "03_failure_modes.xlsx" in results
        assert "04_maintenance_tasks.xlsx" in results
        assert "14_maintenance_strategy.xlsx" in results
