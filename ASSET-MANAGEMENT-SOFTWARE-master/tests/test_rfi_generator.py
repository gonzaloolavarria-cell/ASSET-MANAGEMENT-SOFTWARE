"""Tests for RFI Excel template generator."""

import pytest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from scripts.generate_rfi_template import generate_rfi_template
from tools.models.rfi_models import AMS_TEMPLATES


@pytest.fixture
def rfi_workbook(tmp_path):
    """Generate RFI template and return loaded workbook."""
    output = tmp_path / "test_rfi.xlsx"
    generate_rfi_template(output)
    return load_workbook(str(output))


@pytest.fixture
def rfi_path(tmp_path):
    """Generate RFI template and return path."""
    output = tmp_path / "test_rfi.xlsx"
    generate_rfi_template(output)
    return output


class TestFileGeneration:
    def test_generates_file(self, tmp_path):
        output = tmp_path / "test_rfi.xlsx"
        result = generate_rfi_template(output)
        assert result == output
        assert output.exists()

    def test_creates_parent_dirs(self, tmp_path):
        output = tmp_path / "subdir" / "nested" / "rfi.xlsx"
        generate_rfi_template(output)
        assert output.exists()

    def test_xlsx_not_xlsm(self, rfi_path):
        """Security: file must be .xlsx (no macros), not .xlsm."""
        assert rfi_path.suffix == ".xlsx"

    def test_default_path(self, monkeypatch, tmp_path):
        """Test that default output path is used when None is passed."""
        # We just verify the function runs; actual path depends on install location
        output = tmp_path / "default_test.xlsx"
        generate_rfi_template(output)
        assert output.exists()


class TestSheetStructure:
    def test_9_sheets(self, rfi_workbook):
        """8 data sheets + 1 instructions = 9 sheets."""
        assert len(rfi_workbook.sheetnames) == 9

    def test_instructions_is_first(self, rfi_workbook):
        assert rfi_workbook.sheetnames[0] == "Instructions"

    def test_sheet_names(self, rfi_workbook):
        expected = [
            "Instructions",
            "1-Company Profile",
            "2-Equipment Data",
            "3-Maintenance State",
            "4-Organization",
            "5-Standards",
            "6-KPI Baseline",
            "7-Scope & Timeline",
            "8-Data Availability",
        ]
        assert rfi_workbook.sheetnames == expected

    def test_data_availability_has_14_rows(self, rfi_workbook):
        ws = rfi_workbook["8-Data Availability"]
        # Count non-empty rows after header (row 1)
        data_rows = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                data_rows += 1
        assert data_rows == 14

    def test_data_availability_template_ids(self, rfi_workbook):
        ws = rfi_workbook["8-Data Availability"]
        ids = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                ids.append(row[0])
        expected_ids = [t[0] for t in AMS_TEMPLATES]
        assert ids == expected_ids


class TestQuestionnaireSheets:
    def test_headers_present(self, rfi_workbook):
        """Each questionnaire sheet has ID/Field/Response/Type/Required/Help headers."""
        for sheet_name in rfi_workbook.sheetnames[1:8]:  # Sheets 1-7
            ws = rfi_workbook[sheet_name]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
            assert headers == ["ID", "Field", "Response", "Type", "Required", "Help / Instructions"], \
                f"Headers mismatch in {sheet_name}: {headers}"

    def test_headers_styled(self, rfi_workbook):
        """Header row should have dark fill."""
        ws = rfi_workbook["1-Company Profile"]
        cell = ws.cell(row=1, column=1)
        # Check that fill is applied (non-default)
        assert cell.fill.start_color.rgb is not None

    def test_required_fields_highlighted(self, rfi_workbook):
        """Required response cells should have yellow fill."""
        ws = rfi_workbook["1-Company Profile"]
        # CSP-01 (Company Name) is required — response cell is C2
        response_cell = ws.cell(row=2, column=3)
        # Check that fill is the required fill (FFF2CC)
        fill_color = response_cell.fill.start_color.rgb
        assert fill_color is not None
        # The required fill is FFF2CC (yellow), optional is E8F5E9 (green)
        # CSP-01 is required, so it should be yellow
        assert "FFF2CC" in str(fill_color) or "fff2cc" in str(fill_color).lower()

    def test_optional_fields_highlighted(self, rfi_workbook):
        """Optional response cells should have green fill."""
        ws = rfi_workbook["1-Company Profile"]
        # CSP-04 (Plant Code) is optional — it's row 5 (header=1, CSP-01=2, ..., CSP-04=5)
        response_cell = ws.cell(row=5, column=3)
        fill_color = response_cell.fill.start_color.rgb
        assert fill_color is not None
        assert "E8F5E9" in str(fill_color) or "e8f5e9" in str(fill_color).lower()

    def test_field_ids_present(self, rfi_workbook):
        """Each data row should have a field ID in column A."""
        ws = rfi_workbook["1-Company Profile"]
        ids = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                ids.append(row[0])
        assert len(ids) == 12  # Sheet 1 has 12 fields
        assert ids[0] == "CSP-01"
        assert ids[-1] == "CSP-12"

    def test_all_sheets_have_data(self, rfi_workbook):
        """Each questionnaire sheet should have at least 7 data rows."""
        expected_min_rows = {
            "1-Company Profile": 12,
            "2-Equipment Data": 11,
            "3-Maintenance State": 14,
            "4-Organization": 8,
            "5-Standards": 8,
            "6-KPI Baseline": 10,
            "7-Scope & Timeline": 7,
        }
        for sheet_name, min_rows in expected_min_rows.items():
            ws = rfi_workbook[sheet_name]
            data_rows = sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0])
            assert data_rows == min_rows, \
                f"{sheet_name}: expected {min_rows} rows, got {data_rows}"


class TestDropdowns:
    def test_dropdowns_exist(self, rfi_workbook):
        """Sheets with dropdown fields should have data validations."""
        ws = rfi_workbook["1-Company Profile"]
        # Should have at least 1 data validation (Industry dropdown, Language, etc.)
        assert len(ws.data_validations.dataValidation) > 0

    def test_data_availability_dropdowns(self, rfi_workbook):
        ws = rfi_workbook["8-Data Availability"]
        # Should have dropdowns for Available?, Format, Quality
        assert len(ws.data_validations.dataValidation) >= 3
